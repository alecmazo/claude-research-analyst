"""
FastAPI backend for the DGA Capital Research Analyst iPhone app.

Exposes the existing Python pipeline (SEC → Grok → Word/PPTX) via REST.

Start (from the project root — either works):
    python api/server.py
    python -m uvicorn api.server:app --host 0.0.0.0 --port 8000 --reload
"""

from __future__ import annotations

import sys
import os
import json
import uuid
import shutil
import tempfile
import threading
import traceback
from datetime import datetime
from pathlib import Path
from typing import Any, Optional

import csv
import io
import re
import time
from itertools import groupby
import hashlib
import hmac

# ── yfinance (optional — for live price fetching) ─────────────────────────────
try:
    import yfinance as yf
    _YFINANCE_OK = True
except ImportError:
    _YFINANCE_OK = False

# ── openpyxl (optional — for XLSX cap-table uploads) ─────────────────────────
try:
    import openpyxl
    _OPENPYXL_OK = True
except ImportError:
    _OPENPYXL_OK = False

# ── reportlab (optional — for PDF export) ────────────────────────────────────
try:
    import reportlab  # noqa: F401
    _REPORTLAB_OK = True
except ImportError:
    _REPORTLAB_OK = False

# ── Live-price cache  (TTL: 15 min) ──────────────────────────────────────────
_price_cache: dict = {}  # { symbol: (price, fetched_at) }
_PRICE_CACHE_TTL = 900   # seconds

# Cache for information_schema column checks (avoid slow system-table hits per request)
_col_exists_cache: dict = {}   # { "table.column": bool }

def _fetch_prices(symbols: list) -> dict:
    """Return {symbol: last_price} for the given list, using a 15-min cache.

    Uses yf.download() to batch-fetch all uncached symbols in a single HTTP
    request instead of one Ticker() call per symbol. Falls back to per-ticker
    fast_info for any symbol the batch download misses (e.g. preferred stocks
    with non-standard tickers).
    """
    if not _YFINANCE_OK or not symbols:
        return {}
    now   = time.time()
    out   = {}
    fetch = []   # (original_sym, clean_sym) pairs that need network calls

    MM_FUNDS = ('SPAXX', 'FDRXX', 'SPRXX', 'FZFXX', 'FZDXX')

    for sym in symbols:
        clean = sym.rstrip('*')
        # Money-market: always $1
        if any(mm in clean for mm in MM_FUNDS):
            _price_cache[clean] = (1.0, now)
            out[sym] = 1.0
            continue
        if clean in _price_cache:
            p, ts = _price_cache[clean]
            if now - ts < _PRICE_CACHE_TTL:
                out[sym] = p
                continue
        fetch.append((sym, clean))

    if not fetch:
        return out

    # ── Batch download (one HTTP request for all uncached symbols) ───────────
    clean_syms = list(dict.fromkeys(c for _, c in fetch))  # deduplicated

    def _normalize(sym: str) -> str:
        """Map preferred-stock variants to Yahoo format (BAC-PL → BAC-PL)."""
        import re as _r
        m = _r.match(r'^([A-Z]{1,4})PR([A-Z])$', sym)
        if m:
            return m.group(1) + '-P' + m.group(2)
        return sym

    yahoo_syms = [_normalize(s) for s in clean_syms]
    batch_prices: dict[str, float | None] = {}

    try:
        import pandas as _pd
        data = yf.download(
            " ".join(yahoo_syms),
            period="2d",          # need 2d so market-closed days still have a price
            auto_adjust=True,
            progress=False,
            threads=True,         # yfinance uses threads internally for multi-ticker
        )
        if data is not None and not data.empty:
            close = data["Close"] if "Close" in data.columns else data.get("close")
            if close is not None and not close.empty:
                last_row = close.dropna(how="all").iloc[-1] if len(close.dropna(how="all")) else close.iloc[-1]
                if hasattr(last_row, "items"):
                    # Multi-ticker: Series indexed by ticker symbol
                    for tk, price in last_row.items():
                        tk_str = str(tk)
                        try:
                            p = float(price)
                            if p > 0:
                                batch_prices[tk_str] = p
                        except (TypeError, ValueError):
                            pass
                else:
                    # Single-ticker: scalar
                    try:
                        p = float(last_row)
                        if p > 0 and yahoo_syms:
                            batch_prices[yahoo_syms[0]] = p
                    except (TypeError, ValueError):
                        pass
    except Exception:
        pass  # fall through to per-ticker fallback below

    # Map batch results back to original clean symbols
    for orig_sym, clean in zip(
        [s for s, _ in fetch], [c for _, c in fetch]
    ):
        yahoo = _normalize(clean)
        price = batch_prices.get(yahoo) or batch_prices.get(clean)

        if price is None:
            # Fallback: single Ticker call for anything the batch missed.
            # Also try alternate ticker formats (e.g. BRKB → BRK-B, BRK.B)
            # when the original symbol fails.
            import re as _re2
            variants = [yahoo]
            _m2 = _re2.match(r'^([A-Z]{2,4})([A-Z])$', clean)
            if _m2:
                _pfx, _sfx = _m2.group(1), _m2.group(2)
                variants += [_pfx + '-' + _sfx, _pfx + '.' + _sfx]

            for _variant in variants:
                try:
                    t = yf.Ticker(_variant)
                    p = t.fast_info.last_price
                    price = float(p) if p and float(p) > 0 else None
                    if price:
                        break
                except Exception:
                    price = None

        _price_cache[clean] = (price, now)
        out[orig_sym] = price

    return out


def _fund_market_nav(cur, fid: str) -> float:
    """Return fund NAV = Σ(open_qty × live_price) across all tax_lots.

    Pricing priority:
      1. Live yfinance price (cached 15 min)
      2. Cost-basis fallback when no live price (prevents NAV from going to zero
         on data outages or unlisted securities)

    Money-market positions (asset_class='cash') are priced at $1.00/unit via
    the existing _fetch_prices() logic (SPAXX/FDRXX/SPRXX → 1.0).
    """
    cur.execute("""
        SELECT s.symbol, s.asset_class,
               SUM(tl.quantity)                                       AS total_qty,
               SUM(tl.quantity * tl.cost_basis_per_unit)              AS total_cost,
               SUM(tl.quantity * tl.cost_basis_per_unit)
                 / NULLIF(SUM(tl.quantity), 0)                        AS avg_cost
          FROM tax_lots tl
          JOIN securities s ON s.id = tl.security_id
         WHERE tl.fund_id = %s AND tl.closed_at IS NULL
         GROUP BY s.symbol, s.asset_class
    """, (fid,))
    rows = cur.fetchall()
    if not rows:
        return 0.0

    symbols = [r["symbol"] for r in rows if r["symbol"]]
    prices  = _fetch_prices(symbols)

    total = 0.0
    for r in rows:
        qty      = float(r["total_qty"] or 0)
        avg_cost = float(r["avg_cost"]  or 0)
        sym      = r["symbol"]
        price    = prices.get(sym)
        total   += qty * (price if price is not None else avg_cost)
    return round(total, 2)


def _bulk_fund_market_nav(cur, fids: list) -> dict:
    """Return {fund_id_str: nav_float} for all fund IDs in one SQL query.

    Prices are assumed to be already warmed in _price_cache via a preceding
    _fetch_prices() call.  Falls back to cost-basis per position when a live
    price is unavailable.  Much faster than calling _fund_market_nav() in a loop.
    """
    if not fids:
        return {}

    cur.execute("""
        SELECT tl.fund_id,
               s.symbol,
               s.asset_class,
               SUM(tl.quantity)                                               AS total_qty,
               SUM(tl.quantity * tl.cost_basis_per_unit)
                 / NULLIF(SUM(tl.quantity), 0)                                AS avg_cost
          FROM tax_lots tl
          JOIN securities s ON s.id = tl.security_id
         WHERE tl.fund_id::text = ANY(%s) AND tl.closed_at IS NULL
         GROUP BY tl.fund_id, s.symbol, s.asset_class
    """, (fids,))
    rows = cur.fetchall()

    # Collect all symbols so we can do ONE price fetch for the whole batch
    all_syms = list({r["symbol"] for r in rows if r["symbol"]})
    prices   = _fetch_prices(all_syms)

    totals: dict = {}
    for r in rows:
        fid      = str(r["fund_id"])
        qty      = float(r["total_qty"] or 0)
        avg_cost = float(r["avg_cost"]  or 0)
        sym      = r["symbol"]
        price    = prices.get(sym)
        totals[fid] = round(totals.get(fid, 0.0) + qty * (price if price is not None else avg_cost), 2)

    # Funds with no positions → 0.0
    for fid in fids:
        totals.setdefault(str(fid), 0.0)
    return totals


# ── Embedded chart-of-accounts rows  (replaces reading the seed .sql file) ───
# Matches apps/fund/db/seed/0001_chart_of_accounts.sql exactly.
_FUND_COA = [
    # code,  name,                                    type
    ('1010', 'Cash — Operating Account',              'asset'),
    ('1020', 'Cash — Brokerage',                      'asset'),
    ('1030', 'Cash — Money Market',                   'asset'),
    ('1100', 'Securities at Cost',                    'asset'),
    ('1110', 'Mark-to-Market Adjustment',             'asset'),
    ('1200', 'Subscriptions Receivable',              'asset'),
    ('1210', 'Dividends Receivable',                  'asset'),
    ('1220', 'Interest Receivable',                   'asset'),
    ('1300', 'Prepaid Expenses',                      'asset'),
    ('2010', 'Trade Settlement Payable',              'liability'),
    ('2100', 'Accrued Management Fee',                'liability'),
    ('2110', 'Accrued Performance Fee (Carry)',       'liability'),
    ('2200', 'Distributions Payable',                 'liability'),
    ('2300', 'Accrued Expenses — Audit',              'liability'),
    ('2310', 'Accrued Expenses — Legal',              'liability'),
    ('2320', 'Accrued Expenses — Fund Admin',         'liability'),
    ('2330', 'Accrued Expenses — Other',              'liability'),
    ('3000', 'Capital — General Partner',             'equity'),
    ('3100', 'Capital — Limited Partners (control)',  'equity'),
    ('3900', 'Retained Earnings',                     'equity'),
    ('4100', 'Realized Gain — Long-Term',             'income'),
    ('4110', 'Realized Gain — Short-Term',            'income'),
    ('4200', 'Unrealized Gain (Mark-to-Market)',      'income'),
    ('4300', 'Dividend Income',                       'income'),
    ('4400', 'Interest Income',                       'income'),
    ('4900', 'Other Income',                          'income'),
    ('5100', 'Management Fee Expense',                'expense'),
    ('5200', 'Audit Fee',                             'expense'),
    ('5210', 'Legal Fees',                            'expense'),
    ('5220', 'Fund Administration Fees',              'expense'),
    ('5230', 'Custody Fees',                          'expense'),
    ('5240', 'Brokerage Commissions',                 'expense'),
    ('5300', 'Realized Loss — Long-Term',             'expense'),
    ('5310', 'Realized Loss — Short-Term',            'expense'),
    ('5400', 'Unrealized Loss (Mark-to-Market)',      'expense'),
    ('5900', 'Other Fund Expenses',                   'expense'),
    ('6100', 'Performance Allocation Expense',        'expense'),
    ('6200', 'Performance Allocation Reversal',       'contra'),
]


def _parse_fidelity_csv(content: str) -> list:
    """Parse a Fidelity Account Positions CSV export.
    Returns a list of position dicts ready for DB import."""
    positions = []
    lines = content.splitlines()
    # Fidelity CSVs have a header row but may have extra blank/disclaimer lines.
    # Find the real header row that contains 'Symbol'.
    header_idx = None
    for i, line in enumerate(lines):
        if 'Symbol' in line and 'Account' in line:
            header_idx = i
            break
    if header_idx is None:
        return []

    reader = csv.DictReader(io.StringIO('\n'.join(lines[header_idx:])))

    def parse_dollar(s: str):
        if not s:
            return None
        s = str(s).replace('$', '').replace(',', '').replace('+', '').strip()
        try:
            return float(s)
        except ValueError:
            return None

    for row in reader:
        sym = (row.get('Symbol') or '').strip()
        if not sym:
            continue
        desc = (row.get('Description') or '').strip()
        # Stop at Fidelity disclaimer lines
        if len(sym) > 12 or (desc and 'Brokerage services' in desc):
            break

        qty_str  = (row.get('Quantity') or '').strip()
        is_mm    = 'SPAXX' in sym or 'FDRXX' in sym or not qty_str

        if is_mm:
            val = parse_dollar(row.get('Current Value'))
            if val:
                positions.append({
                    'symbol':     sym.rstrip('*'),
                    'name':       desc or 'Money Market',
                    'quantity':   val,
                    'avg_cost':   1.0,
                    'cost_basis': val,
                    'last_price': 1.0,
                    'lot_type':   (row.get('Type') or 'Cash').strip(),
                    'is_cash':    True,
                })
            continue

        try:
            qty = float(qty_str.replace(',', ''))
        except ValueError:
            continue

        last_price      = parse_dollar(row.get('Last Price'))
        avg_cost        = parse_dollar(row.get('Average Cost Basis'))
        cost_basis_total= parse_dollar(row.get('Cost Basis Total'))
        if cost_basis_total is None and avg_cost and qty:
            cost_basis_total = avg_cost * qty

        positions.append({
            'symbol':     sym,
            'name':       desc,
            'quantity':   qty,
            'avg_cost':   avg_cost   or 0.0,
            'cost_basis': cost_basis_total or 0.0,
            'last_price': last_price,
            'lot_type':   (row.get('Type') or 'Cash').strip(),
            'is_cash':    False,
        })

    # ── Pending Activity detection ─────────────────────────────────────────
    # Some Fidelity exports contain a row where one column reads
    # "Pending Activity" (sold shares settling).  The cash amount is in
    # the "Current Value" column of that same row.  We use DictReader so
    # the lookup is by column name (robust to column order changes) and
    # scan ALL columns for the "pending activity" text so it doesn't
    # matter whether it's in Symbol, Description, or any other field.
    pending_cash = 0.0
    reader2 = csv.DictReader(io.StringIO('\n'.join(lines[header_idx:])))
    for row2 in reader2:
        if any('pending activity' in (v or '').lower() for v in row2.values()):
            val = parse_dollar(row2.get('Current Value'))
            if val and val > 0:
                pending_cash += val
            # don't break — accumulate all pending rows

    if pending_cash > 0:
        spaxx_entry = next(
            (p for p in positions if 'SPAXX' in (p['symbol'] or '').upper()), None
        )
        if spaxx_entry:
            spaxx_entry['quantity']   = float(spaxx_entry['quantity'])   + pending_cash
            spaxx_entry['cost_basis'] = float(spaxx_entry['cost_basis']) + pending_cash
        else:
            positions.append({
                'symbol':     'SPAXX',
                'name':       'Fidelity Government Money Market (Pending)',
                'quantity':   pending_cash,
                'avg_cost':   1.0,
                'cost_basis': pending_cash,
                'last_price': 1.0,
                'lot_type':   'Cash',
                'is_cash':    True,
            })

    return positions


def _parse_balance_history_csv(text: str) -> list:
    rows = list(csv.reader(io.StringIO(text)))
    header_idx = None
    eight_col = False   # True = Fidelity "Balance and Performance Activity" (8-col) format
    for i, row in enumerate(rows):
        if not row:
            continue
        h0 = row[0].strip().lower()
        if h0 == 'monthly':
            header_idx = i
            break
        # "Balance and Performance Activity" export uses "Month" and merges
        # Dividends & Interest into one column → 8 cols instead of 9.
        if h0 == 'month':
            header_idx = i
            eight_col = True
            break
    if header_idx is None:
        return []

    def parse_money(s):
        s2 = str(s or '').replace('$', '').replace(',', '').strip()
        negative = s2.startswith('(') and s2.endswith(')')
        if negative:
            s2 = s2[1:-1]
        if s2 in ('', '-', '—', 'None'):
            return 0.0
        try:
            val = float(s2)
            return -val if negative else val
        except (ValueError, TypeError):
            return 0.0

    MONTH_MAP = {
        'jan': 1, 'feb': 2, 'mar': 3, 'apr': 4, 'may': 5, 'jun': 6,
        'jul': 7, 'aug': 8, 'sep': 9, 'oct': 10, 'nov': 11, 'dec': 12,
    }

    records = []
    for row in rows[header_idx + 1:]:
        if not row:
            continue
        raw_label = row[0].strip()
        if not raw_label:
            continue
        # Strip any trailing "(As of ...)" annotation
        clean = re.sub(r'\(.*?\)', '', raw_label).strip()
        if clean.lower() == 'total':
            break

        # Support two label formats:
        #   "May 2026"  — full month + 4-digit year (first/last rows)
        #   "Apr-26"    — abbreviated Mon-YY (all interior rows)
        month = None
        year = None
        if ' ' in clean:
            parts = clean.split()
            mon_str = parts[0].lower()[:3]
            month = MONTH_MAP.get(mon_str)
            try:
                year = int(parts[1])
            except (ValueError, IndexError):
                pass
        elif '-' in clean:
            dash = clean.split('-', 1)
            mon_str = dash[0].lower()[:3]
            month = MONTH_MAP.get(mon_str)
            try:
                yr_short = int(dash[1])
                year = 2000 + yr_short if yr_short < 50 else 1900 + yr_short
            except (ValueError, IndexError):
                pass

        if month is None or year is None:
            continue

        def g(idx):
            return parse_money(row[idx]) if idx < len(row) else 0.0

        if eight_col:
            # "Balance and Performance Activity" format (8 cols):
            # Month | Beg Bal | Mkt Chg Minus Fees | Div & Interest | Deposits | Withdrawals | Fees | End Bal
            # Note: fees are already netted into mkt_chg in this format, so
            # we add them back so our net_income formula stays consistent.
            beg         = g(1)
            mkt_chg_net = g(2)   # already net of fees
            div_and_int = g(3)   # Dividends & Interest combined
            deposits    = g(4)
            withdrawals = g(5)
            fees        = g(6)
            end         = g(7)
            mkt_chg     = mkt_chg_net + fees   # gross market change
            dividends   = div_and_int
            interest    = 0.0
        else:
            beg       = g(1)
            mkt_chg   = g(2)
            dividends = g(3)
            interest  = g(4)
            deposits  = g(5)
            withdrawals = g(6)
            fees      = g(7)
            end       = g(8)

        net_income = mkt_chg + dividends + interest - fees
        net_flow   = deposits - withdrawals
        denom      = beg + 0.5 * net_flow

        # Detect custodial transfer months: assets wiped to ~$0 then reinstated
        # next month. These are account reconstitutions, not real performance.
        transfer_threshold = 50_000
        is_transfer_out = (end < 100 and beg > transfer_threshold)
        is_transfer_in  = (beg < 100 and end > transfer_threshold)
        skip = is_transfer_out or is_transfer_in

        return_pct = 0.0 if skip else ((net_income / denom * 100) if denom != 0 else 0.0)

        records.append({
            'year':          year,
            'month':         month,
            'label':         clean,
            'beg_balance':   beg,
            'end_balance':   end,
            'market_change': mkt_chg,
            'dividends':     dividends,
            'interest':      interest,
            'deposits':      deposits,
            'withdrawals':   withdrawals,
            'fees':          fees,
            'return_pct':    round(return_pct, 4),
            'skip':          skip,
        })

    records.sort(key=lambda r: (r['year'], r['month']))
    return records


def _parse_captable(content: bytes, filename: str) -> tuple:
    """Parse a cap-table CSV or XLSX.

    Accepted column layouts (case-insensitive):

    Layout A — named commitment column:
        LP Name | Commitment Amount | Entity Type (opt) | Effective Date (opt)

    Layout B — year columns (waterfall / distribution schedule format):
        LP Name | <Entity Type (opt)> | 2022 | 2023 | 2024 | 2025 | …
        The LP's contribution is taken from the MOST RECENT (rightmost highest)
        year column that contains a non-zero value for that row.  If multiple
        year cells are filled the most-recent non-zero one wins; the "effective
        date" is set to Jan-1 of that year.

    Returns list of dicts:
        {legal_name, entity_type, commitment, effective_date, contribution_year}
    where `contribution_year` is None for Layout A.
    """
    out = []
    fname_lower = (filename or '').lower()

    if fname_lower.endswith('.xlsx') or fname_lower.endswith('.xls'):
        if not _OPENPYXL_OK:
            raise HTTPException(400, "openpyxl not installed — upload a CSV instead")
        wb  = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        ws  = wb.active
        raw = [[str(c.value if c.value is not None else '').strip() for c in row]
               for row in ws.iter_rows()]
    else:
        text = content.decode('utf-8', errors='replace')
        raw  = [[cell.strip() for cell in row] for row in csv.reader(io.StringIO(text))]

    if not raw:
        return [], {}, None

    import re as _re

    def _parse_money_inner(s):
        try:
            return float(str(s).replace('$', '').replace(',', '').strip())
        except (ValueError, TypeError):
            return None

    # ── Global scan: "Fund Established | year" (any row, any layout) ─────────
    fund_established_year = None
    for _row in raw:
        if len(_row) < 2:
            continue
        _lbl = str(_row[0]).lower().strip()
        if 'fund established' in _lbl or ('established' in _lbl and 'fund' in _lbl):
            try:
                _yr = int(float(str(_row[1]).replace(',', '').strip()))
                if 2000 <= _yr <= 2099:
                    fund_established_year = _yr
                    break
            except (ValueError, TypeError):
                pass

    # ── Layout C: transposed key-value (single LP, labels in col A, values in col B) ──
    # Detects files like:
    #   Row N:   "LP Name"              | "EM"
    #   Row N+1: "initial contribution" | "$1,400,000"
    #   Row N+2: "Economics"            | "0/25, 5% hurdle"
    #   Row N+3: "Fund Established"     | "2016"
    for i, row in enumerate(raw):
        if len(row) < 2:
            continue
        a = str(row[0]).lower().strip()
        b = str(row[1]).strip() if len(row) > 1 else ''
        if a not in ('lp name', 'legal name', 'investor name', 'investor'):
            continue
        # col B must be a real name, not a header keyword
        skip_kws = ('commitment', 'amount', 'date', 'type', 'entity', 'year', 'nav', 'name')
        if not b or any(kw in b.lower() for kw in skip_kws):
            continue
        # Confirm a contribution label appears within the next 10 rows
        contrib_found = False
        for j in range(i + 1, min(i + 10, len(raw))):
            nrow = raw[j]
            if not nrow or not str(nrow[0]).strip():
                continue
            nl = str(nrow[0]).lower().strip()
            if any(kw in nl for kw in ('contribution', 'committed', 'initial', 'capital commit')):
                contrib_found = True
                break
        if not contrib_found:
            continue

        # ── Parse Layout C ────────────────────────────────────────────────────
        # LP names: scan horizontally from col B onwards until an empty cell.
        # "LP Name | partner1 | partner2 | partner3"  → 3 LPs
        lp_names = []
        for ci in range(1, len(row)):
            nm = str(row[ci]).strip()
            if not nm:
                break       # stop at first empty cell
            lp_names.append(nm)

        # Per-LP contributions live in the NEXT KEY-VALUE row whose label
        # contains "contribution"/"committed"/"initial".  If values are in
        # matching columns they map 1:1; a single value applies to all LPs.
        eff_date   = str(datetime.utcnow().date())
        econ_str   = ''
        lp_amts    = {}   # col_index → amount
        for row2 in raw[i + 1:]:
            if not row2 or not str(row2[0]).strip():
                continue
            lbl = str(row2[0]).lower().strip()
            if any(kw in lbl for kw in ('contribution', 'committed', 'initial', 'capital commit')):
                # Collect amounts from each LP column (col 1…len(lp_names))
                for ci, _nm in enumerate(lp_names, 1):
                    raw_val = row2[ci] if ci < len(row2) else ''
                    v = _parse_money_inner(raw_val)
                    if v and v > 0:
                        lp_amts[ci] = v
                # If only one amount found and multiple LPs, apply to LP at col 1
                if not lp_amts:
                    # Try any cell in the row
                    for ci2 in range(1, len(row2)):
                        v = _parse_money_inner(row2[ci2])
                        if v and v > 0:
                            lp_amts[1] = v
                            break
            elif any(kw in lbl for kw in ('economics', 'econ', 'fee structure', 'terms')):
                econ_str = str(row2[1]).strip() if len(row2) > 1 else ''
            elif any(kw in lbl for kw in ('date', 'effective', 'inception')) and len(row2) > 1:
                raw_date = str(row2[1]).strip()
                if raw_date:
                    # Normalize bare year → YYYY-01-01
                    if _re.fullmatch(r'\d{4}', raw_date):
                        raw_date = raw_date + '-01-01'
                    eff_date = raw_date

        lp_rows_c = []
        for ci, nm in enumerate(lp_names, 1):
            amt = lp_amts.get(ci)
            if not amt or amt <= 0:
                continue
            lp_rows_c.append({
                'legal_name':        nm,
                'entity_type':       'individual',
                'commitment':        amt,
                'effective_date':    eff_date,
                'contribution_year': None,
            })

        # Parse economics string: "0/25, 5% hurdle" | "2 & 20, 8% hurdle" | "2/20"
        economics_c = {}
        if econ_str:
            m_slash = _re.match(r'(\d+(?:\.\d+)?)\s*[/&]\s*(\d+(?:\.\d+)?)', econ_str)
            if m_slash:
                economics_c['mgmt_fee_pct'] = round(float(m_slash.group(1)) / 100, 6)
                economics_c['carry_pct']    = round(float(m_slash.group(2)) / 100, 6)
            hm = _re.search(r'(\d+(?:\.\d+)?)\s*%\s*hurdle', econ_str, _re.IGNORECASE)
            if hm:
                economics_c['hurdle_pct'] = round(float(hm.group(1)) / 100, 6)

        return lp_rows_c, economics_c, fund_established_year

    # ── Find header row ───────────────────────────────────────────────────────
    header = []
    data_start = 0
    for i, row in enumerate(raw):
        joined = ' '.join(row).lower()
        if 'lp name' in joined or 'legal name' in joined or 'name' in joined:
            header = [h.lower().strip() for h in row]
            data_start = i + 1
            break

    if not header:
        raise HTTPException(400, "Could not find header row in cap table. "
            "Expected a column containing 'LP Name', 'Legal Name', or 'Name'.")

    # ── Detect year columns (4-digit int, 2000–2099) ─────────────────────────
    import re as _re
    year_cols = []   # list of (col_index, year_int) sorted ascending by year
    for idx, h in enumerate(header):
        m = _re.fullmatch(r'(20\d{2})', h.strip())
        if m:
            year_cols.append((idx, int(m.group(1))))
    year_cols.sort(key=lambda x: x[1])

    # ── Column helpers ────────────────────────────────────────────────────────
    def col_idx(header, *names):
        """Return index of first header that contains any of the given names."""
        for n in names:
            for i, h in enumerate(header):
                if n.lower() in h:
                    return i
        return None

    def col_val(row, *names):
        idx = col_idx(header, *names)
        return row[idx] if idx is not None and idx < len(row) else ''

    name_idx   = col_idx(header, 'lp name', 'legal name', 'name')
    commit_idx = col_idx(header, 'commitment', 'amount', 'capital')  # Layout A
    etype_idx  = col_idx(header, 'entity type', 'type')
    date_idx   = col_idx(header, 'effective date', 'date')

    valid_etypes = ('individual','joint','llc','trust','ira','corp',
                    'partnership','foundation','other')

    def _parse_money(s):
        try:
            return float(str(s).replace('$','').replace(',','').strip())
        except (ValueError, TypeError):
            return None

    for row in raw[data_start:]:
        if not any(row):
            continue
        if name_idx is None or name_idx >= len(row):
            continue
        name = row[name_idx].strip()
        if not name:
            continue

        etype_raw = (row[etype_idx].strip() if etype_idx is not None and etype_idx < len(row) else '')
        etype     = etype_raw.lower() if etype_raw.lower() in valid_etypes else 'individual'

        # ── Layout B: year columns ────────────────────────────────────────────
        if year_cols:
            # Walk year columns newest → oldest; take first non-zero value
            amt          = None
            contrib_year = None
            for col_i, yr in reversed(year_cols):
                v = _parse_money(row[col_i] if col_i < len(row) else '')
                if v and v > 0:
                    amt          = v
                    contrib_year = yr
                    break

            if amt is None:
                # If the row has only zeros in year cols, fall through to Layout A
                if commit_idx is not None and commit_idx < len(row):
                    amt = _parse_money(row[commit_idx])
                    contrib_year = None

            if not amt or amt <= 0:
                continue

            eff_date = f"{contrib_year}-01-01" if contrib_year else str(datetime.utcnow().date())
            if date_idx is not None and date_idx < len(row) and row[date_idx]:
                eff_date = row[date_idx].strip() or eff_date

            out.append({
                'legal_name':       name,
                'entity_type':      etype,
                'commitment':       amt,
                'effective_date':   eff_date,
                'contribution_year': contrib_year,
            })

        # ── Layout A: named commitment column ─────────────────────────────────
        else:
            if commit_idx is None:
                continue
            amt_str = row[commit_idx] if commit_idx < len(row) else ''
            amt = _parse_money(amt_str)
            if not amt or amt <= 0:
                continue

            eff_date = str(datetime.utcnow().date())
            if date_idx is not None and date_idx < len(row) and row[date_idx]:
                eff_date = row[date_idx].strip() or eff_date

            out.append({
                'legal_name':       name,
                'entity_type':      etype,
                'commitment':       amt,
                'effective_date':   eff_date,
                'contribution_year': None,
            })

    # ── Scan remaining rows for economics (col C = index 2) ──────────────────
    # After the last LP data row (data_start + len(out) ish) look for rows
    # where column A/B contains a fee/carry/hurdle label and column C contains
    # a percentage value.  Handles both "20%" and "0.20" formats.
    economics = {}

    def _parse_pct(s):
        """Return a fraction 0–1 from '20%' or '0.20' or '20', or None."""
        s = str(s or '').replace('%', '').replace('$', '').replace(',', '').strip()
        try:
            v = float(s)
        except (ValueError, TypeError):
            return None
        # If value looks like a whole-number percentage (e.g. 20 → 0.20), convert.
        # We treat values > 1 as already-in-percent form.
        return v / 100 if v > 1 else v

    econ_keywords = {
        'mgmt_fee_pct':  ('management fee', 'mgmt fee', 'management', 'mgmt', 'annual fee'),
        'carry_pct':     ('carry', 'carried interest', 'performance fee', 'incentive fee',
                          'performance allocation'),
        'hurdle_pct':    ('hurdle', 'preferred return', 'pref return', 'preferred',
                          'hurdle rate', 'pref'),
    }

    # scan ALL rows — look for any row where col C looks like a pct
    # and the label (col A or B) matches a known keyword
    for row in raw[data_start:]:
        if len(row) < 3:
            continue
        label_text = ' '.join(str(c).lower() for c in row[:3] if c)
        col_c_val  = row[2].strip() if len(row) > 2 else ''

        pct = _parse_pct(col_c_val)
        if pct is None or pct <= 0 or pct >= 1:
            continue  # only accept 0 < pct < 1

        for key, kws in econ_keywords.items():
            if key not in economics:  # first match wins
                if any(kw in label_text for kw in kws):
                    economics[key] = round(pct, 6)
                    break

    return out, economics, fund_established_year

from fastapi import FastAPI, HTTPException, BackgroundTasks, UploadFile, File, Form, Request
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, RedirectResponse, JSONResponse
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel

# Make the project root importable when running from the repo root.
sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

import claude_analyst as analyst

app = FastAPI(title="DGA Research Analyst API", version="1.0.0")

app.add_middleware(
    CORSMiddleware,
    # Wildcard is intentional — the API is consumed by:
    #   • https://dga-portfolio.up.railway.app       (current Railway URL)
    #   • https://portfolio.dgacapital.com           (new custom domain, ui65+)
    #   • iOS app via Expo runtime                   (mobile)
    # All auth happens via tokens in the x-auth-token / x-auth-v2-token
    # headers, not cookies — so wildcard is safe here and credentials are
    # never sent cross-origin via the browser's credentials channel.
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# ---------------------------------------------------------------------------
# Auth — stateless HMAC token (survives restarts, no DB needed)
# ---------------------------------------------------------------------------
_PUBLIC_PATHS = {
    "/health", "/info", "/api/auth", "/api/build", "/api/diagnostics", "/",
    "/api/auth/v2/login",   # email+password login is unauthenticated by design
}

def _portfolio_password() -> str:
    return os.environ.get("PORTFOLIO_PASSWORD", "dgacapital").strip()

def _token_secret() -> str:
    return os.environ.get("TOKEN_SECRET", "dga-capital-jwt-secret").strip()

def _make_token(password: str) -> str:
    """Derive a deterministic token from password + secret (HMAC-SHA256)."""
    return hmac.new(  # type: ignore[attr-defined]
        _token_secret().encode(), password.encode(), hashlib.sha256
    ).hexdigest()

def _valid_token(token: str) -> bool:
    expected = _make_token(_portfolio_password())
    return hmac.compare_digest(token.strip(), expected)

# Fund-specific auth — second layer on top of main auth.
# FUND_PASSWORD defaults to "dgacapital"; set env var to change it independently.
def _fund_password() -> str:
    return os.environ.get("FUND_PASSWORD", "genesis").strip()

def _make_fund_token() -> str:
    """Deterministic fund token: HMAC of 'fund:<password>' with shared secret."""
    return hmac.new(
        _token_secret().encode(), f"fund:{_fund_password()}".encode(), hashlib.sha256
    ).hexdigest()

def _valid_fund_token(token: str) -> bool:
    return hmac.compare_digest(token.strip(), _make_fund_token())

def _require_fund_token(request: Request) -> None:
    """Allow either a valid fund_token OR a v2 GP token.

    Backward compatible with the old fund portal (which uses x-fund-token),
    while letting the new GP dashboard call the same admin endpoints with
    its v2 JWT (x-auth-v2-token). LPs can never reach these endpoints.
    """
    # First check: v2 GP token (attached by auth_middleware)
    claims = getattr(request.state, 'auth_claims', None)
    if claims and claims.get("role") == "gp":
        return  # GP is authorized for all fund admin operations

    # Fall back to legacy fund token
    token = (request.headers.get("x-fund-token")
             or request.query_params.get("fund_token")
             or "")
    if not _valid_fund_token(token):
        raise HTTPException(status_code=403, detail="Fund access requires GP authentication or fund token")

@app.middleware("http")
async def auth_middleware(request: Request, call_next):
    path = request.url.path
    # Always allow: public API paths, static assets, the web app shell
    if (path in _PUBLIC_PATHS
            or path.startswith("/app/")
            or path.startswith("/branding/")
            or not path.startswith("/api/")):
        return await call_next(request)

    # ── v2 token (email+password JWT) — accepted on every /api/* path.
    # When present, attach the decoded claims to request.state for
    # downstream handlers to use for scope-based filtering.
    v2_tok = (request.headers.get("x-auth-v2-token")
              or request.query_params.get("v2_token")
              or "")
    if v2_tok:
        import auth_v2 as _av2  # local import to avoid module-level cycles
        claims = _av2.verify_token(v2_tok)
        if claims:
            # Refresh fund/account assignments from live user record so GP
            # changes take effect immediately without requiring LP re-login.
            try:
                fresh = _av2.find_user_by_lp_id(claims.get("lp_id", ""))
                if fresh:
                    claims = {
                        **claims,
                        "fund_memberships":    fresh.get("fund_memberships", {}),
                        "managed_account_ids": fresh.get("managed_account_ids", []),
                    }
            except Exception:
                pass
            request.state.auth_claims = claims
            return await call_next(request)

    # ── v1 token (legacy single-password HMAC) — backward compat for the
    # existing web shell and mobile app until they're migrated to v2.
    token = (request.headers.get("x-auth-token")
             or request.query_params.get("token")
             or "")
    if _valid_token(token):
        return await call_next(request)

    # Fund paths may also be accessed with a valid fund token alone
    # (they perform their own _require_fund_token check internally).
    if path.startswith("/api/fund/"):
        fund_tok = (request.headers.get("x-fund-token")
                    or request.query_params.get("fund_token")
                    or "")
        if _valid_fund_token(fund_tok):
            return await call_next(request)

    return JSONResponse(status_code=401, content={"detail": "Unauthorized"})

# In-memory job store: { job_id: { status, ticker, result, error, created_at } }
_jobs: dict[str, dict[str, Any]] = {}
_jobs_lock = threading.Lock()

# ---------------------------------------------------------------------------
# Persistent job-index — survives server restarts on Railway.
# Maps { job_id: { "ticker": str, "type": "analysis"|"portfolio" } }
# Stored next to the stocks folder so it lives on the same volume.
# ---------------------------------------------------------------------------
_JOB_INDEX_PATH      = analyst.STOCKS_FOLDER / "_job_index.json"
# Lean snapshot of the last completed rebalance — persisted to Dropbox so
# both the web client and the mobile app read the same ground truth.
_LAST_JOB_PATH       = analyst.STOCKS_FOLDER / "_portfolio_last_job.json"

def _load_job_index() -> dict:
    try:
        if _JOB_INDEX_PATH.exists():
            return json.loads(_JOB_INDEX_PATH.read_text())
    except Exception:
        pass
    return {}

def _save_job_index_entry(job_id: str, entry: dict) -> None:
    """Append / update one entry in the on-disk job index (best-effort)."""
    try:
        idx = _load_job_index()
        idx[job_id] = entry
        _JOB_INDEX_PATH.write_text(json.dumps(idx, indent=2))
    except Exception:
        pass


# ---------------------------------------------------------------------------
# Pydantic models
# ---------------------------------------------------------------------------

class FundAuthRequest(BaseModel):
    password: str

class CreateFundRequest(BaseModel):
    name:           str                 # 'DGA Capital Fund II, LP'
    short_name:     str                 # 'DGA-II'
    inception_date: str                 # 'YYYY-MM-DD'
    fiscal_year_end: str | None = None  # defaults to Dec 31 of inception year
    mgmt_fee_pct:   float = 0.02
    carry_pct:      float = 0.25
    hurdle_pct:     float = 0.08
    fund_type:      str   = 'lp_fund'   # 'lp_fund' | 'managed_account'

class AnalyzeRequest(BaseModel):
    ticker: str
    generate_gamma: bool = False


class JobStatus(BaseModel):
    job_id: str
    ticker: str
    status: str           # queued | running | done | failed
    created_at: str
    error: str | None = None
    result: dict | None = None
    # Live progress emitted by analyze_ticker — populated on /status polls
    # while status='running'. Frontend uses this to drive a real progress bar
    # instead of simulated step transitions.
    progress: dict | None = None


class PortfolioJobStatus(BaseModel):
    job_id: str
    status: str           # queued | running | done | failed
    created_at: str
    strategy: str
    n_tickers: int
    error: str | None = None
    result: dict | None = None
    # Input weights from the uploaded CSV: {ticker: decimal_weight} so the
    # frontend can render current-weight → target-weight arrows.
    input_weights: dict | None = None
    # Live progress emitted by run_portfolio_rebalance — populated on
    # /portfolio/{id} polls while status='running'. Frontend renders a
    # per-ticker counter ("3 / 12 — analyzing AAPL") and progress bar.
    progress: dict | None = None


class IntelligenceRequest(BaseModel):
    sector: str = "Tech"  # sector focus: Tech, Energy, Healthcare, Financials, Consumer, Industrials, Materials, Real Estate, Best Mix


# ---------------------------------------------------------------------------
# Auth route
# ---------------------------------------------------------------------------

class AuthRequest(BaseModel):
    password: str

@app.post("/api/auth")
def auth(req: AuthRequest):
    if hmac.compare_digest(req.password.strip(), _portfolio_password()):
        return {"token": _make_token(req.password.strip())}
    raise HTTPException(status_code=401, detail="Invalid password")


# ---------------------------------------------------------------------------
# v2 Auth — per-user email+password with role + scope claims
# ---------------------------------------------------------------------------
# Runs alongside /api/auth (which the current mobile app still uses).
# Once the new web + mobile login flows ship, /api/auth can be deprecated.
import auth_v2 as auth_v2_mod   # noqa: E402  — imported here to avoid top-of-file churn


class AuthV2LoginRequest(BaseModel):
    email:    str
    password: str


class AuthV2ChangePasswordRequest(BaseModel):
    old_password: str
    new_password: str


@app.post("/api/auth/v2/login")
def auth_v2_login(req: AuthV2LoginRequest):
    """Per-user email + password login. Returns a signed JWT-style token
    carrying role + scope (fund memberships, managed accounts)."""
    result = auth_v2_mod.login(req.email.strip(), req.password)
    if not result:
        raise HTTPException(status_code=401, detail="Invalid email or password")
    return result


@app.get("/api/auth/v2/me")
def auth_v2_me(request: Request):
    """Return the authenticated user's profile. Read from the
    x-auth-v2-token header (or x-auth-token / token query param as
    fallback so it composes with the existing middleware)."""
    token = (request.headers.get("x-auth-v2-token")
             or request.headers.get("x-auth-token")
             or request.query_params.get("token")
             or "")
    user = auth_v2_mod.whoami(token)
    if not user:
        raise HTTPException(status_code=401, detail="Unauthorized")
    return user


@app.post("/api/auth/v2/change-password")
def auth_v2_change_password(request: Request, body: AuthV2ChangePasswordRequest):
    """Change the current user's password. Requires a valid v2 token."""
    token = (request.headers.get("x-auth-v2-token")
             or request.headers.get("x-auth-token")
             or request.query_params.get("token")
             or "")
    claims = auth_v2_mod.verify_token(token)
    if not claims:
        raise HTTPException(status_code=401, detail="Unauthorized")
    ok = auth_v2_mod.change_password(
        lp_id=claims["lp_id"],
        old_password=body.old_password,
        new_password=body.new_password,
    )
    if not ok:
        raise HTTPException(
            status_code=400,
            detail="Could not change password — check old password and ensure new is ≥ 8 chars",
        )
    return {"ok": True}


# ---------------------------------------------------------------------------
# GP Admin — LP user management
# ---------------------------------------------------------------------------

class LPSetPasswordRequest(BaseModel):
    lp_id:       str
    new_password: str
    must_change:  bool = True   # True = LP must choose their own pw on next login


@app.get("/api/v2/admin/lp/list")
def admin_lp_list(request: Request):
    """Return all LP/GP users (no password hashes). GP-only."""
    claims = _claims_or_401(request)
    if claims.get("role") != "gp":
        raise HTTPException(status_code=403, detail="GP role required")
    return {"users": auth_v2_mod.list_users()}


@app.post("/api/v2/admin/lp/set-password")
def admin_lp_set_password(request: Request, body: LPSetPasswordRequest):
    """GP-only: set an LP's password without requiring the old one.

    Useful for onboarding new LPs and handling forgotten-password requests.
    By default sets must_change_password=True so the LP is prompted to
    choose their own password on first login.
    """
    claims = _claims_or_401(request)
    if claims.get("role") != "gp":
        raise HTTPException(status_code=403, detail="GP role required")

    lp_id = (body.lp_id or "").strip()
    if not lp_id:
        raise HTTPException(status_code=400, detail="lp_id required")

    ok = auth_v2_mod.gp_set_password(
        lp_id=lp_id,
        new_password=body.new_password,
        must_change=body.must_change,
    )
    if not ok:
        raise HTTPException(
            status_code=400,
            detail="Could not set password — check lp_id exists and password is ≥ 6 chars",
        )

    user = auth_v2_mod.find_user_by_lp_id(lp_id)
    return {
        "ok":    True,
        "lp_id": lp_id,
        "email": user.get("email") if user else None,
        "name":  user.get("name")  if user else None,
        "must_change_password": body.must_change,
    }


class LPCreateRequest(BaseModel):
    email:               str
    name:                str
    password:            str
    fund_memberships:    dict  = {}
    managed_account_ids: list  = []


class LPAssignRequest(BaseModel):
    lp_id:               str
    fund_memberships:    dict  = {}
    managed_account_ids: list  = []


@app.post("/api/v2/admin/lp/create")
def admin_lp_create(request: Request, body: LPCreateRequest):
    """GP-only: create a new LP user account."""
    claims = _claims_or_401(request)
    if claims.get("role") != "gp":
        raise HTTPException(403, "GP access required")
    try:
        lp_id = auth_v2_mod.create_user(
            email=body.email,
            name=body.name,
            password=body.password,
            fund_memberships=body.fund_memberships,
            managed_account_ids=body.managed_account_ids,
        )
        return {"ok": True, "lp_id": lp_id, "email": body.email, "name": body.name}
    except ValueError as e:
        raise HTTPException(400, str(e))


@app.post("/api/v2/admin/lp/update-assignments")
def admin_lp_update_assignments(request: Request, body: LPAssignRequest):
    """GP-only: update which funds and managed accounts an LP can see."""
    claims = _claims_or_401(request)
    if claims.get("role") != "gp":
        raise HTTPException(403, "GP access required")
    ok = auth_v2_mod.update_assignments(
        lp_id=body.lp_id,
        fund_memberships=body.fund_memberships,
        managed_account_ids=body.managed_account_ids,
    )
    if not ok:
        raise HTTPException(404, "LP not found")
    return {"ok": True, "lp_id": body.lp_id}


# ===========================================================================
# Fund display settings — GP writes, LP reads
# ===========================================================================

class FundSettingsRequest(BaseModel):
    benchmark: str = "sp500"   # key from _BENCHMARK_DEFS
    period:    str = "all"     # "all" | "5yr" | "3yr"


class ManualAnnualReturnRequest(BaseModel):
    year:       int
    return_pct: Optional[float] = None   # None = delete the override


@app.post("/api/v2/gp/fund/{fund_id}/manual-return")
async def set_manual_annual_return(fund_id: str, body: ManualAnnualReturnRequest, request: Request):
    """GP-only: set or delete a manual annual return override for one year.

    Body: { year: 2022, return_pct: -13.61 }
    Omit return_pct (or pass null) to remove an existing override.
    """
    _require_fund_token(request)
    try:
        conn = _fund_conn()
        try:
            with conn.cursor() as cur:
                if body.return_pct is None:
                    cur.execute(
                        "DELETE FROM manual_annual_returns WHERE fund_id = %s AND year = %s",
                        (fund_id, body.year),
                    )
                else:
                    cur.execute("""
                        INSERT INTO manual_annual_returns (fund_id, year, return_pct, source, updated_at)
                        VALUES (%s, %s, %s, 'gp_manual', now())
                        ON CONFLICT (fund_id, year) DO UPDATE
                            SET return_pct = EXCLUDED.return_pct,
                                source     = EXCLUDED.source,
                                updated_at = now()
                    """, (fund_id, body.year, body.return_pct))
            conn.commit()
        finally:
            conn.close()
    except Exception as exc:
        raise HTTPException(500, f"DB error: {exc}")
    return {"ok": True, "fund_id": fund_id, "year": body.year, "return_pct": body.return_pct}


@app.get("/api/fund/{fund_id}/settings")
async def fund_get_settings(fund_id: str, request: Request):
    """Return GP-configured display settings (benchmark, period). Readable by LP + GP."""
    claims = getattr(request.state, "auth_claims", None)
    if not claims:
        _require_fund_token(request)
    try:
        conn = _fund_conn()
        try:
            with conn.cursor(cursor_factory=_RealDictCursor) as cur:
                fid = _resolve_fund_id(cur, fund_id)
                cur.execute(
                    "SELECT settings_json FROM fund_display_settings WHERE fund_id = %s",
                    (fid,),
                )
                row = cur.fetchone()
                settings = json.loads(row["settings_json"]) if row else {}
        finally:
            conn.close()
    except Exception as exc:
        return {"ok": False, "detail": str(exc), "settings": {}}
    bkey = settings.get("benchmark", "sp500")
    return {
        "ok":              True,
        "fund_id":         fund_id,
        "settings":        settings,
        "benchmark_key":   bkey,
        "benchmark_label": _BENCHMARK_DEFS.get(bkey, {}).get("label", "S&P 500"),
        "period":          settings.get("period", "all"),
    }


@app.post("/api/v2/gp/fund/{fund_id}/settings")
async def fund_save_settings(fund_id: str, request: Request, body: FundSettingsRequest):
    """GP-only: save display settings (benchmark, period) for a fund."""
    claims = _claims_or_401(request)
    if claims.get("role") != "gp":
        raise HTTPException(403, "GP access required")

    bkey   = body.benchmark if body.benchmark in _BENCHMARK_DEFS else "sp500"
    period = body.period    if body.period in ("all", "5yr", "3yr") else "all"
    settings = {"benchmark": bkey, "period": period}

    conn = _fund_conn()
    try:
        with conn.cursor(cursor_factory=_RealDictCursor) as cur:
            fid = _resolve_fund_id(cur, fund_id)
            cur.execute("""
                INSERT INTO fund_display_settings (fund_id, settings_json, updated_at)
                VALUES (%s, %s, now())
                ON CONFLICT (fund_id) DO UPDATE
                    SET settings_json = EXCLUDED.settings_json,
                        updated_at    = now()
            """, (fid, json.dumps(settings)))
        conn.commit()
    finally:
        conn.close()

    return {
        "ok":              True,
        "fund_id":         fund_id,
        "settings":        settings,
        "benchmark_label": _BENCHMARK_DEFS[bkey]["label"],
    }


# ===========================================================================
# Phase B — GP dashboard data endpoints
#
# These endpoints power the new Terminal Pro view at /gp:
#   • /api/market/indices   — top index ribbon (11 instruments, 15s cache)
#   • /api/search/resolve   — company name → ticker
#   • /api/gurufocus/{tk}   — GuruFocus fundamentals snapshot (per-ticker cache)
#   • /api/watchlist        — per-user personal watchlist (v2-token scoped)
# ===========================================================================

# ── Index ribbon ────────────────────────────────────────────────────────────
_INDEX_TICKERS = [
    ("S&P 500",       "^GSPC"),
    ("Dow 30",        "^DJI"),
    ("Nasdaq",        "^IXIC"),
    ("Russell 2000",  "^RUT"),
    ("VIX",           "^VIX"),
    ("10Y Treasury",  "^TNX"),
    ("Dollar Index",  "DX-Y.NYB"),
    ("Gold",          "GC=F"),
    ("Crude Oil",     "CL=F"),
    ("Bitcoin",       "BTC-USD"),
    ("Ethereum",      "ETH-USD"),
]
_INDICES_CACHE: dict[str, Any] = {"data": None, "ts": 0.0}
_INDICES_TTL = 15  # seconds


@app.get("/api/market/indices")
def market_indices():
    """Return live quotes for the 11 instruments shown in the GP index ribbon.
    Cached server-side for 15s to stay well under yfinance rate limits even
    if every GP browser polls aggressively."""
    now = time.time()
    if _INDICES_CACHE["data"] and (now - _INDICES_CACHE["ts"]) < _INDICES_TTL:
        return _INDICES_CACHE["data"]

    if yf is None:
        raise HTTPException(status_code=503, detail="yfinance not installed")

    out = []
    for label, sym in _INDEX_TICKERS:
        try:
            t = yf.Ticker(sym)
            info = t.fast_info
            last  = float(info.last_price)  if info.last_price  is not None else None
            prev  = float(info.previous_close) if info.previous_close is not None else None
            pct   = ((last - prev) / prev * 100.0) if (last and prev and prev > 0) else None
            out.append({
                "label":    label,
                "symbol":   sym,
                "price":    last,
                "prev":     prev,
                "pct":      pct,
            })
        except Exception as exc:
            out.append({"label": label, "symbol": sym, "price": None, "prev": None, "pct": None, "error": str(exc)[:80]})

    payload = {"indices": out, "fetched_at": int(now)}
    _INDICES_CACHE.update(data=payload, ts=now)
    return payload


# ── Universal search: company name / ticker → resolved ticker ───────────────
class SearchResolveResult(BaseModel):
    ticker:   str
    name:     str
    exchange: Optional[str] = None
    score:    float = 1.0


@app.get("/api/search/resolve")
def search_resolve(q: str = ""):
    """Resolve a company name OR ticker symbol to one or more ticker matches.

    Strategy:
      1. If `q` looks like a ticker (≤6 chars, all uppercase letters/digits),
         try yfinance directly — if info has a `longName` we have a hit.
      2. Otherwise, use yfinance's search() helper (which queries Yahoo's
         autocomplete backend). Returns up to 5 results.
    """
    query = (q or "").strip()
    if not query:
        return {"results": []}
    if yf is None:
        raise HTTPException(status_code=503, detail="yfinance not installed")

    results: list[dict] = []

    # Heuristic: looks like a ticker → try direct resolution first
    if len(query) <= 6 and re.fullmatch(r"[A-Za-z0-9.\-]+", query):
        try:
            t = yf.Ticker(query.upper())
            info = getattr(t, "info", {}) or {}
            longname = info.get("longName") or info.get("shortName")
            if longname:
                results.append({
                    "ticker":   query.upper(),
                    "name":     longname,
                    "exchange": info.get("exchange") or info.get("fullExchangeName"),
                    "score":    1.0,
                })
        except Exception:
            pass

    # Yahoo autocomplete via yfinance.Search()
    try:
        search_helper = getattr(yf, "Search", None)
        if search_helper is None:
            # Older yfinance — try the lookup endpoint directly via requests
            import requests as _rq
            r = _rq.get(
                "https://query2.finance.yahoo.com/v1/finance/search",
                params={"q": query, "quotesCount": 5, "newsCount": 0},
                headers={"User-Agent": "Mozilla/5.0 DGA Research"},
                timeout=4,
            )
            if r.ok:
                data = r.json()
                for hit in (data.get("quotes") or []):
                    sym = (hit.get("symbol") or "").strip()
                    if not sym or sym in {x["ticker"] for x in results}:
                        continue
                    results.append({
                        "ticker":   sym,
                        "name":     hit.get("longname") or hit.get("shortname") or sym,
                        "exchange": hit.get("exchDisp") or hit.get("exchange"),
                        "score":    float(hit.get("score") or 0) / 100000.0,
                    })
                    if len(results) >= 5:
                        break
        else:
            s = search_helper(query, max_results=5, news_count=0)
            for hit in (s.quotes or []):
                sym = (hit.get("symbol") or "").strip()
                if not sym or sym in {x["ticker"] for x in results}:
                    continue
                results.append({
                    "ticker":   sym,
                    "name":     hit.get("longname") or hit.get("shortname") or sym,
                    "exchange": hit.get("exchDisp") or hit.get("exchange"),
                    "score":    float(hit.get("score") or 0) / 100000.0,
                })
                if len(results) >= 5:
                    break
    except Exception as exc:
        # Soft-fail — return whatever we got from the direct lookup
        return {"results": results[:5], "warning": str(exc)[:120]}

    return {"results": results[:5]}


# ── GuruFocus fundamentals snapshot ─────────────────────────────────────────
_GURUFOCUS_CACHE: dict[str, tuple] = {}      # { ticker: (data, ts) }
_GURUFOCUS_TTL = 900                          # 15 min — saves API calls


def _gurufocus_token() -> Optional[str]:
    """Read the GuruFocus token from either env var name (supports both the
    legacy GURUFOCUS_TOKEN used in claude_analyst.py and the newer
    GURUFOCUS_API_TOKEN that was added to Railway for the v2 frontend)."""
    return (os.environ.get("GURUFOCUS_API_TOKEN")
            or os.environ.get("GURUFOCUS_TOKEN")
            or "").strip() or None


@app.get("/api/gurufocus/{ticker}")
def gurufocus_snapshot(ticker: str):
    """Return a compact GuruFocus snapshot for a ticker: market cap, PE,
    GF Score, GF Value, financial strength, key ratios. Token never leaves
    the server. Cached per-ticker for 15 min."""
    tk = (ticker or "").strip().upper()
    if not tk:
        raise HTTPException(status_code=400, detail="ticker required")

    token = _gurufocus_token()
    if not token:
        raise HTTPException(status_code=503, detail="GURUFOCUS_API_TOKEN not configured")

    now = time.time()
    cached = _GURUFOCUS_CACHE.get(tk)
    if cached and (now - cached[1]) < _GURUFOCUS_TTL:
        return cached[0]

    import requests as _rq

    # Browser-like headers — GuruFocus often blocks bare/empty UAs on
    # cloud IPs. We rotate through a couple of strategies on failure.
    base = "https://api.gurufocus.com/public/user/{tok}/stock/{tk}/summary"
    strategies = [
        # 1) Full token in URL path with a Chrome-style UA
        {
            "url": base.format(tok=token, tk=tk),
            "headers": {
                "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
                              "AppleWebKit/537.36 (KHTML, like Gecko) "
                              "Chrome/130.0.0.0 Safari/537.36",
                "Accept":     "application/json, text/plain, */*",
                "Accept-Language": "en-US,en;q=0.9",
                "Referer":    "https://www.gurufocus.com/",
            },
        },
    ]
    # If token looks like key:secret, try key-only as another path
    if ":" in token:
        strategies.append({
            "url": base.format(tok=token.split(":", 1)[0], tk=tk),
            "headers": strategies[0]["headers"],
        })

    last_status: list[str] = []
    payload: dict = {"ticker": tk, "ok": False}
    try:
        for strat in strategies:
            r = _rq.get(strat["url"], headers=strat["headers"], timeout=10)
            last_status.append(f"HTTP {r.status_code}")
            if r.status_code == 200:
                try:
                    data = r.json()
                except Exception:
                    data = {"raw": r.text[:2000]}
                payload = {
                    "ticker":     tk,
                    "ok":         True,
                    "fetched_at": int(now),
                    "data":       data,
                }
                break
            else:
                # Capture first chunk of body for diagnostics on failure
                last_status[-1] += f" body={r.text[:120]!r}"
        if not payload.get("ok"):
            payload = {
                "ticker":     tk,
                "ok":         False,
                "error":      "GuruFocus: " + " | ".join(last_status),
                "fetched_at": int(now),
            }
    except Exception as exc:
        payload = {"ticker": tk, "ok": False, "error": str(exc)[:200], "fetched_at": int(now)}

    _GURUFOCUS_CACHE[tk] = (payload, now)
    return payload


# ── Watchlist (per-user, v2-token scoped) ───────────────────────────────────
_WATCHLISTS_FILE = Path(os.environ.get(
    "DGA_WATCHLISTS_PATH",
    str(analyst.STOCKS_FOLDER / "_dga_watchlists.json"),
))
_watchlists_lock = threading.Lock()


def _load_watchlists() -> dict[str, list[str]]:
    try:
        if _WATCHLISTS_FILE.exists():
            return json.loads(_WATCHLISTS_FILE.read_text())
    except Exception:
        pass
    return {}


def _save_watchlists(data: dict[str, list[str]]) -> None:
    _WATCHLISTS_FILE.parent.mkdir(parents=True, exist_ok=True)
    _WATCHLISTS_FILE.write_text(json.dumps(data, indent=2, sort_keys=True))


def _claims_or_401(request: Request) -> dict:
    """Extract v2 auth claims attached by the middleware. 401 if missing
    (i.e. the request only had a legacy v1 token, not a v2 token)."""
    claims = getattr(request.state, "auth_claims", None)
    if not claims:
        raise HTTPException(status_code=401, detail="v2 token required for /api/watchlist")
    return claims


@app.get("/api/watchlist")
def watchlist_get(request: Request):
    claims = _claims_or_401(request)
    lp_id = claims["lp_id"]
    with _watchlists_lock:
        wl = _load_watchlists()
    tickers = wl.get(lp_id, [])

    # Enrich with live quotes so the UI doesn't need a second call
    quotes: dict[str, dict] = {}
    if tickers and yf is not None:
        for tk in tickers:
            try:
                t = yf.Ticker(tk)
                fi = t.fast_info
                last = float(fi.last_price)      if fi.last_price      is not None else None
                prev = float(fi.previous_close)  if fi.previous_close  is not None else None
                pct  = ((last - prev) / prev * 100.0) if (last and prev and prev > 0) else None
                quotes[tk] = {"price": last, "prev": prev, "pct": pct}
            except Exception:
                quotes[tk] = {"price": None, "prev": None, "pct": None}

    return {"tickers": tickers, "quotes": quotes}


class WatchlistAddRequest(BaseModel):
    ticker: str


@app.post("/api/watchlist")
def watchlist_add(request: Request, body: WatchlistAddRequest):
    claims = _claims_or_401(request)
    lp_id = claims["lp_id"]
    tk = (body.ticker or "").strip().upper()
    if not tk or not re.fullmatch(r"[A-Z0-9.\-]+", tk):
        raise HTTPException(status_code=400, detail="invalid ticker")

    with _watchlists_lock:
        wl = _load_watchlists()
        cur = wl.get(lp_id, [])
        if tk not in cur:
            cur.append(tk)
            wl[lp_id] = cur
            _save_watchlists(wl)
    return {"tickers": wl.get(lp_id, [])}


@app.delete("/api/watchlist/{ticker}")
def watchlist_remove(request: Request, ticker: str):
    claims = _claims_or_401(request)
    lp_id = claims["lp_id"]
    tk = (ticker or "").strip().upper()
    with _watchlists_lock:
        wl = _load_watchlists()
        cur = wl.get(lp_id, [])
        wl[lp_id] = [t for t in cur if t != tk]
        _save_watchlists(wl)
    return {"tickers": wl[lp_id]}


# ===========================================================================
# Phase C / A — LP-scoped data endpoints
#
# These endpoints filter by the authenticated user's lp_id + scope claims.
# GPs (role='gp') see ALL data unfiltered; LPs see ONLY their own funds
# and managed accounts.
# ===========================================================================

@app.get("/api/v2/lp/me/overview")
def lp_me_overview(request: Request):
    """Return the authenticated LP's personal overview:
      • profile: lp_id, name, email, role
      • funds[]: one row per fund the LP is a member of, with their
                 commitment + contributed + distributed + current NAV
                 (zeros if those columns aren't populated yet).
      • managed_accounts[]: one row per managed account the LP owns.
    GPs see all funds/accounts (no scope filtering) so the same endpoint
    works as a "fund overview" for them too."""
    claims = _claims_or_401(request)
    role           = claims.get("role", "lp")
    fund_memberships  = claims.get("fund_memberships", {}) or {}
    managed_accts     = claims.get("managed_account_ids", []) or []

    out = {
        "profile": {
            "lp_id": claims["lp_id"],
            "name":  claims.get("name", ""),
            "email": claims.get("email", ""),
            "role":  role,
        },
        "funds":             [],
        "managed_accounts":  [],
        "warnings":          [],
    }

    if not _PSYCOPG2_OK:
        out["warnings"].append("psycopg2 not installed — returning empty data")
        return out

    try:
        with _fund_conn() as conn, conn.cursor(cursor_factory=_RealDictCursor) as cur:

            # ── Funds ────────────────────────────────────────────────
            if role == "gp":
                cur.execute("""
                    SELECT id, name, short_name, fund_type
                      FROM funds
                     WHERE fund_type = 'lp_fund'
                     ORDER BY name, short_name
                """)
                fund_rows = cur.fetchall()
            else:
                names_lower = [n.lower() for n in fund_memberships.keys()]
                if names_lower:
                    cur.execute("""
                        SELECT id, name, short_name, fund_type
                          FROM funds
                         WHERE LOWER(name) = ANY(%s)
                         ORDER BY name, short_name
                    """, (names_lower,))
                    fund_rows = cur.fetchall()
                else:
                    fund_rows = []

            # ── Managed accounts ─────────────────────────────────────
            if role == "gp":
                cur.execute("""
                    SELECT id, name, short_name
                      FROM funds
                     WHERE fund_type = 'managed_account'
                     ORDER BY name, short_name
                """)
                acct_rows = cur.fetchall()
            else:
                accts_lower = [a.lower() for a in managed_accts]
                if accts_lower:
                    cur.execute("""
                        SELECT id, name, short_name
                          FROM funds
                         WHERE fund_type = 'managed_account'
                           AND LOWER(name) = ANY(%s)
                         ORDER BY name, short_name
                    """, (accts_lower,))
                    acct_rows = cur.fetchall()
                else:
                    acct_rows = []

            fund_fids = [str(f["id"]) for f in fund_rows]
            acct_fids = [str(a["id"]) for a in acct_rows]
            all_fids  = fund_fids + acct_fids

            if not all_fids:
                # Nothing to load — return early
                return out

            # ── Bulk YTD cache for managed accounts (fetch FIRST so we can
            #    extract attribution tickers and warm them in the same yfinance call) ──
            ytd_cache_by_fid: dict = {}
            if acct_fids:
                try:
                    cur.execute("""
                        SELECT fund_id::text, nav, ytd_pct, result_json, updated_at
                          FROM managed_account_ytd_cache
                         WHERE fund_id::text = ANY(%s)
                    """, (acct_fids,))
                    for r in cur.fetchall():
                        ytd_cache_by_fid[str(r["fund_id"])] = dict(r)
                except Exception:
                    conn.rollback()

            # ── Collect ALL symbols needing live prices in one shot ──
            # 1. tax_lots symbols (used by LP funds with real positions)
            # 2. attribution tickers from managed-account YTD result_json
            #    (so clicking into IND-I's positions panel hits the cache, not yfinance)
            all_syms_set: set = set()
            try:
                cur.execute("""
                    SELECT DISTINCT s.symbol
                      FROM tax_lots tl
                      JOIN securities s ON s.id = tl.security_id
                     WHERE tl.fund_id::text = ANY(%s) AND tl.closed_at IS NULL
                       AND s.symbol IS NOT NULL
                """, (all_fids,))
                for r in cur.fetchall():
                    if r["symbol"]:
                        all_syms_set.add(r["symbol"])
            except Exception:
                conn.rollback()

            for ytd_row in ytd_cache_by_fid.values():
                rj_raw = ytd_row.get("result_json")
                if not rj_raw:
                    continue
                try:
                    _rj = json.loads(rj_raw) if isinstance(rj_raw, str) else rj_raw
                    for a in (_rj.get("attribution") or []):
                        tk = a.get("ticker")
                        # Include price_missing tickers — variant fallback now
                        # handles cases like BRKB → BRK-B at fetch time.
                        if tk and float(a.get("end_shares") or 0) > 0:
                            all_syms_set.add(tk)
                except Exception:
                    pass

            # ── Single yfinance batch call for everything ──
            if all_syms_set:
                try:
                    _fetch_prices(list(all_syms_set))
                except Exception:
                    pass

            # ── Bulk NAV from positions (uses warmed price cache) ──
            mkt_nav_by_fid = {}
            try:
                mkt_nav_by_fid = _bulk_fund_market_nav(cur, all_fids)
            except Exception:
                conn.rollback()

            # ── Bulk latest NAV snapshots ──
            nav_snap_by_fid: dict = {}
            try:
                cur.execute("""
                    SELECT DISTINCT ON (fund_id) fund_id::text, net_nav, as_of_date
                      FROM nav_snapshots
                     WHERE fund_id::text = ANY(%s)
                     ORDER BY fund_id, as_of_date DESC
                """, (all_fids,))
                for r in cur.fetchall():
                    nav_snap_by_fid[str(r["fund_id"])] = dict(r)
            except Exception:
                conn.rollback()

            # ── Bulk account_balance_history for managed accounts ──
            # Pre-fetch in one query so the assembly loop has zero extra DB calls.
            acct_bal_hist_by_fid: dict = {}
            if acct_fids:
                try:
                    cur.execute("""
                        SELECT fund_id::text, data_json
                          FROM account_balance_history
                         WHERE fund_id::text = ANY(%s)
                    """, (acct_fids,))
                    _cur_year = datetime.utcnow().year
                    for row in cur.fetchall():
                        _raw  = row["data_json"]
                        _recs = json.loads(_raw) if isinstance(_raw, str) else (_raw or [])
                        _ytd  = sorted(
                            [r for r in _recs if r.get("year") == _cur_year and not r.get("skip")],
                            key=lambda r: r.get("month", 0)
                        )
                        if _ytd:
                            acct_bal_hist_by_fid[str(row["fund_id"])] = {
                                "beg":  float(_ytd[0].get("beg_balance") or 0),
                                "deps": round(sum(float(r.get("deposits") or 0) for r in _ytd), 2),
                                "wdrs": round(sum(float(r.get("withdrawals") or 0) for r in _ytd), 2),
                            }
                except Exception:
                    conn.rollback()

            # ── Bulk total committed capital per fund ──
            total_committed_by_fid: dict = {}
            if fund_fids:
                try:
                    cur.execute("""
                        SELECT l.fund_id::text,
                               COALESCE(SUM(c.commitment_amount), 0) AS total_committed
                          FROM commitments c
                          JOIN lps l ON l.id = c.lp_id
                         WHERE l.fund_id::text = ANY(%s) AND c.superseded_by IS NULL
                         GROUP BY l.fund_id
                    """, (fund_fids,))
                    for r in cur.fetchall():
                        total_committed_by_fid[str(r["fund_id"])] = float(r["total_committed"])
                except Exception:
                    conn.rollback()

            # ── Bulk latest annual snapshots (for GP carry calc) ──
            annual_snap_by_fid: dict = {}
            if fund_fids:
                try:
                    cur.execute("""
                        SELECT DISTINCT ON (fund_id) fund_id::text, gp_equity_end, end_nav
                          FROM fund_annual_snapshots
                         WHERE fund_id::text = ANY(%s)
                         ORDER BY fund_id, year DESC
                    """, (fund_fids,))
                    for r in cur.fetchall():
                        annual_snap_by_fid[str(r["fund_id"])] = dict(r)
                except Exception:
                    conn.rollback()

            # ── Bulk LP rows (GP: all funds at once; LP: per fund since alias varies) ──
            lp_rows_by_fid: dict = {}
            if fund_fids:
                if role == "gp":
                    try:
                        cur.execute("""
                            SELECT l.fund_id::text, l.legal_name, l.primary_email,
                                   COALESCE(c.commitment, 0) AS commitment_amount
                              FROM lps l
                              LEFT JOIN (
                                  SELECT lp_id, SUM(commitment_amount) AS commitment
                                    FROM commitments
                                   WHERE superseded_by IS NULL
                                   GROUP BY lp_id
                              ) c ON c.lp_id = l.id
                             WHERE l.fund_id::text = ANY(%s)
                             ORDER BY l.fund_id, l.legal_name
                        """, (fund_fids,))
                        for r in cur.fetchall():
                            fid = str(r["fund_id"])
                            lp_rows_by_fid.setdefault(fid, []).append(dict(r))
                    except Exception:
                        conn.rollback()
                else:
                    # LP: each fund needs its own alias — run one small query per fund
                    # (typically 1-3 funds, so cost is low)
                    for f in fund_rows:
                        fname = f["name"]
                        alias = (fund_memberships.get(fname)
                                 or fund_memberships.get(fname.upper())
                                 or fund_memberships.get(fname.lower())
                                 or None)
                        try:
                            cur.execute("""
                                SELECT l.legal_name, l.primary_email,
                                       COALESCE(c.commitment, 0) AS commitment_amount
                                  FROM lps l
                                  LEFT JOIN (
                                      SELECT lp_id, SUM(commitment_amount) AS commitment
                                        FROM commitments
                                       WHERE superseded_by IS NULL
                                       GROUP BY lp_id
                                  ) c ON c.lp_id = l.id
                                 WHERE l.fund_id = %s
                                   AND LOWER(TRIM(l.legal_name)) = LOWER(%s)
                            """, (f["id"], alias or ""))
                            lp_rows_by_fid[str(f["id"])] = cur.fetchall()
                        except Exception:
                            conn.rollback()
                            lp_rows_by_fid[str(f["id"])] = []

            # (ytd_cache_by_fid was already loaded above before the symbol pre-warm)

            # ── Assemble funds ────────────────────────────────────────
            for f in fund_rows:
                fid   = str(f["id"])
                fname = f["name"]
                alias = (fund_memberships.get(fname)
                         or fund_memberships.get(fname.upper())
                         or fund_memberships.get(fname.lower())
                         or None) if role != "gp" else None

                lp_rows            = lp_rows_by_fid.get(fid, [])
                commitment         = sum((float(r.get("commitment_amount") or 0) for r in lp_rows), 0.0)
                snap               = nav_snap_by_fid.get(fid)
                total_committed_fund = total_committed_by_fid.get(fid, 0.0)
                market_nav_fund    = mkt_nav_by_fid.get(fid, 0.0)
                fund_nav_val       = float(snap["net_nav"]) if snap and snap.get("net_nav") is not None else None
                effective_nav      = fund_nav_val or (market_nav_fund if market_nav_fund > 0 else None)

                gp_accrued_carry = 0.0
                last_wf = annual_snap_by_fid.get(fid)
                if last_wf and effective_nav:
                    last_gp_eq   = float(last_wf["gp_equity_end"] or 0)
                    last_end_nav = float(last_wf["end_nav"] or 0)
                    if last_end_nav > 0:
                        gp_accrued_carry = (last_gp_eq / last_end_nav) * effective_nav

                lp_nav_after_carry = max(0.0, (effective_nav or 0) - gp_accrued_carry)
                stake_value = None
                if lp_nav_after_carry > 0 and total_committed_fund > 0 and commitment > 0:
                    stake_value = round(commitment / total_committed_fund * lp_nav_after_carry, 2)

                out["funds"].append({
                    "fund_id":            fid,
                    "fund_name":          fname,
                    "short_name":         f["short_name"],
                    "lp_alias":           alias if role != "gp" else None,
                    "lp_count":           len(lp_rows),
                    "commitment":         commitment,
                    "total_committed":    total_committed_fund,
                    "fund_nav":           fund_nav_val,
                    "market_nav":         market_nav_fund if market_nav_fund > 0 else None,
                    "effective_nav":      effective_nav,
                    "gp_accrued_carry":   round(gp_accrued_carry, 2) if gp_accrued_carry else None,
                    "lp_nav_after_carry": round(lp_nav_after_carry, 2) if lp_nav_after_carry else None,
                    "stake_value":        stake_value,
                    "fund_nav_as_of":     snap["as_of_date"].isoformat() if snap and snap.get("as_of_date") else None,
                })

            # ── Assemble managed accounts ─────────────────────────────
            for a in acct_rows:
                fid             = str(a["id"])
                snap            = nav_snap_by_fid.get(fid)
                acct_market_nav = mkt_nav_by_fid.get(fid, 0.0)
                ytd_row         = ytd_cache_by_fid.get(fid)

                acct_ytd_pct = None; acct_ytd_upd = None
                ytd_beg_balance = None; ytd_total_deps = None; ytd_total_wdrs = None

                if ytd_row:
                    acct_ytd_pct = float(ytd_row["ytd_pct"] or 0) or None
                    if ytd_row.get("result_json"):
                        try:
                            _rj = json.loads(ytd_row["result_json"])
                            _md = _rj.get("md_return_pct")
                            if not acct_ytd_pct and _md is not None:
                                acct_ytd_pct = float(_md) or None
                            ytd_beg_balance = _rj.get("ytd_beg_balance")
                            ytd_total_deps  = _rj.get("ytd_total_deposits")
                            ytd_total_wdrs  = _rj.get("ytd_total_withdrawals")
                            if ytd_beg_balance is None:
                                _mc = (_rj.get("monthly_chart") or {}).get("monthly") or []
                                if _mc:
                                    ytd_beg_balance = float(_mc[0].get("beg_balance") or 0) or None
                                    ytd_total_deps  = round(sum(float(m.get("perf_detail",{}).get("deposits",0)) for m in _mc), 2)
                                    ytd_total_wdrs  = round(sum(float(m.get("perf_detail",{}).get("withdrawals",0)) for m in _mc), 2)
                        except Exception:
                            pass
                    # Last resort: pre-fetched bulk balance history (no extra query)
                    if ytd_beg_balance is None:
                        _bh = acct_bal_hist_by_fid.get(fid)
                        if _bh:
                            ytd_beg_balance = _bh["beg"] or None
                            ytd_total_deps  = _bh["deps"]
                            ytd_total_wdrs  = _bh["wdrs"]
                    acct_ytd_upd = ytd_row["updated_at"].isoformat()[:10] if ytd_row.get("updated_at") else None

                snap_nav           = float(snap["net_nav"]) if snap and snap.get("net_nav") is not None else None
                effective_acct_nav = snap_nav or (acct_market_nav if acct_market_nav > 0 else None)
                if effective_acct_nav is None and ytd_row and float(ytd_row.get("nav") or 0) > 0:
                    effective_acct_nav = float(ytd_row["nav"])

                out["managed_accounts"].append({
                    "fund_id":               fid,
                    "account_name":          a["name"],
                    "short_name":            a["short_name"],
                    "nav":                   effective_acct_nav,
                    "market_nav":            acct_market_nav if acct_market_nav > 0 else None,
                    "nav_as_of":             snap["as_of_date"].isoformat() if snap and snap.get("as_of_date") else None,
                    "ytd_pct":               acct_ytd_pct,
                    "ytd_as_of":             acct_ytd_upd,
                    "ytd_beg_balance":       ytd_beg_balance,
                    "ytd_total_deposits":    ytd_total_deps,
                    "ytd_total_withdrawals": ytd_total_wdrs,
                })

    except HTTPException:
        raise
    except Exception as exc:
        out["warnings"].append(f"DB query failed: {str(exc)[:200]}")

    return out


# ---------------------------------------------------------------------------
# v2 GP — NAV snapshot entry (simple manual override for period-end valuation)
# ---------------------------------------------------------------------------

class NavSnapshotRequest(BaseModel):
    fund_id:    str
    net_nav:    float          # total fund / account net asset value
    as_of_date: str            # ISO date string, e.g. "2026-03-31"
    period_kind: str = "monthly"   # 'monthly' | 'quarterly' | 'annual'

@app.post("/api/v2/gp/nav")
def gp_add_nav_snapshot(request: Request, body: NavSnapshotRequest):
    """GP-only: upsert a NAV snapshot for any fund or managed account.

    Inserts a minimal nav_snapshots row (net_nav, as_of_date). All component
    fields (cash, securities_mv, etc.) are set to 0 — they can be refined via
    the full import flow later. Uses ON CONFLICT to overwrite any existing
    snapshot for the same (fund_id, as_of_date, period_kind).
    """
    claims = _claims_or_401(request)
    if claims.get("role") != "gp":
        raise HTTPException(403, "GP access required")

    if not _PSYCOPG2_OK:
        raise HTTPException(503, "psycopg2 not available")

    period = body.period_kind if body.period_kind in ("monthly", "quarterly", "annual") else "monthly"

    try:
        with _fund_conn() as conn, conn.cursor(cursor_factory=_RealDictCursor) as cur:
            # Validate fund_id exists
            cur.execute("SELECT id, name, fund_type FROM funds WHERE id = %s", (body.fund_id,))
            fund = cur.fetchone()
            if not fund:
                raise HTTPException(404, f"Fund {body.fund_id} not found")

            net_nav = float(body.net_nav)
            cur.execute("""
                INSERT INTO nav_snapshots
                    (fund_id, as_of_date, period_kind,
                     cash, securities_mv, accrued_income, accrued_expense,
                     accrued_mgmt_fee, accrued_carry,
                     gross_nav, net_nav, status)
                VALUES (%s, %s, %s,
                        0, 0, 0, 0, 0, 0,
                        %s, %s, 'final')
                ON CONFLICT (fund_id, as_of_date, period_kind)
                WHERE restates_id IS NULL
                DO UPDATE SET
                    net_nav    = EXCLUDED.net_nav,
                    gross_nav  = EXCLUDED.gross_nav,
                    status     = 'final'
            """, (body.fund_id, body.as_of_date, period, net_nav, net_nav))
            conn.commit()

        return {
            "ok":       True,
            "fund_id":  body.fund_id,
            "fund_name": fund["name"],
            "net_nav":  net_nav,
            "as_of_date": body.as_of_date,
        }
    except HTTPException:
        raise
    except Exception as exc:
        raise HTTPException(500, f"DB error: {str(exc)[:200]}")


@app.get("/api/v2/gp/fund/{fund_id}/detail")
def gp_fund_detail(fund_id: str, request: Request):
    """GP-only: return full detail for a single fund or managed account.

    Returns fund metadata, LP roster with commitments, and most recent
    NAV snapshot.  Used by the Fund tab drill-down modal.
    """
    claims = _claims_or_401(request)
    if claims.get("role") != "gp":
        raise HTTPException(403, "GP access required")
    if not _PSYCOPG2_OK:
        raise HTTPException(503, "psycopg2 not available")
    try:
        with _fund_conn() as conn, conn.cursor(cursor_factory=_RealDictCursor) as cur:
            cur.execute("SELECT id, name, short_name, fund_type, inception_date, status, mgmt_fee_pct, carry_pct, hurdle_pct FROM funds WHERE id = %s", (fund_id,))
            fund = cur.fetchone()
            if not fund:
                raise HTTPException(404, f"Fund {fund_id} not found")

            # LP roster
            cur.execute("""
                SELECT l.id, l.legal_name, l.primary_email,
                       COALESCE(c.commitment, 0) AS commitment_amount
                  FROM lps l
                  LEFT JOIN (
                      SELECT lp_id, SUM(commitment_amount) AS commitment
                        FROM commitments WHERE superseded_by IS NULL GROUP BY lp_id
                  ) c ON c.lp_id = l.id
                 WHERE l.fund_id = %s ORDER BY l.legal_name
            """, (fund_id,))
            lps = [dict(r) for r in cur.fetchall()]
            for lp in lps:
                lp["commitment_amount"] = float(lp["commitment_amount"] or 0)

            # Most recent NAV
            cur.execute("""
                SELECT net_nav, as_of_date FROM nav_snapshots
                 WHERE fund_id = %s ORDER BY as_of_date DESC LIMIT 1
            """, (fund_id,))
            snap = cur.fetchone()

            # YTD cache for managed accounts
            ytd_pct_val = None
            if fund["fund_type"] == "managed_account":
                try:
                    cur.execute("""
                        SELECT nav, ytd_pct, result_json, updated_at
                          FROM managed_account_ytd_cache
                         WHERE fund_id = %s
                    """, (fund_id,))
                    ytd_row = cur.fetchone()
                    if ytd_row:
                        ytd_pct_val = float(ytd_row["ytd_pct"] or 0) or None
                        # Fallback: read md_return_pct from result_json when ytd_pct=0
                        if ytd_pct_val is None and ytd_row.get("result_json"):
                            try:
                                _rj = json.loads(ytd_row["result_json"])
                                _md = _rj.get("md_return_pct")
                                if _md is not None:
                                    ytd_pct_val = float(_md) or None
                            except Exception:
                                pass
                except Exception:
                    pass

            # Live market NAV from positions (sum of qty × live price)
            market_nav = _fund_market_nav(cur, fund_id)
            snap_nav   = float(snap["net_nav"]) if snap and snap["net_nav"] is not None else None
            effective_nav = snap_nav or (market_nav if market_nav > 0 else None)

            return {
                "fund_id":        fund_id,
                "fund_name":      fund["name"],
                "short_name":     fund["short_name"],
                "fund_type":      fund["fund_type"],
                "inception_date": str(fund["inception_date"]) if fund["inception_date"] else None,
                "status":         fund["status"],
                "mgmt_fee_pct":   float(fund["mgmt_fee_pct"] or 0),
                "carry_pct":      float(fund["carry_pct"] or 0),
                "hurdle_pct":     float(fund["hurdle_pct"] or 0) if fund["hurdle_pct"] is not None else None,
                "lps":            lps,
                "lp_count":       len(lps),
                "total_committed": sum(lp["commitment_amount"] for lp in lps),
                "nav":            effective_nav,
                "market_nav":     market_nav if market_nav > 0 else None,
                "nav_as_of":      snap["as_of_date"].isoformat() if snap and snap["as_of_date"] else None,
                "ytd_pct":        ytd_pct_val,
            }
    except HTTPException:
        raise
    except Exception as exc:
        raise HTTPException(500, f"DB error: {str(exc)[:200]}")


class CreateFundV2Request(BaseModel):
    name:           str
    short_name:     str
    fund_type:      str = "lp_fund"       # 'lp_fund' | 'managed_account'
    inception_date: str = ""              # YYYY-MM-DD
    mgmt_fee_pct:   float = 0.015         # 1.5%
    carry_pct:      float = 0.20          # 20%
    hurdle_pct:     float = 0.08          # 8%

@app.post("/api/v2/gp/fund/create")
def gp_fund_create(request: Request, body: CreateFundV2Request):
    """GP-only: create a new fund or managed account via v2 auth.

    Full creation with all required fields. Idempotent on short_name.
    """
    claims = _claims_or_401(request)
    if claims.get("role") != "gp":
        raise HTTPException(403, "GP access required")
    if not _PSYCOPG2_OK:
        raise HTTPException(503, "psycopg2 not available")

    ftype = body.fund_type if body.fund_type in ("lp_fund", "managed_account") else "lp_fund"
    inc = (body.inception_date or "").strip()
    if not inc:
        inc = str(__import__("datetime").date.today())
    elif re.fullmatch(r'\d{4}', inc):
        inc = f"{inc}-01-01"

    try:
        inc_year = inc.split("-")[0]
        fye = f"{inc_year}-12-31"
    except Exception:
        fye = "2020-12-31"

    structure = "separately_managed" if ftype == "managed_account" else "3c1"
    sn = body.short_name.strip().upper()

    try:
        with _fund_conn() as conn, conn.cursor(cursor_factory=_RealDictCursor) as cur:
            cur.execute("""
                INSERT INTO funds (
                    name, short_name, fund_type, inception_date, fiscal_year_end,
                    structure, domicile, base_ccy,
                    mgmt_fee_pct, mgmt_fee_basis, mgmt_fee_freq,
                    carry_pct, hurdle_pct, status
                ) VALUES (
                    %s, %s, %s, %s, %s,
                    %s, 'DE', 'USD',
                    %s, 'nav', 'quarterly',
                    %s, %s, 'open'
                )
                ON CONFLICT (short_name) DO UPDATE
                    SET name      = EXCLUDED.name,
                        fund_type = EXCLUDED.fund_type,
                        updated_at = NOW()
                RETURNING id, (xmax = 0) AS inserted
            """, (body.name.strip(), sn, ftype, inc, fye, structure,
                  body.mgmt_fee_pct, body.carry_pct, body.hurdle_pct))
            row = cur.fetchone()
            fid = str(row["id"])
            _seed_coa_for_fund(cur, fid)
            conn.commit()
            return {
                "ok":       True,
                "created":  bool(row["inserted"]),
                "fund_id":  fid,
                "fund_name": body.name,
                "fund_type": ftype,
            }
    except HTTPException:
        raise
    except Exception as exc:
        raise HTTPException(500, f"DB error: {str(exc)[:200]}")


class EnsureFundRequest(BaseModel):
    name:       str              # full fund/account name
    short_name: str              # 2-8 char display code
    fund_type:  str = "managed_account"  # 'lp_fund' | 'managed_account'
    inception_date: str = ""     # YYYY-MM-DD; defaults to today

@app.post("/api/v2/gp/fund/ensure")
def gp_ensure_fund(request: Request, body: EnsureFundRequest):
    """GP-only: create a fund or managed account by name if it doesn't exist.

    Idempotent — if a record with that name already exists it is returned as-is.
    Use this to seed LP fund & managed-account records that LPs need to see.
    """
    claims = _claims_or_401(request)
    if claims.get("role") != "gp":
        raise HTTPException(403, "GP access required")
    if not _PSYCOPG2_OK:
        raise HTTPException(503, "psycopg2 not available")

    ftype = body.fund_type if body.fund_type in ("lp_fund", "managed_account") else "managed_account"
    inc = (body.inception_date or "").strip() or str(__import__("datetime").date.today())

    try:
        with _fund_conn() as conn, conn.cursor(cursor_factory=_RealDictCursor) as cur:
            # Check if a record with this name or short_name already exists (case-insensitive)
            sn_upper = body.short_name.strip().upper()
            cur.execute("""
                SELECT id, name, short_name, fund_type FROM funds
                 WHERE LOWER(name) = LOWER(%s) OR short_name = %s
                 LIMIT 1
            """, (body.name, sn_upper))
            existing = cur.fetchone()
            if existing:
                return {"ok": True, "created": False, "fund_id": str(existing["id"]),
                        "fund_name": existing["name"], "fund_type": existing["fund_type"]}

            # Determine fiscal year end from inception date
            try:
                inc_year = inc.split("-")[0]
                fye = f"{inc_year}-12-31"
            except Exception:
                fye = "2020-12-31"

            # Sensible defaults: SMA structure for managed accounts, 3c1 for LP funds
            structure = "separately_managed" if ftype == "managed_account" else "3c1"

            cur.execute("""
                INSERT INTO funds (
                    name, short_name, fund_type, inception_date, fiscal_year_end,
                    structure, domicile, base_ccy,
                    mgmt_fee_pct, mgmt_fee_basis, mgmt_fee_freq,
                    carry_pct, hurdle_pct, status
                )
                VALUES (
                    %s, %s, %s, %s, %s,
                    %s, 'DE', 'USD',
                    0.0150, 'nav', 'quarterly',
                    0.20, 0.08, 'open'
                )
                RETURNING id
            """, (body.name.strip(), sn_upper, ftype, inc, fye, structure))
            new_id = str(cur.fetchone()["id"])
            conn.commit()
            return {"ok": True, "created": True, "fund_id": new_id,
                    "fund_name": body.name, "fund_type": ftype}
    except HTTPException:
        raise
    except Exception as exc:
        raise HTTPException(500, f"DB error: {str(exc)[:200]}")


# ---------------------------------------------------------------------------
# v2 Admin — DESTRUCTIVE ops. Always GP-only, always with backup-first
# safety. Used for one-off cleanup of duplicate / misclassified fund rows.
# ---------------------------------------------------------------------------

class FundDeleteRequest(BaseModel):
    fund_id: str
    confirm: bool = False   # must be True to actually delete
    force:   bool = False   # bypass non-trivial-data safety guards (GP explicit override)


@app.post("/api/v2/admin/fund/delete")
def admin_fund_delete(request: Request, body: FundDeleteRequest):
    """Delete a single fund row + its child rows. GUARDED:
      1. Only GP role can call this.
      2. Dry-run by default (confirm=false) — returns what WOULD be deleted.
      3. Refuses to delete if the fund has non-trivial data:
         • any commitment with amount > 0
         • any non-zero nav_snapshot
         • any capital_call, distribution, transaction, or tax_lot
      4. Writes a JSON backup of the entire fund + its rows to a file
         before the DELETE runs, so you can hand-restore if needed.
      5. Runs DELETE in a transaction — full rollback on any error.
    """
    claims = _claims_or_401(request)
    if claims.get("role") != "gp":
        raise HTTPException(status_code=403, detail="GP role required")
    if not _PSYCOPG2_OK:
        raise HTTPException(status_code=503, detail="psycopg2 not installed")

    fund_id = (body.fund_id or "").strip()
    if not fund_id:
        raise HTTPException(status_code=400, detail="fund_id required")

    inventory:  dict[str, int]  = {}
    backup:     dict[str, list] = {}
    refuse_reasons: list[str]   = []

    conn = _fund_conn()
    try:
        # Phase 1 (inspection) runs in autocommit so a missing-table error
        # doesn't poison the whole transaction. We flip back to false before
        # the destructive DELETEs.
        conn.autocommit = True
        with conn.cursor(cursor_factory=_RealDictCursor) as cur:
            # Discover which tables actually exist in this DB
            cur.execute("""
                SELECT table_name
                  FROM information_schema.tables
                 WHERE table_schema = 'public'
            """)
            existing_tables = {r["table_name"] for r in cur.fetchall()}

            # 1) Load the fund row itself
            cur.execute("SELECT * FROM funds WHERE id = %s", (fund_id,))
            fund_row = cur.fetchone()
            if not fund_row:
                raise HTTPException(status_code=404, detail=f"fund_id {fund_id} not found")
            backup["funds"] = [dict(fund_row)]

            # 2) Inventory every child table that references funds.id
            child_tables = [
                ("lps",                  "fund_id"),
                ("commitments",          "fund_id"),
                ("capital_calls",        "fund_id"),
                ("distributions",        "fund_id"),
                ("nav_snapshots",        "fund_id"),
                ("transactions",         "fund_id"),
                ("tax_lots",             "fund_id"),
                ("mgmt_fee_runs",        "fund_id"),
                ("carry_runs",           "fund_id"),
                ("lp_statements",        "fund_id"),
                ("annual_lp_balances",   "fund_id"),
                ("managed_account_ytd_cache", "fund_id"),
                ("accounts",             "fund_id"),   # chart-of-accounts rows seeded per fund
            ]
            for tbl, col in child_tables:
                if tbl not in existing_tables:
                    inventory[tbl] = 0
                    continue
                cur.execute(f"SELECT * FROM {tbl} WHERE {col} = %s", (fund_id,))
                rows = cur.fetchall()
                inventory[tbl] = len(rows)
                if rows:
                    backup[tbl] = [{k: (str(v) if hasattr(v, 'isoformat') else v)
                                    for k, v in r.items()} for r in rows]

            # transaction_lines has no fund_id — count via transactions + accounts join
            if "transaction_lines" in existing_tables and "transactions" in existing_tables:
                cur.execute("""
                    SELECT COUNT(*)::int AS n FROM transaction_lines
                     WHERE transaction_id IN (SELECT id FROM transactions WHERE fund_id = %s)
                """, (fund_id,))
                inventory["transaction_lines"] = cur.fetchone()["n"]
            else:
                inventory["transaction_lines"] = 0

            # 3) Safety guards
            if "commitments" in existing_tables:
                cur.execute("""
                    SELECT COUNT(*)::int AS n
                      FROM commitments
                     WHERE fund_id = %s AND COALESCE(commitment_amount, 0) > 0
                       AND superseded_by IS NULL
                """, (fund_id,))
                n_real_commits = cur.fetchone()["n"]
                if n_real_commits > 0:
                    refuse_reasons.append(f"{n_real_commits} active commitment(s) with amount > 0")

            if "nav_snapshots" in existing_tables:
                cur.execute("""
                    SELECT COUNT(*)::int AS n
                      FROM nav_snapshots
                     WHERE fund_id = %s
                       AND (COALESCE(net_nav, 0) > 0 OR COALESCE(gross_nav, 0) > 0)
                """, (fund_id,))
                n_real_navs = cur.fetchone()["n"]
                if n_real_navs > 0:
                    refuse_reasons.append(f"{n_real_navs} nav_snapshot(s) with non-zero NAV")

            if "transactions" in existing_tables:
                cur.execute("SELECT COUNT(*)::int AS n FROM transactions WHERE fund_id = %s", (fund_id,))
                n_txns = cur.fetchone()["n"]
                if n_txns > 0:
                    refuse_reasons.append(f"{n_txns} transaction(s) in immutable ledger")

            if "capital_calls" in existing_tables:
                cur.execute("SELECT COUNT(*)::int AS n FROM capital_calls WHERE fund_id = %s", (fund_id,))
                n_calls = cur.fetchone()["n"]
                if n_calls > 0:
                    refuse_reasons.append(f"{n_calls} capital_call(s)")

            if "distributions" in existing_tables:
                cur.execute("SELECT COUNT(*)::int AS n FROM distributions WHERE fund_id = %s", (fund_id,))
                n_dist = cur.fetchone()["n"]
                if n_dist > 0:
                    refuse_reasons.append(f"{n_dist} distribution(s)")

            if refuse_reasons and not body.force:
                return {
                    "ok":              False,
                    "would_delete":    False,
                    "fund":            dict(fund_row),
                    "child_inventory": inventory,
                    "refuse_reasons":  refuse_reasons,
                    "message":         "Refusing to delete fund with non-trivial data. "
                                       "Pass force=true to override (GP explicit).",
                }

            # 4) Dry-run mode → return what would happen
            if not body.confirm:
                return {
                    "ok":              True,
                    "would_delete":    True,
                    "fund":            dict(fund_row),
                    "child_inventory": inventory,
                    "next_step":       "Call again with confirm=true to perform the delete.",
                }

            # 5) Write backup to disk BEFORE any destructive op
            backups_dir = Path(os.environ.get(
                "DGA_FUND_DELETION_BACKUPS",
                str(analyst.STOCKS_FOLDER / "_fund_deletion_backups"),
            ))
            backups_dir.mkdir(parents=True, exist_ok=True)
            stamp = datetime.utcnow().strftime("%Y%m%dT%H%M%SZ")
            short = (fund_row.get("short_name") or fund_id).replace("/", "_").replace(" ", "_")
            backup_path = backups_dir / f"{stamp}__{short}__{fund_id}.json"
            backup_path.write_text(json.dumps({
                "deleted_at":  stamp,
                "deleted_by":  claims.get("email"),
                "lp_id":       claims.get("lp_id"),
                "inventory":   inventory,
                "data":        backup,
            }, indent=2, default=str))

            # 6) Switch to transactional mode for the destructive ops.
            conn.autocommit = False

            # We only DELETE from tables that actually exist (existing_tables
            # set built above). Helper:
            def _delete(cur, sql, params, key):
                cur.execute(sql, params)
                deleted[key] = cur.rowcount

            deleted: dict[str, int] = {}
            if "nav_snapshot_lp" in existing_tables:
                _delete(cur, """
                    DELETE FROM nav_snapshot_lp
                     WHERE nav_snapshot_id IN (
                         SELECT id FROM nav_snapshots WHERE fund_id = %s
                     )
                """, (fund_id,), "nav_snapshot_lp")
            if "capital_call_allocations" in existing_tables:
                _delete(cur, """
                    DELETE FROM capital_call_allocations
                     WHERE capital_call_id IN (
                         SELECT id FROM capital_calls WHERE fund_id = %s
                     )
                """, (fund_id,), "capital_call_allocations")
            if "distribution_allocations" in existing_tables:
                _delete(cur, """
                    DELETE FROM distribution_allocations
                     WHERE distribution_id IN (
                         SELECT id FROM distributions WHERE fund_id = %s
                     )
                """, (fund_id,), "distribution_allocations")
            if "commitments" in existing_tables:
                _delete(cur, """
                    DELETE FROM commitments
                     WHERE lp_id IN (SELECT id FROM lps WHERE fund_id = %s)
                        OR fund_id = %s
                """, (fund_id, fund_id), "commitments")
            if "users" in existing_tables and "lps" in existing_tables:
                _delete(cur, """
                    UPDATE users SET lp_id = NULL
                     WHERE lp_id IN (SELECT id FROM lps WHERE fund_id = %s)
                """, (fund_id,), "users_lp_id_cleared")
            if "lps" in existing_tables:
                _delete(cur, "DELETE FROM lps WHERE fund_id = %s",
                        (fund_id,), "lps")
            # transaction_lines refs accounts.id + transactions.id — must go first
            if "transaction_lines" in existing_tables and "transactions" in existing_tables:
                _delete(cur, """
                    DELETE FROM transaction_lines
                     WHERE transaction_id IN (
                         SELECT id FROM transactions WHERE fund_id = %s
                     )
                """, (fund_id,), "transaction_lines")
            for tbl in ("annual_lp_balances", "managed_account_ytd_cache",
                        "lp_statements", "carry_runs", "mgmt_fee_runs",
                        "accounts"):
                if tbl in existing_tables:
                    _delete(cur, f"DELETE FROM {tbl} WHERE fund_id = %s",
                            (fund_id,), tbl)
            if "nav_snapshots" in existing_tables:
                _delete(cur, "DELETE FROM nav_snapshots WHERE fund_id = %s",
                        (fund_id,), "nav_snapshots")
            if "capital_calls" in existing_tables:
                _delete(cur, "DELETE FROM capital_calls WHERE fund_id = %s",
                        (fund_id,), "capital_calls")
            if "distributions" in existing_tables:
                _delete(cur, "DELETE FROM distributions WHERE fund_id = %s",
                        (fund_id,), "distributions")
            if "tax_lots" in existing_tables:
                _delete(cur, "DELETE FROM tax_lots WHERE fund_id = %s",
                        (fund_id,), "tax_lots")
            if "transactions" in existing_tables:
                _delete(cur, "DELETE FROM transactions WHERE fund_id = %s",
                        (fund_id,), "transactions")

            # Finally the fund row itself
            _delete(cur, "DELETE FROM funds WHERE id = %s",
                    (fund_id,), "funds")

            conn.commit()

            return {
                "ok":              True,
                "deleted":         True,
                "fund_short_name": fund_row.get("short_name"),
                "fund_name":       fund_row.get("name"),
                "rows_deleted":    deleted,
                "backup_path":     str(backup_path),
            }
    except HTTPException:
        try: conn.rollback()
        except: pass
        raise
    except Exception as exc:
        try: conn.rollback()
        except: pass
        raise HTTPException(status_code=500, detail=f"Delete failed (rolled back): {str(exc)[:300]}")
    finally:
        try: conn.close()
        except: pass


# ---------------------------------------------------------------------------
# Background worker
# ---------------------------------------------------------------------------

def _run_analysis(job_id: str, ticker: str, generate_gamma: bool) -> None:
    with _jobs_lock:
        _jobs[job_id]["status"] = "running"
        _jobs[job_id]["progress"] = {"step": "queued", "pct": 0.0, "label": "Starting…"}

    # Progress callback — runs on the worker thread, mutates the shared job dict.
    # Wrapped so a slow lock acquisition can't slow down the analysis itself.
    def _record_progress(step: str, pct: float, label: str) -> None:
        with _jobs_lock:
            if job_id in _jobs:
                _jobs[job_id]["progress"] = {
                    "step": step, "pct": pct, "label": label,
                }

    try:
        system_prompt = analyst.load_system_prompt()
        result = analyst.analyze_ticker(
            ticker,
            system_prompt=system_prompt,
            generate_gamma=generate_gamma,
            verbose=False,
            on_progress=_record_progress,
        )
        with _jobs_lock:
            if result.get("ok"):
                _jobs[job_id]["status"] = "done"
                _jobs[job_id]["progress"] = {"step": "done", "pct": 1.0, "label": "Report ready"}
                # Trim the report text to avoid sending multi-MB payloads in the
                # status response; the full text is available via /report/{ticker}.
                _jobs[job_id]["result"] = {k: v for k, v in result.items()
                                           if k != "report_text"}
                _jobs[job_id]["result"]["has_report"] = bool(result.get("report_text"))
            else:
                _jobs[job_id]["status"] = "failed"
                _jobs[job_id]["error"] = result.get("error", "Unknown error")
    except BaseException as exc:  # noqa: BLE001  # catches SystemExit from any library
        # Log FULL traceback so we can see where the error actually happened.
        tb_str = traceback.format_exc()
        print(f"\n❌ Single-ticker job {job_id} ({ticker}) CRASHED:\n{tb_str}", flush=True)
        with _jobs_lock:
            _jobs[job_id]["status"] = "failed"
            # Include last line of traceback in error so the UI shows something
            # more useful than just the message.
            tb_tail = tb_str.strip().splitlines()[-3:] if tb_str else []
            _jobs[job_id]["error"] = f"{exc} | {' | '.join(tb_tail)}"


# In-memory portfolio job store.
_pjobs: dict[str, dict[str, Any]] = {}
_pjobs_lock = threading.Lock()

# In-memory scan job store.
_sjobs: dict[str, dict[str, Any]] = {}
_sjobs_lock = threading.Lock()

# In-memory intelligence job store.
_ijobs: dict[str, dict[str, Any]] = {}
_ijobs_lock = threading.Lock()

# In-memory daily-brief job store (Goldman-style PM morning notes).
_bjobs: dict[str, dict[str, Any]] = {}
_bjobs_lock = threading.Lock()


def _run_portfolio(
    job_id: str,
    portfolio_records: list[dict],
    strategy: str,
    generate_gamma: bool,
    reuse_existing: bool,
    xlsx_out_path: str,
) -> None:
    with _pjobs_lock:
        _pjobs[job_id]["status"] = "running"
        _pjobs[job_id]["progress"] = {
            "step": "queued", "pct": 0.0,
            "label": "Starting portfolio run…",
            "ticker_index": 0, "ticker_total": 0,
            "current_ticker": None,
            "ok": [], "failed": [],
        }

    # Progress callback — runs on the worker thread, mutates the shared
    # job dict. The portfolio pipeline emits per-ticker progress so the
    # frontend can render a "3 / 12 — analyzing AAPL" counter.
    def _record_progress(step: str, pct: float, label: str, extra: dict) -> None:
        with _pjobs_lock:
            if job_id in _pjobs:
                _pjobs[job_id]["progress"] = {
                    "step": step,
                    "pct": pct,
                    "label": label,
                    "ticker_index": extra.get("ticker_index", 0),
                    "ticker_total": extra.get("ticker_total", 0),
                    "current_ticker": extra.get("ticker"),
                    "ok": list(extra.get("ok", []) or []),
                    "failed": list(extra.get("failed", []) or []),
                }

    try:
        result = analyst.run_portfolio_rebalance(
            portfolio_records=portfolio_records,
            primary_strategy=strategy,
            generate_gamma=generate_gamma,
            reuse_existing=reuse_existing,
            output_path=xlsx_out_path,
            on_progress=_record_progress,
        )
        with _pjobs_lock:
            _pjobs[job_id]["status"] = "done" if result.get("ok") else "failed"
            _pjobs[job_id]["result"] = result
            if not result.get("ok"):
                _pjobs[job_id]["error"] = "No tickers could be analyzed."
            # Mark progress as done so the UI doesn't keep showing "Analyzing X"
            # after the result lands.
            _pjobs[job_id]["progress"] = {
                **(_pjobs[job_id].get("progress") or {}),
                "step": "done" if result.get("ok") else "failed",
                "pct": 1.0,
                "label": (f"Done — {len(result.get('tickers_ok') or [])} ok, "
                          f"{len(result.get('tickers_failed') or [])} failed")
                          if result.get("ok") else "Failed",
            }
            # ── Persist lean payload so web + mobile share the same last run ──
            if result.get("ok"):
                try:
                    lean_result = {k: v for k, v in result.items()
                                   if k != "xlsx_path"}   # strip local server path
                    last_job_payload = {
                        "job_id":       job_id,
                        "n_tickers":    _pjobs[job_id]["n_tickers"],
                        "strategy":     strategy,
                        "completed_at": datetime.utcnow().isoformat() + "Z",
                        "result":       lean_result,
                        "input_weights": _pjobs[job_id].get("input_weights") or {},
                    }
                    _LAST_JOB_PATH.write_text(json.dumps(last_job_payload, default=str))
                except Exception as _e:
                    print(f"⚠️  Could not write last-job snapshot: {_e}")
        # Auto-promote the input holdings as the new "live portfolio" benchmark
        # so the Paper Tracker can compare idea baskets against your real book.
        if result.get("ok"):
            try:
                analyst.promote_live_portfolio(portfolio_records)
            except Exception as exc:  # noqa: BLE001
                print(f"⚠️  Live-portfolio promotion failed: {exc}")
    except BaseException as exc:  # noqa: BLE001  # catches SystemExit from any library
        with _pjobs_lock:
            _pjobs[job_id]["status"] = "failed"
            _pjobs[job_id]["error"] = str(exc)


# ---------------------------------------------------------------------------
# Routes
# ---------------------------------------------------------------------------

WEB_DIR = Path(__file__).resolve().parent.parent / "web"
BRANDING_DIR = Path(__file__).resolve().parent.parent / "branding"


@app.get("/")
async def root():
    """Serve the new production login page (portfolio.dgacapital.com entry point).

    The legacy single-password app shell is still reachable at /app/ for
    backward compat — only the root route has been re-pointed.
    """
    path = WEB_DIR / "portfolio.html"
    if path.exists():
        return FileResponse(str(path), headers=_NOCACHE)
    # Fallback if portfolio.html missing (shouldn't happen post-deploy)
    return RedirectResponse(url="/app/")


@app.get("/gp")
async def serve_gp_dashboard():
    """GP dashboard (Terminal Pro). Client-side auth-guard.js redirects
    unauthenticated requests to /. Server doesn't enforce here — the
    page is just static HTML; all sensitive data goes through /api/*
    which has its own token check."""
    path = WEB_DIR / "portfolio-gp.html"
    if not path.exists():
        raise HTTPException(status_code=404, detail="portfolio-gp.html not found")
    return FileResponse(str(path), headers=_NOCACHE)


@app.get("/lp")
async def serve_lp_dashboard():
    """LP dashboard. Same auth-guard pattern as /gp."""
    path = WEB_DIR / "portfolio-lp.html"
    if not path.exists():
        raise HTTPException(status_code=404, detail="portfolio-lp.html not found")
    return FileResponse(str(path), headers=_NOCACHE)


# Hard-coded no-cache headers applied to every shell response.
# Using an explicit FileResponse endpoint (not the StaticFiles mount) is the
# ONLY reliable way to guarantee the browser + CDN edge never serve a stale
# index.html.  StaticFiles + middleware is not guaranteed to override Railway's
# edge layer; a dedicated route with headers= on the FileResponse is.
_NOCACHE = {
    "Cache-Control":              "no-cache, no-store, must-revalidate",
    "Pragma":                     "no-cache",
    "Expires":                    "0",
    "Surrogate-Control":          "no-store",
    "CDN-Cache-Control":          "no-store",
    "Cloudflare-CDN-Cache-Control": "no-store",
}


@app.get("/app/")
@app.get("/app/index.html")
async def serve_shell():
    """Serve the web app shell with strict no-cache headers.

    Defined as explicit routes so they take precedence over the
    StaticFiles mount — guarantees Cache-Control is set on every
    response regardless of CDN/edge behaviour.
    """
    path = WEB_DIR / "index.html"
    if not path.exists():
        raise HTTPException(status_code=404, detail="index.html not found")
    return FileResponse(str(path), headers=_NOCACHE)


@app.get("/info")
def info():
    return {"service": "DGA Research Analyst API", "status": "ok"}


# ── Mockup preview routes — convenience short URLs ───────────────────────────
# These serve the front-page redesign mockups from the project root so they're
# accessible at /mockup-a.html (in addition to /app/mockup-a.html via the
# StaticFiles mount). Used to preview proposed layouts before shipping.
@app.get("/mockup-a.html")
async def serve_mockup_a():
    path = WEB_DIR / "mockup-a.html"
    if not path.exists():
        raise HTTPException(status_code=404, detail="mockup-a.html not found")
    return FileResponse(str(path), headers=_NOCACHE)


@app.get("/mockup-b.html")
async def serve_mockup_b():
    path = WEB_DIR / "mockup-b.html"
    if not path.exists():
        raise HTTPException(status_code=404, detail="mockup-b.html not found")
    return FileResponse(str(path), headers=_NOCACHE)


@app.get("/mockup-hybrid.html")
@app.get("/preview")
async def serve_mockup_hybrid():
    """Unified login + role-router preview. Two-password gate (GP/LP)
    routes to the appropriate layout mockup. Used to validate the
    portfolio.dgacapital.com flow before shipping to production."""
    path = WEB_DIR / "mockup-hybrid.html"
    if not path.exists():
        raise HTTPException(status_code=404, detail="mockup-hybrid.html not found")
    return FileResponse(str(path), headers=_NOCACHE)


# ── Build/version endpoint ────────────────────────────────────────────────────
# The web client polls this to detect deploys and force a hard reload of
# stale iOS PWA / Safari caches. Bumped on every UI deploy.
WEB_BUILD_VERSION = "ui76-20260511"


@app.get("/api/build")
def build_version():
    """Return the current web UI build identifier.

    The web client compares this to its embedded BUILD constant; on mismatch
    it forces a `location.reload(true)` with a cache-bust query param so even
    home-screen PWAs see the latest UI without manual cache clearing.
    """
    return {"build": WEB_BUILD_VERSION}


@app.get("/health")
def health():
    return {"status": "ok", "timestamp": datetime.utcnow().isoformat()}


@app.get("/api/diagnostics")
def diagnostics():
    """Return non-secret config info so we can verify env vars are set
    without exposing the actual values.  Used to debug Gamma / Dropbox /
    Drive integrations on Railway."""
    def _is_set(name: str) -> bool:
        return bool((os.environ.get(name, "") or "").strip())
    return {
        "gamma_api_key_set":     _is_set("GAMMA_API_KEY"),
        "gamma_folder_id_set":   _is_set("GAMMA_FOLDER_ID"),
        "dropbox_configured":    all(_is_set(k) for k in ("DROPBOX_APP_KEY","DROPBOX_APP_SECRET","DROPBOX_REFRESH_TOKEN")),
        "google_drive_set":      _is_set("GOOGLE_SERVICE_ACCOUNT_JSON") or Path(__file__).resolve().parent.parent.joinpath("dga-service-account.json").exists(),
        "xai_api_key_set":       _is_set("XAI_API_KEY"),
        "build":                 WEB_BUILD_VERSION,
        "stocks_folder":         str(analyst.STOCKS_FOLDER),
        "stocks_pptx_count":     len(list(analyst.STOCKS_FOLDER.glob("*_DGA_Presentation.pptx"))),
        "stocks_docx_count":     len(list(analyst.STOCKS_FOLDER.glob("*_DGA_Report.docx"))),
        "stocks_md_count":       len(list(analyst.STOCKS_FOLDER.glob("*_DGA_Report.md"))),
        "gamma_index_entries":   len(analyst._load_gamma_index()),
    }


@app.post("/api/analyze", response_model=JobStatus)
def start_analysis(req: AnalyzeRequest, background_tasks: BackgroundTasks):
    """Kick off an async analysis for *ticker*. Returns a job_id to poll."""
    ticker = req.ticker.strip().upper()
    if not ticker or not ticker.isalpha() or len(ticker) > 10:
        raise HTTPException(status_code=422, detail="Invalid ticker symbol")

    job_id = str(uuid.uuid4())
    now = datetime.utcnow().isoformat()
    with _jobs_lock:
        _jobs[job_id] = {
            "job_id": job_id,
            "ticker": ticker,
            "status": "queued",
            "created_at": now,
            "error": None,
            "result": None,
            # Progress reported by the analysis pipeline. {step, pct, label}.
            # `step` is one of: queued | sec_filings | financials | market_data
            #                 | grok | rendering | gamma | upload | done
            "progress": {"step": "queued", "pct": 0.0, "label": "Queued — starting shortly…"},
        }

    # Persist mapping so we can recover after a server restart.
    _save_job_index_entry(job_id, {"ticker": ticker, "type": "analysis", "created_at": now})

    background_tasks.add_task(_run_analysis, job_id, ticker, req.generate_gamma)
    return _jobs[job_id]


@app.get("/api/jobs/{job_id}", response_model=JobStatus)
def get_job_status(job_id: str):
    """Poll for the status of a previously submitted job.

    If the job is not in memory (server restarted), we fall back to the
    on-disk job-index: if the report file already exists we return a
    synthetic "done" response so the mobile app can navigate to the report.
    If neither the in-memory job nor the report file exists we return 404.
    """
    with _jobs_lock:
        job = _jobs.get(job_id)
    if job:
        return job

    # --- Recovery path after server restart ---
    idx = _load_job_index()
    entry = idx.get(job_id)
    if entry and entry.get("type") == "analysis":
        ticker = entry.get("ticker", "").upper()
        md_path = analyst.STOCKS_FOLDER / f"{ticker}_DGA_Report.md"
        if md_path.exists():
            # Report was completed before the restart — synthesize a done response.
            has_docx = (analyst.STOCKS_FOLDER / f"{ticker}_DGA_Report.docx").exists()
            has_pptx = (analyst.STOCKS_FOLDER / f"{ticker}_DGA_Presentation.pptx").exists()
            recovered = {
                "job_id": job_id,
                "ticker": ticker,
                "status": "done",
                "created_at": entry.get("created_at", ""),
                "error": None,
                "result": {
                    "ok": True,
                    "has_report": True,
                    "has_docx": has_docx,
                    "has_pptx": has_pptx,
                    "gamma_url": None,
                    "gamma_error": None,
                    "recovered": True,   # flag so client knows it was recovered
                },
            }
            # Re-hydrate in memory so subsequent polls are fast.
            with _jobs_lock:
                _jobs[job_id] = recovered
            return recovered
        else:
            # Index entry exists but the report never finished — the job was lost.
            raise HTTPException(
                status_code=404,
                detail=f"Job was lost in a server restart before completing. Please re-run the analysis for {ticker}.",
            )

    raise HTTPException(status_code=404, detail="Job not found")


@app.get("/api/jobs")
def list_jobs():
    """Return all jobs (newest first), without full result payloads."""
    with _jobs_lock:
        jobs = list(_jobs.values())
    jobs.sort(key=lambda j: j["created_at"], reverse=True)
    return [
        {k: v for k, v in j.items() if k != "result"}
        for j in jobs
    ]


@app.get("/api/report/{ticker}")
def get_report(ticker: str):
    """Return the full markdown report text for the most-recently-analyzed ticker."""
    ticker = ticker.strip().upper()
    md_path = analyst.STOCKS_FOLDER / f"{ticker}_DGA_Report.md"
    if not md_path.exists():
        raise HTTPException(status_code=404, detail=f"No report found for {ticker}")
    folder = analyst.STOCKS_FOLDER
    has_docx = (folder / f"{ticker}_DGA_Report.docx").exists()
    has_pptx = (folder / f"{ticker}_DGA_Presentation.pptx").exists()
    # Look up Gamma URL from the on-disk index so the mobile UI can show
    # the "View Gamma" button even days/weeks after the original run.
    gamma_url = None
    gamma_generated_at = None
    try:
        idx = analyst._load_gamma_index()
        entry = idx.get(ticker)
        if entry:
            gamma_url = entry.get("gamma_url")
            gamma_generated_at = entry.get("generated_at")
    except Exception:
        pass
    return {
        "ticker": ticker,
        "report_md": md_path.read_text(),
        "generated_at": datetime.utcfromtimestamp(md_path.stat().st_mtime).isoformat(),
        "has_docx": has_docx,
        "has_pptx": has_pptx,
        "gamma_url": gamma_url,
        "gamma_generated_at": gamma_generated_at,
    }


@app.get("/api/download/{ticker}/docx")
def download_docx(ticker: str):
    """Download the Word report for *ticker*."""
    ticker = ticker.strip().upper()
    path = analyst.STOCKS_FOLDER / f"{ticker}_DGA_Report.docx"
    if not path.exists():
        raise HTTPException(status_code=404, detail="Word report not found")
    return FileResponse(
        path=str(path),
        media_type="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        filename=path.name,
    )


@app.get("/api/download/{ticker}/pptx")
def download_pptx(ticker: str):
    """Download the PowerPoint presentation for *ticker*."""
    ticker = ticker.strip().upper()
    path = analyst.STOCKS_FOLDER / f"{ticker}_DGA_Presentation.pptx"
    if not path.exists():
        raise HTTPException(status_code=404, detail="Presentation not found")
    return FileResponse(
        path=str(path),
        media_type="application/vnd.openxmlformats-officedocument.presentationml.presentation",
        filename=path.name,
    )


@app.get("/api/quote/{ticker}")
def get_quote(ticker: str):
    """Return live market price for *ticker* from Yahoo Finance."""
    ticker = ticker.strip().upper()
    snapshot = analyst.fetch_market_snapshot(ticker)
    return {"ticker": ticker, **snapshot}


# ── SPY YTD cache (TTL: 15 min) ───────────────────────────────────────────────
_spy_ytd_cache: dict = {}   # {"ytd_pct": float, "first_close": float,
                             #  "last_close": float, "as_of": str, "ts": float}
_SPY_YTD_TTL = 900          # seconds — matches _PRICE_CACHE_TTL

# ── Benchmark annual return cache ─────────────────────────────────────────────
# Keyed by (benchmark_key, year) → float (return %).  Persists for the lifetime
# of the process (changes once per year; no TTL needed).
_benchmark_annual_cache: dict = {}

_BENCHMARK_DEFS: dict = {
    "sp500":      {"label": "S&P 500",     "tickers": [("SPY",  1.00)]},
    "dow30":      {"label": "Dow 30",      "tickers": [("DIA",  1.00)]},
    "nasdaq":     {"label": "Nasdaq",      "tickers": [("QQQ",  1.00)]},
    "msci_world": {"label": "MSCI World",  "tickers": [("URTH", 1.00)]},
    "bonds":      {"label": "Bond Index",  "tickers": [("AGG",  1.00)]},
    "60_40":      {"label": "60/40 Blend", "tickers": [("SPY",  0.60), ("AGG", 0.40)]},
    "85_15":      {"label": "85/15 Blend", "tickers": [("SPY",  0.85), ("AGG", 0.15)]},
}

# ── Curated historical benchmark returns (2000-2025) ──────────────────────────
# S&P 500: price return (Dec 31 to Dec 31), matching MacroTrends methodology.
# 2016-2025 values confirmed directly from MacroTrends screenshot.
# 2000-2015 values consistent with MacroTrends price-return series.
# Other benchmarks: ETF annual price returns (DIA, QQQ, URTH, AGG)
# MSCI World 2000-2011: MSCI World Index net returns (pre-URTH)
# Bonds 2000-2003: Bloomberg US Aggregate Bond Index (pre-AGG)
# 60/40 and 85/15: computed as weighted blend of sp500 + bonds
_CURATED_BENCHMARK_RETURNS: dict[str, dict[int, tuple[float, bool, str]]] = {}

def _build_curated_benchmark_returns() -> None:
    """Compute and populate _CURATED_BENCHMARK_RETURNS once at import time."""
    sp500 = {
        # 2000-2015: MacroTrends price return (consistent methodology)
        2000: (-10.14, True, "macrotrends"), 2001: (-13.04, True, "macrotrends"),
        2002: (-23.37, True, "macrotrends"), 2003: (26.38,  True, "macrotrends"),
        2004: (8.99,   True, "macrotrends"), 2005: (3.00,   True, "macrotrends"),
        2006: (13.62,  True, "macrotrends"), 2007: (3.53,   True, "macrotrends"),
        2008: (-38.49, True, "macrotrends"), 2009: (23.45,  True, "macrotrends"),
        2010: (12.78,  True, "macrotrends"), 2011: (0.00,   True, "macrotrends"),
        2012: (13.41,  True, "macrotrends"), 2013: (29.60,  True, "macrotrends"),
        2014: (11.39,  True, "macrotrends"), 2015: (-0.73,  True, "macrotrends"),
        # 2016-2025: confirmed from MacroTrends screenshot
        2016: (9.54,   True, "macrotrends"), 2017: (19.42,  True, "macrotrends"),
        2018: (-6.24,  True, "macrotrends"), 2019: (28.88,  True, "macrotrends"),
        2020: (16.26,  True, "macrotrends"), 2021: (26.89,  True, "macrotrends"),
        2022: (-19.44, True, "macrotrends"), 2023: (24.23,  True, "macrotrends"),
        2024: (23.31,  True, "macrotrends"), 2025: (16.39,  True, "macrotrends"),
    }
    dow30 = {
        # DIA ETF price return (Dec 31 to Dec 31). Pre-DIA years use DJIA price return.
        2000: (-6.17,  False, "dia_price"), 2001: (-5.35,  False, "dia_price"),
        2002: (-14.97, False, "dia_price"), 2003: (28.28,  False, "dia_price"),
        2004: (5.32,   False, "dia_price"), 2005: (2.31,   False, "dia_price"),
        2006: (19.47,  False, "dia_price"), 2007: (8.88,   False, "dia_price"),
        2008: (-31.93, False, "dia_price"), 2009: (22.68,  False, "dia_price"),
        2010: (14.06,  False, "dia_price"), 2011: (8.38,   False, "dia_price"),
        2012: (10.24,  False, "dia_price"), 2013: (32.37,  False, "dia_price"),
        2014: (10.04,  False, "dia_price"), 2015: (0.21,   False, "dia_price"),
        2016: (16.50,  False, "dia_price"), 2017: (28.11,  False, "dia_price"),
        2018: (-3.48,  False, "dia_price"), 2019: (25.34,  False, "dia_price"),
        2020: (9.72,   False, "dia_price"), 2021: (20.95,  False, "dia_price"),
        2022: (-6.86,  False, "dia_price"), 2023: (16.18,  False, "dia_price"),
        2024: (15.51,  False, "dia_price"), 2025: (4.75,   False, "estimate"),
    }
    nasdaq = {
        # QQQ ETF price return
        2000: (-41.73, False, "qqq_price"), 2001: (-33.05, False, "qqq_price"),
        2002: (-37.30, False, "qqq_price"), 2003: (49.72,  False, "qqq_price"),
        2004: (10.46,  False, "qqq_price"), 2005: (2.40,   False, "qqq_price"),
        2006: (6.79,   False, "qqq_price"), 2007: (19.18,  False, "qqq_price"),
        2008: (-41.73, False, "qqq_price"), 2009: (54.69,  False, "qqq_price"),
        2010: (19.16,  False, "qqq_price"), 2011: (2.72,   False, "qqq_price"),
        2012: (18.12,  False, "qqq_price"), 2013: (36.63,  False, "qqq_price"),
        2014: (19.20,  False, "qqq_price"), 2015: (9.50,   False, "qqq_price"),
        2016: (7.47,   False, "qqq_price"), 2017: (32.66,  False, "qqq_price"),
        2018: (-3.92,  False, "qqq_price"), 2019: (38.79,  False, "qqq_price"),
        2020: (47.58,  False, "qqq_price"), 2021: (27.42,  False, "qqq_price"),
        2022: (-32.58, False, "qqq_price"), 2023: (54.85,  False, "qqq_price"),
        2024: (25.61,  False, "qqq_price"), 2025: (-8.39,  False, "estimate"),
    }
    msci_world = {
        # 2000-2011: MSCI World Index net total return; 2012+: URTH ETF price return
        2000: (-13.18, False, "msci_index"), 2001: (-16.52, False, "msci_index"),
        2002: (-19.54, False, "msci_index"), 2003: (33.76,  False, "msci_index"),
        2004: (15.25,  False, "msci_index"), 2005: (10.02,  False, "msci_index"),
        2006: (20.65,  False, "msci_index"), 2007: (9.57,   False, "msci_index"),
        2008: (-40.33, False, "msci_index"), 2009: (29.99,  False, "msci_index"),
        2010: (11.76,  False, "msci_index"), 2011: (-5.02,  False, "msci_index"),
        2012: (16.54,  False, "urth_price"), 2013: (27.37,  False, "urth_price"),
        2014: (5.50,   False, "urth_price"), 2015: (-1.93,  False, "urth_price"),
        2016: (8.15,   False, "urth_price"), 2017: (23.07,  False, "urth_price"),
        2018: (-8.91,  False, "urth_price"), 2019: (28.40,  False, "urth_price"),
        2020: (16.50,  False, "urth_price"), 2021: (22.35,  False, "urth_price"),
        2022: (-17.73, False, "urth_price"), 2023: (24.42,  False, "urth_price"),
        2024: (19.62,  False, "urth_price"), 2025: (5.80,   False, "estimate"),
    }
    bonds = {
        # 2000-2003: Bloomberg US Agg Bond Index; 2004+: AGG ETF
        2000: (11.63,  False, "bbg_agg"), 2001: (8.44,   False, "bbg_agg"),
        2002: (10.26,  False, "bbg_agg"), 2003: (4.10,   False, "bbg_agg"),
        2004: (4.34,   False, "agg_price"), 2005: (2.43,  False, "agg_price"),
        2006: (4.33,   False, "agg_price"), 2007: (6.97,  False, "agg_price"),
        2008: (5.24,   False, "agg_price"), 2009: (5.93,  False, "agg_price"),
        2010: (6.54,   False, "agg_price"), 2011: (7.84,  False, "agg_price"),
        2012: (4.21,   False, "agg_price"), 2013: (-2.02, False, "agg_price"),
        2014: (5.97,   False, "agg_price"), 2015: (0.55,  False, "agg_price"),
        2016: (2.65,   False, "agg_price"), 2017: (3.54,  False, "agg_price"),
        2018: (0.01,   False, "agg_price"), 2019: (8.72,  False, "agg_price"),
        2020: (7.51,   False, "agg_price"), 2021: (-1.54, False, "agg_price"),
        2022: (-13.01, False, "agg_price"), 2023: (5.53,  False, "agg_price"),
        2024: (1.25,   False, "agg_price"), 2025: (2.39,  False, "estimate"),
    }

    # Blends computed from sp500 + bonds components
    def _blend(w_sp, w_b):
        out = {}
        for yr in range(2000, 2026):
            sp = sp500.get(yr, (None,))[0]
            bo = bonds.get(yr, (None,))[0]
            if sp is not None and bo is not None:
                out[yr] = (round(sp * w_sp + bo * w_b, 2), False, "computed")
        return out

    _CURATED_BENCHMARK_RETURNS["sp500"]      = sp500
    _CURATED_BENCHMARK_RETURNS["dow30"]      = dow30
    _CURATED_BENCHMARK_RETURNS["nasdaq"]     = nasdaq
    _CURATED_BENCHMARK_RETURNS["msci_world"] = msci_world
    _CURATED_BENCHMARK_RETURNS["bonds"]      = bonds
    _CURATED_BENCHMARK_RETURNS["60_40"]      = _blend(0.60, 0.40)
    _CURATED_BENCHMARK_RETURNS["85_15"]      = _blend(0.85, 0.15)

_build_curated_benchmark_returns()


def _get_benchmark_annual_from_db(benchmark_key: str, years: list) -> dict:
    """Return {year: return_pct} from the benchmark_annual_returns DB table.
    Returns only years that exist in the table."""
    if not _PSYCOPG2_OK or not os.environ.get("DATABASE_URL"):
        return {}
    try:
        conn = _fund_conn()
        with conn.cursor() as cur:
            cur.execute(
                """SELECT year, return_pct FROM benchmark_annual_returns
                    WHERE benchmark_key = %s AND year = ANY(%s)""",
                (benchmark_key, years),
            )
            return {row[0]: float(row[1]) for row in cur.fetchall()}
    except Exception:
        return {}


# YTD cache for the current (in-progress) year — 15-minute TTL so it stays fresh
# keyed by benchmark_key → (return_pct, fetched_at)
_ytd_bmark_cache: dict = {}
_YTD_BMARK_TTL = 900  # seconds


def _get_ytd_benchmark_return(benchmark_key: str) -> Optional[float]:
    """Fetch YTD return for *benchmark_key* for the current calendar year.

    Uses daily closes: first trading day of year → latest available close.
    Results are cached for 15 minutes. Returns None if yfinance unavailable
    or the fetch fails.
    """
    import datetime as _dt
    if not _YFINANCE_OK:
        return None
    defn = _BENCHMARK_DEFS.get(benchmark_key)
    if not defn:
        return None

    now = time.time()
    cached = _ytd_bmark_cache.get(benchmark_key)
    if cached and (now - cached[1]) < _YTD_BMARK_TTL:
        return cached[0]

    year = _dt.date.today().year
    start_of_year = f"{year}-01-01"
    total_w, total_r = 0.0, 0.0
    for ticker, weight in defn["tickers"]:
        try:
            hist = yf.Ticker(ticker).history(start=start_of_year, interval="1d")
            if hist.empty:
                continue
            first_close = float(hist["Close"].iloc[0])
            last_close  = float(hist["Close"].iloc[-1])
            ytd = (last_close / first_close - 1) * 100
            total_r += ytd * weight
            total_w += weight
        except Exception:
            continue

    if total_w > 0:
        result = round(total_r / total_w, 2)
        _ytd_bmark_cache[benchmark_key] = (result, now)
        return result
    return None


def _get_benchmark_annual(benchmark_key: str, years: list) -> dict:
    """Return {year: return_pct} for the given benchmark and years.

    Strategy (Option C hybrid):
      - Historical years (≤ 2025): DB first (curated/verified data seeded at startup),
        falls back to in-memory curated dict if DB unavailable.
      - Current year (2026): YTD from daily closes (first trading day → latest),
        cached for 15 minutes so the column stays fresh without hammering Yahoo.
    """
    if not years:
        return {}
    defn = _BENCHMARK_DEFS.get(benchmark_key)
    if not defn:
        return {}

    import datetime as _dt
    current_year = _dt.date.today().year
    HISTORICAL_CUTOFF = current_year - 1  # everything before this year is historical

    historical_years = [yr for yr in years if yr <= HISTORICAL_CUTOFF]
    live_years       = [yr for yr in years if yr >  HISTORICAL_CUTOFF]

    result: dict[int, float] = {}

    # ── Historical: DB → in-memory curated fallback ───────────────────────────
    if historical_years:
        db_result = _get_benchmark_annual_from_db(benchmark_key, historical_years)
        result.update(db_result)
        curated = _CURATED_BENCHMARK_RETURNS.get(benchmark_key, {})
        for yr in historical_years:
            if yr not in result and yr in curated:
                result[yr] = curated[yr][0]

    # ── Live: YTD daily-close calculation for the current year ────────────────
    for yr in live_years:
        ytd = _get_ytd_benchmark_return(benchmark_key)
        if ytd is not None:
            result[yr] = ytd

    return result


@app.get("/api/market/spy-ytd")
def get_spy_ytd():
    """Return S&P 500 (SPY) year-to-date return % in real-time from Yahoo Finance.

    Cached for 15 minutes so rapid page reloads don't hammer Yahoo.
    Returns:
        ytd_pct     — YTD return as a decimal percentage (e.g. -4.37)
        first_close — SPY closing price on the first trading day of the year
        last_close  — SPY closing price on the most recent trading day
        as_of       — date string of the last_close bar (YYYY-MM-DD)
    """
    import datetime as _dt

    now = time.time()
    cached = _spy_ytd_cache.get("ts")
    if cached and (now - cached) < _SPY_YTD_TTL:
        return {k: v for k, v in _spy_ytd_cache.items() if k != "ts"}

    if not _YFINANCE_OK:
        raise HTTPException(status_code=503, detail="yfinance not installed on this server")

    try:
        start_of_year = _dt.date(_dt.date.today().year, 1, 1).isoformat()
        spy = yf.Ticker("SPY")
        hist = spy.history(start=start_of_year, interval="1d")
        if hist.empty:
            raise ValueError("Empty history returned for SPY")
        first_close = float(hist["Close"].iloc[0])
        last_close  = float(hist["Close"].iloc[-1])
        ytd_pct     = round((last_close / first_close - 1) * 100, 2)
        as_of       = str(hist.index[-1].date())
        result = {
            "ytd_pct":     ytd_pct,
            "first_close": round(first_close, 2),
            "last_close":  round(last_close, 2),
            "as_of":       as_of,
        }
        _spy_ytd_cache.update({**result, "ts": now})
        return result
    except HTTPException:
        raise
    except Exception as exc:
        # Return stale cache if available rather than erroring
        if cached:
            return {k: v for k, v in _spy_ytd_cache.items() if k != "ts"}
        raise HTTPException(status_code=503, detail=f"SPY YTD fetch failed: {exc}")


@app.get("/api/benchmark-annual")
def get_benchmark_annual(key: str = "sp500", years: str = ""):
    """Return stored annual returns for *key* benchmark for the given *years*.

    Query params:
      key   — benchmark key (sp500 | dow30 | nasdaq | msci_world | bonds | 60_40 | 85_15)
      years — comma-separated list of years, e.g. "2020,2021,2022"

    Returns:
      { ok: true, key, label, returns: { "2020": 18.40, "2021": 28.71, ... } }

    Historical years (≤ 2025) are served from the DB (curated data seeded at startup).
    2026+ are fetched live from yfinance and cached for the process lifetime.
    No auth required — this is public index data.
    """
    if key not in _BENCHMARK_DEFS:
        raise HTTPException(400, f"Unknown benchmark key: {key}")

    try:
        year_list = [int(y.strip()) for y in years.split(",") if y.strip().isdigit()]
    except Exception:
        year_list = []

    if not year_list:
        raise HTTPException(400, "Provide at least one year via ?years=2020,2021,...")

    returns_map = _get_benchmark_annual(key, year_list)
    return {
        "ok":      True,
        "key":     key,
        "label":   _BENCHMARK_DEFS[key]["label"],
        "returns": {str(yr): v for yr, v in returns_map.items()},
    }


# ── Per-ticker sector + recent headline (for rebalance table columns) ─────────
_ticker_meta_cache: dict[str, dict] = {}   # ticker → {"sector", "industry", "recent_dev", "ts"}
_TICKER_META_TTL = 900                     # 15 min — same as price/SPY caches


@app.get("/api/market/ticker-meta/{ticker}")
def get_ticker_meta(ticker: str):
    """Return sector, industry, and most-recent news headline for *ticker*.

    Cached for 15 minutes.  Used by the rebalance table to fill the
    Category and Recent Development columns via async client-side injection.

    Returns:
        sector     — e.g. "Technology"
        industry   — e.g. "Semiconductors"
        recent_dev — title of the most-recent Yahoo Finance news item, or ""
    """
    t = ticker.strip().upper()
    now = time.time()
    cached = _ticker_meta_cache.get(t)
    if cached and (now - cached["ts"]) < _TICKER_META_TTL:
        return {k: v for k, v in cached.items() if k != "ts"}

    # Sector / industry — analyst module has a robust multi-fallback implementation
    try:
        sector, industry = analyst.fetch_sector_and_industry(t)
    except Exception:
        sector, industry = "Unknown", "Unknown"

    # Most-recent news headline via yfinance
    recent_dev = ""
    news_list: list[dict] = []
    if _YFINANCE_OK:
        try:
            news_items = yf.Ticker(t).news or []
            for item in news_items[:8]:        # fetch up to 8, keep best 5
                # yfinance ≥ 0.2.x wraps fields inside "content" sub-dict
                content = item.get("content") or {}
                title = (
                    item.get("title") or content.get("title") or ""
                ).strip()
                summary = (
                    item.get("summary") or item.get("description")
                    or content.get("summary") or content.get("description") or ""
                ).strip()
                publisher = (
                    item.get("publisher")
                    or (content.get("provider") or {}).get("displayName", "")
                ).strip()
                url = (
                    item.get("link") or item.get("url")
                    or (content.get("canonicalUrl") or {}).get("url", "")
                ).strip()
                pub_ts = item.get("providerPublishTime") or None
                if not title:
                    continue
                news_list.append({
                    "title":     title,
                    "summary":   summary,   # full, untruncated
                    "publisher": publisher,
                    "url":       url,
                    "pub_ts":    pub_ts,    # Unix timestamp or None
                })
                if len(news_list) >= 5:
                    break

            # Compact inline string for the table cell (2 stories, short excerpt)
            snippets: list[str] = []
            for story in news_list[:2]:
                text = story["title"]
                if story["summary"] and story["summary"].lower() != story["title"].lower():
                    first = story["summary"].split(".")[0].strip()
                    excerpt = first if len(first) <= 120 else first[:117] + "…"
                    if excerpt:
                        text += " — " + excerpt
                if story["publisher"]:
                    text += f"  ({story['publisher']})"
                snippets.append(text)
            recent_dev = "  ·  ".join(snippets)
        except Exception:
            pass

    result = {
        "sector":     sector,
        "industry":   industry,
        "recent_dev": recent_dev,   # compact string for the inline cell
        "news":       news_list,    # structured array for the expanded modal
    }
    _ticker_meta_cache[t] = {**result, "ts": now}
    return result


@app.delete("/api/cache")
def clear_local_cache():
    """Delete all locally-cached _DGA_Report.md files from /stocks.

    Dropbox / Drive files are NOT touched — only this server instance's
    local copy is cleared. After the user deletes files from Dropbox,
    calling this endpoint ensures the Research tab reflects the true state
    instead of showing stale reports that were hydrated at startup.
    """
    cleared: list[str] = []
    for md in analyst.STOCKS_FOLDER.glob("*_DGA_Report.md"):
        try:
            md.unlink()
            cleared.append(md.name.replace("_DGA_Report.md", ""))
        except OSError:
            pass
    return {"cleared": cleared, "count": len(cleared)}


@app.delete("/api/report/{ticker}")
def delete_report(ticker: str):
    """Delete the locally-cached report files for a single ticker.

    Removes the .md, .docx, and .pptx files from /stocks for *ticker*.
    Dropbox / Drive copies are NOT touched — only this server instance's
    local cache is cleared. The next portfolio or analysis run will
    re-hydrate or regenerate as needed.
    """
    ticker = ticker.strip().upper()
    folder = analyst.STOCKS_FOLDER
    targets = [
        folder / f"{ticker}_DGA_Report.md",
        folder / f"{ticker}_DGA_Report.docx",
        folder / f"{ticker}_DGA_Presentation.pptx",
    ]
    cleared: list[str] = []
    for path in targets:
        if path.exists():
            try:
                path.unlink()
                cleared.append(path.name)
            except OSError:
                pass
    if not cleared:
        raise HTTPException(status_code=404, detail=f"No cached files found for {ticker}")
    return {"ticker": ticker, "cleared": cleared, "count": len(cleared)}


# In-memory summary cache, keyed by md_file.stat().st_mtime_ns. Avoids
# re-parsing the same markdown on every /api/reports call. Capped implicitly
# by the number of tickers a user has cached locally.
_summary_cache: dict[str, dict] = {}


def _extract_summary_cached(md_file: Path) -> dict:
    """Return {rating, price_target, upside_pct} for a saved report.

    Parsing happens once per (ticker, mtime) pair — re-running an analysis
    bumps mtime which invalidates the cache. We only read the first 12KB
    of the file because every field extract_summary_from_report cares about
    is in the report header / summary table.
    """
    try:
        key = f"{md_file.name}:{md_file.stat().st_mtime_ns}"
    except OSError:
        return {}
    cached = _summary_cache.get(key)
    if cached is not None:
        return cached
    try:
        with open(md_file, "rb") as fh:
            head = fh.read(12 * 1024).decode("utf-8", errors="replace")
        full = analyst.extract_summary_from_report(head)
    except Exception:  # noqa: BLE001
        full = {}
    summary = {
        "rating":        full.get("rating"),
        "price_target":  full.get("price_target"),
        "current_price": full.get("current_price"),
        "upside_pct":    full.get("upside_pct"),
    }
    _summary_cache[key] = summary
    return summary


@app.get("/api/reports")
def list_reports():
    """Return all tickers that have saved reports.

    Each entry now includes an extracted summary (`rating`, `price_target`,
    `upside_pct`) so the Research tab can show a target-vs-price chip in
    the saved-reports list without an extra API round-trip per row.
    """
    folder = analyst.STOCKS_FOLDER
    try:
        gamma_idx = analyst._load_gamma_index()
    except Exception:
        gamma_idx = {}
    reports = []
    for md_file in sorted(folder.glob("*_DGA_Report.md"), key=lambda p: p.stat().st_mtime, reverse=True):
        ticker = md_file.name.replace("_DGA_Report.md", "")
        has_docx = (folder / f"{ticker}_DGA_Report.docx").exists()
        has_pptx = (folder / f"{ticker}_DGA_Presentation.pptx").exists()
        gamma_entry = gamma_idx.get(ticker) or {}
        summary = _extract_summary_cached(md_file)
        reports.append({
            "ticker": ticker,
            "generated_at": datetime.utcfromtimestamp(md_file.stat().st_mtime).isoformat(),
            "has_docx": has_docx,
            "has_pptx": has_pptx,
            "gamma_url": gamma_entry.get("gamma_url"),
            # Extracted summary fields (may be None for older reports without
            # the standard header).
            "rating":        summary.get("rating"),
            "price_target":  summary.get("price_target"),
            "current_price": summary.get("current_price"),
            "upside_pct":    summary.get("upside_pct"),
        })
    return reports


# ---------------------------------------------------------------------------
# Portfolio endpoints
# ---------------------------------------------------------------------------
PORTFOLIO_OUT_DIR = Path(__file__).resolve().parent.parent / "portfolio_runs"
PORTFOLIO_OUT_DIR.mkdir(exist_ok=True)


@app.get("/api/strategies")
def list_strategies():
    """Return the three rebalance strategies with metadata for the GUI."""
    return [
        {
            "key": k,
            "label": cfg["label"],
            "description": cfg["description"],
            "min_names": cfg["min_names"],
            "max_names": cfg["max_names"],
            "max_position": cfg["max_position"],
            "max_sector": cfg["max_sector"],
        }
        for k, cfg in analyst.STRATEGIES.items()
    ]


@app.post("/api/portfolio", response_model=PortfolioJobStatus)
async def start_portfolio(
    background_tasks: BackgroundTasks,
    strategy: str = Form("current"),
    generate_gamma: bool = Form(False),
    reuse_existing: bool = Form(True),
    file: UploadFile = File(...),
):
    """Upload a portfolio CSV/XLSX and kick off a rebalance run.

    Accepts multipart/form-data. Returns a job_id to poll.
    """
    if strategy not in analyst.STRATEGIES:
        raise HTTPException(status_code=422, detail=f"Unknown strategy: {strategy}")

    # Persist the upload to a temp file (loader reads from disk).
    suffix = Path(file.filename or "portfolio.xlsx").suffix.lower() or ".xlsx"
    if suffix not in (".csv", ".xlsx", ".xls"):
        raise HTTPException(status_code=422, detail="File must be .csv or .xlsx")
    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=suffix)
    try:
        content = await file.read()
        tmp.write(content)
        tmp.close()
        try:
            records = analyst.load_portfolio_file(tmp.name)
        except Exception as exc:  # noqa: BLE001
            raise HTTPException(status_code=422, detail=f"Could not parse portfolio: {exc}")
    finally:
        try:
            os.unlink(tmp.name)
        except OSError:
            pass

    if not records:
        raise HTTPException(status_code=422, detail="Portfolio file has no rows")

    job_id = str(uuid.uuid4())
    now = datetime.utcnow().isoformat()
    date_str = datetime.utcnow().strftime("%Y%m%d")
    xlsx_out = PORTFOLIO_OUT_DIR / f"Rebalance{date_str}.xlsx"

    # Store input weights so the UI can render current→target arrows
    input_weights = {}
    for r in records:
        t = (r.get("ticker") or r.get("Ticker") or "").strip().upper()
        w = r.get("weight") or r.get("Weight") or r.get("current_weight") or 0
        try:
            w = float(w)
        except (TypeError, ValueError):
            w = 0.0
        if t:
            input_weights[t] = w

    with _pjobs_lock:
        _pjobs[job_id] = {
            "job_id": job_id,
            "status": "queued",
            "created_at": now,
            "strategy": strategy,
            "n_tickers": len(records),
            "error": None,
            "result": None,
            # Input weights (ticker → decimal weight) for current→target display
            "input_weights": input_weights,
            # Seeded so a poll right after submit gets a useful payload
            # rather than null. Real values arrive once the worker thread
            # starts firing on_progress.
            "progress": {
                "step": "queued", "pct": 0.0,
                "label": "Queued — analyzing tickers shortly…",
                "ticker_index": 0, "ticker_total": len(records),
                "current_ticker": None, "ok": [], "failed": [],
            },
        }

    # Persist so we can recover the xlsx path after a server restart.
    _save_job_index_entry(job_id, {
        "type": "portfolio",
        "xlsx_path": str(xlsx_out),
        "created_at": now,
    })

    background_tasks.add_task(
        _run_portfolio,
        job_id,
        records,
        strategy,
        generate_gamma,
        reuse_existing,
        str(xlsx_out),
    )
    return _pjobs[job_id]


# ---------------------------------------------------------------------------
# Watchlist endpoints
# ---------------------------------------------------------------------------

class WatchlistUpdate(BaseModel):
    tickers: list[str]


@app.get("/api/watchlist")
def get_watchlist():
    """Return the current watchlist."""
    return {"tickers": analyst.load_watchlist()}


@app.put("/api/watchlist")
def set_watchlist(body: WatchlistUpdate):
    """Replace the entire watchlist."""
    clean = [t.strip().upper() for t in body.tickers if t.strip()]
    analyst.save_watchlist(clean)
    return {"tickers": analyst.load_watchlist()}


@app.post("/api/watchlist/{ticker}")
def add_watchlist_ticker(ticker: str):
    """Add a single ticker to the watchlist."""
    t = ticker.strip().upper()
    if not t or not t.replace(".", "").isalnum() or len(t) > 10:
        raise HTTPException(status_code=422, detail="Invalid ticker")
    tickers = analyst.add_to_watchlist(t)
    return {"tickers": tickers}


@app.delete("/api/watchlist/{ticker}")
def remove_watchlist_ticker(ticker: str):
    """Remove a single ticker from the watchlist."""
    tickers = analyst.remove_from_watchlist(ticker.strip().upper())
    return {"tickers": tickers}


# ---------------------------------------------------------------------------
# Scan job worker
# ---------------------------------------------------------------------------

def _run_scan(job_id: str, tickers: list[str]) -> None:
    with _sjobs_lock:
        _sjobs[job_id]["status"] = "running"

    completed: dict[str, Any] = {}

    def on_progress(ticker: str, result: dict) -> None:
        with _sjobs_lock:
            _sjobs[job_id]["results"][ticker] = result
            _sjobs[job_id]["tickers_done"] = list(_sjobs[job_id]["results"].keys())

    try:
        final = analyst.run_portfolio_scan(tickers, on_progress=on_progress, verbose=True)
        with _sjobs_lock:
            _sjobs[job_id]["status"] = "done"
            _sjobs[job_id]["results"] = final["results"]
            _sjobs[job_id]["scanned_at"] = final["scanned_at"]
            _sjobs[job_id]["tickers_done"] = list(final["results"].keys())
    except BaseException as exc:  # noqa: BLE001
        tb = traceback.format_exc()
        print(f"\n❌ Scan job {job_id} CRASHED:\n{tb}", flush=True)
        with _sjobs_lock:
            _sjobs[job_id]["status"] = "failed"
            _sjobs[job_id]["error"] = str(exc)


# ---------------------------------------------------------------------------
# Scan endpoints
# ---------------------------------------------------------------------------

@app.post("/api/scan")
def start_scan(background_tasks: BackgroundTasks):
    """Kick off a live-search news scan for all watchlist tickers."""
    tickers = analyst.load_watchlist()
    if not tickers:
        raise HTTPException(status_code=422, detail="Watchlist is empty — add tickers first")

    job_id = str(uuid.uuid4())
    now = datetime.utcnow().isoformat()
    with _sjobs_lock:
        _sjobs[job_id] = {
            "job_id": job_id,
            "status": "queued",
            "created_at": now,
            "tickers": tickers,
            "tickers_done": [],
            "results": {},
            "scanned_at": None,
            "error": None,
        }
    background_tasks.add_task(_run_scan, job_id, tickers)
    return _sjobs[job_id]


@app.get("/api/scan/latest")
def get_latest_scan():
    """Return the most-recently-completed scan (persisted to disk)."""
    path = analyst.SCAN_RESULTS_FILE
    if not path.exists():
        return {"exists": False}
    try:
        data = json.loads(path.read_text())
        data["exists"] = True
        return data
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"Could not read scan results: {exc}")


@app.get("/api/scan/{job_id}")
def get_scan_status(job_id: str):
    """Poll a running or completed scan job."""
    with _sjobs_lock:
        job = _sjobs.get(job_id)
    if not job:
        raise HTTPException(status_code=404, detail="Scan job not found")
    return job


# ---------------------------------------------------------------------------
# Intelligence — macro → sector → company idea generation
# ---------------------------------------------------------------------------

def _run_intelligence(job_id: str, sector: str) -> None:
    with _ijobs_lock:
        _ijobs[job_id]["status"] = "running"
    try:
        result = analyst.run_market_intelligence(sector)
        with _ijobs_lock:
            _ijobs[job_id]["status"] = "done" if result.get("ok") else "failed"
            _ijobs[job_id]["result"] = result
            if not result.get("ok"):
                _ijobs[job_id]["error"] = result.get("error", "Unknown error")
    except BaseException as exc:  # noqa: BLE001
        tb = traceback.format_exc()
        print(f"\n❌ Intelligence job {job_id} CRASHED:\n{tb}", flush=True)
        with _ijobs_lock:
            _ijobs[job_id]["status"] = "failed"
            _ijobs[job_id]["error"] = str(exc)


@app.post("/api/intelligence")
def start_intelligence(req: IntelligenceRequest, background_tasks: BackgroundTasks):
    """Start a market intelligence run for the given sector focus."""
    sector = (req.sector or "Tech").strip()
    job_id = str(uuid.uuid4())
    now = datetime.utcnow().isoformat()
    with _ijobs_lock:
        _ijobs[job_id] = {
            "job_id": job_id,
            "status": "queued",
            "created_at": now,
            "sector": sector,
            "result": None,
            "error": None,
        }
    background_tasks.add_task(_run_intelligence, job_id, sector)
    return _ijobs[job_id]


@app.get("/api/intelligence/latest")
def get_latest_intelligence():
    """Return the most-recently-completed intelligence run (persisted to disk)."""
    path = analyst.INTEL_FILE
    if not path.exists():
        return {"exists": False}
    try:
        data = json.loads(path.read_text())
        data["exists"] = True
        return data
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"Could not read intelligence: {exc}")


@app.get("/api/intelligence/{job_id}")
def get_intelligence_status(job_id: str):
    """Poll a running or completed intelligence job."""
    with _ijobs_lock:
        job = _ijobs.get(job_id)
    if not job:
        raise HTTPException(status_code=404, detail="Intelligence job not found")
    return job


# ---------------------------------------------------------------------------
# Daily Brief — Goldman-style PM morning note (Grok 4.x w/ live X + web)
# ---------------------------------------------------------------------------

def _run_daily_brief(job_id: str) -> None:
    with _bjobs_lock:
        _bjobs[job_id]["status"] = "running"
    try:
        result = analyst.run_daily_brief()
        with _bjobs_lock:
            _bjobs[job_id]["status"] = "done" if result.get("ok") else "failed"
            _bjobs[job_id]["result"] = result
            if not result.get("ok"):
                _bjobs[job_id]["error"] = result.get("error", "Unknown error")
    except BaseException as exc:  # noqa: BLE001
        tb = traceback.format_exc()
        print(f"\n❌ Daily Brief job {job_id} CRASHED:\n{tb}", flush=True)
        with _bjobs_lock:
            _bjobs[job_id]["status"] = "failed"
            _bjobs[job_id]["error"] = str(exc)


@app.post("/api/daily-brief")
def start_daily_brief(background_tasks: BackgroundTasks):
    """Start a Goldman-style PM morning brief (Grok 4.30-beta w/ live search)."""
    job_id = str(uuid.uuid4())
    now = datetime.utcnow().isoformat()
    with _bjobs_lock:
        _bjobs[job_id] = {
            "job_id": job_id,
            "status": "queued",
            "created_at": now,
            "result": None,
            "error": None,
        }
    background_tasks.add_task(_run_daily_brief, job_id)
    return _bjobs[job_id]


@app.get("/api/daily-brief/latest")
def get_latest_daily_brief():
    """Return the most-recently-completed daily brief (persisted to disk)."""
    path = analyst.DAILY_BRIEF_FILE
    if not path.exists():
        return {"exists": False}
    try:
        data = json.loads(path.read_text())
        data["exists"] = True
        return data
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"Could not read daily brief: {exc}")


@app.get("/api/daily-brief/{job_id}")
def get_daily_brief_status(job_id: str):
    """Poll a running or completed daily-brief job."""
    with _bjobs_lock:
        job = _bjobs.get(job_id)
    if not job:
        raise HTTPException(status_code=404, detail="Daily brief job not found")
    return job


# ---------------------------------------------------------------------------
# Paper Portfolio Tracker
# ---------------------------------------------------------------------------

class TrackerHolding(BaseModel):
    ticker: str
    weight: float   # 0..1 (or 0..100; coerced server-side)


class TrackerCreateRequest(BaseModel):
    name: str
    holdings: list[TrackerHolding]
    source: dict | None = None


@app.post("/api/track")
def create_tracker(req: TrackerCreateRequest):
    """Lock in a new paper portfolio (idea basket) for forward tracking."""
    try:
        portfolio = analyst.create_idea_portfolio(
            name=req.name,
            holdings_input=[h.model_dump() for h in req.holdings],
            source=req.source or {},
        )
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc))
    return portfolio


@app.get("/api/track")
def list_trackers():
    """List all paper portfolios with computed performance metrics."""
    return {"portfolios": analyst.list_idea_portfolios()}


@app.get("/api/track/live")
def get_tracker_live():
    """Return the auto-promoted live portfolio (the benchmark)."""
    state = analyst._load_tracker_state()
    return {"live_portfolio": state.get("live_portfolio")}


@app.get("/api/track/live/detail")
def get_tracker_live_detail(snapshot_id: str | None = None):
    """YTD attribution for the live portfolio (year-start baseline + SPY benchmark).

    Per-holding return uses each ticker's first close of the calendar year as
    the entry baseline, so this surfaces who is driving annual outperformance
    versus the SPY YTD constant.

    Optional `snapshot_id` query param — when provided, returns the YTD detail
    using that historical snapshot's holdings and attribution instead of the
    current live state, so the user can re-open any past run.
    """
    return analyst.compute_live_ytd_detail(snapshot_id=snapshot_id)


@app.post("/api/track/live/ytd")
async def compute_live_ytd_unified(
    positions_file:     UploadFile = File(...),
    activity_file:      UploadFile = File(...),
    begin_value:        float | None = Form(None),
    monthly_perf_file:  UploadFile   = File(None),
):
    """Single unified YTD endpoint: Modified Dietz + TWRR + per-stock attribution.

    Multipart inputs:
      - positions_file:    Fidelity Positions CSV
      - activity_file:     Fidelity Activity / History CSV
      - begin_value:       Jan 1 portfolio total ($).  Optional when
                           monthly_perf_file is supplied — the first month's
                           beginning balance is used automatically.
      - monthly_perf_file: (optional) Fidelity monthly performance summary CSV.
                           Provides exact month-end balances + per-component breakdown.
                           When supplied begin_value may be omitted; the YTD-by-month
                           chart uses Fidelity's exact values and TWRR matches exactly.
    """

    try:
        pos_content    = await positions_file.read()
        positions_text = pos_content.decode("utf-8", errors="replace")
    except Exception as exc:
        raise HTTPException(status_code=422, detail=f"Could not read positions file: {exc}")

    try:
        act_content   = await activity_file.read()
        activity_text = act_content.decode("utf-8", errors="replace")
    except Exception as exc:
        raise HTTPException(status_code=422, detail=f"Could not read activity file: {exc}")

    monthly_perf_text: str | None = None
    if monthly_perf_file and monthly_perf_file.filename:
        try:
            mp_content        = await monthly_perf_file.read()
            monthly_perf_text = mp_content.decode("utf-8", errors="replace")
        except Exception as exc:
            raise HTTPException(status_code=422, detail=f"Could not read monthly performance file: {exc}")

    try:
        result = analyst.compute_unified_ytd(
            positions_text, activity_text, begin_value,
            monthly_perf_text=monthly_perf_text,
        )
    except Exception as exc:
        tb = traceback.format_exc()
        raise HTTPException(
            status_code=500,
            detail=f"compute_unified_ytd raised: {exc}\n\nTraceback:\n{tb}",
        )

    if not result.get("ok"):
        raise HTTPException(status_code=422, detail=result.get("error", "Unknown error"))
    return result


@app.get("/api/track/live/snapshots")
def list_live_ytd_snapshots():
    """List all stored YTD snapshots (newest first) for the live portfolio."""
    return analyst.list_ytd_snapshots()


@app.get("/api/track/live/snapshots/{snapshot_id}")
def get_live_ytd_snapshot(snapshot_id: str):
    """Get one stored YTD snapshot by id (full attribution + holdings)."""
    result = analyst.get_ytd_snapshot(snapshot_id)
    if not result.get("ok"):
        raise HTTPException(status_code=404, detail=result.get("error"))
    return result


@app.delete("/api/track/live/snapshots/{snapshot_id}")
def delete_live_ytd_snapshot(snapshot_id: str):
    """Delete a stored YTD snapshot by id."""
    result = analyst.delete_ytd_snapshot(snapshot_id)
    if not result.get("ok"):
        raise HTTPException(status_code=404, detail=result.get("error"))
    return result


class EmailYtdRequest(BaseModel):
    email: str
    snapshot_id: str | None = None


@app.post("/api/track/live/ytd/email")
def email_live_ytd_report(body: EmailYtdRequest):
    """Email the YTD report (live benchmark + attribution) to the given address.

    If snapshot_id is omitted, sends the most recent stored snapshot.
    """
    if not body.email or "@" not in body.email:
        raise HTTPException(status_code=422, detail="A valid email address is required.")
    try:
        result = analyst.email_ytd_report(body.email, body.snapshot_id)
    except Exception as exc:
        tb = traceback.format_exc()
        raise HTTPException(
            status_code=500,
            detail=f"email_ytd_report raised: {exc}\n\nTraceback:\n{tb}",
        )
    if not result.get("ok"):
        raise HTTPException(status_code=422, detail=result.get("error", "Email failed"))
    return result


@app.post("/api/track/live/ytd/set-current/{snapshot_id}")
def set_current_ytd_snapshot(snapshot_id: str):
    """Promote a past snapshot to the current account_history (live benchmark view).

    When the user selects a different run from the Past YTD Runs list, calling
    this makes the Live Benchmark card and YTD detail reflect that run.
    """
    try:
        result = analyst.set_current_ytd_snapshot(snapshot_id)
    except Exception as exc:
        tb = traceback.format_exc()
        raise HTTPException(
            status_code=500,
            detail=f"set_current_ytd_snapshot raised: {exc}\n\nTraceback:\n{tb}",
        )
    if not result.get("ok"):
        raise HTTPException(status_code=404, detail=result.get("error"))
    return result


@app.post("/api/track/snapshot")
def trigger_tracker_snapshot():
    """Manually trigger a daily snapshot (admin / debug)."""
    return analyst.take_daily_snapshot(force=True)


@app.get("/api/track/{portfolio_id}")
def get_tracker(portfolio_id: str):
    """Return one portfolio with full daily series + benchmark series for charting."""
    p = analyst.get_idea_portfolio(portfolio_id)
    if not p:
        raise HTTPException(status_code=404, detail="Paper portfolio not found")
    return p


@app.post("/api/track/{portfolio_id}/close")
def close_tracker(portfolio_id: str):
    """Mark a paper portfolio as closed (stops daily snapshotting)."""
    if not analyst.close_idea_portfolio(portfolio_id):
        raise HTTPException(status_code=404, detail="Paper portfolio not found")
    return {"ok": True}


@app.delete("/api/track/{portfolio_id}")
def delete_tracker(portfolio_id: str):
    """Permanently delete a paper portfolio."""
    if not analyst.delete_idea_portfolio(portfolio_id):
        raise HTTPException(status_code=404, detail="Paper portfolio not found")
    return {"ok": True}


@app.get("/api/portfolio/exclude")
def get_portfolio_exclude(request: Request):
    """Return the current portfolio analysis exclude list."""
    _require_auth(request)
    return {"tickers": analyst.load_portfolio_exclude()}


class PortfolioExcludeBody(BaseModel):
    tickers: list[str]


@app.put("/api/portfolio/exclude")
def set_portfolio_exclude(body: PortfolioExcludeBody, request: Request):
    """Replace the portfolio analysis exclude list.

    Pass an empty list to clear all exclusions (money-market defaults will
    be re-added on the next save cycle unless you explicitly manage them).
    """
    _require_auth(request)
    analyst.save_portfolio_exclude(body.tickers)
    return {"tickers": analyst.load_portfolio_exclude()}


@app.get("/api/portfolio/last")
def get_last_portfolio():
    """Return metadata about the most recent portfolio run (for the Research page link).

    Looks at the most-recently-modified Portfolio_Summary.md in /stocks as the
    ground truth for "when was a portfolio last run?" — this file is written
    by :func:`run_portfolio_summary` on every successful multi-ticker run and
    survives Railway redeploys via Dropbox hydration.
    """
    md_path = analyst.STOCKS_FOLDER / "Portfolio_Summary.md"
    if not md_path.exists():
        return {"exists": False}
    return {
        "exists": True,
        "generated_at": datetime.utcfromtimestamp(md_path.stat().st_mtime).isoformat(),
        "title": "Portfolio Review",
    }


@app.get("/api/portfolio/last-job")
def get_last_portfolio_job():
    """Return the full payload of the most recent completed rebalance run.

    Reads from _portfolio_last_job.json in STOCKS_FOLDER (Dropbox-synced),
    so both the web client and the mobile app see the same result regardless
    of which device triggered the run.  Returns 404 if no run has completed yet.
    """
    if not _LAST_JOB_PATH.exists():
        raise HTTPException(status_code=404, detail="No completed portfolio run yet")
    try:
        return json.loads(_LAST_JOB_PATH.read_text())
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"Could not read last-job file: {exc}")


@app.get("/api/portfolio/summary")
def get_portfolio_summary_md():
    """Return the full markdown of the last Portfolio_Summary.md."""
    md_path = analyst.STOCKS_FOLDER / "Portfolio_Summary.md"
    if not md_path.exists():
        raise HTTPException(status_code=404, detail="No portfolio summary available yet")
    return {
        "summary_md": md_path.read_text(),
        "generated_at": datetime.utcfromtimestamp(md_path.stat().st_mtime).isoformat(),
    }


@app.get("/api/portfolio/{job_id}", response_model=PortfolioJobStatus)
def get_portfolio_status(job_id: str):
    """Poll a portfolio run.  Falls back to the on-disk index after a restart."""
    with _pjobs_lock:
        job = _pjobs.get(job_id)
    if job:
        return job

    # Recovery path after server restart.
    idx = _load_job_index()
    entry = idx.get(job_id)
    if entry and entry.get("type") == "portfolio":
        xlsx_path = entry.get("xlsx_path", "")
        if xlsx_path and Path(xlsx_path).exists():
            recovered = {
                "job_id": job_id,
                "status": "done",
                "created_at": entry.get("created_at", ""),
                "strategy": "current",
                "n_tickers": 0,
                "error": None,
                "result": {
                    "ok": True,
                    "xlsx_path": xlsx_path,
                    "gamma_url": None,
                    "gamma_error": None,
                    "recovered": True,
                },
            }
            with _pjobs_lock:
                _pjobs[job_id] = recovered
            return recovered
        raise HTTPException(
            status_code=404,
            detail="Portfolio job was lost in a server restart. Please re-run.",
        )

    raise HTTPException(status_code=404, detail="Portfolio job not found")


@app.get("/api/portfolio/{job_id}/download")
def download_portfolio_xlsx(job_id: str):
    """Download the DGA-portfolio.xlsx produced by a portfolio run."""
    with _pjobs_lock:
        job = _pjobs.get(job_id)
    if not job:
        raise HTTPException(status_code=404, detail="Portfolio job not found")
    result = job.get("result") or {}
    xlsx_path = result.get("xlsx_path")
    if not xlsx_path or not Path(xlsx_path).exists():
        raise HTTPException(status_code=404, detail="Portfolio xlsx not ready yet")
    return FileResponse(
        path=xlsx_path,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename="DGA-portfolio.xlsx",
    )


# ---------------------------------------------------------------------------
# Dropbox startup hydration — runs once in a background thread so it doesn't
# block the server from accepting requests. Lists every *_DGA_Report.md in the
# Dropbox app folder and downloads any that are missing from the local /stocks
# folder. This repopulates the Saved Reports list after a Railway redeploy.
# ---------------------------------------------------------------------------
def _hydrate_from_dropbox() -> None:
    """Restore all report/analysis files from Dropbox into the container's
    /stocks folder on startup so the app survives Railway redeploys (which
    wipe local files).

    Files are now spread across four Dropbox subfolders:
      <base>/                  → metadata JSON files (tracker, gamma index, etc.)
      <base>/Presentations/    → .pptx PowerPoint files
      <base>/Reports/          → .docx Word reports
      <base>/MD cached/        → .md markdown reports
      <base>/Rebalanced/       → .xlsx portfolio rebalance files
    """
    dbx = analyst._dropbox_client()
    if dbx is None:
        return
    folder = analyst._dropbox_folder()

    def _sub(name: str) -> str:
        return f"{folder}/{name}" if folder else f"/{name}"

    pres_folder      = _sub(analyst.DROPBOX_PRESENTATIONS_SUBFOLDER)
    reports_folder   = _sub(analyst.DROPBOX_REPORTS_SUBFOLDER)
    md_folder        = _sub(analyst.DROPBOX_MD_SUBFOLDER)
    rebalanced_folder = _sub(analyst.DROPBOX_REBALANCED_SUBFOLDER)

    def _list(path: str):
        try:
            result = dbx.files_list_folder(path if path else "")
            ents = list(result.entries)
            while result.has_more:
                result = dbx.files_list_folder_continue(result.cursor)
                ents += result.entries
            return ents
        except Exception:
            return []

    downloaded = 0

    def _hydrate_entries(entries: list, source_folder: str,
                         accept_md: bool = False,
                         accept_docx: bool = False,
                         accept_pptx: bool = False,
                         accept_xlsx: bool = False,
                         accept_special: bool = False) -> int:
        n = 0
        for entry in entries:
            name = getattr(entry, "name", "")
            keep = False
            if accept_md and (name.endswith("_DGA_Report.md")
                              or name in {"Portfolio_Summary.md"}):
                keep = True
            if accept_docx and (name.endswith("_DGA_Report.docx")
                                or name == "Portfolio_Summary.docx"):
                keep = True
            if accept_pptx and (name.endswith("_DGA_Presentation.pptx")
                                or name == "Portfolio_Summary.pptx"):
                keep = True
            if accept_xlsx and name.lower().endswith(".xlsx"):
                keep = True
            if accept_special and name in {
                "tracker.json", "intelligence.json", "daily_brief.json",
                "_gamma_index.json", "_job_index.json",
            }:
                keep = True
            if not keep:
                continue
            local = analyst.STOCKS_FOLDER / name
            if local.exists():
                continue
            try:
                path = f"{source_folder}/{name}" if source_folder else f"/{name}"
                _, resp = dbx.files_download(path)
                local.write_bytes(resp.content)
                n += 1
            except Exception:
                pass
        return n

    # Base folder — metadata JSON + legacy .md/.docx that haven't been moved yet
    base_ents = _list(folder)
    downloaded += _hydrate_entries(base_ents, folder,
                                   accept_special=True,
                                   accept_md=True,    # legacy fallback
                                   accept_docx=True)  # legacy fallback
    # Subfolders — canonical locations after the folder reorganisation
    downloaded += _hydrate_entries(_list(pres_folder),       pres_folder,
                                   accept_pptx=True)
    downloaded += _hydrate_entries(_list(reports_folder),    reports_folder,
                                   accept_docx=True)
    downloaded += _hydrate_entries(_list(md_folder),         md_folder,
                                   accept_md=True)
    downloaded += _hydrate_entries(_list(rebalanced_folder), rebalanced_folder,
                                   accept_xlsx=True)

    if downloaded:
        print(f"☁️  Hydrated {downloaded} file(s) from Dropbox into /stocks "
              f"(base: {folder or '/'}, subfolders: Presentations, Reports, "
              f"MD cached, Rebalanced)")


threading.Thread(target=_hydrate_from_dropbox, daemon=True).start()

# Start the daily snapshot worker (runs once per day after market close)
analyst._start_tracker_snapshot_worker()


# ---------------------------------------------------------------------------
# Fund Admin — read-only endpoints (queries Railway Postgres via psycopg2)
# ---------------------------------------------------------------------------
try:
    import psycopg2
    from psycopg2.extras import RealDictCursor as _RealDictCursor
    _PSYCOPG2_OK = True
except ImportError:
    _PSYCOPG2_OK = False

def _fund_conn():
    if not _PSYCOPG2_OK:
        raise HTTPException(status_code=503, detail="psycopg2 not installed")
    url = os.environ.get("DATABASE_URL")
    if not url:
        raise HTTPException(status_code=503, detail="DATABASE_URL not configured")
    # Ensure self-migrations have run before the first real query.
    # Cheap no-op after the first successful run (guarded by _MIGRATIONS_APPLIED).
    try:
        _apply_self_migrations()
    except Exception:
        pass
    try:
        return psycopg2.connect(url)
    except Exception as e:
        raise HTTPException(status_code=503, detail=f"Fund DB unavailable: {e}")


# ---------------------------------------------------------------------------
# Idempotent self-applying migrations.
# Each ALTER uses IF NOT EXISTS so it's safe to run on every startup.
# Failures are logged but don't crash the app.
# ---------------------------------------------------------------------------
_MIGRATIONS_APPLIED = False

# ---------------------------------------------------------------------------
# SQL statement splitter — handles dollar-quoted strings ($$...$$, $tag$...$tag$)
# so we can split a .sql migration file into individual statements.
# ---------------------------------------------------------------------------
def _iter_sql_statements(sql: str):
    """Yield individual SQL statements, properly handling $$ dollar-quoted blocks."""
    buf: list[str] = []
    dollar_tag: str | None = None
    i = 0
    n = len(sql)
    while i < n:
        # Skip single-line comments outside dollar quotes
        if dollar_tag is None and sql[i] == '-' and i + 1 < n and sql[i + 1] == '-':
            while i < n and sql[i] != '\n':
                i += 1
            continue
        # Detect dollar-quote open / close
        if sql[i] == '$':
            j = sql.find('$', i + 1)
            if j >= 0:
                tag = sql[i: j + 1]
                if dollar_tag is None:
                    # Opening a dollar-quote block
                    dollar_tag = tag
                    buf.append(tag)
                    i = j + 1
                    continue
                elif tag == dollar_tag:
                    # Closing the dollar-quote block
                    buf.append(tag)
                    dollar_tag = None
                    i = j + 1
                    continue
        # Statement terminator (only outside dollar quotes)
        if dollar_tag is None and sql[i] == ';':
            stmt = ''.join(buf).strip()
            if stmt and not stmt.startswith('--'):
                yield stmt
            buf = []
            i += 1
            continue
        buf.append(sql[i])
        i += 1
    # Trailing statement (no trailing semicolon)
    stmt = ''.join(buf).strip()
    if stmt and not stmt.startswith('--'):
        yield stmt


def _exec_sql_file(conn, path: Path) -> tuple[int, int]:
    """Execute every statement in *path* against *conn*.

    Returns (ok_count, err_count). Errors are logged but do not abort the run
    so we get as many tables created as possible even if, e.g., an extension
    isn't available.
    """
    try:
        sql = path.read_text(encoding='utf-8')
    except Exception as e:
        print(f"[migration] cannot read {path.name}: {e}")
        return 0, 1

    ok = err = 0
    for stmt in _iter_sql_statements(sql):
        # Skip pure comment blocks
        stripped = '\n'.join(
            line for line in stmt.splitlines()
            if not line.strip().startswith('--')
        ).strip()
        if not stripped:
            continue
        try:
            with conn.cursor() as cur:
                cur.execute(stripped)
            conn.commit()
            ok += 1
        except Exception as e:
            conn.rollback()
            err += 1
            print(f"[migration] non-fatal ({path.name}): {e!s:.120}")
    print(f"[migration] {path.name}: {ok} ok, {err} errors")
    return ok, err


def _ensure_balance_history_table(conn) -> None:
    with conn.cursor() as cur:
        cur.execute("""
            CREATE TABLE IF NOT EXISTS account_balance_history (
                fund_id    UUID PRIMARY KEY,
                data_json  TEXT NOT NULL,
                updated_at TIMESTAMPTZ DEFAULT now()
            )
        """)
    conn.commit()


def _ensure_lp_creds_table(conn) -> None:
    with conn.cursor() as cur:
        cur.execute("""
            CREATE TABLE IF NOT EXISTS lp_credentials_kv (
                lp_id      TEXT PRIMARY KEY,
                data_json  TEXT NOT NULL,
                updated_at TIMESTAMPTZ DEFAULT now()
            )
        """)
    conn.commit()


def _ensure_fund_display_settings_table(conn) -> None:
    with conn.cursor() as cur:
        cur.execute("""
            CREATE TABLE IF NOT EXISTS fund_display_settings (
                fund_id       UUID PRIMARY KEY,
                settings_json TEXT NOT NULL DEFAULT '{}',
                updated_at    TIMESTAMPTZ DEFAULT now()
            )
        """)
    conn.commit()


def _ensure_benchmark_annual_returns_table(conn) -> None:
    with conn.cursor() as cur:
        cur.execute("""
            CREATE TABLE IF NOT EXISTS benchmark_annual_returns (
                benchmark_key TEXT    NOT NULL,
                year          INTEGER NOT NULL,
                return_pct    NUMERIC(8,4) NOT NULL,
                verified      BOOLEAN NOT NULL DEFAULT false,
                source        TEXT,
                updated_at    TIMESTAMPTZ DEFAULT now(),
                PRIMARY KEY (benchmark_key, year)
            )
        """)
    conn.commit()


def _seed_benchmark_historical(conn) -> None:
    """Upsert curated historical benchmark returns into the DB.

    Rows with verified=true are never overwritten; rows with verified=false
    may be updated by a live yfinance fetch later.  Running this more than
    once is safe (ON CONFLICT DO NOTHING for verified rows).
    """
    rows = []
    for bkey, year_data in _CURATED_BENCHMARK_RETURNS.items():
        for yr, (ret, verified, source) in year_data.items():
            rows.append((bkey, yr, ret, verified, source))

    if not rows:
        return

    with conn.cursor() as cur:
        for bkey, yr, ret, verified, source in rows:
            cur.execute("""
                INSERT INTO benchmark_annual_returns
                    (benchmark_key, year, return_pct, verified, source, updated_at)
                VALUES (%s, %s, %s, %s, %s, now())
                ON CONFLICT (benchmark_key, year) DO UPDATE
                    SET return_pct = EXCLUDED.return_pct,
                        verified   = EXCLUDED.verified,
                        source     = EXCLUDED.source,
                        updated_at = now()
            """, (bkey, yr, ret, verified, source))
    conn.commit()
    print(f"[migration] seeded {len(rows)} benchmark return rows")


def _ensure_manual_annual_returns_table(conn) -> None:
    with conn.cursor() as cur:
        cur.execute("""
            CREATE TABLE IF NOT EXISTS manual_annual_returns (
                fund_id    UUID    NOT NULL,
                year       INTEGER NOT NULL,
                return_pct NUMERIC(8,4) NOT NULL,
                source     TEXT,
                updated_at TIMESTAMPTZ DEFAULT now(),
                PRIMARY KEY (fund_id, year)
            )
        """)
    conn.commit()


# Hardwired Fidelity-confirmed annual returns for specific accounts.
# Each entry: list of name tokens to match (any of short_name / fund_name
# containing the token, case-insensitive) + year→return_pct dict.
_HARDWIRED_ANNUAL_RETURNS: list[dict] = [
    {
        "match_tokens": ["EM-DEF", "EMDEF", "EM DEFENSIVE"],
        "returns": {
            2019: 16.21, 2020: -4.58, 2021: 24.61, 2022: -13.61,
            2023: 12.99, 2024: 14.86, 2025: 34.95,
        },
    },
]


def _seed_manual_annual_returns(conn) -> None:
    """Upsert hardwired annual returns into manual_annual_returns.

    Resolves each account by matching any of its name tokens against both
    short_name and fund_name (case-insensitive ILIKE).  Safe to re-run.
    """
    if not _HARDWIRED_ANNUAL_RETURNS:
        return
    with conn.cursor() as cur:
        for entry in _HARDWIRED_ANNUAL_RETURNS:
            tokens   = entry["match_tokens"]
            year_map = entry["returns"]
            fid = None
            for token in tokens:
                cur.execute(
                    """SELECT fund_id FROM funds
                        WHERE (short_name ILIKE %s OR fund_name ILIKE %s)
                          AND fund_type = 'managed_account'
                        LIMIT 1""",
                    (f"%{token}%", f"%{token}%"),
                )
                row = cur.fetchone()
                if row:
                    fid = row[0]
                    print(f"[manual_returns] matched account {fid} via token {token!r}")
                    break
            if not fid:
                print(f"[manual_returns] no account matched tokens {tokens} — skipping")
                continue
            for yr, ret in year_map.items():
                cur.execute("""
                    INSERT INTO manual_annual_returns (fund_id, year, return_pct, source, updated_at)
                    VALUES (%s, %s, %s, 'hardwired', now())
                    ON CONFLICT (fund_id, year) DO UPDATE
                        SET return_pct = EXCLUDED.return_pct,
                            source     = EXCLUDED.source,
                            updated_at = now()
                """, (fid, yr, ret))
            print(f"[manual_returns] seeded {len(year_map)} rows for {fid}")
    conn.commit()


def _load_manual_annual_returns(fid: str) -> dict[int, float]:
    """Return {year: return_pct} of manual overrides for *fid*, or {} if none."""
    if not _PSYCOPG2_OK or not os.environ.get("DATABASE_URL"):
        # Fall back to in-memory hardwired dict (dev/local mode)
        try:
            conn = _fund_conn()
            with conn.cursor() as cur:
                cur.execute("SELECT short_name FROM funds WHERE fund_id = %s", (fid,))
                row = cur.fetchone()
                if row:
                    sname = row[0]
                    return {yr: ret for (sn, yr), ret in _HARDWIRED_ANNUAL_RETURNS.items()
                            if sn == sname}
        except Exception:
            pass
        return {}
    try:
        conn = _fund_conn()
        with conn.cursor() as cur:
            cur.execute(
                "SELECT year, return_pct FROM manual_annual_returns WHERE fund_id = %s",
                (fid,),
            )
            return {row[0]: float(row[1]) for row in cur.fetchall()}
    except Exception:
        return {}


def _load_fund_settings(fid: str) -> dict:
    """Return the GP-configured display settings for a fund (benchmark, period)."""
    try:
        conn = _fund_conn()
        try:
            with conn.cursor(cursor_factory=_RealDictCursor) as cur:
                cur.execute(
                    "SELECT settings_json FROM fund_display_settings WHERE fund_id = %s",
                    (fid,),
                )
                row = cur.fetchone()
                return json.loads(row["settings_json"]) if row else {}
        finally:
            conn.close()
    except Exception:
        return {}


def _db_load_lp_overlay() -> dict:
    """Load LP credential overlay from PostgreSQL."""
    try:
        conn = _fund_conn()
        try:
            with conn.cursor(cursor_factory=_RealDictCursor) as cur:
                cur.execute("SELECT lp_id, data_json FROM lp_credentials_kv")
                rows = cur.fetchall()
                return {r["lp_id"]: json.loads(r["data_json"]) for r in rows}
        finally:
            conn.close()
    except Exception:
        return {}


def _db_save_lp_overlay(overlay: dict) -> None:
    """Upsert LP credential overlay into PostgreSQL."""
    if not overlay:
        return
    try:
        conn = _fund_conn()
        try:
            with conn.cursor() as cur:
                for lp_id, data in overlay.items():
                    cur.execute("""
                        INSERT INTO lp_credentials_kv (lp_id, data_json, updated_at)
                        VALUES (%s, %s, now())
                        ON CONFLICT (lp_id) DO UPDATE
                            SET data_json  = EXCLUDED.data_json,
                                updated_at = now()
                    """, (lp_id, json.dumps(data)))
            conn.commit()
        finally:
            conn.close()
    except Exception as e:
        print(f"[lp-creds-db] save failed: {e}")


def _bootstrap_fund_schema(conn) -> None:
    """Apply all numbered migration files in order to create/update fund tables."""
    base = Path(__file__).parent.parent / "apps" / "fund" / "db" / "migrations"
    for fname in (
        "0001_initial_schema.sql",
        "0002_annual_snapshots.sql",
        "0004_ytd_cache.sql",
    ):
        p = base / fname
        if p.exists():
            _exec_sql_file(conn, p)
        else:
            print(f"[migration] WARNING: {fname} not found at {p}")


def _fix_fund_type_column(conn) -> None:
    """Ensure the fund_type column uses ('lp_fund','managed_account').

    Handles three cases:
      A) Column does not exist yet — add it with correct constraint.
      B) Column exists with the old CHECK ('open_ended',…) — drop + re-add constraint.
      C) Column exists with correct constraint — no-op.
    """
    statements = [
        # Drop old check constraint (name auto-assigned by PG; IF EXISTS is safe)
        "ALTER TABLE funds DROP CONSTRAINT IF EXISTS funds_fund_type_check",
        # Add column (idempotent — IF NOT EXISTS)
        ("ALTER TABLE funds "
         "ADD COLUMN IF NOT EXISTS fund_type TEXT NOT NULL DEFAULT 'lp_fund'"),
        # Re-add correct constraint (may already exist if we just added the column)
        ("ALTER TABLE funds "
         "ADD CONSTRAINT funds_fund_type_check "
         "CHECK (fund_type IN ('lp_fund','managed_account'))"),
        # Fix any stale rows that have the old enum values
        ("UPDATE funds SET fund_type = 'lp_fund' "
         "WHERE fund_type NOT IN ('lp_fund','managed_account')"),
    ]
    for stmt in statements:
        try:
            with conn.cursor() as cur:
                cur.execute(stmt)
            conn.commit()
        except Exception as e:
            conn.rollback()
            print(f"[migration] fund_type fixup non-fatal: {e!s:.120}")


def _apply_self_migrations() -> None:
    """Bootstrap the full fund schema if missing; then fix fund_type column.

    Idempotent — safe to call multiple times; exits immediately after the
    first successful run (guarded by _MIGRATIONS_APPLIED).
    """
    global _MIGRATIONS_APPLIED
    if _MIGRATIONS_APPLIED or not _PSYCOPG2_OK:
        return
    if not os.environ.get("DATABASE_URL"):
        return
    try:
        conn = psycopg2.connect(os.environ["DATABASE_URL"])
        try:
            # Check whether the funds table already exists
            with conn.cursor() as cur:
                cur.execute("""
                    SELECT 1 FROM information_schema.tables
                     WHERE table_schema = 'public' AND table_name = 'funds'
                """)
                funds_exists = cur.fetchone() is not None

            if not funds_exists:
                print("[migration] funds table missing — bootstrapping schema from SQL files")
                _bootstrap_fund_schema(conn)
            else:
                print("[migration] funds table exists — running incremental migrations")
                # Run incremental migrations (each is idempotent via IF NOT EXISTS)
                base = Path(__file__).parent.parent / "apps" / "fund" / "db" / "migrations"
                for fname in ("0002_annual_snapshots.sql", "0004_ytd_cache.sql"):
                    p = base / fname
                    if p.exists():
                        _exec_sql_file(conn, p)

            # Always fix the fund_type column / constraint (idempotent)
            _fix_fund_type_column(conn)
            _ensure_balance_history_table(conn)
            _ensure_lp_creds_table(conn)
            _ensure_fund_display_settings_table(conn)
            _ensure_benchmark_annual_returns_table(conn)
            _seed_benchmark_historical(conn)
            _ensure_manual_annual_returns_table(conn)
            _seed_manual_annual_returns(conn)

            _MIGRATIONS_APPLIED = True
            print("[migration] self-migrations complete")
        finally:
            conn.close()
    except Exception as e:
        print(f"[migration] could not connect: {e}")


@app.on_event("startup")
async def _on_startup_run_migrations() -> None:
    _apply_self_migrations()
    # Wire DB-backed overlay into auth_v2 so LP assignments persist
    # in PostgreSQL rather than relying on the overlay file alone.
    if _PSYCOPG2_OK and os.environ.get("DATABASE_URL"):
        try:
            import auth_v2 as _av2
            _av2.register_db_backend(_db_load_lp_overlay, _db_save_lp_overlay)
            print("[auth_v2] DB backend registered")
        except Exception as _e:
            print(f"[auth_v2] DB backend registration failed: {_e}")

def _fund_id(cur) -> str:
    fid = os.environ.get("FUND_ID")
    if fid:
        return fid
    cur.execute("SELECT id FROM funds WHERE short_name = 'DGA-I' LIMIT 1")
    row = cur.fetchone()
    if not row:
        raise HTTPException(status_code=404, detail="Fund not found — run apply_schema.py first")
    return str(row["id"])


def _resolve_fund_id(cur, fund_id: str = None) -> str:
    """Return fund UUID: use caller-supplied fund_id if provided and valid,
    otherwise fall back to the default (env FUND_ID / DGA-I)."""
    if fund_id:
        cur.execute("SELECT id FROM funds WHERE id = %s", (fund_id,))
        row = cur.fetchone()
        if not row:
            raise HTTPException(status_code=404, detail=f"Fund {fund_id!r} not found")
        return str(row["id"])
    return _fund_id(cur)


@app.get("/api/fund/diagnostic")
async def fund_diagnostic(request: Request):
    """Diagnostic — returns raw fund table state and migration status.
    Used to debug list/create issues."""
    _require_fund_token(request)
    out: dict = {
        "migrations_applied_flag": _MIGRATIONS_APPLIED,
        "psycopg2_ok":             _PSYCOPG2_OK,
    }
    try:
        # Force-run migration
        _apply_self_migrations()
        out["migrations_applied_after_run"] = _MIGRATIONS_APPLIED
    except Exception as e:
        out["migration_error"] = str(e)
    try:
        conn = _fund_conn()
        try:
            with conn.cursor(cursor_factory=_RealDictCursor) as cur:
                # Does fund_type column exist?
                cur.execute("""
                    SELECT column_name FROM information_schema.columns
                     WHERE table_schema = 'public' AND table_name = 'funds'
                """)
                cols = sorted([r["column_name"] for r in cur.fetchall()])
                out["funds_columns"] = cols
                out["has_fund_type"] = "fund_type" in cols

                # Count funds
                cur.execute("SELECT COUNT(*) AS n FROM funds")
                out["fund_count"] = cur.fetchone()["n"]

                # Sample of funds
                if "fund_type" in cols:
                    cur.execute("""
                        SELECT id, name, short_name, fund_type, status
                          FROM funds ORDER BY created_at DESC LIMIT 10
                    """)
                else:
                    cur.execute("""
                        SELECT id, name, short_name, status
                          FROM funds ORDER BY created_at DESC LIMIT 10
                    """)
                out["funds_sample"] = [
                    {k: str(v) if v else v for k, v in dict(r).items()}
                    for r in cur.fetchall()
                ]
        finally:
            conn.close()
    except Exception as e:
        out["query_error"] = str(e)
    return out


@app.get("/api/fund/list")
async def fund_list(request: Request, fund_type: str = None):
    """Return a lightweight summary of funds in the DB, optionally filtered
    by fund_type ('lp_fund' | 'managed_account').
    Used by the multi-fund selector UI to show all funds before drilling in.

    Defensive: if the fund_type column doesn't exist (migration didn't run),
    we still return funds — treating every row as 'lp_fund' implicitly.

    Performance: all per-fund stats (lp_count, commitments, positions, nav,
    ytd_cache) are fetched in bulk queries then assembled in Python — no
    N-sequential-query loops.
    """
    _require_fund_token(request)
    conn = _fund_conn()
    try:
        with conn.cursor(cursor_factory=_RealDictCursor) as cur:
            # ── 1. Check for fund_type column (cached to avoid slow system-table hit) ──
            cache_key = "funds.fund_type"
            has_fund_type = _col_exists_cache.get(cache_key)
            if has_fund_type is None:
                cur.execute("""
                    SELECT 1 FROM information_schema.columns
                     WHERE table_schema = 'public'
                       AND table_name   = 'funds'
                       AND column_name  = 'fund_type'
                """)
                has_fund_type = cur.fetchone() is not None
                _col_exists_cache[cache_key] = has_fund_type

            if has_fund_type and fund_type in ('lp_fund', 'managed_account'):
                cur.execute("""
                    SELECT id, name, short_name, inception_date, status,
                           mgmt_fee_pct, carry_pct, hurdle_pct, fund_type
                      FROM funds
                     WHERE fund_type = %s
                     ORDER BY inception_date ASC
                """, (fund_type,))
            elif has_fund_type:
                cur.execute("""
                    SELECT id, name, short_name, inception_date, status,
                           mgmt_fee_pct, carry_pct, hurdle_pct, fund_type
                      FROM funds
                     ORDER BY inception_date ASC
                """)
            else:
                # Column doesn't exist — run migration now and return all funds
                try:
                    _apply_self_migrations()
                except Exception:
                    pass
                if fund_type == 'managed_account':
                    return []
                cur.execute("""
                    SELECT id, name, short_name, inception_date, status,
                           mgmt_fee_pct, carry_pct, hurdle_pct
                      FROM funds
                     ORDER BY inception_date ASC
                """)
            funds = [dict(r) for r in cur.fetchall()]
            if not funds:
                return []

            fids     = [str(f["id"]) for f in funds]
            fids_pg  = fids  # list works with ANY(%s) via psycopg2

            # ── 2a. Bulk YTD cache (loaded EARLY so attribution tickers can
            #        be included in the price pre-warm — clicking into an
            #        account then hits the cache instead of yfinance) ──
            ytd_by_fid: dict = {}
            try:
                cur.execute("""
                    SELECT fund_id::text, nav, ytd_pct, result_json
                      FROM managed_account_ytd_cache
                     WHERE fund_id::text = ANY(%s)
                """, (fids_pg,))
                for r in cur.fetchall():
                    ytd_by_fid[str(r["fund_id"])] = dict(r)
            except Exception:
                conn.rollback()

            # ── 2b. Collect all symbols needing live prices (one batch) ──
            all_syms_set: set = set()
            try:
                cur.execute("""
                    SELECT DISTINCT s.symbol
                      FROM tax_lots tl
                      JOIN securities s ON s.id = tl.security_id
                     WHERE tl.fund_id::text = ANY(%s) AND tl.closed_at IS NULL
                       AND s.symbol IS NOT NULL
                """, (fids_pg,))
                for r in cur.fetchall():
                    if r["symbol"]:
                        all_syms_set.add(r["symbol"])
            except Exception:
                conn.rollback()

            for ytd_row in ytd_by_fid.values():
                rj_raw = ytd_row.get("result_json")
                if not rj_raw:
                    continue
                try:
                    _rj = json.loads(rj_raw) if isinstance(rj_raw, str) else rj_raw
                    for a in (_rj.get("attribution") or []):
                        tk = a.get("ticker")
                        # Include price_missing tickers — variant fallback now
                        # handles cases like BRKB → BRK-B at fetch time.
                        if tk and float(a.get("end_shares") or 0) > 0:
                            all_syms_set.add(tk)
                except Exception:
                    pass

            if all_syms_set:
                try:
                    _fetch_prices(list(all_syms_set))
                except Exception:
                    pass

            # ── 3. Bulk NAV (one SQL + cached prices, no N yfinance calls) ──
            nav_by_fid = {}
            try:
                nav_by_fid = _bulk_fund_market_nav(cur, fids_pg)
            except Exception:
                conn.rollback()

            # ── 4. Bulk LP counts ──
            lp_count_by_fid: dict = {fid: 0 for fid in fids}
            try:
                cur.execute("""
                    SELECT fund_id::text, COUNT(*) AS n
                      FROM lps
                     WHERE fund_id::text = ANY(%s) AND status = 'active'
                     GROUP BY fund_id
                """, (fids_pg,))
                for r in cur.fetchall():
                    lp_count_by_fid[str(r["fund_id"])] = int(r["n"])
            except Exception:
                conn.rollback()

            # ── 5. Bulk committed capital ──
            contrib_by_fid: dict = {fid: 0.0 for fid in fids}
            try:
                cur.execute("""
                    SELECT l.fund_id::text,
                           COALESCE(SUM(c.commitment_amount), 0) AS total
                      FROM commitments c
                      JOIN lps l ON l.id = c.lp_id
                     WHERE l.fund_id::text = ANY(%s) AND c.superseded_by IS NULL
                     GROUP BY l.fund_id
                """, (fids_pg,))
                for r in cur.fetchall():
                    contrib_by_fid[str(r["fund_id"])] = float(r["total"])
            except Exception:
                conn.rollback()

            # ── 6. Bulk position counts ──
            pos_count_by_fid: dict = {fid: 0 for fid in fids}
            try:
                cur.execute("""
                    SELECT fund_id::text, COUNT(DISTINCT security_id) AS n
                      FROM tax_lots
                     WHERE fund_id::text = ANY(%s) AND closed_at IS NULL
                     GROUP BY fund_id
                """, (fids_pg,))
                for r in cur.fetchall():
                    pos_count_by_fid[str(r["fund_id"])] = int(r["n"])
            except Exception:
                conn.rollback()

            # (ytd_by_fid was loaded earlier in step 2a, before price pre-warm)

            # ── 7b. Bulk account_balance_history for managed accounts ──
            # ONE query for all managed-account fund IDs so we can compute
            # the positions-based YTD return without any per-fund DB round-trips.
            acct_fids = [str(f["id"]) for f in funds if f.get("fund_type") == "managed_account"]
            bal_hist_by_fid: dict = {}
            if acct_fids:
                try:
                    cur.execute("""
                        SELECT fund_id::text, data_json
                          FROM account_balance_history
                         WHERE fund_id::text = ANY(%s)
                    """, (acct_fids,))
                    _cur_yr = datetime.utcnow().year
                    for row in cur.fetchall():
                        _raw = row["data_json"]
                        _recs = json.loads(_raw) if isinstance(_raw, str) else (_raw or [])
                        _ytd_r = sorted(
                            [r for r in _recs if r.get("year") == _cur_yr and not r.get("skip")],
                            key=lambda r: r.get("month", 0)
                        )
                        if _ytd_r:
                            bal_hist_by_fid[str(row["fund_id"])] = {
                                "beg":  float(_ytd_r[0].get("beg_balance") or 0),
                                "deps": round(sum(float(r.get("deposits") or 0) for r in _ytd_r), 2),
                                "wdrs": round(sum(float(r.get("withdrawals") or 0) for r in _ytd_r), 2),
                            }
                except Exception:
                    conn.rollback()

            # ── 8. Assemble result ──
            result = []
            for f in funds:
                fid           = str(f["id"])
                nav           = nav_by_fid.get(fid, 0.0)
                lp_count      = lp_count_by_fid.get(fid, 0)
                contributions = contrib_by_fid.get(fid, 0.0)
                position_count = pos_count_by_fid.get(fid, 0)
                ytd_pct       = None
                market_nav_val = nav_by_fid.get(fid, 0.0) or 0.0
                ytd_pos_pct   = None   # positions-based return (matches Fidelity)

                if f.get("fund_type") == "managed_account":
                    ytd_row = ytd_by_fid.get(fid)
                    _beg = _deps = _wdrs = 0.0
                    if ytd_row:
                        ytd_nav = float(ytd_row["nav"] or 0) or None
                        ytd_pct = float(ytd_row["ytd_pct"] or 0) or None
                        if ytd_nav:
                            nav = ytd_nav
                        if ytd_row.get("result_json"):
                            try:
                                cached = json.loads(ytd_row["result_json"])
                                if not ytd_pct:
                                    _md = cached.get("md_return_pct")
                                    if _md is not None:
                                        ytd_pct = float(_md) or None
                                attr = cached.get("attribution") or []
                                if attr:
                                    position_count = len(attr)
                                # Try result_json fields first
                                _beg  = float(cached.get("ytd_beg_balance") or 0)
                                _deps = float(cached.get("ytd_total_deposits") or 0)
                                _wdrs = float(cached.get("ytd_total_withdrawals") or 0)
                                # Fallback: mc_monthly
                                if not _beg:
                                    _mc = (cached.get("monthly_chart") or {}).get("monthly") or []
                                    if _mc:
                                        _beg  = float(_mc[0].get("beg_balance") or 0)
                                        _deps = round(sum(float(m.get("perf_detail", {}).get("deposits", 0)) for m in _mc), 2)
                                        _wdrs = round(sum(float(m.get("perf_detail", {}).get("withdrawals", 0)) for m in _mc), 2)
                            except Exception:
                                pass

                    # Last resort: pre-fetched bulk balance history (no extra query)
                    if not _beg:
                        bh = bal_hist_by_fid.get(fid)
                        if bh:
                            _beg, _deps, _wdrs = bh["beg"], bh["deps"], bh["wdrs"]

                    # Mirror what the detail view does: use live tax_lots nav when
                    # available, fall back to cached nav from the YTD cache.
                    # (Detail view: currentMV = a.market_nav || a.nav)
                    _ytd_cached_nav = float((ytd_row or {}).get("nav") or 0)
                    _effective_nav  = market_nav_val if market_nav_val > 0 else _ytd_cached_nav
                    if _effective_nav > 0 and _beg > 0:
                        ytd_pos_pct = round((_effective_nav + _wdrs - _deps - _beg) / _beg * 100, 4)

                gain     = nav - contributions
                gain_pct = (gain / contributions * 100) if contributions else 0.0

                result.append({
                    "id":             fid,
                    "name":           f["name"],
                    "short_name":     f["short_name"],
                    "inception_date": str(f["inception_date"]),
                    "status":         f["status"],
                    "fund_type":      f.get("fund_type") or "lp_fund",
                    "mgmt_fee_pct":   float(f["mgmt_fee_pct"]),
                    "carry_pct":      float(f["carry_pct"]),
                    "hurdle_pct":     float(f["hurdle_pct"]),
                    "nav":            round(nav, 2),
                    "contributions":  round(contributions, 2),
                    "total_gain":     round(gain, 2),
                    "gain_pct":       round(gain_pct, 2),
                    "lp_count":       lp_count,
                    "position_count": position_count,
                    "ytd_pct":        round(ytd_pct, 4) if ytd_pct is not None else None,
                    "ytd_pos_pct":    ytd_pos_pct,          # positions-based, matches Fidelity
                    "market_nav":     round(market_nav_val, 2) if market_nav_val > 0 else None,
                })
            return result
    finally:
        conn.close()


class YtdCacheSaveRequest(BaseModel):
    nav:         float
    ytd_pct:     float
    result_json: str | None = None   # full JSON from /api/track/live/ytd


@app.put("/api/fund/account/{fund_id}/ytd-cache")
async def save_account_ytd_cache(fund_id: str, body: YtdCacheSaveRequest, request: Request):
    """Persist the latest YTD result for a managed account in the DB.

    Called by the frontend after a successful YTD calculation so the result
    survives Railway redeploys (no longer localStorage-only).
    """
    _require_fund_token(request)
    conn = _fund_conn()
    try:
        with conn.cursor() as cur:
            cur.execute("""
                INSERT INTO managed_account_ytd_cache
                       (fund_id, nav, ytd_pct, result_json, updated_at)
                VALUES (%s, %s, %s, %s, now())
                ON CONFLICT (fund_id) DO UPDATE
                   SET nav         = EXCLUDED.nav,
                       ytd_pct     = EXCLUDED.ytd_pct,
                       result_json = EXCLUDED.result_json,
                       updated_at  = now()
            """, (fund_id, body.nav, body.ytd_pct, body.result_json))
        conn.commit()
        return {"ok": True}
    except Exception as e:
        conn.rollback()
        raise HTTPException(500, f"Failed to save YTD cache: {e}")
    finally:
        conn.close()


@app.get("/api/fund/account/{fund_id}/ytd-cache")
async def get_account_ytd_cache(fund_id: str, request: Request):
    """Return the persisted YTD result for a managed account, augmented with
    the real beginning-of-year balance and YTD cash flows pulled directly
    from account_balance_history.data_json — the source-of-truth uploaded
    by the user. This guarantees `ytd_beg_balance` / `ytd_total_deposits` /
    `ytd_total_withdrawals` are populated even when the cached result_json
    predates those fields.

    Returns 404 if no YTD result has been saved yet.
    """
    _require_fund_token(request)
    conn = _fund_conn()
    try:
        with conn.cursor(cursor_factory=_RealDictCursor) as cur:
            cur.execute("""
                SELECT fund_id, nav, ytd_pct, result_json, updated_at
                  FROM managed_account_ytd_cache
                 WHERE fund_id = %s
            """, (fund_id,))
            row = cur.fetchone()
            if not row:
                raise HTTPException(404, "No YTD cache for this account")

            # ── Pull raw monthly records from the uploaded CSV ──
            # These ALWAYS contain beg_balance/deposits/withdrawals per month,
            # regardless of when the YTD result_json was last generated.
            cur.execute("""
                SELECT data_json
                  FROM account_balance_history
                 WHERE fund_id = %s
            """, (fund_id,))
            hist = cur.fetchone()

        # Parse the cached result_json (it's stored as a JSON string)
        rj_raw = row["result_json"]
        rj = None
        if rj_raw:
            try:
                rj = json.loads(rj_raw) if isinstance(rj_raw, str) else rj_raw
            except Exception:
                rj = None
        if rj is None:
            rj = {}

        # Compute the real BoY balance + YTD flows from the raw history
        if hist and hist.get("data_json"):
            try:
                import datetime as _dt
                _records = hist["data_json"]
                if isinstance(_records, str):
                    _records = json.loads(_records)
                cur_year = _dt.date.today().year
                ytd_recs = sorted(
                    [r for r in _records if r.get("year") == cur_year and not r.get("skip")],
                    key=lambda r: (r.get("year", 0), r.get("month", 0)),
                )
                if ytd_recs:
                    beg = float(ytd_recs[0].get("beg_balance") or 0)
                    if beg > 0:
                        rj["ytd_beg_balance"] = beg
                    rj["ytd_total_deposits"]    = round(sum(float(r.get("deposits")    or 0) for r in ytd_recs), 2)
                    rj["ytd_total_withdrawals"] = round(sum(float(r.get("withdrawals") or 0) for r in ytd_recs), 2)
            except Exception:
                pass  # non-fatal — return whatever result_json already had

        return {
            "fund_id":     str(row["fund_id"]),
            "nav":         float(row["nav"] or 0),
            "ytd_pct":     float(row["ytd_pct"] or 0),
            "result_json": json.dumps(rj),  # re-serialize with augmented fields
            "updated_at":  str(row["updated_at"]),
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(500, f"Failed to load YTD cache: {e}")
    finally:
        conn.close()


@app.post("/api/fund/auth")
async def fund_auth_endpoint(body: FundAuthRequest):
    """Exchange the fund password for a fund-specific access token.
    Requires the main app token (handled by middleware) + correct FUND_PASSWORD."""
    if not hmac.compare_digest(body.password.strip(), _fund_password()):
        # Return 403, NOT 401 — 401 triggers the mobile client's main-auth retry loop.
        raise HTTPException(status_code=403, detail="Incorrect fund password")
    return {"fund_token": _make_fund_token()}


@app.get("/api/fund/overview")
async def fund_overview(request: Request, fund_id: str = None):
    _require_fund_token(request)
    conn = _fund_conn()
    try:
        with conn.cursor(cursor_factory=_RealDictCursor) as cur:
            fid = _resolve_fund_id(cur, fund_id)
            cur.execute("""
                SELECT name, short_name, inception_date, status,
                       mgmt_fee_pct, carry_pct, hurdle_pct, catch_up_pct
                  FROM funds WHERE id = %s
            """, (fid,))
            fund = dict(cur.fetchone())

            # NAV = net balance of all asset accounts
            # NAV = live market value of all open positions
            nav = _fund_market_nav(cur, fid)

            # Total LP committed capital
            cur.execute("""
                SELECT COALESCE(SUM(c.commitment_amount), 0) AS total
                  FROM commitments c
                  JOIN lps l ON l.id = c.lp_id
                 WHERE l.fund_id = %s AND c.superseded_by IS NULL
            """, (fid,))
            contributions = float(cur.fetchone()["total"])

            cur.execute("SELECT COUNT(*) AS n FROM lps WHERE fund_id=%s AND status='active'", (fid,))
            lp_count = cur.fetchone()["n"]

            cur.execute("SELECT COUNT(DISTINCT security_id) AS n FROM tax_lots WHERE fund_id=%s AND closed_at IS NULL", (fid,))
            position_count = cur.fetchone()["n"]

            gain = nav - contributions
            gain_pct = (gain / contributions * 100) if contributions else 0

            return {
                "fund_name":      fund["name"],
                "short_name":     fund["short_name"],
                "inception_date": str(fund["inception_date"]),
                "status":         fund["status"],
                "mgmt_fee_pct":   float(fund["mgmt_fee_pct"]),
                "carry_pct":      float(fund["carry_pct"]),
                "hurdle_pct":     float(fund["hurdle_pct"]),
                "catch_up_pct":   float(fund["catch_up_pct"]) if fund["catch_up_pct"] else None,
                "nav":            nav,
                "contributions":  contributions,
                "total_gain":     gain,
                "gain_pct":       gain_pct,
                "lp_count":       lp_count,
                "position_count": position_count,
            }
    finally:
        conn.close()


@app.get("/api/fund/lps")
async def fund_lps(request: Request, fund_id: str = None):
    _require_fund_token(request)
    conn = _fund_conn()
    try:
        with conn.cursor(cursor_factory=_RealDictCursor) as cur:
            fid = _resolve_fund_id(cur, fund_id)

            # Market NAV for the fund
            market_nav = _fund_market_nav(cur, fid)

            cur.execute("""
                SELECT l.id, l.legal_name, l.entity_type, l.onboarded_at,
                       COALESCE(c.commitment, 0) AS commitment
                  FROM lps l
                  LEFT JOIN (
                      SELECT lp_id, SUM(commitment_amount) AS commitment
                        FROM commitments WHERE superseded_by IS NULL GROUP BY lp_id
                  ) c ON c.lp_id = l.id
                 WHERE l.fund_id = %s AND l.status = 'active'
                 ORDER BY COALESCE(c.commitment, 0) DESC
            """, (fid,))
            rows = cur.fetchall()

            # Total committed determines each LP's ownership share
            total_committed = sum(float(r["commitment"]) for r in rows) or 0

            result = []
            for r in rows:
                commitment = float(r["commitment"])
                share      = (commitment / total_committed) if total_committed > 0 else 0.0
                cur_val    = round(market_nav * share, 2)
                result.append({
                    "id":           str(r["id"]),
                    "legal_name":   r["legal_name"],
                    "entity_type":  r["entity_type"],
                    "onboarded_at": str(r["onboarded_at"]) if r["onboarded_at"] else None,
                    "commitment":   commitment,
                    "current_value":cur_val,
                    "gain":         round(cur_val - commitment, 2),
                    "share_pct":    round(share * 100, 2),
                })
            return result
    finally:
        conn.close()


@app.get("/api/fund/positions")
async def fund_positions(request: Request, fund_id: str = None):
    _require_fund_token(request)
    conn = _fund_conn()
    try:
        with conn.cursor(cursor_factory=_RealDictCursor) as cur:
            fid = _resolve_fund_id(cur, fund_id)
            cur.execute("""
                SELECT
                    s.symbol, s.name, s.issuer,
                    COUNT(tl.id)                                              AS lot_count,
                    SUM(tl.quantity)                                          AS total_qty,
                    SUM(tl.quantity * tl.cost_basis_per_unit)
                        / SUM(tl.quantity)                                    AS avg_cost,
                    SUM(tl.quantity * tl.cost_basis_per_unit)                 AS total_cost,
                    MIN(tl.acquired_at)                                       AS first_acquired
                  FROM tax_lots tl
                  JOIN securities s ON s.id = tl.security_id
                 WHERE tl.fund_id = %s AND tl.closed_at IS NULL
                 GROUP BY s.id, s.symbol, s.name, s.issuer
                 ORDER BY SUM(tl.quantity * tl.cost_basis_per_unit) DESC
            """, (fid,))
            rows = cur.fetchall()
            total_cost = sum(float(r["total_cost"]) for r in rows) or 1

            # Fetch live prices for all symbols
            symbols = [r["symbol"] for r in rows if r["symbol"]]
            prices  = _fetch_prices(symbols)

            result = []
            total_mkt = 0.0
            for r in rows:
                sym       = r["symbol"]
                qty       = float(r["total_qty"])
                avg_cost  = float(r["avg_cost"])
                tot_cost  = float(r["total_cost"])
                last_p    = prices.get(sym)
                mkt_val   = (qty * last_p) if last_p else None
                if mkt_val:
                    total_mkt += mkt_val
                result.append({
                    "symbol":        sym,
                    "name":          r["name"],
                    "issuer":        r["issuer"],
                    "lot_count":     r["lot_count"],
                    "total_qty":     qty,
                    "avg_cost":      avg_cost,
                    "total_cost":    tot_cost,
                    "last_price":    round(last_p, 4) if last_p else None,
                    "market_value":  round(mkt_val, 2) if mkt_val else None,
                    "unrealized_gain": round(mkt_val - tot_cost, 2) if mkt_val else None,
                    "weight_pct":    tot_cost / total_cost * 100,
                    "first_acquired":str(r["first_acquired"])[:10] if r["first_acquired"] else None,
                })

            # Patch market-based weight_pct if we have prices
            if total_mkt > 0:
                for item in result:
                    if item["market_value"] is not None:
                        item["market_weight_pct"] = round(item["market_value"] / total_mkt * 100, 2)

            # Fallback for managed accounts: if tax_lots empty, serve from YTD cache
            if not result:
                cur.execute("SELECT fund_type FROM funds WHERE id = %s", (fid,))
                frow = cur.fetchone()
                if frow and frow["fund_type"] == "managed_account":
                    try:
                        cur.execute("""
                            SELECT result_json FROM managed_account_ytd_cache
                             WHERE fund_id = %s
                        """, (fid,))
                        cache_row = cur.fetchone()
                        if cache_row and cache_row["result_json"]:
                            import json as _jj
                            cached = _jj.loads(cache_row["result_json"])
                            attr = cached.get("attribution") or []
                            # Fetch live prices for ALL tickers (including those
                            # previously marked price_missing) — variant fallback
                            # (BRKB → BRK-B) in _fetch_prices now handles tickers
                            # that yfinance couldn't resolve at YTD-run time.
                            symbols_c = [a["ticker"] for a in attr
                                         if a.get("ticker")
                                         and float(a.get("end_shares") or 0) > 0]
                            prices_c  = _fetch_prices(symbols_c) if symbols_c else {}
                            total_mv_c = 0.0
                            for a in attr:
                                tk  = a.get("ticker") or ""
                                qty = float(a.get("end_shares") or 0)
                                if qty <= 0:
                                    continue
                                avg  = float(a.get("jan1_price") or 0)
                                ep   = float(a.get("end_price") or 0)
                                live = prices_c.get(tk) or ep or None
                                mv   = round(qty * live, 2) if live else None
                                cb   = round(qty * avg, 2) if avg else 0.0
                                if mv: total_mv_c += mv
                                result.append({
                                    "symbol":          tk,
                                    "name":            tk,
                                    "issuer":          None,
                                    "lot_count":       1,
                                    "total_qty":       qty,
                                    "avg_cost":        avg,
                                    "total_cost":      cb,
                                    "last_price":      round(live, 4) if live else None,
                                    "market_value":    mv,
                                    "unrealized_gain": round(mv - cb, 2) if mv else None,
                                    "weight_pct":      0,
                                    "first_acquired":  None,
                                    "_from_cache":     True,
                                })

                            # ── Synthetic SPAXX entry ──────────────────────
                            # If no money-market position survived (end_shares=0
                            # because all cash was in Pending Activity at run
                            # time), infer the cash balance as:
                            #   end_value (cached total) − sum of equity MVs
                            # and show it as a SPAXX row so cash isn't invisible.
                            has_mm = any(a.get("is_mm") and float(a.get("end_shares") or 0) > 0
                                         for a in attr)
                            if not has_mm:
                                cached_end_val = float(cached.get("end_value") or 0)
                                implied_cash   = round(cached_end_val - total_mv_c, 2)
                                if implied_cash > 0:
                                    result.append({
                                        "symbol":          "SPAXX",
                                        "name":            "Cash / Money Market (est.)",
                                        "issuer":          None,
                                        "lot_count":       1,
                                        "total_qty":       implied_cash,
                                        "avg_cost":        1.0,
                                        "total_cost":      implied_cash,
                                        "last_price":      1.0,
                                        "market_value":    implied_cash,
                                        "unrealized_gain": 0.0,
                                        "weight_pct":      0,
                                        "first_acquired":  None,
                                        "_from_cache":     True,
                                    })
                                    total_mv_c += implied_cash

                            result.sort(key=lambda x: (x["market_value"] or 0), reverse=True)
                            if total_mv_c > 0:
                                for item in result:
                                    if item["market_value"]:
                                        item["market_weight_pct"] = round(
                                            item["market_value"] / total_mv_c * 100, 2)
                    except Exception:
                        pass  # non-fatal fallback

            return result
    finally:
        conn.close()


@app.get("/api/fund/activity")
async def fund_activity(request: Request, fund_id: str = None):
    _require_fund_token(request)
    conn = _fund_conn()
    try:
        with conn.cursor(cursor_factory=_RealDictCursor) as cur:
            fid = _resolve_fund_id(cur, fund_id)
            cur.execute("""
                SELECT
                    t.id, t.effective_date, t.category, t.description, t.posted_at,
                    COALESCE(SUM(tl.debit),  0) AS total_debit,
                    COALESCE(SUM(tl.credit), 0) AS total_credit
                  FROM transactions t
                  JOIN transaction_lines tl ON tl.transaction_id = t.id
                 WHERE t.fund_id = %s
                 GROUP BY t.id, t.effective_date, t.category, t.description, t.posted_at
                 ORDER BY t.posted_at DESC
                 LIMIT 25
            """, (fid,))
            rows = cur.fetchall()
            return [{
                "id":             str(r["id"]),
                "effective_date": str(r["effective_date"]),
                "category":       r["category"],
                "description":    r["description"],
                "posted_at":      r["posted_at"].isoformat() if r["posted_at"] else None,
                "amount":         max(float(r["total_debit"]), float(r["total_credit"])),
            } for r in rows]
    finally:
        conn.close()


@app.get("/api/fund/waterfall")
async def fund_waterfall(request: Request, fund_id: str = None):
    """GP carry / LP waterfall — high-watermark model.

    Carry model:
      • The hurdle ($100 K/yr fixed) is a per-year GATE: if the year's gross
        profit exceeds the hurdle, the GP is entitled to carry.
      • Carry amount  = (year_end_NAV − high_watermark) × carry_pct (25 %).
      • High-watermark starts at  contributions + hurdle  (inception year),
        then advances to the highest end-NAV ever recorded.
      • HWM only moves up; years where end_NAV < HWM earn zero carry even if
        the year was profitable.
      • All carry is rolled into GP fractional equity (never paid out as cash),
        so the GP's dollar stake = accum_fraction × current_NAV at any moment.

    Data source priority:
      1. fund_annual_snapshots (exact year-by-year actuals — preferred)
      2. Approximation from contribution + current NAV (shows warning)
    """
    _require_fund_token(request)
    from datetime import date as _date
    conn = _fund_conn()
    try:
        with conn.cursor(cursor_factory=_RealDictCursor) as cur:
            fid = _resolve_fund_id(cur, fund_id)

            cur.execute("""
                SELECT inception_date, carry_pct, hurdle_pct,
                       COALESCE(catch_up_pct, 1.0) AS catch_up_pct
                  FROM funds WHERE id = %s
            """, (fid,))
            fund        = dict(cur.fetchone())
            inception   = fund["inception_date"]
            carry_pct   = float(fund["carry_pct"])
            hurdle_pct  = float(fund["hurdle_pct"])
            catch_up    = float(fund["catch_up_pct"])
            today       = _date.today()

            # ── Check for annual snapshots ────────────────────────────────
            has_snapshots = False
            annual_rows   = []
            try:
                cur.execute("""
                    SELECT year, start_nav, end_nav, contributions,
                           hurdle_amount, gross_profit, carry_earned,
                           carry_paid, carry_rolled, gp_equity_end, notes
                      FROM fund_annual_snapshots
                     WHERE fund_id = %s
                     ORDER BY year ASC
                """, (fid,))
                annual_rows = [dict(r) for r in cur.fetchall()]
                has_snapshots = len(annual_rows) > 0
            except Exception:
                pass  # table doesn't exist yet

            # ── LP commitments ────────────────────────────────────────────
            cur.execute("""
                SELECT COALESCE(SUM(c.commitment_amount), 0) AS total
                  FROM commitments c JOIN lps l ON l.id = c.lp_id
                 WHERE l.fund_id = %s AND c.superseded_by IS NULL
            """, (fid,))
            contributions = float(cur.fetchone()["total"])

            cur.execute("""
                SELECT l.legal_name,
                       COALESCE(SUM(c.commitment_amount), 0) AS commitment
                  FROM lps l
                  LEFT JOIN commitments c ON c.lp_id = l.id AND c.superseded_by IS NULL
                 WHERE l.fund_id = %s AND l.status = 'active'
                 GROUP BY l.id, l.legal_name
                HAVING COALESCE(SUM(c.commitment_amount), 0) > 0
                 ORDER BY COALESCE(SUM(c.commitment_amount), 0) DESC
            """, (fid,))
            lp_rows = cur.fetchall()

            # NAV = live market value of all open positions
            nav = _fund_market_nav(cur, fid)

            if has_snapshots:
                # ── Exact annual calculation from year-by-year snapshots ─────
                #
                # GP equity model: carry is issued as fractional ownership each
                # year carry is earned.  The GP's dollar stake at any date is:
                #   gp_equity = accum_fraction × current_NAV
                # (co-investment: the stake moves with NAV, not a fixed dollar).
                #
                # High-watermark:
                #   HWM_initial = contributions + hurdle_amount  (inception gate)
                #   HWM advances to end_NAV whenever end_NAV > HWM
                #
                annual_hurdle_fixed = float(annual_rows[0]["hurdle_amount"])   # $100 K
                last_row            = annual_rows[-1]
                last_year           = last_row["year"]
                last_end_nav        = float(last_row["end_nav"])
                last_gp_equity      = float(last_row["gp_equity_end"])
                last_accum_frac     = last_gp_equity / last_end_nav if last_end_nav else 0.0

                years_since         = (today - inception).days / 365.25
                total_gain          = nav - contributions

                # ── Reconstruct HWM per row for the year-by-year table ───────
                # HWM_initial = first year start_nav + hurdle  (= 2,000,000 + 100,000 = 2,100,000)
                # Using start_nav (total fund including GP) rather than LP contributions alone.
                hwm = float(annual_rows[0]["start_nav"]) + annual_hurdle_fixed   # inception gate
                snapshot_rows_out = []
                hwm_at_end = hwm  # tracks the running HWM
                for r in annual_rows:
                    end_nav_r  = float(r["end_nav"])
                    gp_eq_r    = float(r["gp_equity_end"])
                    accum_frac_r = gp_eq_r / end_nav_r if end_nav_r else 0.0
                    row_hwm    = hwm  # HWM at start of this year (threshold to exceed)
                    snapshot_rows_out.append({
                        "year":             r["year"],
                        "start_nav":        float(r["start_nav"]),
                        "end_nav":          end_nav_r,
                        "gross_profit":     float(r["gross_profit"]),
                        "hurdle_amount":    float(r["hurdle_amount"]),
                        "carry_earned":     float(r["carry_earned"]),
                        "gp_equity_end":    round(gp_eq_r, 2),
                        "accum_gp_pct":     round(accum_frac_r * 100, 4),
                        "hwm_threshold":    round(row_hwm, 2),
                    })
                    if end_nav_r > hwm:
                        hwm = end_nav_r
                hwm_at_end = hwm   # current HWM after all completed years

                # ── Current year (partial) carry estimate ────────────────────
                # HWM = hwm_at_end (highest end_NAV in recorded history)
                # New carry is earned only if nav > HWM AND gain > pro-rata hurdle
                cur_year_days      = (today - _date(last_year + 1, 1, 1)).days
                cur_year_frac      = max(0.0, cur_year_days / 365.25)
                cur_year_hurdle    = annual_hurdle_fixed * cur_year_frac
                cur_year_gain      = nav - last_end_nav
                cur_year_new_carry = 0.0
                if cur_year_gain > cur_year_hurdle and nav > hwm_at_end:
                    cur_year_new_carry = (nav - hwm_at_end) * carry_pct

                # ── GP equity at current NAV ──────────────────────────────────
                # Base: last accum fraction × current NAV (stake moves with NAV)
                # Plus: any new carry earned in partial current year
                gp_equity_base   = last_accum_frac * nav
                gp_accrued_carry = gp_equity_base + cur_year_new_carry
                gp_equity_pct    = (gp_accrued_carry / nav * 100) if nav else 0.0
                lp_nav_after_carry = nav - gp_accrued_carry

                # ── Carry history summary ─────────────────────────────────────
                total_carry_earned = sum(float(r["carry_earned"]) for r in annual_rows)
                carry_years        = [r["year"] for r in annual_rows if float(r["carry_earned"]) > 0]
                hurdle_cleared     = len(carry_years) > 0

                per_lp = []
                for row in lp_rows:
                    commitment = float(row["commitment"])
                    share      = commitment / contributions if contributions else 0
                    per_lp.append({
                        "legal_name":      row["legal_name"],
                        "commitment":      commitment,
                        "share_pct":       round(share * 100, 4),
                        "carry_charge":    round(gp_accrued_carry * share, 2),
                        "nav_after_carry": round(lp_nav_after_carry * share, 2),
                    })

                return {
                    "as_of":                 today.isoformat(),
                    "inception_date":        str(inception),
                    "data_source":           "annual_snapshots",
                    "years_since_inception": round(years_since, 2),
                    "nav":                   round(nav, 2),
                    "contributions":         round(contributions, 2),
                    "total_gain":            round(total_gain, 2),
                    "hurdle_pct":            hurdle_pct,
                    "carry_pct":             carry_pct,
                    "catch_up_pct":          catch_up,
                    # Waterfall summary
                    "hurdle_cleared":        hurdle_cleared,
                    "carry_years":           carry_years,
                    "high_watermark":        round(hwm_at_end, 2),
                    "total_carry_earned_hist": round(total_carry_earned, 2),
                    "gp_accrued_carry":      round(gp_accrued_carry, 2),
                    "gp_equity_pct":         round(gp_equity_pct, 4),
                    "lp_nav_after_carry":    round(lp_nav_after_carry, 2),
                    # Current-year estimates
                    "cur_year_gain":         round(cur_year_gain, 2),
                    "cur_year_hurdle":       round(cur_year_hurdle, 2),
                    "cur_year_new_carry":    round(cur_year_new_carry, 2),
                    # Year-by-year table
                    "annual_snapshots":      snapshot_rows_out,
                    "per_lp":                per_lp,
                }

            else:
                # ── Approximation (no annual snapshots yet) ───────────────
                # Annual hurdle: each year the fund must return hurdle_pct on
                # starting NAV before carry applies.  Without year-by-year data
                # we approximate using simple annual hurdle on contributions.
                years           = (today - inception).days / 365.25
                total_gain      = nav - contributions
                annual_hurdle_amount = contributions * hurdle_pct
                # Simple sum: hurdle_pct × contributions × years (not compound)
                cumulative_hurdle = annual_hurdle_amount * years
                hurdle_cleared  = total_gain >= cumulative_hurdle
                carry_pool      = max(0.0, total_gain - cumulative_hurdle)
                gp_carry        = carry_pool * carry_pct if hurdle_cleared else 0.0
                lp_net          = total_gain - gp_carry

                per_lp = []
                for row in lp_rows:
                    commitment = float(row["commitment"])
                    share      = commitment / contributions if contributions else 0
                    per_lp.append({
                        "legal_name":      row["legal_name"],
                        "commitment":      commitment,
                        "share_pct":       round(share * 100, 4),
                        "carry_charge":    round(gp_carry * share, 2),
                        "net_gain":        round(lp_net * share, 2),
                        "nav_after_carry": round(commitment + lp_net * share, 2),
                    })

                return {
                    "as_of":                today.isoformat(),
                    "inception_date":       str(inception),
                    "data_source":          "approximation",
                    "data_source_warning":  "Annual snapshots not yet entered. Figures are approximate — provide year-end NAVs for exact calculation.",
                    "years_since_inception":round(years, 2),
                    "nav":                  round(nav, 2),
                    "contributions":        round(contributions, 2),
                    "total_gain":           round(total_gain, 2),
                    "hurdle_pct":           hurdle_pct,
                    "carry_pct":            carry_pct,
                    "annual_hurdle_amount": round(annual_hurdle_amount, 2),
                    "cumulative_hurdle":    round(cumulative_hurdle, 2),
                    "hurdle_cleared":       hurdle_cleared,
                    "carry_pool":           round(carry_pool, 2),
                    "gp_accrued_carry":     round(gp_carry, 2),
                    "lp_net_gain":          round(lp_net, 2),
                    "lp_nav_after_carry":   round(contributions + lp_net, 2),
                    "annual_snapshots":     [],
                    "per_lp":               per_lp,
                }
    finally:
        conn.close()


# ---------------------------------------------------------------------------
# Fund Admin — management endpoints (create, import positions, import cap table)
# ---------------------------------------------------------------------------

def _seed_coa_for_fund(cur, fund_id: str) -> None:
    """Insert the standard chart-of-accounts for a new fund (idempotent)."""
    cur.execute("SELECT COUNT(*) AS n FROM accounts WHERE fund_id = %s", (fund_id,))
    if cur.fetchone()["n"] > 0:
        return  # already seeded
    for code, name, atype in _FUND_COA:
        cur.execute("""
            INSERT INTO accounts (fund_id, code, name, type)
            VALUES (%s, %s, %s, %s)
            ON CONFLICT (fund_id, code) DO NOTHING
        """, (fund_id, code, name, atype))


@app.post("/api/fund/admin/create")
async def fund_admin_create(request: Request, body: CreateFundRequest):
    """Create a new fund + seed its chart of accounts.
    Idempotent on short_name — re-calling with the same short_name updates
    the fund name but does not touch the CoA."""
    _require_fund_token(request)

    # Force-run migration up-front (creates tables if missing, fixes fund_type column).
    # If startup failed to migrate (e.g. DB wasn't ready yet), this gives it a
    # second chance right before the first fund creation.
    _apply_self_migrations()

    # Normalize the inception date — accept "2017", "2017-1-1", "2017-01-01"
    inc = (body.inception_date or '').strip()
    if re.fullmatch(r'\d{4}', inc):
        inc = f"{inc}-01-01"
    else:
        m = re.fullmatch(r'(\d{4})-(\d{1,2})-(\d{1,2})', inc)
        if m:
            inc = f"{m.group(1)}-{int(m.group(2)):02d}-{int(m.group(3)):02d}"
    if not re.fullmatch(r'\d{4}-\d{2}-\d{2}', inc):
        raise HTTPException(400, f"Invalid inception_date {body.inception_date!r} — use YYYY-MM-DD")

    fy_end = body.fiscal_year_end
    if not fy_end:
        fy_end = f"{inc[:4]}-12-31"
    ftype = body.fund_type if body.fund_type in ('lp_fund', 'managed_account') else 'lp_fund'

    conn = _fund_conn()
    try:
        with conn.cursor(cursor_factory=_RealDictCursor) as cur:
            try:
                cur.execute("""
                    INSERT INTO funds (
                        name, short_name, structure, domicile, base_ccy,
                        inception_date, fiscal_year_end,
                        mgmt_fee_pct, mgmt_fee_basis, mgmt_fee_freq,
                        carry_pct, hurdle_pct, catch_up_pct,
                        max_lps, status, fund_type
                    ) VALUES (
                        %s, %s, '3c1', 'DE', 'USD',
                        %s, %s,
                        %s, 'committed', 'quarterly',
                        %s, %s, 1.00,
                        99, 'open', %s
                    )
                    ON CONFLICT (short_name) DO UPDATE
                        SET name      = EXCLUDED.name,
                            fund_type = EXCLUDED.fund_type,
                            updated_at = NOW()
                    RETURNING id
                """, (body.name, body.short_name, inc, fy_end,
                      body.mgmt_fee_pct, body.carry_pct, body.hurdle_pct, ftype))
            except Exception as e:
                # Most likely causes:
                #   A) fund_type column missing — fall back to no-fund_type INSERT
                #   B) Table doesn't exist — re-raise with informative message
                conn.rollback()
                msg = str(e).lower()
                if 'does not exist' in msg or 'relation' in msg:
                    raise HTTPException(500,
                        f"Database table missing — schema may not have been created: {e}")
                if 'fund_type' in msg or 'column' in msg:
                    cur.execute("""
                        INSERT INTO funds (
                            name, short_name, structure, domicile, base_ccy,
                            inception_date, fiscal_year_end,
                            mgmt_fee_pct, mgmt_fee_basis, mgmt_fee_freq,
                            carry_pct, hurdle_pct, catch_up_pct,
                            max_lps, status
                        ) VALUES (
                            %s, %s, '3c1', 'DE', 'USD',
                            %s, %s,
                            %s, 'committed', 'quarterly',
                            %s, %s, 1.00,
                            99, 'open'
                        )
                        ON CONFLICT (short_name) DO UPDATE
                            SET name = EXCLUDED.name,
                                updated_at = NOW()
                        RETURNING id
                    """, (body.name, body.short_name, inc, fy_end,
                          body.mgmt_fee_pct, body.carry_pct, body.hurdle_pct))
                else:
                    raise HTTPException(500, f"Insert failed: {e}")

            row = cur.fetchone()
            if not row:
                raise HTTPException(500, "Insert returned no row")
            fid = str(row["id"])
            _seed_coa_for_fund(cur, fid)
        conn.commit()
        return {"fund_id": fid, "name": body.name, "short_name": body.short_name,
                "fund_type": ftype}
    except HTTPException:
        raise
    except Exception as e:
        conn.rollback()
        raise HTTPException(500, f"Create fund failed: {e}")
    finally:
        conn.close()


@app.post("/api/fund/import-positions")
async def fund_import_positions(
    request: Request,
    fund_id: str = Form(None),
    file: UploadFile = File(...),
):
    """Upload a Fidelity Account Positions CSV (or XLSX) to refresh the
    fund's open tax-lot positions.

    Flow:
     1. Parse CSV → list of {symbol, name, qty, avg_cost, last_price, ...}
     2. Upsert securities records
     3. Close all existing open lots (without a hard delete — immutable ledger)
     4. Create a single 'adjustment' transaction with balanced double-entry:
          Dr. 1100 Securities at Cost  (one line per security)
          Dr. 1030 Cash — Money Market (for MM positions)
          Cr. 3000 Capital — GP        (single balancing credit)
     5. Insert new tax_lots pointing at that transaction
    """
    _require_fund_token(request)
    raw = await file.read()
    # Handle both CSV and XLSX uploads
    if (file.filename or '').lower().endswith(('.xlsx', '.xls')):
        if not _OPENPYXL_OK:
            raise HTTPException(400, "openpyxl not installed — upload a .csv file")
        wb = openpyxl.load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
        ws = wb.active
        header = [str(c.value or '').strip() for c in next(ws.iter_rows())]
        lines  = [header] + [
            [str(c.value or '').strip() for c in row] for row in ws.iter_rows(min_row=2)
        ]
        content = '\n'.join(','.join(r) for r in lines)
    else:
        content = raw.decode('utf-8', errors='replace')

    positions = _parse_fidelity_csv(content)
    if not positions:
        raise HTTPException(400, "No valid positions found in file. "
            "Expected a Fidelity Account Positions CSV export.")

    conn = _fund_conn()
    try:
        with conn.cursor(cursor_factory=_RealDictCursor) as cur:
            fid = _resolve_fund_id(cur, fund_id)

            # ── Fetch CoA accounts ─────────────────────────────────────────
            cur.execute("""
                SELECT code, id FROM accounts
                 WHERE fund_id = %s AND code IN ('1020','1030','1100','3000')
            """, (fid,))
            acct_map = {r["code"]: str(r["id"]) for r in cur.fetchall()}
            sec_acct  = acct_map.get("1100")
            mm_acct   = acct_map.get("1030") or acct_map.get("1020")
            cap_acct  = acct_map.get("3000")
            if not sec_acct or not cap_acct:
                raise HTTPException(500, "Chart of accounts not found — "
                    "run /api/fund/admin/create first")

            # ── Upsert securities ─────────────────────────────────────────
            sec_ids = {}
            for p in positions:
                sym = p["symbol"]
                cur.execute("""
                    INSERT INTO securities (symbol, name, asset_class, is_public)
                    VALUES (%s, %s, %s, TRUE)
                    ON CONFLICT (symbol) DO UPDATE SET name = EXCLUDED.name
                    RETURNING id
                """, (sym, p["name"] or sym,
                      "cash" if p["is_cash"] else "equity"))
                sec_ids[sym] = str(cur.fetchone()["id"])

            # ── Close all existing open lots ──────────────────────────────
            cur.execute("""
                UPDATE tax_lots SET closed_at = NOW()
                 WHERE fund_id = %s AND closed_at IS NULL
            """, (fid,))

            # ── Create balancing adjustment transaction ────────────────────
            today = datetime.utcnow().date().isoformat()
            cur.execute("""
                INSERT INTO transactions
                    (fund_id, effective_date, category, description)
                VALUES (%s, %s, 'adjustment', 'Position import from Fidelity CSV')
                RETURNING id
            """, (fid, today))
            txn_id = str(cur.fetchone()["id"])

            # Build lines: one Dr per position, one Cr total
            total_cost = 0.0
            line_num   = 1
            for p in positions:
                cost = float(p["cost_basis"])
                if cost <= 0:
                    continue
                total_cost += cost
                acct = mm_acct if p["is_cash"] else sec_acct
                if not acct:
                    continue
                cur.execute("""
                    INSERT INTO transaction_lines
                        (transaction_id, line_number, account_id, debit, security_id)
                    VALUES (%s, %s, %s, %s, %s)
                """, (txn_id, line_num, acct, round(cost, 4),
                      sec_ids.get(p["symbol"])))
                line_num += 1

            # Balancing credit to Capital — GP
            if total_cost > 0:
                cur.execute("""
                    INSERT INTO transaction_lines
                        (transaction_id, line_number, account_id, credit)
                    VALUES (%s, %s, %s, %s)
                """, (txn_id, line_num, cap_acct, round(total_cost, 4)))

            # ── Insert new tax_lots ───────────────────────────────────────
            for p in positions:
                cur.execute("""
                    INSERT INTO tax_lots
                        (fund_id, security_id, acquired_at,
                         quantity, cost_basis_per_unit, open_transaction_id)
                    VALUES (%s, %s, NOW(), %s, %s, %s)
                """, (fid, sec_ids[p["symbol"]],
                      float(p["quantity"]),
                      float(p["avg_cost"]) if p["avg_cost"] else 0.0,
                      txn_id))

        conn.commit()

        # Fetch prices for imported symbols
        symbols = [p["symbol"] for p in positions if not p["is_cash"]]
        prices  = _fetch_prices(symbols)
        market_value = sum(
            float(p["quantity"]) * (prices.get(p["symbol"]) or float(p.get("last_price") or 0))
            for p in positions
        )

        return {
            "fund_id":            fid,
            "imported":           len(positions),
            "positions_imported": len(positions),
            "market_value_total": round(market_value, 2),
            "message": f"Successfully imported {len(positions)} position lots.",
        }
    finally:
        conn.close()


@app.post("/api/fund/import-captable")
async def fund_import_captable(
    request: Request,
    fund_id: str = Form(None),
    file: UploadFile = File(...),
):
    """Upload a cap-table CSV or XLSX to set/update LP records and commitments.

    Expected columns (case-insensitive):
        LP Name, Commitment Amount, Entity Type (opt), Effective Date (opt)

    On re-upload, existing LPs are matched by legal_name (case-insensitive).
    New commitments supersede the previous commitment for each LP.
    """
    _require_fund_token(request)
    raw  = await file.read()
    rows, economics, fund_estab_year = _parse_captable(raw, file.filename or '')
    if not rows:
        raise HTTPException(400, "No valid LP rows found in file. "
            "Expected a column containing 'LP Name' and contribution amounts "
            "(either a 'Commitment Amount' column or year columns like '2024', '2025').")
    # Also parse annual NAV rows from the same file (waterfall data)
    nav_rows = _parse_annual_nav(raw, file.filename or '')

    conn = _fund_conn()
    try:
        with conn.cursor(cursor_factory=_RealDictCursor) as cur:
            fid = _resolve_fund_id(cur, fund_id)

            # Look up CoA accounts needed for contribution journal entries
            cur.execute("""
                SELECT code, id FROM accounts
                 WHERE fund_id = %s AND code IN ('1020','1010','3100','3000')
            """, (fid,))
            acct_map = {r["code"]: str(r["id"]) for r in cur.fetchall()}
            cash_acct = acct_map.get("1020") or acct_map.get("1010")
            lp_cap_acct = acct_map.get("3100")  # Capital — LP

            # ── Wipe existing LP + commitment rows for this fund ─────────────
            # Delete in dependency order (children before parents) to avoid FK
            # violations. Only touch tables that the cap-table import owns.
            # Nullable lp_id references (transaction_lines, accounts, users)
            # are NULLed rather than deleted so those rows are preserved.
            cur.execute("""
                UPDATE transaction_lines SET lp_id = NULL
                 WHERE lp_id IN (SELECT id FROM lps WHERE fund_id = %s)
            """, (fid,))
            cur.execute("""
                UPDATE accounts SET lp_id = NULL
                 WHERE lp_id IN (SELECT id FROM lps WHERE fund_id = %s)
            """, (fid,))
            cur.execute("""
                UPDATE users SET lp_id = NULL
                 WHERE lp_id IN (SELECT id FROM lps WHERE fund_id = %s)
            """, (fid,))
            # Delete leaf tables with NOT NULL lp_id FK (deepest first)
            for tbl in ("lp_statements", "carry_allocations",
                        "mgmt_fee_allocations", "nav_snapshot_lp",
                        "distribution_allocations", "capital_call_allocations",
                        "lp_annual_snapshots"):
                cur.execute(
                    f"DELETE FROM {tbl} WHERE lp_id IN "
                    f"(SELECT id FROM lps WHERE fund_id = %s)",
                    (fid,)
                )
            # commitments has its own fund_id — delete directly
            cur.execute("DELETE FROM commitments WHERE fund_id = %s", (fid,))
            # Now lps has no remaining FK children — safe to delete
            cur.execute("DELETE FROM lps WHERE fund_id = %s", (fid,))

            # ── Insert fresh LP rows ──────────────────────────────────────────
            imported = 0
            for row in rows:
                commitment = float(row["commitment"])
                eff_date   = row["effective_date"]

                cur.execute("""
                    INSERT INTO lps (fund_id, legal_name, entity_type,
                                     accred_type, status)
                    VALUES (%s, %s, %s, 'net_worth', 'active')
                    RETURNING id
                """, (fid, row["legal_name"], row["entity_type"]))
                r = cur.fetchone()
                if not r:
                    continue
                lp_id = str(r["id"])

                # ── Commitment ────────────────────────────────────────────────
                cur.execute("""
                    INSERT INTO commitments
                        (lp_id, fund_id, commitment_amount, effective_date)
                    VALUES (%s, %s, %s, %s)
                """, (lp_id, fid, commitment, eff_date))

                # ── Capital contribution journal entry ────────────────────────
                if cash_acct and lp_cap_acct and commitment > 0:
                    cur.execute("""
                        INSERT INTO transactions
                            (fund_id, effective_date, category, description)
                        VALUES (%s, %s, 'contribution', %s)
                        RETURNING id
                    """, (fid, eff_date,
                          f"Capital contribution — {row['legal_name']}"))
                    txn_id = str(cur.fetchone()["id"])
                    cur.execute("""
                        INSERT INTO transaction_lines
                            (transaction_id, line_number, account_id, debit)
                        VALUES (%s, 1, %s, %s)
                    """, (txn_id, cash_acct, round(commitment, 4)))
                    cur.execute("""
                        INSERT INTO transaction_lines
                            (transaction_id, line_number, account_id, credit)
                        VALUES (%s, 2, %s, %s)
                    """, (txn_id, lp_cap_acct, round(commitment, 4)))

                imported += 1

            # ── Apply fund economics + Fund Established date ─────────────────
            econ_applied = {}
            fund_updates = []
            fund_vals    = []
            for col, key in [
                ("mgmt_fee_pct", "mgmt_fee_pct"),
                ("carry_pct",    "carry_pct"),
                ("hurdle_pct",   "hurdle_pct"),
            ]:
                if key in economics:
                    fund_updates.append(f"{col} = %s")
                    fund_vals.append(economics[key])
                    econ_applied[key] = economics[key]
            if fund_estab_year:
                fund_updates.append("inception_date = %s")
                fund_vals.append(f"{fund_estab_year}-01-01")
            if fund_updates:
                fund_vals.append(fid)
                cur.execute(
                    f"UPDATE funds SET {', '.join(fund_updates)} WHERE id = %s",
                    fund_vals,
                )

            # ── Import annual NAV rows (waterfall) ───────────────────────────
            nav_imported = 0
            for r in nav_rows:
                cur.execute("""
                    INSERT INTO fund_annual_snapshots
                        (fund_id, year, start_nav, end_nav, contributions,
                         hurdle_amount, gross_profit, carry_earned,
                         carry_paid, carry_rolled, gp_equity_end)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                    ON CONFLICT (fund_id, year) DO UPDATE
                        SET start_nav      = EXCLUDED.start_nav,
                            end_nav        = EXCLUDED.end_nav,
                            contributions  = EXCLUDED.contributions,
                            hurdle_amount  = EXCLUDED.hurdle_amount,
                            gross_profit   = EXCLUDED.gross_profit,
                            carry_earned   = EXCLUDED.carry_earned,
                            carry_paid     = EXCLUDED.carry_paid,
                            carry_rolled   = EXCLUDED.carry_rolled,
                            gp_equity_end  = EXCLUDED.gp_equity_end,
                            updated_at     = NOW()
                """, (fid, r['year'], r['start_nav'], r['end_nav'],
                      r['contributions'], r['hurdle_amount'], r['gross_profit'],
                      r['carry_earned'], r['carry_paid'], r['carry_rolled'],
                      r['gp_equity_end']))
                nav_imported += 1

        conn.commit()
        year_note = ""
        contrib_years = [r["contribution_year"] for r in rows if r.get("contribution_year")]
        if contrib_years:
            year_note = f" (contributions from {min(contrib_years)}–{max(contrib_years)})"
        resp: dict = {
            "fund_id":           fid,
            "imported":          imported,
            "lps_imported":      imported,
            "nav_rows_imported": nav_imported,
            "message":           f"Imported {imported} LP record(s){year_note}.",
        }
        if nav_imported:
            resp["message"] += f" {nav_imported} annual NAV rows loaded."
        if fund_estab_year:
            resp["message"] += f" Fund established {fund_estab_year}."
        if econ_applied:
            resp["economics_applied"] = econ_applied
            parts = []
            if "mgmt_fee_pct" in econ_applied:
                parts.append(f"mgmt fee {econ_applied['mgmt_fee_pct']*100:.2f}%")
            if "carry_pct" in econ_applied:
                parts.append(f"carry {econ_applied['carry_pct']*100:.0f}%")
            if "hurdle_pct" in econ_applied:
                parts.append(f"hurdle {econ_applied['hurdle_pct']*100:.0f}%")
            resp["message"] += f" Fund economics updated: {', '.join(parts)}."
        return resp
    finally:
        conn.close()


# ---------------------------------------------------------------------------
# Annual NAV import — populates fund_annual_snapshots for waterfall display
# ---------------------------------------------------------------------------

def _parse_annual_nav(content: bytes, filename: str) -> list:
    """Parse an annual NAV / waterfall spreadsheet.

    Expected columns (detected by header keyword matching):
        Year | Jan 1 NAV | Dec 31 NAV | Contributions | Hurdle Amount |
        Carry Owed | GP Equity Allocated | LP Allocations | … |
        accum GP equity in fund

    Returns list of dicts matching fund_annual_snapshots columns.
    Rows without a valid Dec 31 NAV (e.g. current partial year) are skipped.
    """
    fname_lower = (filename or '').lower()
    if fname_lower.endswith(('.xlsx', '.xls')):
        if not _OPENPYXL_OK:
            raise HTTPException(400, "openpyxl not installed — upload a CSV instead")
        wb  = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        ws  = wb.active
        raw = [[str(c.value if c.value is not None else '').strip() for c in row]
               for row in ws.iter_rows()]
    else:
        text = content.decode('utf-8', errors='replace')
        raw  = [[cell.strip() for cell in row] for row in csv.reader(io.StringIO(text))]

    if not raw:
        return []

    def _pm(s):
        """Parse money string: '$1,474,741' or '-' → float (dash / blank → 0)."""
        s2 = str(s or '').replace('$', '').replace(',', '').strip()
        if s2 in ('', '-', '—', 'None'):
            return 0.0
        try:
            return float(s2)
        except (ValueError, TypeError):
            return None

    def _pct_val(s):
        """Parse percent string '13.37%' or '0.1337' → fraction 0–1, or None."""
        s2 = str(s or '').strip()
        # Strip everything after a space (e.g. "0.08% EM: 99.02%" → "0.08%")
        s2 = s2.split()[0] if s2 else ''
        s2 = s2.replace('%', '').strip()
        try:
            v = float(s2)
            return v / 100.0 if v > 1 else v   # "13.37" → 0.1337; "0.1337" → 0.001337?
            # Heuristic: values > 1 are already percent form
        except (ValueError, TypeError):
            return None

    # ── Locate the header row (col A = "year") ───────────────────────────────
    header_idx = None
    for i, row in enumerate(raw):
        if row and str(row[0]).strip().lower() == 'year':
            header_idx = i
            break

    if header_idx is None:
        return []

    header = [str(c).lower().strip() for c in raw[header_idx]]

    def _find_col(*kws):
        for kw in kws:
            for j, h in enumerate(header):
                if kw in h:
                    return j
        return None

    c_start  = _find_col('jan 1', 'jan1', 'start nav', 'beginning') or 1
    c_end    = _find_col('dec 31', 'dec31', 'end nav', 'ending')    or 2
    c_contr  = _find_col('contribution')                             or 3
    c_hurdle = _find_col('hurdle amount', 'hurdle')                  or 4
    c_carry  = _find_col('carry owed', 'carry earned', 'carry')      or 5
    c_accum  = _find_col('accum gp', 'accumulated gp', 'accum gp equity')

    # accum GP column is often the last column (J = index 9 in typical layout)
    if c_accum is None:
        c_accum = max(9, len(header) - 1)

    out = []
    for row in raw[header_idx + 1:]:
        if not row or not str(row[0]).strip():
            continue
        try:
            year = int(float(str(row[0]).strip()))
        except (ValueError, TypeError):
            continue
        if year < 2000 or year > 2099:
            continue

        start_nav = _pm(row[c_start]  if c_start  < len(row) else '') or 0.0
        end_nav   = _pm(row[c_end]    if c_end    < len(row) else '')
        contrib   = _pm(row[c_contr]  if c_contr  < len(row) else '') or 0.0
        hurdle    = _pm(row[c_hurdle] if c_hurdle < len(row) else '') or 0.0
        carry     = _pm(row[c_carry]  if c_carry  < len(row) else '') or 0.0
        accum_raw = row[c_accum]      if c_accum  < len(row) else ''

        # Skip partial / current year rows (no Dec 31 NAV yet)
        if end_nav is None or end_nav <= 0:
            continue

        accum_pct    = _pct_val(accum_raw)
        gp_equity_end = round(accum_pct * end_nav, 2) if accum_pct is not None else 0.0
        gross_profit  = round(end_nav - start_nav - contrib, 2)

        out.append({
            'year':          year,
            'start_nav':     round(start_nav, 2),
            'end_nav':       round(end_nav,   2),
            'contributions': round(contrib,   2),
            'hurdle_amount': round(hurdle,    2),
            'gross_profit':  gross_profit,
            'carry_earned':  round(carry,     2),
            'carry_paid':    0.0,
            'carry_rolled':  round(carry,     2),
            'gp_equity_end': gp_equity_end,
        })

    return out


@app.post("/api/fund/import-annual-nav")
async def fund_import_annual_nav(
    request: Request,
    fund_id: str = Form(None),
    file: UploadFile = File(...),
):
    """SINGLE LP-FUND ADMIN UPLOAD. Parses one spreadsheet that contains:
      • Year-by-year NAV / hurdle / carry / GP equity rows  → fund_annual_snapshots
      • LP roster + initial contributions                   → lps + commitments
      • Economics ("0/25, 5% hurdle")                        → funds.mgmt_fee_pct, carry_pct, hurdle_pct
      • "Fund Established | <year>"                          → funds.inception_date

    Everything is keyword-detected; rows can be in any order. Re-uploading
    safely overwrites all four destinations.
    """
    _require_fund_token(request)
    raw       = await file.read()
    nav_rows  = _parse_annual_nav(raw, file.filename or '')
    lp_rows, economics, fund_estab_year = _parse_captable(raw, file.filename or '')

    if not nav_rows and not lp_rows:
        raise HTTPException(400,
            "No valid data found. Expected the annual NAV worksheet to contain "
            "either a 'Year | Jan 1 NAV | Dec 31 NAV | …' header block, or an "
            "'LP Name | <partners…>' / 'initial contribution | <amounts…>' block, "
            "or both.")

    conn = _fund_conn()
    try:
        with conn.cursor(cursor_factory=_RealDictCursor) as cur:
            fid = _resolve_fund_id(cur, fund_id)

            # ── 1. LP roster + commitments (only if present in the file) ─────
            lp_imported = 0
            if lp_rows:
                # Wipe existing LPs/commitments — fresh import is the contract.
                cur.execute("""
                    DELETE FROM transaction_lines
                     WHERE transaction_id IN (
                         SELECT id FROM transactions
                          WHERE fund_id = %s AND category = 'contribution')
                """, (fid,))
                cur.execute("DELETE FROM transactions WHERE fund_id = %s AND category = 'contribution'", (fid,))
                cur.execute("DELETE FROM commitments WHERE fund_id = %s", (fid,))
                cur.execute("DELETE FROM lps WHERE fund_id = %s", (fid,))
                for row in lp_rows:
                    cur.execute("""
                        INSERT INTO lps (fund_id, legal_name, entity_type, accred_type, status)
                        VALUES (%s, %s, %s, 'net_worth', 'active')
                        RETURNING id
                    """, (fid, row["legal_name"], row.get("entity_type", "individual")))
                    r2 = cur.fetchone()
                    if not r2: continue
                    cur.execute("""
                        INSERT INTO commitments (lp_id, fund_id, commitment_amount, effective_date)
                        VALUES (%s, %s, %s, %s)
                    """, (str(r2["id"]), fid, float(row["commitment"]), row["effective_date"]))
                    lp_imported += 1

            # ── 2. Fund economics + inception (only if present in file) ──────
            econ_applied = {}
            updates, vals = [], []
            for col, key in [("mgmt_fee_pct", "mgmt_fee_pct"),
                             ("carry_pct",    "carry_pct"),
                             ("hurdle_pct",   "hurdle_pct")]:
                if key in economics:
                    updates.append(f"{col} = %s")
                    vals.append(economics[key])
                    econ_applied[key] = economics[key]
            if fund_estab_year:
                updates.append("inception_date = %s")
                vals.append(f"{fund_estab_year}-01-01")
            if updates:
                vals.append(fid)
                cur.execute(f"UPDATE funds SET {', '.join(updates)} WHERE id = %s", vals)

            # ── 3. Year-by-year NAV snapshots (only if present in file) ──────
            nav_imported = 0
            for r in nav_rows:
                cur.execute("""
                    INSERT INTO fund_annual_snapshots
                        (fund_id, year, start_nav, end_nav, contributions,
                         hurdle_amount, gross_profit, carry_earned,
                         carry_paid, carry_rolled, gp_equity_end)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                    ON CONFLICT (fund_id, year) DO UPDATE
                        SET start_nav      = EXCLUDED.start_nav,
                            end_nav        = EXCLUDED.end_nav,
                            contributions  = EXCLUDED.contributions,
                            hurdle_amount  = EXCLUDED.hurdle_amount,
                            gross_profit   = EXCLUDED.gross_profit,
                            carry_earned   = EXCLUDED.carry_earned,
                            carry_paid     = EXCLUDED.carry_paid,
                            carry_rolled   = EXCLUDED.carry_rolled,
                            gp_equity_end  = EXCLUDED.gp_equity_end,
                            updated_at     = NOW()
                """, (fid, r['year'], r['start_nav'], r['end_nav'],
                      r['contributions'], r['hurdle_amount'], r['gross_profit'],
                      r['carry_earned'], r['carry_paid'], r['carry_rolled'],
                      r['gp_equity_end']))
                nav_imported += 1

        conn.commit()
        years = sorted(r['year'] for r in nav_rows) if nav_rows else []
        parts = []
        if lp_imported:   parts.append(f"{lp_imported} LP{'s' if lp_imported != 1 else ''}")
        if nav_imported:  parts.append(f"{nav_imported} annual NAV row{'s' if nav_imported != 1 else ''} ({min(years)}–{max(years)})")
        if econ_applied:  parts.append("economics")
        if fund_estab_year: parts.append(f"inception {fund_estab_year}")
        return {
            "fund_id":       fid,
            "nav_imported":  nav_imported,
            "lp_imported":   lp_imported,
            "economics":     econ_applied,
            "inception_year": fund_estab_year,
            "years":         years,
            "message":       "Imported " + ", ".join(parts) + "." if parts else "Nothing imported.",
        }
    finally:
        conn.close()


# ─────────────────────────────────────────────────────────────────────────────
# Benchmark monthly closes — for overlay on managed-account balance chart
# Supports SPY (default) and any other ETF/index the UI requests.
# ─────────────────────────────────────────────────────────────────────────────
_bench_monthly_cache: dict = {}   # keyed by ticker symbol

_ALLOWED_BENCH = {"SPY", "QQQ", "DIA", "URTH", "EFA", "AGG",
                  "IWM", "VTI", "GLD", "TLT", "VNQ", "VXUS"}

@app.get("/api/market/spy-monthly")
def get_spy_monthly(ticker: str = "SPY"):
    """Return month-end closes for the requested benchmark ticker for the
    current calendar year, indexed to the first trading day's close.
    Cached 15 min per ticker. Defaults to SPY."""
    sym = (ticker or "SPY").upper().strip()
    if sym not in _ALLOWED_BENCH:
        sym = "SPY"
    import time as _time
    now    = _time.time()
    cached = _bench_monthly_cache.get(sym, {})
    if cached.get("ts") and (now - cached["ts"]) < 900:
        return {k: v for k, v in cached.items() if k != "ts"}
    try:
        import yfinance as yf
        from datetime import date as _date
        tkr   = yf.Ticker(sym)
        start = f"{_date.today().year}-01-01"
        hist  = tkr.history(start=start, interval="1d")
        if hist.empty:
            raise ValueError(f"Empty history for {sym}")
        first_close = float(hist.iloc[0]["Close"])
        monthly = hist.resample("ME").last().dropna(subset=["Close"])
        points = []
        for ts, row in monthly.iterrows():
            close = float(row["Close"])
            points.append({
                "month":   ts.strftime("%Y-%m"),
                "close":   round(close, 2),
                "ytd_pct": round((close / first_close - 1.0) * 100, 4),
                "norm":    round(close / first_close, 6),
            })
        result = {"ticker": sym, "first_close": round(first_close, 2), "points": points}
        _bench_monthly_cache[sym] = {**result, "ts": now}
        return result
    except Exception as exc:
        if _bench_monthly_cache.get(sym):
            return {k: v for k, v in _bench_monthly_cache[sym].items() if k != "ts"}
        # Legacy fallback key for old _spy_monthly_cache callers
        raise HTTPException(status_code=503, detail=f"{sym} monthly fetch failed: {exc}")


def get_spy_monthly_data():
    """Internal helper — returns SPY monthly dict (same shape as the endpoint)."""
    try:
        import yfinance as yf
        from datetime import date as _date
        now = __import__("time").time()
        cached = _bench_monthly_cache.get("SPY", {})
        if cached.get("ts") and (now - cached["ts"]) < 900:
            return {k: v for k, v in cached.items() if k != "ts"}
        sym = "SPY"
        tkr = yf.Ticker(sym)
        start = f"{_date.today().year}-01-01"
        hist  = tkr.history(start=start, interval="1d")
        if hist.empty:
            raise ValueError("Empty SPY history")
        first_close = float(hist.iloc[0]["Close"])
        monthly = hist.resample("ME").last().dropna(subset=["Close"])
        points = []
        for ts, row in monthly.iterrows():
            close = float(row["Close"])
            points.append({
                "month":   ts.strftime("%Y-%m"),
                "close":   round(close, 2),
                "ytd_pct": round((close / first_close - 1.0) * 100, 4),
                "norm":    round(close / first_close, 6),
            })
        result = {"ticker": sym, "first_close": round(first_close, 2), "points": points}
        _bench_monthly_cache[sym] = {**result, "ts": now}
        return result
    except Exception:
        return None


# ─────────────────────────────────────────────────────────────────────────────
@app.post("/api/fund/{fund_id}/import-balance-history")
async def fund_import_balance_history(
    fund_id: str,
    request: Request,
    file: UploadFile = File(...),
):
    _require_fund_token(request)
    raw = await file.read()
    text = raw.decode('utf-8', errors='replace')
    records = _parse_balance_history_csv(text)
    if not records:
        raise HTTPException(400, "No valid monthly rows found in file.")
    # Derive YTD from the LATEST year present in the data, not the wall-clock
    # year — the CSV's most recent month tells us what "current year" means for
    # this account. This prevents stale or backdated demo data from showing
    # zero/missing months and prevents collisions when multiple years share
    # the same month number.
    import datetime as _dt
    years_in_data = sorted({r["year"] for r in records})
    cur_year = years_in_data[-1] if years_in_data else _dt.date.today().year
    ytd_months = [r for r in records if r["year"] == cur_year and not r.get("skip")]
    ytd_chain = 1.0
    for m in ytd_months:
        ytd_chain *= (1 + m["return_pct"] / 100)
    ytd_pct = round((ytd_chain - 1) * 100, 4)
    nav = records[-1]["end_balance"] if records else 0.0
    # Short month names for YTD chart labels (tooltip appends the year)
    _MN = {1:'Jan',2:'Feb',3:'Mar',4:'Apr',5:'May',6:'Jun',
           7:'Jul',8:'Aug',9:'Sep',10:'Oct',11:'Nov',12:'Dec'}
    # Include ALL current-year months (skip months shown as N/A placeholder in chart)
    mc_monthly = [
        {
            "label":       _MN.get(r["month"], r["label"]),
            "month":       r["month"],
            "return_pct":  r["return_pct"],
            "beg_balance": float(r.get("beg_balance") or 0),
            "end_balance": r["end_balance"],
            "skip":        r.get("skip", False),
            "perf_detail": {
                "deposits":    r.get("deposits",    0),
                "withdrawals": r.get("withdrawals", 0),
                "net_flow":    r.get("deposits", 0) - r.get("withdrawals", 0),
            },
        }
        for r in records if r["year"] == cur_year
    ]
    # Summary totals for positions-based return calculation on the client
    ytd_beg_balance      = float(ytd_months[0].get("beg_balance") or 0) if ytd_months else None
    ytd_total_deposits   = round(sum(float(m.get("deposits")    or 0) for m in ytd_months), 2)
    ytd_total_withdrawals = round(sum(float(m.get("withdrawals") or 0) for m in ytd_months), 2)
    result_json = {
        "md_return_pct":       ytd_pct,
        "monthly_chart":       {"monthly": mc_monthly},
        "ytd_beg_balance":     ytd_beg_balance,
        "ytd_total_deposits":  ytd_total_deposits,
        "ytd_total_withdrawals": ytd_total_withdrawals,
    }

    conn = _fund_conn()
    try:
        with conn.cursor(cursor_factory=_RealDictCursor) as cur:
            fid = _resolve_fund_id(cur, fund_id)
            cur.execute("""
                INSERT INTO account_balance_history (fund_id, data_json, updated_at)
                VALUES (%s, %s, now())
                ON CONFLICT (fund_id) DO UPDATE
                  SET data_json = EXCLUDED.data_json, updated_at = now()
            """, (fid, json.dumps(records)))
            cur.execute("""
                INSERT INTO managed_account_ytd_cache (fund_id, nav, ytd_pct, result_json, updated_at)
                VALUES (%s, %s, %s, %s, now())
                ON CONFLICT (fund_id) DO UPDATE
                  SET nav = EXCLUDED.nav, ytd_pct = EXCLUDED.ytd_pct,
                      result_json = EXCLUDED.result_json, updated_at = now()
            """, (fid, nav, ytd_pct, json.dumps(result_json)))
        conn.commit()
    finally:
        conn.close()
    ytd_sign = "+" if ytd_pct >= 0 else ""
    return {
        "ok": True, "months": len(records),
        "range": f"{records[0]['label']} – {records[-1]['label']}",
        "ytd_pct": ytd_pct, "ytd_label": f"{ytd_sign}{ytd_pct:.2f}%",
    }


@app.get("/api/fund/{fund_id}/balance-history")
async def fund_balance_history(fund_id: str, request: Request):
    # GPs and LPs (viewing their own accounts) are both allowed.
    claims = getattr(request.state, 'auth_claims', None)
    if not claims:
        _require_fund_token(request)  # legacy fund token fallback
    conn = _fund_conn()
    try:
        with conn.cursor(cursor_factory=_RealDictCursor) as cur:
            fid = _resolve_fund_id(cur, fund_id)
            cur.execute(
                "SELECT data_json, updated_at FROM account_balance_history WHERE fund_id = %s",
                (fid,)
            )
            row = cur.fetchone()
    finally:
        conn.close()

    if not row:
        return {"ok": False, "monthly": [], "quarterly": [], "annual": [], "updated_at": None}

    monthly = json.loads(row["data_json"])
    updated_at = row["updated_at"].isoformat() if row["updated_at"] else None

    def chain_returns(pts):
        result = 1.0
        for p in pts:
            if p.get("skip"):
                continue
            result *= (1 + p["return_pct"] / 100)
        return round((result - 1) * 100, 4)

    # Compute running cash-only balance for each monthly period.
    # cash_only = what the balance would be with 0% investment return:
    # starts at initial beg_balance, then each period adds net deposits only.
    cash_only_running = 0.0
    for p in monthly:
        if cash_only_running == 0.0 and p.get("beg_balance"):
            cash_only_running = float(p["beg_balance"])
        net = float(p.get("deposits") or 0) - float(p.get("withdrawals") or 0)
        cash_only_running += net
        p["cash_only_balance"] = round(cash_only_running, 2)

    quarterly = []
    for (yr, q), grp in groupby(monthly, key=lambda r: (r["year"], (r["month"] - 1) // 3 + 1)):
        pts = list(grp)
        quarterly.append({
            "year": yr, "quarter": q,
            "label": f"Q{q} {yr}",
            "beg_balance":       pts[0]["beg_balance"],
            "end_balance":       pts[-1]["end_balance"],
            "cash_only_balance": pts[-1].get("cash_only_balance"),
            "deposits":          round(sum(float(p.get("deposits") or 0) for p in pts), 2),
            "withdrawals":       round(sum(float(p.get("withdrawals") or 0) for p in pts), 2),
            "return_pct":        chain_returns(pts),
        })

    # Load fund display settings (benchmark choice set by GP)
    fund_settings  = _load_fund_settings(fid)
    benchmark_key  = fund_settings.get("benchmark", "sp500")
    benchmark_defn = _BENCHMARK_DEFS.get(benchmark_key, _BENCHMARK_DEFS["sp500"])
    benchmark_label = benchmark_defn["label"]

    # Fetch benchmark annual returns for all years in history
    years = sorted({p["year"] for p in monthly}) if monthly else []
    bmark_annual: dict[int, float] = _get_benchmark_annual(benchmark_key, years)

    # Manual overrides take precedence over Modified Dietz calculation
    manual_returns = _load_manual_annual_returns(fid)

    annual = []
    for yr, grp in groupby(monthly, key=lambda r: r["year"]):
        pts = list(grp)
        bmark_ret = bmark_annual.get(yr)
        port_ret  = manual_returns.get(yr) if yr in manual_returns else chain_returns(pts)
        annual.append({
            "year":                 yr,
            "label":                str(yr),
            "beg_balance":          pts[0]["beg_balance"],
            "end_balance":          pts[-1]["end_balance"],
            "cash_only_balance":    pts[-1].get("cash_only_balance"),
            "deposits":             round(sum(float(p.get("deposits") or 0) for p in pts), 2),
            "withdrawals":          round(sum(float(p.get("withdrawals") or 0) for p in pts), 2),
            "return_pct":           port_ret,
            "benchmark_return_pct": bmark_ret,
            "alpha":                round(port_ret - bmark_ret, 2) if bmark_ret is not None else None,
            "return_source":        "manual" if yr in manual_returns else "computed",
        })

    return {
        "ok":             True,
        "monthly":        monthly,
        "quarterly":      quarterly,
        "annual":         annual,
        "benchmark_key":  benchmark_key,
        "benchmark_label": benchmark_label,
        "period":         fund_settings.get("period", "all"),
        "updated_at":     updated_at,
    }


# Managed-account YTD calc — server-side wrapper that runs the existing
# unified Modified Dietz + attribution + monthly-balance computation,
# overlays the SPY benchmark series, and persists everything to the cache.
# ─────────────────────────────────────────────────────────────────────────────
@app.post("/api/fund/account/{fund_id}/ytd-run")
async def fund_account_ytd_run(
    fund_id: str,
    request: Request,
    positions_file:    UploadFile     = File(...),
    activity_file:     UploadFile     = File(...),
    monthly_perf_file: UploadFile     = File(None),
    begin_value:       float | None   = Form(None),
):
    """Run YTD for a managed account from 3 Fidelity CSVs, attach SPY
    benchmark, persist to managed_account_ytd_cache, return the merged
    result."""
    _require_fund_token(request)

    try:
        pos_text = (await positions_file.read()).decode("utf-8", errors="replace")
        act_text = (await activity_file.read()).decode("utf-8", errors="replace")
    except Exception as exc:
        raise HTTPException(422, f"Could not read Fidelity files: {exc}")

    mp_text = None
    if monthly_perf_file and monthly_perf_file.filename:
        try:
            mp_text = (await monthly_perf_file.read()).decode("utf-8", errors="replace")
        except Exception as exc:
            raise HTTPException(422, f"Could not read monthly performance file: {exc}")

    try:
        result = analyst.compute_unified_ytd(
            pos_text, act_text, begin_value, monthly_perf_text=mp_text,
        )
    except Exception as exc:
        raise HTTPException(500, f"YTD computation failed: {exc}\n\n{traceback.format_exc()}")

    if not result.get("ok"):
        raise HTTPException(422, result.get("error", "YTD computation returned no data"))

    # Attach SPY benchmark for chart overlay
    try:
        spy = get_spy_monthly_data()
        result["spy_monthly"] = spy
    except Exception:
        result["spy_monthly"] = None

    # Persist into managed_account_ytd_cache
    nav     = result.get("end_value") or result.get("end_nav") or 0.0
    ytd_pct = (result.get("md_return_pct") or result.get("ytd_pct")
               or result.get("modified_dietz") or 0.0)
    import json as _json
    conn = _fund_conn()
    try:
        with conn.cursor(cursor_factory=_RealDictCursor) as cur:
            # Resolve fund ID
            fid = _resolve_fund_id(cur, fund_id)

            cur.execute("""
                INSERT INTO managed_account_ytd_cache
                       (fund_id, nav, ytd_pct, result_json, updated_at)
                VALUES (%s, %s, %s, %s, now())
                ON CONFLICT (fund_id) DO UPDATE
                   SET nav         = EXCLUDED.nav,
                       ytd_pct     = EXCLUDED.ytd_pct,
                       result_json = EXCLUDED.result_json,
                       updated_at  = now()
            """, (fid, float(nav), float(ytd_pct), _json.dumps(result)))

            # Also sync positions into tax_lots so the Positions panel is populated.
            # Ensure CoA exists (managed accounts created via v2 API may not have it).
            try:
                _seed_coa_for_fund(cur, fid)
                positions_parsed = _parse_fidelity_csv(pos_text)
                if positions_parsed:
                    # Ensure COA exists (backfill accounts created before seeding was wired)
                    _seed_coa_for_fund(cur, fid)
                    # Fetch CoA account IDs
                    cur.execute("""
                        SELECT code, id FROM accounts
                         WHERE fund_id = %s AND code IN ('1020','1030','1100','3000')
                    """, (fid,))
                    acct_map  = {r["code"]: str(r["id"]) for r in cur.fetchall()}
                    sec_acct  = acct_map.get("1100")
                    mm_acct   = acct_map.get("1030") or acct_map.get("1020")
                    cap_acct  = acct_map.get("3000")

                    if sec_acct and cap_acct:
                        # Upsert securities
                        sec_ids = {}
                        for p in positions_parsed:
                            cur.execute("""
                                INSERT INTO securities (symbol, name, asset_class, is_public)
                                VALUES (%s, %s, %s, TRUE)
                                ON CONFLICT (symbol) DO UPDATE SET name = EXCLUDED.name
                                RETURNING id
                            """, (p["symbol"], p["name"] or p["symbol"],
                                  "cash" if p["is_cash"] else "equity"))
                            sec_ids[p["symbol"]] = str(cur.fetchone()["id"])

                        # Close existing open lots
                        cur.execute("""
                            UPDATE tax_lots SET closed_at = NOW()
                             WHERE fund_id = %s AND closed_at IS NULL
                        """, (fid,))

                        # Create adjustment transaction
                        today_iso = datetime.utcnow().date().isoformat()
                        cur.execute("""
                            INSERT INTO transactions
                                (fund_id, effective_date, category, description)
                            VALUES (%s, %s, 'adjustment', 'Position import from YTD run')
                            RETURNING id
                        """, (fid, today_iso))
                        txn_id = str(cur.fetchone()["id"])

                        # Transaction lines (Dr positions, Cr capital)
                        total_cost_lots = 0.0
                        ln = 1
                        for p in positions_parsed:
                            cost = float(p["cost_basis"] or 0)
                            if cost <= 0:
                                continue
                            total_cost_lots += cost
                            acct = mm_acct if p["is_cash"] else sec_acct
                            if not acct:
                                continue
                            cur.execute("""
                                INSERT INTO transaction_lines
                                    (transaction_id, line_number, account_id, debit, security_id)
                                VALUES (%s, %s, %s, %s, %s)
                            """, (txn_id, ln, acct, round(cost, 4),
                                  sec_ids.get(p["symbol"])))
                            ln += 1
                        if total_cost_lots > 0:
                            cur.execute("""
                                INSERT INTO transaction_lines
                                    (transaction_id, line_number, account_id, credit)
                                VALUES (%s, %s, %s, %s)
                            """, (txn_id, ln, cap_acct, round(total_cost_lots, 4)))

                        # Insert tax lots
                        for p in positions_parsed:
                            cur.execute("""
                                INSERT INTO tax_lots
                                    (fund_id, security_id, acquired_at,
                                     quantity, cost_basis_per_unit, open_transaction_id)
                                VALUES (%s, %s, NOW(), %s, %s, %s)
                            """, (fid, sec_ids[p["symbol"]],
                                  float(p["quantity"]),
                                  float(p["avg_cost"]) if p["avg_cost"] else 0.0,
                                  txn_id))
            except Exception as _pos_exc:
                # Non-fatal: YTD data was already saved; positions sync is best-effort
                print(f"⚠️  positions sync failed for {fund_id}: {_pos_exc}")

            # If Bal.Detail (monthly_perf_file) was provided, also persist as
            # all-time balance history so the All-Time chart is populated.
            if mp_text:
                try:
                    bh_records = _parse_balance_history_csv(mp_text)
                    if bh_records:
                        cur.execute("""
                            INSERT INTO account_balance_history (fund_id, data_json, updated_at)
                            VALUES (%s, %s, now())
                            ON CONFLICT (fund_id) DO UPDATE
                              SET data_json = EXCLUDED.data_json, updated_at = now()
                        """, (fid, _json.dumps(bh_records)))
                except Exception as _bh_exc:
                    print(f"⚠️  balance history sync failed for {fund_id}: {_bh_exc}")

        conn.commit()
    finally:
        conn.close()

    return result


# ---------------------------------------------------------------------------
# Fund administration — delete
# ---------------------------------------------------------------------------

def _safe_delete(cur, sql: str, params: tuple) -> None:
    """Execute a DELETE inside a savepoint.

    Silently skips if the table (or a referenced table in a subquery) does not
    exist — psycopg2 puts the SQLSTATE code on exc.pgcode, NOT in str(exc), so
    we check that attribute directly.

    SQLSTATE codes we treat as non-fatal:
      42P01 — undefined_table
      42703 — undefined_column
    All other exceptions are re-raised so real errors surface to the caller.
    """
    _SKIP_CODES = ("42P01", "42703")
    cur.execute("SAVEPOINT _del_sp")
    try:
        cur.execute(sql, params)
        cur.execute("RELEASE SAVEPOINT _del_sp")
    except Exception as exc:
        cur.execute("ROLLBACK TO SAVEPOINT _del_sp")
        pgcode = getattr(exc, "pgcode", "") or ""
        if pgcode not in _SKIP_CODES:
            raise  # real error — bubble up with full context


@app.delete("/api/fund/admin/delete")
async def fund_admin_delete(request: Request, fund_id: str):
    """Permanently delete a fund and ALL its associated data.

    Requires the x-delete-password header to match the FUND_DELETE_PASSWORD
    environment variable.  If FUND_DELETE_PASSWORD is not set the endpoint
    is disabled (returns 503).
    """
    _require_fund_token(request)

    # ── Verify FUND_DELETE_PASSWORD ──────────────────────────────────────────
    required_pw = os.environ.get("FUND_DELETE_PASSWORD", "").strip()
    if not required_pw:
        raise HTTPException(
            status_code=503,
            detail="FUND_DELETE_PASSWORD is not configured on this server.",
        )
    submitted_pw = request.headers.get("x-delete-password", "").strip()
    if not hmac.compare_digest(submitted_pw, required_pw):
        raise HTTPException(status_code=403, detail="Incorrect delete password.")

    conn = _fund_conn()
    try:
        with conn.cursor(cursor_factory=_RealDictCursor) as cur:
            fid = _resolve_fund_id(cur, fund_id)
            cur.execute("SELECT name FROM funds WHERE id = %s", (fid,))
            row = cur.fetchone()
            if not row:
                raise HTTPException(404, "Fund not found")
            name = row["name"]

            # ── 1. Grandchild rows (FK → non-funds parent) ────────────────────
            _safe_delete(cur, """
                DELETE FROM transaction_lines
                 WHERE transaction_id IN (
                       SELECT id FROM transactions WHERE fund_id = %s)
            """, (fid,))
            _safe_delete(cur, """
                DELETE FROM capital_call_allocations
                 WHERE capital_call_id IN (
                       SELECT id FROM capital_calls WHERE fund_id = %s)
            """, (fid,))
            _safe_delete(cur, """
                DELETE FROM distribution_allocations
                 WHERE distribution_id IN (
                       SELECT id FROM distributions WHERE fund_id = %s)
            """, (fid,))
            _safe_delete(cur, """
                DELETE FROM nav_snapshot_lp
                 WHERE snapshot_id IN (
                       SELECT id FROM nav_snapshots WHERE fund_id = %s)
            """, (fid,))
            _safe_delete(cur, """
                DELETE FROM carry_allocations
                 WHERE carry_run_id IN (
                       SELECT id FROM carry_runs WHERE fund_id = %s)
            """, (fid,))
            _safe_delete(cur, """
                DELETE FROM mgmt_fee_allocations
                 WHERE mgmt_fee_run_id IN (
                       SELECT id FROM mgmt_fee_runs WHERE fund_id = %s)
            """, (fid,))
            _safe_delete(cur, """
                DELETE FROM commitments
                 WHERE lp_id IN (SELECT id FROM lps WHERE fund_id = %s)
            """, (fid,))
            _safe_delete(cur, """
                DELETE FROM lp_statements
                 WHERE lp_id IN (SELECT id FROM lps WHERE fund_id = %s)
            """, (fid,))

            # ── 2. Direct fund_id children (order matters for remaining FKs) ──
            for tbl in (
                "transactions",
                "capital_calls",
                "distributions",
                "nav_snapshots",
                "carry_runs",
                "mgmt_fee_runs",
                "lps",
                "tax_lots",
                "accounts",
                "managed_account_ytd_cache",
                "audit_log",
            ):
                _safe_delete(cur, f"DELETE FROM {tbl} WHERE fund_id = %s", (fid,))  # noqa: S608

            # ── 3. Finally remove the fund itself ─────────────────────────────
            cur.execute("DELETE FROM funds WHERE id = %s", (fid,))
        conn.commit()
        return {"deleted": fid, "name": name}
    except HTTPException:
        raise
    except Exception as exc:
        try:
            conn.rollback()
        except Exception:
            pass
        raise HTTPException(
            status_code=500,
            detail=f"Delete failed: {type(exc).__name__}: {exc}",
        )
    finally:
        conn.close()


# ---------------------------------------------------------------------------
# GP-auth'd hard purge — no password, no backup, deletes everything
# ---------------------------------------------------------------------------

@app.delete("/api/v2/gp/fund/{fund_id}/purge")
async def fund_purge(fund_id: str, request: Request):
    """Permanently purge a fund + all child rows. GP JWT required. No backup kept."""
    claims = _claims_or_401(request)
    if claims.get("role") != "gp":
        raise HTTPException(403, "GP only")
    conn = _fund_conn()
    try:
        with conn.cursor(cursor_factory=_RealDictCursor) as cur:
            fid = _resolve_fund_id(cur, fund_id)
            cur.execute("SELECT name, short_name FROM funds WHERE id = %s", (fid,))
            row = cur.fetchone()
            if not row:
                raise HTTPException(404, "Fund not found")
            name = row["name"]
            # Grandchild rows (FK chains)
            _safe_delete(cur, "DELETE FROM transaction_lines WHERE transaction_id IN (SELECT id FROM transactions WHERE fund_id = %s)", (fid,))
            _safe_delete(cur, "DELETE FROM capital_call_allocations WHERE capital_call_id IN (SELECT id FROM capital_calls WHERE fund_id = %s)", (fid,))
            _safe_delete(cur, "DELETE FROM distribution_allocations WHERE distribution_id IN (SELECT id FROM distributions WHERE fund_id = %s)", (fid,))
            _safe_delete(cur, "DELETE FROM nav_snapshot_lp WHERE snapshot_id IN (SELECT id FROM nav_snapshots WHERE fund_id = %s)", (fid,))
            _safe_delete(cur, "DELETE FROM carry_allocations WHERE carry_run_id IN (SELECT id FROM carry_runs WHERE fund_id = %s)", (fid,))
            _safe_delete(cur, "DELETE FROM mgmt_fee_allocations WHERE mgmt_fee_run_id IN (SELECT id FROM mgmt_fee_runs WHERE fund_id = %s)", (fid,))
            _safe_delete(cur, "DELETE FROM commitments WHERE lp_id IN (SELECT id FROM lps WHERE fund_id = %s)", (fid,))
            _safe_delete(cur, "DELETE FROM lp_statements WHERE lp_id IN (SELECT id FROM lps WHERE fund_id = %s)", (fid,))
            # Direct fund_id children
            for tbl in ("account_balance_history", "managed_account_ytd_cache",
                        "transactions", "capital_calls", "distributions",
                        "nav_snapshots", "carry_runs", "mgmt_fee_runs",
                        "lps", "tax_lots", "accounts", "audit_log",
                        "annual_lp_balances", "fund_annual_snapshots"):
                _safe_delete(cur, f"DELETE FROM {tbl} WHERE fund_id = %s", (fid,))  # noqa: S608
            cur.execute("DELETE FROM funds WHERE id = %s", (fid,))
        conn.commit()
        return {"ok": True, "deleted": name}
    except HTTPException:
        raise
    except Exception as exc:
        conn.rollback()
        raise HTTPException(500, f"Purge failed: {exc}")
    finally:
        conn.close()


# ---------------------------------------------------------------------------
# Merge two balance-history CSVs → download combined CSV
# ---------------------------------------------------------------------------

@app.post("/api/fund/merge-balance-history")
async def merge_balance_history(
    request: Request,
    file1: UploadFile = File(...),
    file2: UploadFile = File(...),
):
    """Parse two Fidelity Investment Income & Balance Detail CSVs and return
    a merged CSV where overlapping months are summed field-by-field."""
    _require_fund_token(request)
    t1 = (await file1.read()).decode("utf-8", errors="replace")
    t2 = (await file2.read()).decode("utf-8", errors="replace")
    r1 = _parse_balance_history_csv(t1)
    r2 = _parse_balance_history_csv(t2)
    if not r1 and not r2:
        raise HTTPException(400, "No valid monthly rows found in either file.")
    map1 = {(r["year"], r["month"]): r for r in r1}
    map2 = {(r["year"], r["month"]): r for r in r2}
    SUM_FIELDS = ("beg_balance", "end_balance", "market_change",
                  "dividends", "interest", "deposits", "withdrawals", "fees")
    merged = []
    for key in sorted(set(map1) | set(map2), reverse=True):
        a, b = map1.get(key), map2.get(key)
        if a and b:
            row = {"label": a["label"], "year": a["year"], "month": a["month"]}
            for f in SUM_FIELDS:
                row[f] = (a.get(f) or 0.0) + (b.get(f) or 0.0)
        else:
            src = a or b
            row = {f: src.get(f, 0.0) for f in ("label", "year", "month") + SUM_FIELDS}
        merged.append(row)

    def _fm(v):
        return f'-${abs(v):,.2f}' if v < 0 else f'${v:,.2f}'

    lines = [
        "Investment income Export",
        '"Income For: COMBINED"',
        f'"Investment income - ({merged[-1]["label"]} - {merged[0]["label"]})"',
        '"Monthly","Beginning balance","Market change","Dividends","Interest","Deposits","Withdrawals","Net advisory fees","Ending balance"',
    ]
    for r in merged:
        lines.append(",".join([
            f'"{r["label"]}"', f'"{_fm(r["beg_balance"])}"',
            f'"{_fm(r["market_change"])}"', f'"{_fm(r["dividends"])}"',
            f'"{_fm(r["interest"])}"', f'"{_fm(r["deposits"])}"',
            f'"{_fm(r["withdrawals"])}"', f'"{_fm(r["fees"])}"',
            f'"{_fm(r["end_balance"])}"',
        ]))
    from fastapi.responses import Response as _Resp
    return _Resp(
        content="\n".join(lines),
        media_type="text/csv",
        headers={"Content-Disposition": 'attachment; filename="merged_balance_history.csv"'},
    )


# ---------------------------------------------------------------------------
# Fund export — Excel workbook
# ---------------------------------------------------------------------------

@app.get("/api/fund/export-excel")
async def fund_export_excel(request: Request, fund_id: str = None):
    """Generate a comprehensive Excel workbook for the fund and return it
    as a downloadable .xlsx file.

    Sheets:
      1. Fund Summary    — key metrics & economics overview
      2. Portfolio       — all positions with live prices & P/L
      3. Limited Partners — LP commitments & current values
      4. Annual Waterfall — year-by-year hurdle / carry / GP equity
      5. Transactions    — recent activity log
    """
    _require_fund_token(request)
    if not _OPENPYXL_OK:
        raise HTTPException(400, "openpyxl not installed on this server")

    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from fastapi.responses import StreamingResponse

    conn = _fund_conn()
    try:
        with conn.cursor(cursor_factory=_RealDictCursor) as cur:
            fid = _resolve_fund_id(cur, fund_id)

            # ── Fund metadata ─────────────────────────────────────────────────
            cur.execute("""
                SELECT name, short_name, inception_date, status,
                       mgmt_fee_pct, carry_pct, hurdle_pct, fund_type
                  FROM funds WHERE id = %s
            """, (fid,))
            fund = dict(cur.fetchone())
            is_acct = fund.get('fund_type') == 'managed_account'

            # ── Market NAV ────────────────────────────────────────────────────
            nav = _fund_market_nav(cur, fid)

            # ── Positions ─────────────────────────────────────────────────────
            cur.execute("""
                SELECT s.symbol, s.name AS sec_name, s.asset_class,
                       SUM(tl.quantity) AS qty,
                       SUM(tl.quantity * tl.cost_basis_per_unit)
                         / NULLIF(SUM(tl.quantity), 0)  AS avg_cost,
                       SUM(tl.quantity * tl.cost_basis_per_unit) AS cost_basis
                  FROM tax_lots tl
                  JOIN securities s ON s.id = tl.security_id
                 WHERE tl.fund_id = %s AND tl.closed_at IS NULL
                 GROUP BY s.symbol, s.name, s.asset_class
                 ORDER BY SUM(tl.quantity * tl.cost_basis_per_unit) DESC
            """, (fid,))
            positions = [dict(r) for r in cur.fetchall()]
            symbols   = [p['symbol'] for p in positions if p['symbol']]
            prices    = _fetch_prices(symbols)

            # ── YTD cache (managed accounts only) ────────────────────────────
            ytd_cache = None
            if is_acct:
                try:
                    cur.execute("""
                        SELECT nav, ytd_pct, result_json, updated_at
                          FROM managed_account_ytd_cache WHERE fund_id = %s
                    """, (fid,))
                    row = cur.fetchone()
                    if row:
                        import json as _json2
                        ytd_cache = dict(row)
                        if ytd_cache.get('result_json'):
                            ytd_cache['result'] = _json2.loads(ytd_cache['result_json'])
                except Exception:
                    ytd_cache = None

            # ── LPs (LP funds only) ───────────────────────────────────────────
            lps             = []
            total_committed = 0.0
            snapshots       = []
            txns            = []
            if not is_acct:
                cur.execute("""
                    SELECT l.legal_name, l.entity_type,
                           COALESCE(SUM(c.commitment_amount), 0) AS commitment
                      FROM lps l
                      LEFT JOIN commitments c ON c.lp_id = l.id AND c.superseded_by IS NULL
                     WHERE l.fund_id = %s AND l.status = 'active'
                     GROUP BY l.id, l.legal_name, l.entity_type
                    HAVING COALESCE(SUM(c.commitment_amount), 0) > 0
                     ORDER BY commitment DESC
                """, (fid,))
                lps = [dict(r) for r in cur.fetchall()]
                total_committed = sum(float(lp['commitment']) for lp in lps) or 0.0

                try:
                    cur.execute("""
                        SELECT year, start_nav, end_nav, contributions,
                               hurdle_amount, gross_profit, carry_earned,
                               carry_paid, carry_rolled, gp_equity_end
                          FROM fund_annual_snapshots
                         WHERE fund_id = %s ORDER BY year ASC
                    """, (fid,))
                    snapshots = [dict(r) for r in cur.fetchall()]
                except Exception:
                    snapshots = []

                cur.execute("""
                    SELECT t.effective_date, t.category, t.description,
                           ROUND(SUM(CASE WHEN tl.debit  > 0 THEN tl.debit  ELSE 0 END), 2) AS total_debit,
                           ROUND(SUM(CASE WHEN tl.credit > 0 THEN tl.credit ELSE 0 END), 2) AS total_credit
                      FROM transactions t
                      JOIN transaction_lines tl ON tl.transaction_id = t.id
                     WHERE t.fund_id = %s
                     GROUP BY t.id, t.effective_date, t.category, t.description
                     ORDER BY t.effective_date DESC
                     LIMIT 200
                """, (fid,))
                txns = [dict(r) for r in cur.fetchall()]

    finally:
        conn.close()

    # ── Managed account: separate workbook ───────────────────────────────────
    if is_acct:
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        from fastapi.responses import StreamingResponse as _SR2

        NAVY2  = '0E1D38'; GOLD2  = 'C9A84C'; LGRAY2 = 'F4F6F9'
        WHITE2 = 'FFFFFF'; DKGRAY2= '444444'; GREEN2 = '1A7F40'; RED2 = 'CC3333'
        thin2 = Side(style='thin', color='CCCCCC')
        bdr2  = Border(left=thin2, right=thin2, top=thin2, bottom=thin2)

        def _fill2(h): return PatternFill('solid', fgColor=h)
        def _hf2(sz=10, c=GOLD2): return Font(bold=True, color=c, size=sz, name='Calibri')
        def _bf2(sz=10, bold=False, c=DKGRAY2): return Font(bold=bold, size=sz, color=c, name='Calibri')
        def _ctr2(): return Alignment(horizontal='center', vertical='center')
        def _rgt2(): return Alignment(horizontal='right', vertical='center')

        def _arow2(ws, ri, vals, fmts=None, bg=None, bold=False, fc=DKGRAY2):
            for ci, val in enumerate(vals, 1):
                cell = ws.cell(row=ri, column=ci, value=val)
                cell.font   = Font(bold=bold, size=10, color=fc, name='Calibri')
                cell.border = bdr2
                if bg: cell.fill = _fill2(bg)
                if fmts and ci-1 < len(fmts) and fmts[ci-1]:
                    cell.number_format = fmts[ci-1]; cell.alignment = _rgt2()
                else:
                    cell.alignment = Alignment(vertical='center')

        def _whdr2(ws, ri, cols, bg=NAVY2):
            for ci, c in enumerate(cols, 1):
                cell = ws.cell(row=ri, column=ci, value=c)
                cell.font = _hf2(); cell.fill = _fill2(bg)
                cell.border = bdr2; cell.alignment = _ctr2()
            ws.row_dimensions[ri].height = 18

        def _aw2(ws, mn=10, mx=40):
            for col in ws.columns:
                best = mn
                for cell in col:
                    try: best = max(best, len(str(cell.value or '')) + 2)
                    except: pass
                ws.column_dimensions[get_column_letter(col[0].column)].width = min(best, mx)

        def _sec2(ws, ri, text, span):
            cell = ws.cell(row=ri, column=1, value=text)
            cell.font = Font(bold=True, size=11, color=NAVY2, name='Calibri')
            cell.fill = _fill2('E8F0F8'); cell.border = bdr2
            cell.alignment = Alignment(horizontal='left', vertical='center')
            ws.merge_cells(start_row=ri, start_column=1, end_row=ri, end_column=span)
            ws.row_dimensions[ri].height = 16

        FMT_USD = '$#,##0.00'; FMT_USD0 = '$#,##0'
        FMT_PCT = '0.00%';    FMT_PCT1 = '0.0%'
        FMT_NUM = '#,##0.##'; FMT_DATE = 'YYYY-MM-DD'
        today_str2 = datetime.utcnow().strftime('%B %d, %Y')

        result   = (ytd_cache or {}).get('result') or {}
        ytd_pct  = float((ytd_cache or {}).get('ytd_pct') or result.get('md_return_pct') or 0) / 100.0
        begin_v  = float(result.get('begin_value') or 0)
        end_v    = float(result.get('end_value') or nav or 0)
        net_flow = float(result.get('net_flow') or 0)
        attribution = result.get('attribution') or []
        monthly_raw = (result.get('monthly_chart') or {}).get('monthly') or []
        # spy_monthly is stored as {ticker, first_close, points: [...]} dict
        spy_data    = result.get('spy_monthly') or {}
        spy_points  = (spy_data.get('points') or []) if isinstance(spy_data, dict) else []
        flows       = result.get('flows') or []
        updated_at  = str((ytd_cache or {}).get('updated_at') or '')[:10] or today_str2

        # Build SPY lookup by month integer (1-12)
        spy_by_month_int = {}
        for pt in spy_points:
            try:
                m_int = int(str(pt.get('month', '')).split('-')[1])
                spy_by_month_int[m_int] = pt
            except Exception:
                pass

        spy_ytd = float(spy_points[-1].get('ytd_pct') or 0) / 100.0 if spy_points else 0.0
        alpha = ytd_pct - spy_ytd

        wb2 = openpyxl.Workbook(); wb2.remove(wb2.active)

        # ── Sheet 1: Account Summary ──────────────────────────────────────────
        s1 = wb2.create_sheet('Account Summary')
        s1.sheet_view.showGridLines = False
        t = s1.cell(row=1, column=1, value=fund['name'])
        t.font = Font(bold=True, size=16, color=NAVY2, name='Calibri')
        s1.merge_cells('A1:D1'); s1.row_dimensions[1].height = 28
        sub2 = s1.cell(row=2, column=1,
                       value=f'DGA Capital — Managed Account Report  |  As of {today_str2}')
        sub2.font = Font(size=10, color='888888', name='Calibri')
        s1.merge_cells('A2:D2')

        r = 4
        _sec2(s1, r, 'ACCOUNT INFORMATION', 4); r += 1
        info_rows = [
            ('Account Name',   fund['name']),
            ('Short Name',     fund.get('short_name', '')),
            ('Inception Date', str(fund.get('inception_date', ''))[:10]),
            ('Status',         (fund.get('status') or '').title()),
            ('YTD Data As Of', updated_at),
        ]
        for i, (lbl, val) in enumerate(info_rows):
            bg = LGRAY2 if i % 2 == 0 else WHITE2
            _arow2(s1, r, [lbl, val, '', ''], bg=bg)
            s1.cell(row=r, column=1).font = Font(bold=True, size=10, color=DKGRAY2, name='Calibri')
            r += 1

        r += 1
        _sec2(s1, r, 'YTD PERFORMANCE', 4); r += 1
        perf_data = [
            ('Account NAV (End Value)', end_v,    FMT_USD0),
            ('Jan 1 Value (Begin)',     begin_v,  FMT_USD0),
            ('Net External Flows',      net_flow, FMT_USD0),
            ('YTD Return (Mod Dietz)',  ytd_pct,  FMT_PCT),
            ('SPY YTD Return',          spy_ytd,  FMT_PCT),
            ('Alpha vs SPY',            alpha,    FMT_PCT),
        ]
        for i, (lbl, val, fmt) in enumerate(perf_data):
            bg = LGRAY2 if i % 2 == 0 else WHITE2
            _arow2(s1, r, [lbl, val, '', ''], fmts=[None, fmt, None, None], bg=bg)
            s1.cell(row=r, column=1).font = Font(bold=True, size=10, color=DKGRAY2, name='Calibri')
            vc = GREEN2 if lbl in ('Alpha vs SPY', 'YTD Return (Mod Dietz)') and val >= 0 \
                 else RED2 if lbl in ('Alpha vs SPY', 'YTD Return (Mod Dietz)') and val < 0 \
                 else DKGRAY2
            s1.cell(row=r, column=2).font = Font(bold=True, size=10, color=vc, name='Calibri')
            r += 1

        s1.column_dimensions['A'].width = 30
        s1.column_dimensions['B'].width = 22
        s1.sheet_properties.tabColor = GOLD2

        # ── Sheet 2: Portfolio Positions ──────────────────────────────────────
        s2 = wb2.create_sheet('Portfolio Positions')
        s2.sheet_view.showGridLines = False; s2.freeze_panes = 'A2'
        pos_cols2 = ['Symbol', 'Security Name', 'Asset Class', 'Shares',
                     'Avg Cost', 'Cost Basis', 'Last Price', 'Market Value',
                     'Unrealized P/L', 'P/L %', 'Weight %']
        _whdr2(s2, 1, pos_cols2)
        total_mv2 = total_cb2 = total_unrl2 = 0.0
        for i, p in enumerate(positions):
            qty = float(p['qty'] or 0); avg = float(p['avg_cost'] or 0)
            cb  = float(p['cost_basis'] or 0)
            last = prices.get(p['symbol']) or avg
            mv  = qty * last; unrl = mv - cb
            pl_pct = (unrl / cb) if cb else 0.0
            wt_pct = (mv / end_v) if end_v else 0.0
            total_mv2 += mv; total_cb2 += cb; total_unrl2 += unrl
            bg = WHITE2 if i % 2 == 0 else LGRAY2
            _arow2(s2, i+2, [p['symbol'], p['sec_name'], p['asset_class'],
                             qty, avg, cb, last, mv, unrl, pl_pct, wt_pct],
                   fmts=[None, None, None, FMT_NUM, FMT_USD, FMT_USD0,
                         FMT_USD, FMT_USD0, FMT_USD0, FMT_PCT1, FMT_PCT1], bg=bg)
            c = s2.cell(row=i+2, column=9)
            c.font = Font(size=10, color=GREEN2 if unrl >= 0 else RED2, name='Calibri')
        tr2 = len(positions) + 2
        _arow2(s2, tr2, ['TOTAL', '', '', '', '', total_cb2, '', total_mv2, total_unrl2, '', ''],
               fmts=[None]*5 + [FMT_USD0, None, FMT_USD0, FMT_USD0, None, None],
               bg=NAVY2, bold=True, fc=GOLD2)
        _aw2(s2); s2.column_dimensions['B'].width = 38
        s2.sheet_properties.tabColor = NAVY2

        # ── Sheet 3: Monthly Performance ──────────────────────────────────────
        if monthly_raw:
            s3 = wb2.create_sheet('Monthly Performance')
            s3.sheet_view.showGridLines = False; s3.freeze_panes = 'A2'
            _whdr2(s3, 1, ['Month', 'Start Value', 'End Value', 'Dollar Gain',
                           'Monthly Return', 'SPY Monthly Return', 'Alpha'])
            for i, mo in enumerate(monthly_raw):
                m_int  = int(mo.get('month') or 0)
                label  = str(mo.get('label') or m_int)
                sv     = float(mo.get('start_value') or 0)
                ev     = float(mo.get('end_value') or 0)
                dg     = float(mo.get('dollar_gain') or 0)
                ret_pct = float(mo.get('return_pct') or 0) / 100.0
                spy_pt  = spy_by_month_int.get(m_int) or {}
                # compute SPY monthly return from ytd_pct series
                prev_pt = spy_by_month_int.get(m_int - 1) if m_int > 1 else None
                spy_ytd_m  = float(spy_pt.get('ytd_pct') or 0)
                spy_ytd_pr = float(prev_pt.get('ytd_pct') or 0) if prev_pt else 0.0
                spy_mo_ret = ((spy_ytd_m - spy_ytd_pr) / 100.0) if spy_pt else None
                alpha_mo   = (ret_pct - spy_mo_ret) if spy_mo_ret is not None else None
                bg = WHITE2 if i % 2 == 0 else LGRAY2
                _arow2(s3, i+2,
                       [label, sv, ev, dg, ret_pct, spy_mo_ret, alpha_mo],
                       fmts=[None, FMT_USD0, FMT_USD0, FMT_USD0,
                             FMT_PCT1, FMT_PCT1, FMT_PCT1], bg=bg)
                rc = s3.cell(row=i+2, column=5)
                rc.font = Font(size=10, color=GREEN2 if ret_pct >= 0 else RED2, name='Calibri')
                if alpha_mo is not None:
                    ac = s3.cell(row=i+2, column=7)
                    ac.font = Font(size=10, color=GREEN2 if alpha_mo >= 0 else RED2, name='Calibri')
            _aw2(s3)
            s3.sheet_properties.tabColor = NAVY2

        # ── Sheet 4: YTD Attribution ──────────────────────────────────────────
        if attribution:
            s4 = wb2.create_sheet('YTD Attribution')
            s4.sheet_view.showGridLines = False; s4.freeze_panes = 'A2'
            _whdr2(s4, 1, ['Ticker', 'End Shares', 'Jan 1 Price', 'End Price',
                           'Dollar Gain', 'Portfolio Contribution %', 'Stock Return %'])
            for i, a in enumerate(sorted(attribution,
                                         key=lambda x: float(x.get('contribution_pct') or 0),
                                         reverse=True)):
                dg      = float(a.get('dollar_gain') or 0)
                contrib = float(a.get('contribution_pct') or 0) / 100.0
                tk_ret  = a.get('ticker_return_pct')
                tk_ret_f = float(tk_ret) / 100.0 if tk_ret is not None else None
                bg = WHITE2 if i % 2 == 0 else LGRAY2
                _arow2(s4, i+2,
                       [a.get('ticker', ''),
                        float(a.get('end_shares') or 0),
                        float(a.get('jan1_price') or 0),
                        float(a.get('end_price') or 0),
                        dg, contrib, tk_ret_f],
                       fmts=[None, FMT_NUM, FMT_USD, FMT_USD,
                             FMT_USD0, FMT_PCT, FMT_PCT1],
                       bg=bg)
                c4 = s4.cell(row=i+2, column=5)
                c4.font = Font(size=10, color=GREEN2 if dg >= 0 else RED2, name='Calibri')
                c4b = s4.cell(row=i+2, column=6)
                c4b.font = Font(size=10, color=GREEN2 if contrib >= 0 else RED2, name='Calibri')
            _aw2(s4)
            s4.sheet_properties.tabColor = NAVY2

        # ── Sheet 5: Cash Flow History ────────────────────────────────────────
        if flows:
            s5 = wb2.create_sheet('Cash Flow History')
            s5.sheet_view.showGridLines = False; s5.freeze_panes = 'A2'
            _whdr2(s5, 1, ['Date', 'Description', 'Amount', 'Type'])
            for i, f in enumerate(sorted(flows, key=lambda x: x.get('date', ''), reverse=True)):
                amt = float(f.get('amount') or 0)
                bg  = WHITE2 if i % 2 == 0 else LGRAY2
                _arow2(s5, i+2,
                       [f.get('date', ''), f.get('description', ''), amt,
                        'Inflow' if amt > 0 else 'Outflow'],
                       fmts=[None, None, FMT_USD0, None], bg=bg)
                s5.cell(row=i+2, column=3).font = Font(
                    size=10, color=GREEN2 if amt > 0 else RED2, name='Calibri')
            _aw2(s5); s5.column_dimensions['B'].width = 38
            s5.sheet_properties.tabColor = NAVY2

        buf2 = io.BytesIO(); wb2.save(buf2); buf2.seek(0)
        safe2 = fund.get('short_name', 'Account').replace('/', '-').replace('\\', '-')
        fname2 = f"{safe2}_AccountReport_{datetime.utcnow().strftime('%Y%m%d')}.xlsx"
        return _SR2(
            buf2,
            media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={'Content-Disposition': f'attachment; filename="{fname2}"'},
        )

    # ── Build workbook (LP fund) ──────────────────────────────────────────────
    wb = openpyxl.Workbook()
    wb.remove(wb.active)   # remove default sheet

    # ── Style helpers ─────────────────────────────────────────────────────────
    NAVY  = '0E1D38'
    GOLD  = 'C9A84C'
    LGRAY = 'F4F6F9'
    WHITE = 'FFFFFF'
    DKGRAY= '444444'
    GREEN = '1A7F40'
    RED   = 'CC3333'

    thin = Side(style='thin', color='CCCCCC')
    bdr  = Border(left=thin, right=thin, top=thin, bottom=thin)

    def fill(hex_color):
        return PatternFill('solid', fgColor=hex_color)

    def hdr_font(size=10, color=GOLD):
        return Font(bold=True, color=color, size=size, name='Calibri')

    def body_font(size=10, bold=False, color=DKGRAY):
        return Font(bold=bold, size=size, color=color, name='Calibri')

    def center():
        return Alignment(horizontal='center', vertical='center')

    def right():
        return Alignment(horizontal='right', vertical='center')

    def apply_row(ws, row_idx, values, formats=None, bg=None, bold=False,
                  font_color=DKGRAY):
        for ci, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=ci, value=val)
            cell.font   = Font(bold=bold, size=10, color=font_color, name='Calibri')
            cell.border = bdr
            if bg:
                cell.fill = fill(bg)
            if formats and ci - 1 < len(formats) and formats[ci - 1]:
                cell.number_format = formats[ci - 1]
                cell.alignment = right()
            else:
                cell.alignment = Alignment(vertical='center', wrap_text=False)

    def write_header(ws, row_idx, cols, bg=NAVY):
        for ci, col in enumerate(cols, 1):
            cell = ws.cell(row=row_idx, column=ci, value=col)
            cell.font      = hdr_font()
            cell.fill      = fill(bg)
            cell.border    = bdr
            cell.alignment = center()
        ws.row_dimensions[row_idx].height = 18

    def auto_width(ws, min_w=10, max_w=40):
        for col in ws.columns:
            best = min_w
            for cell in col:
                try:
                    best = max(best, len(str(cell.value or '')) + 2)
                except Exception:
                    pass
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(best, max_w)

    def section_title(ws, row_idx, text, span):
        cell = ws.cell(row=row_idx, column=1, value=text)
        cell.font      = Font(bold=True, size=11, color=NAVY, name='Calibri')
        cell.fill      = fill('E8F0F8')
        cell.border    = bdr
        cell.alignment = Alignment(horizontal='left', vertical='center')
        ws.merge_cells(start_row=row_idx, start_column=1,
                       end_row=row_idx, end_column=span)
        ws.row_dimensions[row_idx].height = 16

    today_str = datetime.utcnow().strftime('%B %d, %Y')
    FMT_USD   = '$#,##0.00'
    FMT_USD0  = '$#,##0'
    FMT_PCT   = '0.00%'
    FMT_PCT1  = '0.0%'
    FMT_NUM   = '#,##0.##'
    FMT_DATE  = 'YYYY-MM-DD'

    # ── Sheet 1: Fund Summary ─────────────────────────────────────────────────
    ws1 = wb.create_sheet('Fund Summary')
    ws1.sheet_view.showGridLines = False

    # Big title
    t = ws1.cell(row=1, column=1, value=fund['name'])
    t.font = Font(bold=True, size=16, color=NAVY, name='Calibri')
    ws1.merge_cells('A1:D1')
    ws1.row_dimensions[1].height = 28

    sub = ws1.cell(row=2, column=1, value=f'DGA Capital — Fund Report  |  As of {today_str}')
    sub.font = Font(size=10, color='888888', name='Calibri')
    ws1.merge_cells('A2:D2')

    # Fund Information section
    r = 4
    section_title(ws1, r, 'FUND INFORMATION', 4); r += 1
    rows_info = [
        ('Fund Name',       fund['name'],          None, None),
        ('Short Name',      fund['short_name'],     None, None),
        ('Inception Date',  str(fund.get('inception_date', '')), None, None),
        ('Status',          fund.get('status', '').title(), None, None),
    ]
    for label, val, _, __ in rows_info:
        apply_row(ws1, r, [label, val, '', ''], bg=WHITE if r % 2 == 0 else LGRAY)
        ws1.cell(row=r, column=1).font = Font(bold=True, size=10, color=DKGRAY, name='Calibri')
        r += 1

    # Performance section
    r += 1
    section_title(ws1, r, 'PERFORMANCE', 4); r += 1
    total_gain = nav - total_committed
    gain_pct   = (total_gain / total_committed) if total_committed else 0.0
    perf_rows = [
        ('Current NAV (Market Value)',  nav,             FMT_USD0),
        ('Total LP Contributions',      total_committed, FMT_USD0),
        ('Total Fund Gain / (Loss)',    total_gain,      FMT_USD0),
        ('Gain % on Contributions',     gain_pct,        FMT_PCT),
    ]
    for i, (label, val, fmt) in enumerate(perf_rows):
        bg = LGRAY if i % 2 == 0 else WHITE
        apply_row(ws1, r, [label, val, '', ''],
                  formats=[None, fmt, None, None], bg=bg)
        ws1.cell(row=r, column=1).font = Font(bold=True, size=10, color=DKGRAY, name='Calibri')
        vcolor = GREEN if (label == 'Total Fund Gain / (Loss)' and total_gain >= 0) else \
                 RED   if (label == 'Total Fund Gain / (Loss)' and total_gain < 0)  else DKGRAY
        ws1.cell(row=r, column=2).font = Font(bold=True, size=10, color=vcolor, name='Calibri')
        r += 1

    # Economics section
    r += 1
    section_title(ws1, r, 'FUND ECONOMICS', 4); r += 1
    econ_rows = [
        ('Management Fee',   float(fund.get('mgmt_fee_pct', 0)), FMT_PCT),
        ('Carried Interest', float(fund.get('carry_pct', 0)),    FMT_PCT),
        ('Hurdle Rate',      float(fund.get('hurdle_pct', 0)),   FMT_PCT),
    ]
    for i, (label, val, fmt) in enumerate(econ_rows):
        bg = LGRAY if i % 2 == 0 else WHITE
        apply_row(ws1, r, [label, val, '', ''],
                  formats=[None, fmt, None, None], bg=bg)
        ws1.cell(row=r, column=1).font = Font(bold=True, size=10, color=DKGRAY, name='Calibri')
        r += 1

    # Overview section
    r += 1
    section_title(ws1, r, 'PORTFOLIO OVERVIEW', 4); r += 1
    total_cost = sum(float(p['cost_basis'] or 0) for p in positions)
    total_mktval = sum(
        float(p['qty'] or 0) * (prices.get(p['symbol']) or float(p['avg_cost'] or 0))
        for p in positions
    )
    ov_rows = [
        ('Number of LPs',           len(lps),           None),
        ('Number of Positions',     len(positions),      None),
        ('Total Cost Basis',        total_cost,          FMT_USD0),
        ('Total Market Value',      total_mktval,        FMT_USD0),
        ('Unrealized P/L',          total_mktval - total_cost, FMT_USD0),
    ]
    for i, (label, val, fmt) in enumerate(ov_rows):
        bg = LGRAY if i % 2 == 0 else WHITE
        fmts = [None, fmt, None, None] if fmt else None
        apply_row(ws1, r, [label, val, '', ''], formats=fmts, bg=bg)
        ws1.cell(row=r, column=1).font = Font(bold=True, size=10, color=DKGRAY, name='Calibri')
        r += 1

    ws1.column_dimensions['A'].width = 28
    ws1.column_dimensions['B'].width = 22
    ws1.column_dimensions['C'].width = 18
    ws1.column_dimensions['D'].width = 18

    # ── Sheet 2: Portfolio Positions ─────────────────────────────────────────
    ws2 = wb.create_sheet('Portfolio Positions')
    ws2.sheet_view.showGridLines = False
    ws2.freeze_panes = 'A2'

    pos_cols = ['Symbol', 'Security Name', 'Asset Class', 'Shares',
                'Avg Cost', 'Cost Basis', 'Last Price', 'Market Value',
                'Unrealized P/L', 'P/L %', 'Weight %']
    write_header(ws2, 1, pos_cols)

    total_mv   = 0.0
    total_cb   = 0.0
    total_unrl = 0.0
    for i, p in enumerate(positions):
        qty      = float(p['qty'] or 0)
        avg_cost = float(p['avg_cost'] or 0)
        cb       = float(p['cost_basis'] or 0)
        last     = prices.get(p['symbol']) or avg_cost
        mv       = qty * last
        unrl     = mv - cb
        pl_pct   = (unrl / cb) if cb else 0.0
        wt_pct   = (mv / nav) if nav else 0.0
        total_mv   += mv
        total_cb   += cb
        total_unrl += unrl
        bg = WHITE if i % 2 == 0 else LGRAY
        vals  = [p['symbol'], p['sec_name'], p['asset_class'],
                 qty, avg_cost, cb, last, mv, unrl, pl_pct, wt_pct]
        fmts  = [None, None, None, FMT_NUM,
                 FMT_USD, FMT_USD0, FMT_USD, FMT_USD0,
                 FMT_USD0, FMT_PCT1, FMT_PCT1]
        apply_row(ws2, i + 2, vals, formats=fmts, bg=bg)
        pl_c = ws2.cell(row=i + 2, column=9)
        pl_c.font = Font(size=10, color=GREEN if unrl >= 0 else RED, name='Calibri')

    # Totals row
    tr = len(positions) + 2
    totals = ['TOTAL', '', '', '', '', total_cb, '', total_mv, total_unrl, '', '']
    apply_row(ws2, tr, totals,
              formats=[None, None, None, None, None, FMT_USD0, None,
                       FMT_USD0, FMT_USD0, None, None],
              bg=NAVY, bold=True, font_color=GOLD)

    auto_width(ws2)
    ws2.column_dimensions['B'].width = 38

    # ── Sheet 3: Limited Partners ─────────────────────────────────────────────
    ws3 = wb.create_sheet('Limited Partners')
    ws3.sheet_view.showGridLines = False
    ws3.freeze_panes = 'A2'

    lp_cols = ['LP Name', 'Entity Type', 'Commitment ($)',
               'Current Value ($)', 'NAV Share %']
    write_header(ws3, 1, lp_cols)

    for i, lp in enumerate(lps):
        cmt   = float(lp['commitment'])
        share = (cmt / total_committed) if total_committed else 0.0
        cur_v = nav * share
        bg = WHITE if i % 2 == 0 else LGRAY
        apply_row(ws3, i + 2,
                  [lp['legal_name'], lp.get('entity_type', '').title(), cmt, cur_v, share],
                  formats=[None, None, FMT_USD0, FMT_USD0, FMT_PCT],
                  bg=bg)

    if lps:
        tr3 = len(lps) + 2
        apply_row(ws3, tr3,
                  ['TOTAL', '', total_committed, nav, 1.0],
                  formats=[None, None, FMT_USD0, FMT_USD0, FMT_PCT],
                  bg=NAVY, bold=True, font_color=GOLD)

    auto_width(ws3)
    ws3.column_dimensions['A'].width = 30

    # ── Sheet 4: Annual Waterfall ─────────────────────────────────────────────
    if snapshots:
        ws4 = wb.create_sheet('Annual Waterfall')
        ws4.sheet_view.showGridLines = False
        ws4.freeze_panes = 'A2'

        wf_cols = ['Year', 'Jan 1 NAV', 'Dec 31 NAV', 'Contributions',
                   'Hurdle Amount', 'Gross Profit', 'Carry Earned',
                   'Carry Rolled', 'GP Equity (Year End)', 'Accum GP %']
        write_header(ws4, 1, wf_cols)

        for i, s in enumerate(snapshots):
            end_nav  = float(s['end_nav'] or 0)
            gp_eq    = float(s['gp_equity_end'] or 0)
            accum_pct = (gp_eq / end_nav) if end_nav else 0.0
            bg = WHITE if i % 2 == 0 else LGRAY
            apply_row(ws4, i + 2,
                      [s['year'],
                       float(s['start_nav']),   float(s['end_nav']),
                       float(s['contributions']), float(s['hurdle_amount']),
                       float(s['gross_profit']),  float(s['carry_earned']),
                       float(s['carry_rolled']),  gp_eq,
                       accum_pct],
                      formats=[None, FMT_USD0, FMT_USD0, FMT_USD0,
                               FMT_USD0, FMT_USD0, FMT_USD0,
                               FMT_USD0, FMT_USD0, FMT_PCT1],
                      bg=bg)
            carry_c = ws4.cell(row=i + 2, column=7)
            carry_c.font = Font(size=10, name='Calibri',
                                color=GREEN if float(s['carry_earned']) > 0 else DKGRAY)

        auto_width(ws4)

    # ── Sheet 5: Transaction History ─────────────────────────────────────────
    ws5 = wb.create_sheet('Transaction History')
    ws5.sheet_view.showGridLines = False
    ws5.freeze_panes = 'A2'

    tx_cols = ['Date', 'Category', 'Description', 'Debit', 'Credit']
    write_header(ws5, 1, tx_cols)

    for i, t in enumerate(txns):
        bg = WHITE if i % 2 == 0 else LGRAY
        apply_row(ws5, i + 2,
                  [str(t['effective_date']),
                   str(t.get('category', '') or '').replace('_', ' ').title(),
                   t.get('description', '') or '',
                   float(t.get('total_debit', 0) or 0),
                   float(t.get('total_credit', 0) or 0)],
                  formats=[None, None, None, FMT_USD0, FMT_USD0],
                  bg=bg)

    auto_width(ws5)
    ws5.column_dimensions['C'].width = 40

    # ── Freeze, set tab colors ────────────────────────────────────────────────
    ws1.sheet_properties.tabColor = GOLD
    for ws in [ws2, ws3, ws5]:
        ws.sheet_properties.tabColor = NAVY

    # ── Output ────────────────────────────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    safe = fund.get('short_name', 'Fund').replace('/', '-').replace('\\', '-')
    fname = f"{safe}_Report_{datetime.utcnow().strftime('%Y%m%d')}.xlsx"
    return StreamingResponse(
        buf,
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename="{fname}"'},
    )


@app.get("/api/fund/export-pdf")
async def fund_export_pdf(request: Request, fund_id: str = None):
    """Generate a PDF report for the fund and return it as a downloadable file.

    Uses reportlab for layout. Falls back with a 400 if reportlab is missing.
    """
    _require_fund_token(request)
    if not _REPORTLAB_OK:
        raise HTTPException(400, "reportlab not installed on this server")

    from reportlab.lib.pagesizes import letter
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.lib import colors
    from reportlab.platypus import (
        SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, HRFlowable, Image as RLImage
    )
    from reportlab.lib.enums import TA_CENTER, TA_LEFT
    from fastapi.responses import StreamingResponse

    conn = _fund_conn()
    try:
        with conn.cursor(cursor_factory=_RealDictCursor) as cur:
            fid = _resolve_fund_id(cur, fund_id)
            cur.execute("""
                SELECT name, short_name, inception_date, status,
                       mgmt_fee_pct, carry_pct, hurdle_pct, fund_type
                  FROM funds WHERE id = %s
            """, (fid,))
            fund = dict(cur.fetchone())
            is_acct_pdf = fund.get('fund_type') == 'managed_account'
            nav = _fund_market_nav(cur, fid)

            cur.execute("""
                SELECT s.symbol, s.name AS sec_name,
                       SUM(tl.quantity) AS qty,
                       SUM(tl.quantity * tl.cost_basis_per_unit)
                         / NULLIF(SUM(tl.quantity), 0) AS avg_cost,
                       SUM(tl.quantity * tl.cost_basis_per_unit) AS cost_basis
                  FROM tax_lots tl
                  JOIN securities s ON s.id = tl.security_id
                 WHERE tl.fund_id = %s AND tl.closed_at IS NULL
                 GROUP BY s.symbol, s.name
                 ORDER BY SUM(tl.quantity * tl.cost_basis_per_unit) DESC
            """, (fid,))
            positions = [dict(r) for r in cur.fetchall()]
            symbols   = [p['symbol'] for p in positions if p['symbol']]
            prices    = _fetch_prices(symbols)

            # YTD cache for managed accounts
            ytd_cache_pdf = None
            if is_acct_pdf:
                try:
                    cur.execute("""
                        SELECT nav, ytd_pct, result_json, updated_at
                          FROM managed_account_ytd_cache WHERE fund_id = %s
                    """, (fid,))
                    row = cur.fetchone()
                    if row:
                        import json as _json3
                        ytd_cache_pdf = dict(row)
                        if ytd_cache_pdf.get('result_json'):
                            ytd_cache_pdf['result'] = _json3.loads(ytd_cache_pdf['result_json'])
                except Exception:
                    ytd_cache_pdf = None

            lps             = []
            total_committed = 0.0
            snapshots_pdf   = []
            if not is_acct_pdf:
                cur.execute("""
                    SELECT l.legal_name, l.entity_type,
                           COALESCE(SUM(c.commitment_amount), 0) AS commitment
                      FROM lps l
                      LEFT JOIN commitments c ON c.lp_id = l.id AND c.superseded_by IS NULL
                     WHERE l.fund_id = %s AND l.status = 'active'
                     GROUP BY l.id, l.legal_name, l.entity_type
                    HAVING COALESCE(SUM(c.commitment_amount), 0) > 0
                     ORDER BY commitment DESC
                """, (fid,))
                lps = [dict(r) for r in cur.fetchall()]
                total_committed = sum(float(lp['commitment']) for lp in lps) or 0.0
                try:
                    cur.execute("""
                        SELECT year, start_nav, end_nav, contributions,
                               hurdle_amount, gross_profit, carry_earned,
                               carry_paid, carry_rolled, gp_equity_end
                          FROM fund_annual_snapshots
                         WHERE fund_id = %s ORDER BY year ASC
                    """, (fid,))
                    snapshots_pdf = [dict(r) for r in cur.fetchall()]
                except Exception:
                    snapshots_pdf = []
    finally:
        conn.close()

    # ── Shared PDF helpers ────────────────────────────────────────────────────
    _LOGO_PATH = Path(__file__).parent.parent / 'branding' / 'dga_logo_small.png'

    def _logo_img(height=0.45*inch):
        if _LOGO_PATH.exists():
            img = RLImage(str(_LOGO_PATH))
            aspect = img.imageWidth / img.imageHeight if img.imageHeight else 1.0
            return RLImage(str(_LOGO_PATH), width=height*aspect, height=height)
        return None

    _NAVY_PDF  = colors.HexColor('#0a1628')
    _GOLD_PDF  = colors.HexColor('#c9a84c')
    _LGRAY_PDF = colors.HexColor('#e8ecf2')
    _GREEN_PDF = colors.HexColor('#1a7f40')
    _RED_PDF   = colors.HexColor('#cc3333')
    _WHITE_PDF = colors.white
    _styles    = getSampleStyleSheet()
    today_pdf  = datetime.utcnow().strftime('%B %d, %Y')

    _h1  = ParagraphStyle('h1p', parent=_styles['Heading1'],
                           fontSize=18, textColor=_NAVY_PDF, spaceAfter=4)
    _h2  = ParagraphStyle('h2p', parent=_styles['Heading2'],
                           fontSize=12, textColor=_NAVY_PDF, spaceAfter=4, spaceBefore=14)
    _sub = ParagraphStyle('subp', parent=_styles['Normal'],
                           fontSize=9, textColor=colors.HexColor('#4a6080'), spaceAfter=2)

    def money(v):
        if v is None: return '—'
        v = float(v)
        if abs(v) >= 1e6: return f'${v/1e6:,.2f}M'
        if abs(v) >= 1e3: return f'${v/1e3:,.1f}K'
        return f'${v:,.2f}'

    def pct(v, decimals=2):
        if v is None: return '—'
        return f'{float(v)*100:+.{decimals}f}%'

    def _tbl_pdf(data, col_widths, header_rows=1):
        t = Table(data, colWidths=col_widths)
        sty = [
            ('BACKGROUND', (0,0), (-1, header_rows-1), _NAVY_PDF),
            ('TEXTCOLOR',  (0,0), (-1, header_rows-1), _WHITE_PDF),
            ('FONTNAME',   (0,0), (-1, header_rows-1), 'Helvetica-Bold'),
            ('FONTSIZE',   (0,0), (-1,-1), 8.5),
            ('ROWBACKGROUNDS', (0, header_rows), (-1,-1), [_WHITE_PDF, _LGRAY_PDF]),
            ('GRID',       (0,0), (-1,-1), 0.4, colors.HexColor('#c0ccd8')),
            ('VALIGN',     (0,0), (-1,-1), 'MIDDLE'),
            ('TOPPADDING', (0,0), (-1,-1), 3),
            ('BOTTOMPADDING', (0,0), (-1,-1), 3),
        ]
        t.setStyle(TableStyle(sty))
        return t

    # ── Managed account PDF ───────────────────────────────────────────────────
    if is_acct_pdf:
        from fastapi.responses import StreamingResponse as _SR3

        result_pdf   = (ytd_cache_pdf or {}).get('result') or {}
        ytd_pct_pdf  = float((ytd_cache_pdf or {}).get('ytd_pct') or
                             result_pdf.get('md_return_pct') or 0) / 100.0
        begin_v_pdf  = float(result_pdf.get('begin_value') or 0)
        end_v_pdf    = float(result_pdf.get('end_value') or nav or 0)
        net_flow_pdf = float(result_pdf.get('net_flow') or 0)
        attr_pdf     = result_pdf.get('attribution') or []
        monthly_pdf  = (result_pdf.get('monthly_chart') or {}).get('monthly') or []
        spy_raw_pdf  = result_pdf.get('spy_monthly') or {}
        spy_pts_pdf  = (spy_raw_pdf.get('points') or []) if isinstance(spy_raw_pdf, dict) else []
        flows_pdf    = result_pdf.get('flows') or []
        upd_pdf      = str((ytd_cache_pdf or {}).get('updated_at') or '')[:10] or today_pdf

        # SPY lookup by month int
        spy_by_m_int_pdf = {}
        for pt in spy_pts_pdf:
            try:
                mi = int(str(pt.get('month', '')).split('-')[1])
                spy_by_m_int_pdf[mi] = pt
            except Exception:
                pass

        spy_ytd_pdf  = float(spy_pts_pdf[-1].get('ytd_pct') or 0) / 100.0 if spy_pts_pdf else 0.0
        alpha_pdf    = ytd_pct_pdf - spy_ytd_pdf

        buf_a = io.BytesIO()
        doc_a = SimpleDocTemplate(buf_a, pagesize=letter,
                                  leftMargin=0.75*inch, rightMargin=0.75*inch,
                                  topMargin=0.75*inch, bottomMargin=0.75*inch)
        usable = 7.0 * inch
        story_a = []

        # Cover header with logo
        logo_img_a = _logo_img(0.5*inch)
        if logo_img_a:
            hdr_data = [[logo_img_a, Paragraph(fund['name'], _h1)]]
            hdr_tbl = Table(hdr_data, colWidths=[2.0*inch, usable - 2.0*inch])
            hdr_tbl.setStyle(TableStyle([
                ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
                ('LEFTPADDING', (0,0), (-1,-1), 0),
                ('RIGHTPADDING', (0,0), (-1,-1), 0),
            ]))
            story_a.append(hdr_tbl)
        else:
            story_a.append(Paragraph(fund['name'], _h1))
        story_a.append(Paragraph(
            f"DGA Capital  ·  Managed Account Report  ·  As of {today_pdf}  ·  "
            f"YTD data updated {upd_pdf}", _sub))
        story_a.append(HRFlowable(width='100%', thickness=1.5,
                                  color=_GOLD_PDF, spaceAfter=8))

        # Summary tiles
        gain_color = _GREEN_PDF if ytd_pct_pdf >= 0 else _RED_PDF
        alpha_color = _GREEN_PDF if alpha_pdf >= 0 else _RED_PDF
        metrics_a = [
            ['Account NAV', 'Jan 1 Value', 'YTD Return', 'SPY YTD', 'Alpha vs SPY'],
            [money(end_v_pdf), money(begin_v_pdf),
             f'{ytd_pct_pdf*100:+.2f}%', f'{spy_ytd_pdf*100:+.2f}%',
             f'{alpha_pdf*100:+.2f}%'],
        ]
        t_metrics = _tbl_pdf(metrics_a, [usable/5]*5)
        t_metrics.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), _NAVY_PDF),
            ('TEXTCOLOR',  (0,0), (-1,0), _GOLD_PDF),
            ('TEXTCOLOR',  (2,1), (2,1), gain_color),
            ('TEXTCOLOR',  (4,1), (4,1), alpha_color),
            ('FONTNAME',   (0,0), (-1,0), 'Helvetica-Bold'),
            ('FONTNAME',   (0,1), (-1,1), 'Helvetica-Bold'),
            ('FONTSIZE',   (0,0), (-1,-1), 9),
            ('ALIGN',      (0,0), (-1,-1), 'CENTER'),
            ('VALIGN',     (0,0), (-1,-1), 'MIDDLE'),
            ('TOPPADDING', (0,0), (-1,-1), 6),
            ('BOTTOMPADDING', (0,0), (-1,-1), 6),
            ('GRID',       (0,0), (-1,-1), 0.4, colors.HexColor('#c0ccd8')),
        ]))
        story_a.append(t_metrics)
        story_a.append(Spacer(1, 12))

        # Positions
        if positions:
            story_a.append(Paragraph('Portfolio Positions', _h2))
            rows_a = [['Ticker', 'Name', 'Qty', 'Avg Cost', 'Cost Basis',
                       'Live Price', 'Market Value', 'P/L']]
            for p in positions:
                sym  = p['symbol'] or ''
                qty  = float(p['qty'] or 0)
                cost = float(p['cost_basis'] or 0)
                price = prices.get(sym)
                mv   = qty * price if price else None
                pl   = (mv - cost) if mv else None
                rows_a.append([sym, (p['sec_name'] or '')[:22],
                                f'{qty:,.0f}', money(p['avg_cost']),
                                money(cost), money(price), money(mv), money(pl)])
            cw_a = [0.6, 2.0, 0.65, 0.75, 0.85, 0.75, 0.8, 0.8]
            story_a.append(_tbl_pdf(rows_a, [w*inch for w in cw_a]))
            story_a.append(Spacer(1, 12))

        # YTD Attribution
        if attr_pdf:
            story_a.append(Paragraph('YTD Attribution by Position', _h2))
            rows_attr = [['Ticker', 'End Shares', 'Jan 1 Price', 'End Price',
                          'Dollar Gain', 'Contribution %', 'Stock Return %']]
            for a in sorted(attr_pdf,
                            key=lambda x: float(x.get('contribution_pct') or 0),
                            reverse=True):
                dg_a = float(a.get('dollar_gain') or 0)
                cp_a = float(a.get('contribution_pct') or 0)
                tr_a = a.get('ticker_return_pct')
                rows_attr.append([
                    a.get('ticker', ''),
                    f"{float(a.get('end_shares') or 0):,.3f}",
                    money(a.get('jan1_price')),
                    money(a.get('end_price')),
                    money(dg_a),
                    f"{cp_a:+.2f}%",
                    f"{float(tr_a):+.2f}%" if tr_a is not None else '—',
                ])
            story_a.append(_tbl_pdf(rows_attr,
                                    [0.65*inch, 0.85*inch, 0.85*inch, 0.85*inch,
                                     0.85*inch, 0.9*inch, 0.9*inch]))
            story_a.append(Spacer(1, 12))

        # Monthly performance
        if monthly_pdf:
            story_a.append(Paragraph('Monthly Performance', _h2))
            rows_mo = [['Month', 'Start Value', 'End Value', 'Dollar Gain',
                        'Monthly Return', 'SPY Monthly', 'Alpha']]
            for mo in monthly_pdf:
                m_int_p  = int(mo.get('month') or 0)
                lbl_p    = str(mo.get('label') or m_int_p)
                sv_p     = float(mo.get('start_value') or 0)
                ev_p     = float(mo.get('end_value') or 0)
                dg_p     = float(mo.get('dollar_gain') or 0)
                ret_p    = float(mo.get('return_pct') or 0)
                spy_pt_p = spy_by_m_int_pdf.get(m_int_p) or {}
                prev_p   = spy_by_m_int_pdf.get(m_int_p - 1) if m_int_p > 1 else None
                spy_ytd_p = float(spy_pt_p.get('ytd_pct') or 0)
                spy_prv_p = float(prev_p.get('ytd_pct') or 0) if prev_p else 0.0
                spy_mo_p  = spy_ytd_p - spy_prv_p if spy_pt_p else None
                alpha_p   = (ret_p - spy_mo_p) if spy_mo_p is not None else None
                rows_mo.append([
                    lbl_p, money(sv_p), money(ev_p), money(dg_p),
                    f"{ret_p:+.2f}%",
                    f"{spy_mo_p:+.2f}%" if spy_mo_p is not None else '—',
                    f"{alpha_p:+.2f}%" if alpha_p is not None else '—',
                ])
            story_a.append(_tbl_pdf(rows_mo,
                                    [0.55*inch, 1.0*inch, 1.0*inch, 0.9*inch,
                                     0.85*inch, 0.85*inch, 0.75*inch]))
            story_a.append(Spacer(1, 12))

        # Cash flows
        if flows_pdf:
            story_a.append(Paragraph('Cash Flow History', _h2))
            rows_fl = [['Date', 'Description', 'Amount', 'Type']]
            for f in sorted(flows_pdf, key=lambda x: x.get('date', ''), reverse=True):
                amt = float(f.get('amount') or 0)
                rows_fl.append([f.get('date', ''), (f.get('description', '') or '')[:38],
                                 money(amt), 'Inflow' if amt > 0 else 'Outflow'])
            story_a.append(_tbl_pdf(rows_fl,
                                    [0.9*inch, 3.5*inch, 1.0*inch, 0.9*inch]))
            story_a.append(Spacer(1, 12))

        story_a.append(HRFlowable(width='100%', thickness=0.5,
                                  color=_LGRAY_PDF, spaceAfter=4))
        story_a.append(Paragraph(
            f'Report generated {datetime.utcnow().strftime("%Y-%m-%d %H:%M")} UTC  ·  '
            f'DGA Capital  ·  Confidential', _sub))

        doc_a.build(story_a)
        buf_a.seek(0)
        safe_a = fund.get('short_name', 'Account').replace('/', '-').replace('\\', '-')
        fname_a = f"{safe_a}_AccountReport_{datetime.utcnow().strftime('%Y%m%d')}.pdf"
        return _SR3(
            buf_a,
            media_type='application/pdf',
            headers={'Content-Disposition': f'attachment; filename="{fname_a}"'},
        )

    # ── Build PDF (LP fund) ───────────────────────────────────────────────────
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=letter,
                            leftMargin=0.75*inch, rightMargin=0.75*inch,
                            topMargin=0.75*inch, bottomMargin=0.75*inch)

    styles = getSampleStyleSheet()
    NAVY   = _NAVY_PDF
    GOLD   = _GOLD_PDF
    LGRAY  = _LGRAY_PDF
    WHITE  = _WHITE_PDF

    h1  = _h1
    h2  = _h2
    sub = _sub

    def tbl(data, col_widths, header_rows=1):
        t = Table(data, colWidths=col_widths)
        style = [
            ('BACKGROUND', (0,0), (-1, header_rows-1), NAVY),
            ('TEXTCOLOR',  (0,0), (-1, header_rows-1), WHITE),
            ('FONTNAME',   (0,0), (-1, header_rows-1), 'Helvetica-Bold'),
            ('FONTSIZE',   (0,0), (-1,-1), 9),
            ('ROWBACKGROUNDS', (0, header_rows), (-1,-1), [WHITE, LGRAY]),
            ('GRID',       (0,0), (-1,-1), 0.4, colors.HexColor('#c0ccd8')),
            ('VALIGN',     (0,0), (-1,-1), 'MIDDLE'),
            ('TOPPADDING', (0,0), (-1,-1), 4),
            ('BOTTOMPADDING', (0,0), (-1,-1), 4),
        ]
        t.setStyle(TableStyle(style))
        return t

    story = []
    logo_img_lp = _logo_img(0.5*inch)
    if logo_img_lp:
        hdr_data_lp = [[logo_img_lp, Paragraph(fund['name'], h1)]]
        hdr_tbl_lp = Table(hdr_data_lp, colWidths=[2.0*inch, usable_w - 2.0*inch])
        hdr_tbl_lp.setStyle(TableStyle([
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('LEFTPADDING', (0,0), (-1,-1), 0),
            ('RIGHTPADDING', (0,0), (-1,-1), 0),
        ]))
        story.append(hdr_tbl_lp)
    else:
        story.append(Paragraph(fund['name'], h1))
    story.append(Paragraph(
        f"Short name: {fund['short_name']}  ·  "
        f"Inception: {str(fund['inception_date'])[:10]}  ·  "
        f"Status: {fund['status'].upper()}  ·  "
        f"Mgmt fee: {float(fund['mgmt_fee_pct'])*100:.1f}%  ·  "
        f"Carry: {float(fund['carry_pct'])*100:.0f}%  ·  "
        f"Hurdle: {float(fund['hurdle_pct'])*100:.0f}%",
        sub))
    story.append(HRFlowable(width='100%', thickness=1.5, color=GOLD, spaceAfter=8))

    # Summary metrics
    usable_w = 7.0 * inch
    gain = nav - total_committed
    gain_pct = (gain / total_committed * 100) if total_committed else 0.0
    metrics = [
        ['Market NAV', 'Total Committed', 'Total Gain', 'Gain %', 'LP Count'],
        [money(nav), money(total_committed), money(gain),
         f'{gain_pct:+.2f}%', str(len(lps))],
    ]
    story.append(tbl(metrics, [usable_w/5]*5))
    story.append(Spacer(1, 12))

    # Positions
    if positions:
        story.append(Paragraph('Portfolio Positions', h2))
        rows = [['Ticker', 'Name', 'Qty', 'Avg Cost', 'Cost Basis', 'Live Price', 'MV', 'P/L']]
        for p in positions:
            sym   = p['symbol'] or ''
            qty   = float(p['qty'] or 0)
            cost  = float(p['cost_basis'] or 0)
            price = prices.get(sym)
            mv    = qty * price if price else None
            pl    = (mv - cost) if mv else None
            rows.append([
                sym, (p['sec_name'] or '')[:22],
                f'{qty:,.0f}', money(p['avg_cost']),
                money(cost), money(price), money(mv), money(pl),
            ])
        cw = [0.6, 2.0, 0.65, 0.75, 0.85, 0.75, 0.8, 0.8]
        story.append(tbl(rows, [w*inch for w in cw]))
        story.append(Spacer(1, 12))

    # LPs
    if lps:
        story.append(Paragraph('Limited Partners', h2))
        rows = [['LP Name', 'Type', 'Commitment', '% of Total']]
        for lp in lps:
            comm = float(lp['commitment'])
            pct  = comm / total_committed * 100 if total_committed else 0
            rows.append([
                (lp['legal_name'] or '')[:38],
                lp['entity_type'] or '',
                money(comm),
                f'{pct:.1f}%',
            ])
        story.append(tbl(rows, [3.5*inch, 1.2*inch, 1.2*inch, 1.1*inch]))
        story.append(Spacer(1, 12))

    # Annual Waterfall & Carry
    if snapshots_pdf:
        story.append(Paragraph('Annual Waterfall & Carry', h2))
        wf_rows = [['Year', 'Start NAV', 'Contributions', 'Gross Profit',
                    'Hurdle Amt', 'Carry Earned', 'Carry Paid', 'GP Equity', 'End NAV']]
        for s in snapshots_pdf:
            wf_rows.append([
                str(s.get('year', '')),
                money(s.get('start_nav')),
                money(s.get('contributions')),
                money(s.get('gross_profit')),
                money(s.get('hurdle_amount')),
                money(s.get('carry_earned')),
                money(s.get('carry_paid')),
                money(s.get('gp_equity_end')),
                money(s.get('end_nav')),
            ])
        story.append(tbl(wf_rows, [0.5*inch, 0.8*inch, 0.85*inch, 0.8*inch,
                                    0.8*inch, 0.8*inch, 0.75*inch, 0.75*inch, 0.8*inch]))
        story.append(Spacer(1, 12))

    # Per-LP Allocation After Carry
    if lps and total_committed and nav > 0:
        story.append(Paragraph('Per-LP Allocation After Carry', h2))
        # Cumulative GP equity: carry accrued = (last gp_equity_end / last end_nav) * current nav
        gp_accrued_pdf = 0.0
        if snapshots_pdf:
            last_s = snapshots_pdf[-1]
            last_gp_eq  = float(last_s.get('gp_equity_end') or 0)
            last_end_nav = float(last_s.get('end_nav') or 0)
            if last_end_nav > 0:
                gp_accrued_pdf = (last_gp_eq / last_end_nav) * nav
        lp_distributable = max(0.0, nav - gp_accrued_pdf)
        alloc_rows = [['LP Name', 'Commitment', 'Share %', 'Carry Offset', 'Net Allocation']]
        for lp in lps:
            comm = float(lp['commitment'])
            share = comm / total_committed if total_committed else 0.0
            carry_offset = gp_accrued_pdf * share
            alloc = lp_distributable * share
            alloc_rows.append([
                (lp['legal_name'] or '')[:32],
                money(comm),
                f'{share*100:.2f}%',
                money(carry_offset),
                money(alloc),
            ])
        story.append(tbl(alloc_rows, [2.8*inch, 1.2*inch, 0.8*inch, 1.0*inch, 1.2*inch]))
        story.append(Spacer(1, 12))

    story.append(Spacer(1, 20))
    story.append(Paragraph(
        f'Report generated {datetime.utcnow().strftime("%Y-%m-%d %H:%M")} UTC  ·  DGA Capital Research',
        sub))

    doc.build(story)
    buf.seek(0)
    safe  = fund.get('short_name', 'Fund').replace('/', '-').replace('\\', '-')
    fname = f"{safe}_Report_{datetime.utcnow().strftime('%Y%m%d')}.pdf"
    return StreamingResponse(
        buf,
        media_type='application/pdf',
        headers={'Content-Disposition': f'attachment; filename="{fname}"'},
    )


# ---------------------------------------------------------------------------
# Static web UI — mount last so API routes take precedence.
# ---------------------------------------------------------------------------

@app.middleware("http")
async def no_cache_shell_middleware(request: Request, call_next):
    """Force the browser to re-fetch the HTML/CSS/JS shell on every request.

    Railway redeploys don't change the file URL, so default browser caching
    (which can keep static files for hours) would leave users staring at
    the previous build. We want the shell *itself* to refresh so the
    ``?v=`` pins on the CSS/JS tags get honored.
    """
    response = await call_next(request)
    path = request.url.path
    if path.startswith("/app/") or path == "/":
        response.headers["Cache-Control"] = "no-cache, no-store, must-revalidate"
        response.headers["Pragma"] = "no-cache"
        response.headers["Expires"] = "0"
        # Stop Railway's edge layer / Cloudflare from caching the shell
        response.headers["Surrogate-Control"] = "no-store"
        response.headers["CDN-Cache-Control"] = "no-store"
    return response


if BRANDING_DIR.exists():
    app.mount("/branding", StaticFiles(directory=str(BRANDING_DIR)), name="branding")

if WEB_DIR.exists():
    app.mount("/app", StaticFiles(directory=str(WEB_DIR), html=True), name="web")


if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)
