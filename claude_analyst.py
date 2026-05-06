#!/usr/bin/env python3
"""
DGA Capital Research Analyst — Claude Edition
----------------------------------------------

Terminal entry-point that:
  1. Prompts for a single ticker or a portfolio CSV/XLSX file.
  2. Asks whether to generate Gamma.app presentations (credits guard).
  3. Pulls authoritative financial data from SEC EDGAR XBRL (companyfacts).
  4. Sends the DGA_SYSTEM_PROMPT + verified data to xAI Grok for analysis.
  5. Renders Grok's markdown into a bordered-table Word report.
  6. Optionally creates Gamma presentations for each stock.
  7. For portfolios: builds a roll-up Word summary + Gamma deck that ranks
     what to trim/sell vs add/buy based on valuations.

Data source: SEC EDGAR XBRL only (free, authoritative).

Usage:
    python3 claude_analyst.py
    python3 claude_analyst.py AAPL
    python3 claude_analyst.py --portfolio /path/to/portfolio.csv
    python3 claude_analyst.py --portfolio /path/to/portfolio.xlsx --no-gamma
    python3 claude_analyst.py --portfolio portfolio_watchlist.xlsx --strategy pro
    python3 claude_analyst.py --portfolio portfolio_watchlist.xlsx --strategy concentrated --reuse

Portfolio file schema (CSV or XLSX) — exactly three columns, first row headers:
    Ticker | Weight | Optimized
The "Optimized" column is intentionally IGNORED when loaded, so the output of
one run (DGA-portfolio.xlsx) can be fed back in as the input for the next run.
Weights can be expressed either as decimals (0.05) or whole-number percents (5).
"""

from __future__ import annotations

import argparse
import json
import mimetypes
import os
import re
import smtplib
import ssl
import sys
import time
import traceback
from datetime import datetime, timedelta
from email.message import EmailMessage
from pathlib import Path
from typing import Any

import pandas as pd
import requests

import sec_edgar_xbrl as edgar  # legacy companyfacts fallback
import excel_financials as xlsx_edgar  # primary: filing-accurate XBRL via edgartools
import pull_sec_financials  # downloader (edgartools-backed)
from word_report import render_report

# ============================================================================
# CONFIG — everything sensitive lives in .env (or real environment variables)
# ============================================================================
SCRIPT_DIR = Path(__file__).resolve().parent
STOCKS_FOLDER = SCRIPT_DIR / "stocks"
STOCKS_FOLDER.mkdir(parents=True, exist_ok=True)

# Watchlist & scan persistence
WATCHLIST_FILE = STOCKS_FOLDER / "watchlist.json"
SCAN_RESULTS_FILE = STOCKS_FOLDER / "scan_results.json"
INTEL_FILE = STOCKS_FOLDER / "intelligence.json"
DAILY_BRIEF_FILE = STOCKS_FOLDER / "daily_brief.json"  # most-recent Goldman-style brief
TRACKER_FILE = STOCKS_FOLDER / "tracker.json"   # paper portfolios + live + SPY series

# Gamma metadata index — maps ticker → { gamma_url, generated_at, ... }.
# Survives Railway restarts via Dropbox hydration so the "View Gamma"
# button keeps working days/weeks after the original analysis ran.
GAMMA_INDEX_FILE = STOCKS_FOLDER / "_gamma_index.json"
# How long a Gamma deck is considered "fresh" — re-runs of the same ticker
# inside this window reuse the existing presentation instead of burning
# credits on a duplicate.
GAMMA_FRESH_DAYS = 30

# Default recipient for portfolio-analysis emails (override via PORTFOLIO_EMAIL_TO env var).
DEFAULT_PORTFOLIO_EMAIL_TO = "alecmazo1@gmail.com"


def _load_dotenv() -> None:
    """Load a .env file sitting next to this script, if present.

    Uses python-dotenv when installed, otherwise falls back to a tiny parser
    so the script still runs on a minimal install.
    """
    env_path = SCRIPT_DIR / ".env"
    if not env_path.exists():
        return
    try:
        from dotenv import load_dotenv  # type: ignore
        load_dotenv(env_path)
        return
    except Exception:
        pass
    # Minimal fallback parser: KEY=VALUE, ignores blanks and '#' comments.
    for raw in env_path.read_text().splitlines():
        line = raw.strip()
        if not line or line.startswith("#"):
            continue
        if "=" not in line:
            continue
        k, v = line.split("=", 1)
        k = k.strip()
        v = v.strip()
        # Strip matching surrounding quotes.
        if len(v) >= 2 and ((v[0] == v[-1] == '"') or (v[0] == v[-1] == "'")):
            v = v[1:-1]
        # Don't overwrite anything the shell has already set.
        os.environ.setdefault(k, v)


_load_dotenv()


def _require_env(name: str, *, hint: str = "") -> str:
    val = os.environ.get(name, "").strip()
    if not val:
        msg = (
            f"Missing required environment variable: {name}\n"
            f"   Set it in your shell or in {SCRIPT_DIR / '.env'}"
        )
        if hint:
            msg += f"\n   Hint: {hint}"
        # Raise RuntimeError (subclass of Exception) so background threads and
        # API handlers can catch it with `except Exception`.  CLI entry-point
        # catches it separately and exits cleanly.
        raise RuntimeError(msg)
    return val


def _optional_env(name: str, default: str = "") -> str:
    return os.environ.get(name, default).strip() or default


# xAI API (Grok) — required at call time (not at import; keeps unit-testability)
# All analysis, research, intelligence, and daily brief runs on grok-4.3.
# Override via GROK_MODEL in .env to pin a specific version.
GROK_MODEL = _optional_env("GROK_MODEL", "grok-4.3-latest")
# Intelligence / Daily Brief use the same model as analysis.
GROK_INTEL_MODEL = GROK_MODEL

# Gamma.app folder ID is optional; API key is only required if Gamma generation
# is actually requested at runtime.
GAMMA_FOLDER_ID = _optional_env("GAMMA_FOLDER_ID", "")


def get_sec_user_agent() -> str:
    """Resolved on demand so a missing UA fails cleanly when we actually need it."""
    return _require_env(
        "SEC_USER_AGENT",
        hint="Format: 'Your Name your.email@example.com'. SEC blocks anonymous scrapers.",
    )


def get_grok_api_key() -> str:
    return _require_env("XAI_API_KEY", hint="Get yours at https://console.x.ai/")


def get_gamma_api_key() -> str:
    return _require_env("GAMMA_API_KEY", hint="Get yours at https://gamma.app/account")


# ============================================================================
# Portfolio-analysis email notification
# ----------------------------------------------------------------------------
# Triggered ONLY for multi-ticker (portfolio) runs. Single-ticker runs do not
# send mail. Configure via .env:
#   PORTFOLIO_EMAIL_TO=alecmazo1@gmail.com   (override default recipient)
#   GMAIL_USER=you@gmail.com                  (Gmail address to send from)
#   GMAIL_APP_PASSWORD=xxxx xxxx xxxx xxxx    (Gmail App Password — NOT your
#                                              normal Google password. Create at
#                                              https://myaccount.google.com/apppasswords
#                                              with 2FA enabled.)
# If credentials are missing, the function still composes the email body and
# writes it to disk under stocks/Portfolio_Email.eml so nothing is lost.
# ============================================================================

def _ranked_table_text(rows: list[dict] | None, limit: int = 25) -> str:
    if not rows:
        return "(ranked table unavailable)"
    headers = ["Ticker", "Rating", "Price", "12M Target", "Upside %", "Sector"]
    lines = [" | ".join(headers), " | ".join("---" for _ in headers)]
    for r in rows[:limit]:
        price = r.get("price") or r.get("current_price") or ""
        target = r.get("price_target") or r.get("target_price") or ""
        upside = r.get("upside_pct") or ""
        try:
            upside_str = f"{float(upside):.1f}%"
        except Exception:
            upside_str = str(upside) or ""
        lines.append(" | ".join([
            str(r.get("ticker", ""))[:6],
            str(r.get("rating", ""))[:12],
            f"${price}" if price else "",
            f"${target}" if target else "",
            upside_str,
            str(r.get("sector", ""))[:24],
        ]))
    return "\n".join(lines)


def _strategy_weights_text(strategy_results: dict[str, dict] | None) -> str:
    if not strategy_results:
        return "(no strategy weights computed)"
    blocks: list[str] = []
    for skey, res in strategy_results.items():
        weights = res.get("weights", {}) or {}
        held = {t: w for t, w in weights.items() if w and w > 0}
        if not held:
            blocks.append(f"[{skey}] (no positions held)")
            continue
        rows = sorted(held.items(), key=lambda kv: kv[1], reverse=True)
        body = "\n".join(f"  {t:<6} {w*100:6.2f}%" for t, w in rows)
        blocks.append(f"[{skey}]  ({len(rows)} positions)\n{body}")
    return "\n\n".join(blocks)


def _rating_color(rating: str) -> str:
    r = (rating or "").lower()
    if "strong buy" in r:  return "#1a7a3c"
    if "buy" in r:         return "#2e9e54"
    if "hold" in r:        return "#b07d00"
    if "sell" in r:        return "#c0392b"
    return "#555555"


def _upside_color(upside: Any) -> str:
    try:
        v = float(str(upside).replace("%", ""))
        if v >= 30:   return "#1a7a3c"
        if v >= 10:   return "#2e9e54"
        if v >= 0:    return "#b07d00"
        return "#c0392b"
    except Exception:
        return "#555555"


def _md_to_html(md: str) -> str:
    """Convert the most common markdown patterns to HTML for email display."""
    import re
    # Escape HTML special chars first
    md = md.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")
    # Bold + italic
    md = re.sub(r"\*\*\*(.+?)\*\*\*", r"<strong><em>\1</em></strong>", md)
    md = re.sub(r"\*\*(.+?)\*\*", r"<strong>\1</strong>", md)
    md = re.sub(r"\*(.+?)\*", r"<em>\1</em>", md)
    # Headers
    md = re.sub(r"^### (.+)$", r"<h4 style='margin:16px 0 6px;color:#0A1628'>\1</h4>", md, flags=re.MULTILINE)
    md = re.sub(r"^## (.+)$",  r"<h3 style='margin:18px 0 8px;color:#0A1628'>\1</h3>", md, flags=re.MULTILINE)
    md = re.sub(r"^# (.+)$",   r"<h2 style='margin:20px 0 10px;color:#0A1628'>\1</h2>", md, flags=re.MULTILINE)
    # Pipe tables → HTML tables
    def _pipe_table(m: re.Match) -> str:
        lines = [l.strip() for l in m.group(0).strip().splitlines() if l.strip()]
        rows = [[c.strip() for c in l.strip("|").split("|")] for l in lines]
        # Second row is separator (---)
        header, body_rows = rows[0], rows[2:]
        th = "".join(f"<th style='padding:6px 10px;background:#0A1628;color:#fff;text-align:left;font-size:12px'>{c}</th>" for c in header)
        trs = ""
        for i, row in enumerate(body_rows):
            bg = "#f8f9fb" if i % 2 == 0 else "#fff"
            tds = "".join(f"<td style='padding:5px 10px;border-bottom:1px solid #e8eaed;font-size:12px'>{c}</td>" for c in row)
            trs += f"<tr style='background:{bg}'>{tds}</tr>"
        return (
            "<table style='border-collapse:collapse;width:100%;margin:10px 0'>"
            f"<thead><tr>{th}</tr></thead><tbody>{trs}</tbody></table>"
        )
    md = re.sub(r"(\|.+\|\n\|[-| :]+\|\n(?:\|.+\|\n?)+)", _pipe_table, md)
    # Bullet lists
    md = re.sub(r"^[-*] (.+)$", r"<li style='margin:3px 0'>\1</li>", md, flags=re.MULTILINE)
    md = re.sub(r"(<li.*</li>\n?)+", lambda m: f"<ul style='margin:8px 0 8px 20px;padding:0'>{m.group(0)}</ul>", md)
    # Newlines → paragraphs (two newlines = paragraph break)
    paragraphs = re.split(r"\n{2,}", md)
    parts = []
    for p in paragraphs:
        p = p.strip()
        if not p:
            continue
        if p.startswith("<h") or p.startswith("<table") or p.startswith("<ul"):
            parts.append(p)
        else:
            p = p.replace("\n", "<br>")
            parts.append(f"<p style='margin:6px 0;line-height:1.5'>{p}</p>")
    return "\n".join(parts)


def _html_ranked_table(rows: list[dict] | None) -> str:
    """Render ranked positions table. Rows come from strategy_results[key]['rows']."""
    if not rows:
        return "<p style='color:#888;font-style:italic'>Ranked table unavailable</p>"

    def _num(v):
        try:
            return float(v) if v is not None and v != "" else None
        except (TypeError, ValueError):
            return None

    # Compute a display upside for each row (prefer provided, else price+target).
    def _effective_upside(r):
        u = _num(r.get("upside_pct"))
        if u is not None:
            return u
        p = _num(r.get("price") or r.get("current_price") or r.get("market_price"))
        t = _num(r.get("price_target") or r.get("target_price"))
        if p and t:
            return (t - p) / p * 100.0
        return None

    # Sort by computed upside desc; rows with no upside land at the bottom.
    sortable = [(r, _effective_upside(r)) for r in rows]
    sortable.sort(key=lambda pair: (-999 if pair[1] is None else -pair[1]))
    sorted_rows = [r for r, _ in sortable]

    rows_html = ""
    for i, r in enumerate(sorted_rows[:25]):
        rating = str(r.get("rating") or "—")
        price = _num(r.get("price") or r.get("current_price") or r.get("market_price"))
        target = _num(r.get("price_target") or r.get("target_price"))
        upside = _effective_upside(r)
        sector = r.get("sector") or "—"
        if sector == "Unknown":
            sector = "—"
        price_str = f"${price:,.2f}" if price is not None else "—"
        target_str = f"${target:,.2f}" if target is not None else "—"
        upside_str = f"{upside:+.1f}%" if upside is not None else "—"
        bg = "#f8f9fb" if i % 2 == 0 else "#fff"
        rows_html += (
            f"<tr style='background:{bg}'>"
            f"<td style='padding:7px 10px;font-weight:700;color:#0A1628;border-bottom:1px solid #e8eaed'>{r.get('ticker','')}</td>"
            f"<td style='padding:7px 10px;border-bottom:1px solid #e8eaed;color:{_rating_color(rating)};font-weight:600'>{rating}</td>"
            f"<td style='padding:7px 10px;border-bottom:1px solid #e8eaed;text-align:right;font-family:monospace'>{price_str}</td>"
            f"<td style='padding:7px 10px;border-bottom:1px solid #e8eaed;text-align:right;font-family:monospace'>{target_str}</td>"
            f"<td style='padding:7px 10px;border-bottom:1px solid #e8eaed;text-align:right;color:{_upside_color(upside)};font-weight:700'>{upside_str}</td>"
            f"<td style='padding:7px 10px;border-bottom:1px solid #e8eaed;color:#555;font-size:12px'>{sector}</td>"
            f"</tr>"
        )
    return (
        "<table style='border-collapse:collapse;width:100%'>"
        "<thead><tr>"
        "<th style='padding:8px 10px;background:#0A1628;color:#C9A84C;text-align:left'>Ticker</th>"
        "<th style='padding:8px 10px;background:#0A1628;color:#C9A84C;text-align:left'>Rating</th>"
        "<th style='padding:8px 10px;background:#0A1628;color:#C9A84C;text-align:right'>Price</th>"
        "<th style='padding:8px 10px;background:#0A1628;color:#C9A84C;text-align:right'>12M Target</th>"
        "<th style='padding:8px 10px;background:#0A1628;color:#C9A84C;text-align:right'>Upside</th>"
        "<th style='padding:8px 10px;background:#0A1628;color:#C9A84C;text-align:left'>Sector</th>"
        "</tr></thead>"
        f"<tbody>{rows_html}</tbody></table>"
    )


def _html_strategy_weights(strategy_results: dict[str, dict] | None) -> str:
    if not strategy_results:
        return "<p style='color:#888;font-style:italic'>No strategy weights computed</p>"
    blocks = []
    for skey, res in strategy_results.items():
        weights = {t: w for t, w in (res.get("weights") or {}).items() if w and w > 0}
        if not weights:
            continue
        label = res.get("label") or skey
        sorted_w = sorted(weights.items(), key=lambda kv: kv[1], reverse=True)
        rows_html = ""
        for i, (ticker, w) in enumerate(sorted_w):
            bg = "#f8f9fb" if i % 2 == 0 else "#fff"
            bar_w = int(w * 400)
            rows_html += (
                f"<tr style='background:{bg}'>"
                f"<td style='padding:6px 10px;font-weight:700;color:#0A1628;width:70px'>{ticker}</td>"
                f"<td style='padding:6px 10px'>"
                f"<div style='background:#e8eaed;border-radius:3px;height:10px;width:100%'>"
                f"<div style='background:#C9A84C;border-radius:3px;height:10px;width:{min(bar_w,400)}px'></div></div></td>"
                f"<td style='padding:6px 10px;text-align:right;font-weight:600;color:#0A1628;width:60px'>{w*100:.1f}%</td>"
                f"</tr>"
            )
        blocks.append(
            f"<div style='flex:1;min-width:200px;margin:0 8px 16px'>"
            f"<div style='background:#0A1628;color:#C9A84C;padding:8px 10px;font-weight:700;font-size:13px;border-radius:6px 6px 0 0'>"
            f"{label} &nbsp;<span style='color:#fff;font-weight:400;font-size:11px'>({len(sorted_w)} positions)</span></div>"
            f"<table style='border-collapse:collapse;width:100%;border:1px solid #e8eaed;border-top:none;border-radius:0 0 6px 6px'>"
            f"<tbody>{rows_html}</tbody></table></div>"
        )
    return f"<div style='display:flex;flex-wrap:wrap;margin:0 -8px'>{''.join(blocks)}</div>"


def build_portfolio_email(
    *,
    tickers_ok: list[str],
    tickers_failed: list[str],
    summary_markdown: str,
    ranked_rows: list[dict] | None,
    strategy_results: dict[str, dict] | None,
    output_xlsx: Path | None,
    portfolio_docx: Path | None,
    gamma_url: str | None,
) -> EmailMessage:
    """Compose the EmailMessage for a portfolio-analysis run (HTML + plain text)."""
    import base64 as _b64

    to_addr = _optional_env("PORTFOLIO_EMAIL_TO", DEFAULT_PORTFOLIO_EMAIL_TO)
    from_addr = _optional_env("GMAIL_USER", to_addr)
    today = datetime.now().strftime("%Y-%m-%d")
    generated = datetime.now().strftime("%Y-%m-%d %H:%M")

    subject = f"DGA Portfolio Analysis — {today} — {len(tickers_ok)} positions"

    # Use strategy rows (always complete) for the ranked table, sorted by upside desc.
    primary_strategy = next(iter(strategy_results)) if strategy_results else None
    email_rows = []
    if strategy_results and primary_strategy:
        email_rows = sorted(
            strategy_results[primary_strategy].get("rows", []),
            key=lambda r: -(r.get("upside_pct") or 0),
        )

    # ---- Plain-text fallback ------------------------------------------------
    plain_parts = [
        "DGA Capital — Portfolio Analysis Run",
        f"Generated: {generated}",
        "",
        f"Tickers analyzed: {', '.join(tickers_ok) or '(none)'}",
    ]
    if tickers_failed:
        plain_parts.append(f"Failed: {', '.join(tickers_failed)}")
    if gamma_url:
        plain_parts.append(f"Gamma deck: {gamma_url}")
    plain_parts += ["", "RANKED TABLE", "=" * 60, _ranked_table_text(email_rows),
                    "", "STRATEGY WEIGHTS", "=" * 60, _strategy_weights_text(strategy_results),
                    "", "PORTFOLIO ROLL-UP", "=" * 60,
                    (summary_markdown or "(no roll-up generated)").strip()]
    plain_body = "\n".join(plain_parts)

    # ---- Logo (base64 embedded so it shows in all email clients) ------------
    # Logos are RGBA with transparent bg — wrap in a white pill so the dark
    # DGA lettering is visible on the navy header background.
    logo_img_tag = ""
    for logo_name in ("DGAlogo-webFINAL-68.png", "dga_logo_small.png", "DGAlogo-web184.png", "dga_logo.png"):
        logo_path = SCRIPT_DIR / "branding" / logo_name
        if logo_path.exists():
            logo_b64 = _b64.b64encode(logo_path.read_bytes()).decode()
            logo_img_tag = (
                "<div style='background:#ffffff;border-radius:8px;padding:6px 14px;"
                "display:inline-block;line-height:0'>"
                f"<img src='data:image/png;base64,{logo_b64}' "
                f"alt='DGA Capital' style='height:40px;width:auto;display:block'>"
                "</div>"
            )
            break

    # ---- HTML body ----------------------------------------------------------
    failed_row = ""
    if tickers_failed:
        failed_row = (
            f"<tr><td style='color:#888;width:130px'>Failed</td>"
            f"<td style='color:#c0392b'>{', '.join(tickers_failed)}</td></tr>"
        )
    gamma_row = ""
    if gamma_url:
        gamma_row = (
            f"<tr><td style='color:#888'>Gamma deck</td>"
            f"<td><a href='{gamma_url}' style='color:#C9A84C'>{gamma_url}</a></td></tr>"
        )
    md_html = _md_to_html((summary_markdown or "").strip()[:60_000]
                          or "<em>No roll-up generated.</em>")

    html_body = f"""<!DOCTYPE html>
<html><head><meta charset="UTF-8"></head>
<body style="margin:0;padding:0;background:#f4f6f8;font-family:'Helvetica Neue',Arial,sans-serif">
<table width="100%" cellpadding="0" cellspacing="0" style="background:#f4f6f8;padding:24px 0">
<tr><td align="center">
<table width="640" cellpadding="0" cellspacing="0" style="max-width:640px;width:100%">

  <!-- Header -->
  <tr><td style="background:#0A1628;border-radius:10px 10px 0 0;padding:24px 32px">
    <table cellpadding="0" cellspacing="0"><tr>
      <td style="vertical-align:middle">{logo_img_tag if logo_img_tag else '<span style="color:#C9A84C;font-size:22px;font-weight:700;letter-spacing:2px">DGA CAPITAL</span>'}</td>
      <td style="vertical-align:middle;padding-left:16px;color:rgba(255,255,255,0.5);font-size:13px;letter-spacing:1px">Portfolio Analysis</td>
    </tr></table>
  </td></tr>

  <!-- Meta -->
  <tr><td style="background:#fff;padding:20px 32px;border-bottom:2px solid #C9A84C">
    <table width="100%" cellpadding="0" cellspacing="0" style="font-size:13px;color:#333">
      <tr>
        <td style="color:#888;width:130px">Generated</td>
        <td style="font-weight:600">{generated}</td>
      </tr>
      <tr>
        <td style="color:#888;padding-top:4px">Tickers</td>
        <td style="padding-top:4px">{', '.join(tickers_ok) or '(none)'}</td>
      </tr>
      {failed_row}
      {gamma_row}
    </table>
  </td></tr>

  <!-- Ranked Table -->
  <tr><td style="background:#fff;padding:24px 32px">
    <h2 style="margin:0 0 14px;font-size:15px;font-weight:700;color:#0A1628;
               letter-spacing:1px;text-transform:uppercase;border-left:4px solid #C9A84C;padding-left:10px">
      Ranked Positions
    </h2>
    {_html_ranked_table(email_rows)}
  </td></tr>

  <!-- Strategy Weights -->
  <tr><td style="background:#f8f9fb;padding:24px 32px;border-top:1px solid #e8eaed">
    <h2 style="margin:0 0 16px;font-size:15px;font-weight:700;color:#0A1628;
               letter-spacing:1px;text-transform:uppercase;border-left:4px solid #C9A84C;padding-left:10px">
      Strategy Weights
    </h2>
    {_html_strategy_weights(strategy_results)}
  </td></tr>

  <!-- Roll-Up -->
  <tr><td style="background:#fff;padding:24px 32px;border-top:1px solid #e8eaed">
    <h2 style="margin:0 0 14px;font-size:15px;font-weight:700;color:#0A1628;
               letter-spacing:1px;text-transform:uppercase;border-left:4px solid #C9A84C;padding-left:10px">
      Portfolio Roll-Up
    </h2>
    <div style="font-size:13px;color:#333;line-height:1.6">
      {md_html}
    </div>
  </td></tr>

  <!-- Footer -->
  <tr><td style="background:#0A1628;border-radius:0 0 10px 10px;padding:16px 32px;text-align:center">
    <span style="color:#888;font-size:11px">DGA Capital · Portfolio Analysis · {today}</span>
    {"&nbsp;·&nbsp;<a href='" + gamma_url + "' style='color:#C9A84C;font-size:11px'>View Gamma Deck</a>" if gamma_url else ""}
  </td></tr>

</table>
</td></tr></table>
</body></html>"""

    msg = EmailMessage()
    msg["Subject"] = subject
    msg["From"] = from_addr
    msg["To"] = to_addr
    msg.set_content(plain_body)
    msg.add_alternative(html_body, subtype="html")

    # Attach xlsx and Word report.
    for path in (portfolio_docx, output_xlsx):
        if path and Path(path).is_file():
            ctype, _ = mimetypes.guess_type(str(path))
            maintype, subtype = (ctype.split("/", 1) if ctype else ("application", "octet-stream"))
            with open(path, "rb") as fh:
                msg.add_attachment(
                    fh.read(), maintype=maintype, subtype=subtype, filename=Path(path).name,
                )

    return msg


def _email_msg_to_resend_payload(msg: EmailMessage, from_override: str | None = None) -> dict:
    """Convert a stdlib EmailMessage into Resend's JSON schema.

    Extracts text body, html body, and base64-encodes every attachment.
    """
    import base64
    text_body = ""
    html_body = ""
    attachments: list[dict] = []

    # Walk every MIME part. set_content + add_alternative + add_attachment
    # produce a multipart tree, so we iterate over all parts.
    for part in msg.walk():
        if part.is_multipart():
            continue
        ctype = part.get_content_type()
        disp = (part.get("Content-Disposition") or "").lower()
        if "attachment" in disp or part.get_filename():
            try:
                payload = part.get_payload(decode=True) or b""
                attachments.append({
                    "filename": part.get_filename() or "attachment.bin",
                    "content": base64.b64encode(payload).decode("ascii"),
                    "content_type": ctype,
                })
            except Exception as exc:  # noqa: BLE001
                print(f"   ⚠️  Could not encode attachment {part.get_filename()}: {exc}")
            continue
        if ctype == "text/plain" and not text_body:
            text_body = part.get_content() if hasattr(part, "get_content") else str(part.get_payload(decode=True) or "", "utf-8", "replace")
        elif ctype == "text/html" and not html_body:
            html_body = part.get_content() if hasattr(part, "get_content") else str(part.get_payload(decode=True) or "", "utf-8", "replace")

    to_hdr = msg["To"] or ""
    to_list = [addr.strip() for addr in to_hdr.split(",") if addr.strip()]

    payload = {
        "from": from_override or msg["From"] or "onboarding@resend.dev",
        "to": to_list,
        "subject": msg["Subject"] or "(no subject)",
    }
    if html_body:
        payload["html"] = html_body
    if text_body:
        payload["text"] = text_body
    if attachments:
        payload["attachments"] = attachments
    return payload


def _send_via_resend(msg: EmailMessage) -> dict | None:
    """Send via Resend HTTPS API. Returns result dict, or None if RESEND_API_KEY absent.

    Railway blocks outbound SMTP on port 465/587 by default, so we need an
    HTTPS-based transport. Resend's free tier (3k emails/month) works and
    doesn't require domain verification when sending from onboarding@resend.dev.
    """
    api_key = _optional_env("RESEND_API_KEY", "")
    if not api_key:
        return None

    # Resend requires the From address to be on a verified domain.
    # Use RESEND_FROM env var (Railway) if set; otherwise default to the
    # verified dgacapital.com domain so emails reach any recipient.
    from_override = _optional_env("RESEND_FROM", "") or "DGA Capital <reports@dgacapital.com>"
    payload = _email_msg_to_resend_payload(msg, from_override=from_override)

    try:
        resp = requests.post(
            "https://api.resend.com/emails",
            headers={
                "Authorization": f"Bearer {api_key}",
                "Content-Type": "application/json",
            },
            json=payload,
            timeout=30,
        )
    except Exception as exc:  # noqa: BLE001
        return {"ok": False, "transport": "resend", "error": f"Resend network error: {exc}"}

    if resp.status_code in (200, 202):
        try:
            body = resp.json()
        except Exception:  # noqa: BLE001
            body = {}
        return {
            "ok": True,
            "transport": "resend",
            "sent_to": msg["To"],
            "resend_id": body.get("id", ""),
        }

    # Error path — surface the Resend error message so the user can fix the key.
    try:
        err_body = resp.json()
        err_msg = err_body.get("message") or err_body.get("error") or resp.text
    except Exception:  # noqa: BLE001
        err_msg = resp.text

    # Special-case the most common gotcha: Resend's free tier (no verified
    # domain) only lets you send to your own account email. Translate the raw
    # 403 into something actionable so the UI can show clear next steps.
    is_test_only = (
        resp.status_code == 403
        and ("verify a domain" in err_msg.lower() or "testing emails" in err_msg.lower())
    )
    if is_test_only:
        err_msg = (
            "Resend rejected the send — the From address may not be verified. "
            "Set the Railway env var RESEND_FROM to a verified sender "
            "(e.g. RESEND_FROM=reports@dgacapital.com)."
        )
    return {
        "ok": False,
        "transport": "resend",
        "error": f"Resend API {resp.status_code}: {err_msg}",
    }


def send_portfolio_email(msg: EmailMessage) -> dict:
    """Send portfolio email with smart multi-transport fallback.

    Transport priority:
      1. Resend HTTPS API (RESEND_API_KEY) — works on Railway
      2. Gmail SMTP (GMAIL_USER + GMAIL_APP_PASSWORD) — works locally
      3. Save .eml file to disk so the artifact always survives

    Returns {ok, sent_to, transport, fallback_path, error?}.
    """
    fallback = STOCKS_FOLDER / "Portfolio_Email.eml"
    try:
        with open(fallback, "wb") as fh:
            fh.write(bytes(msg))
    except Exception as exc:  # pragma: no cover — disk write should not fail in practice
        print(f"   ⚠️  Could not write fallback .eml: {exc}")

    errors: list[str] = []

    # --- Transport 1: Resend (HTTPS-based, Railway-compatible)
    resend_result = _send_via_resend(msg)
    if resend_result is not None:
        if resend_result.get("ok"):
            resend_result["fallback_path"] = str(fallback)
            print(f"   📧 Email sent via Resend → {msg['To']} "
                  f"(id: {resend_result.get('resend_id','')})")
            return resend_result
        # Resend returned an error; remember it and try SMTP too.
        errors.append(resend_result.get("error", "Resend unknown error"))

    # --- Transport 2: Gmail SMTP (works locally; blocked on Railway)
    user = _optional_env("GMAIL_USER", "")
    pwd = _optional_env("GMAIL_APP_PASSWORD", "")
    if user and pwd:
        try:
            ctx = ssl.create_default_context()
            with smtplib.SMTP_SSL("smtp.gmail.com", 465, context=ctx, timeout=30) as smtp:
                smtp.login(user, pwd)
                smtp.send_message(msg)
            print(f"   📧 Email sent via Gmail SMTP → {msg['To']}")
            return {
                "ok": True,
                "transport": "gmail_smtp",
                "sent_to": msg["To"],
                "fallback_path": str(fallback),
            }
        except Exception as exc:  # noqa: BLE001
            errors.append(f"Gmail SMTP failed: {exc}")

    # --- All transports failed (or none configured).
    if not errors:
        errors.append(
            "No email transport configured — set RESEND_API_KEY (recommended "
            "for Railway) or GMAIL_USER + GMAIL_APP_PASSWORD (for local use)."
        )
    return {
        "ok": False,
        "transport": "none",
        "sent_to": msg["To"],
        "fallback_path": str(fallback),
        "error": f"{' | '.join(errors)}. Email body saved to {fallback}.",
    }


# ============================================================================
# DGA_SYSTEM_PROMPT  (reads from /stocks/dga_system_prompt.txt if present, else uses default)
# ============================================================================
_DEFAULT_SYSTEM_PROMPT_PATH = STOCKS_FOLDER / "dga_system_prompt.txt"

DEFAULT_DGA_SYSTEM_PROMPT = """You are DGA Capital Analyst, a senior equity research analyst at Goldman Sachs level. You produce formal, institutional-quality research reports.

You are operating on today's date: Always base every analysis on the most recent market data, earnings, news, filings, and analyst updates available right now.
Use ONLY the price, previous close, and market cap numbers explicitly given in the user message. Never use outdated numbers.

FINANCIAL DATA EXTRACTION RULES - MANDATORY AND NON-NEGOTIABLE FOR ALL STOCK ANALYSIS:

You MUST treat financial numbers as the single most important part of the report. Any error here invalidates the entire analysis.

MANDATORY DATA VERIFICATION STEP (perform this at the VERY START):
1. You are given verified financial data below from official SEC EDGAR XBRL filings (companyfacts + submissions APIs).
2. You MUST use ONLY these exact numbers for ALL tables and calculations.
3. The verified block contains:
   - LATEST_FILING_TYPE (10-K or 10-Q)
   - ANNUAL DATA table: first row is TTM (pre-computed), followed by FY annual rows
   - QUARTERLY DATA (latest quarter + same quarter prior year + YTD both years) from the latest 10-Q
4. The TTM row in the ANNUAL DATA table is already computed — use it directly for the TTM column. Do NOT attempt to recalculate TTM from quarterly data.
5. Use the ANNUAL DATA for the main Key Metrics table (TTM column = TTM row, FY columns = annual rows).
6. If LATEST_FILING_TYPE is 10-Q, also create the Latest Quarterly YoY Analysis table using the QUARTERLY DATA.
7. If LATEST_FILING_TYPE is 10-K, skip the quarterly subsection entirely.
8. If any number is 'N/A', write exactly: "N/A".
9. All values in the VERIFIED block are in $ millions (already converted). Use them directly — do NOT divide by 1,000,000 again. EPS is in $ per share (use as-is).

TABLE FORMAT REQUIREMENT (use this exact markdown format - no deviations):
| Metric                          | TTM (as of [Date]) | FY[Year] (ended [Date]) | FY[Year-1] (ended [Date]) | FY[Year-2] (ended [Date]) |
|---------------------------------|--------------------|--------------------------|---------------------------|---------------------------|
| Revenue ($M)                    | exact_number      | exact_number            | exact_number             | exact_number             |
| Operating Income ($M)           | exact_number      | exact_number            | exact_number             | exact_number             |
| Operating Margin (%)            | xx.x%             | xx.x%                   | xx.x%                    | xx.x%                    |
| Net Income ($M)                 | exact_number      | exact_number            | exact_number             | exact_number             |
| Net Profit Margin (%)           | xx.x%             | xx.x%                   | xx.x%                    | xx.x%                    |
| Diluted EPS                     | exact_number      | exact_number            | exact_number             | exact_number             |
| Free Cash Flow ($M)             | exact_number      | exact_number            | exact_number             | exact_number             |
| Total Debt ($M)                 | exact_number      | exact_number            | exact_number             | exact_number             |
| Cash & Equivalents ($M)         | exact_number      | exact_number            | exact_number             | exact_number             |
| Net Debt ($M)                   | exact_number      | exact_number            | exact_number             | exact_number             |

At the top of the financial section you MUST write exactly:
"Data Verification: Official company FY[Year] 10-K + 10-Q filings (verified via SEC EDGAR XBRL). Numbers used are exact filing figures only. TTM pre-computed via bridge formula (FY annual + YTD delta)."

Sources must always be stated at the bottom exactly as:
"Sources: Official [Company Name] FY[Year] 10-K + 10-Q filings (verified via SEC EDGAR XBRL). Data Verification completed."

All other sections of the report (Executive Summary, Valuation, etc.) must be consistent with these verified financial numbers. Never contradict them.

GENERAL RULES:
- Put "DGA Capital Research" as the header and add today's date.

SECTION 1 — Executive Summary:
→ Investment thesis: Why should someone care about this stock right now?
→ Overall rating: Strong Buy / Buy / Hold / Sell
→ 12-month price target with the methodology used to calculate it
→ The single biggest reason to own this stock and the single biggest risk
→ The 30-second elevator pitch: If you had one paragraph to pitch this stock to a portfolio manager, what would you say?

SECTION 2 — Business Overview:
→ What the company does in plain English
→ Revenue breakdown by segment, product, and geography (with percentages)
→ Business model: How they make money and what drives repeat revenue
→ Competitive moat: What makes this company hard to replicate?

SECTION 3 — Financial Deep Dive:
→ Data Verification & Financial Foundation – Always begin with the MANDATORY DATA VERIFICATION STEP and full financial tables from the FINANCIAL DATA EXTRACTION RULES above.
→ Key Metrics: go to FINANCIAL DATA EXTRACTION RULES - MANDATORY FOR ALL STOCK ANALYSIS, and follow every step exactly as written, including table format.
→ Balance sheet health: cash, debt, current ratio, debt-to-equity
→ Cash flow quality: operating cash flow vs. net income ratio (flag if significantly different)
→ Capital allocation: How is management spending money? Buybacks, dividends, M&A, R&D?

→ Latest Quarterly YoY Analysis:
   ONLY include this subsection if the verified block says LATEST_FILING_TYPE: 10-Q.
   If LATEST_FILING_TYPE: 10-K, skip this subsection entirely.
   When included, create a dedicated Quarterly YoY Comparison Table using this exact column order and markdown format.

   TABLE FORMAT REQUIREMENT (use this exact markdown format - no deviations):
   | Metric                          | Latest Quarter (ended [Date]) | Same Quarter Last Year (ended [Date]) | YoY % Change | YTD Current Fiscal (ended [Date]) | Prior YTD | YoY % Change (YTD) |
   |---------------------------------|-------------------------------|---------------------------------------|--------------|-----------------------------------|-----------|--------------------|
   | Revenue ($M)                    | exact_number                 | exact_number                         | +xx.x%      | exact_number                     | exact_number | +xx.x%            |
   | Operating Income ($M)           | exact_number                 | exact_number                         | +xx.x%      | exact_number                     | exact_number | +xx.x%            |
   | Operating Margin (%)            | xx.x%                        | xx.x%                                | +xx.xppt    | xx.x%                            | xx.x%        | +xx.xppt          |
   | Net Income ($M)                 | exact_number                 | exact_number                         | +xx.x%      | exact_number                     | exact_number | +xx.x%            |
   | Net Profit Margin (%)           | xx.x%                        | xx.x%                                | +xx.xppt    | xx.x%                            | xx.x%        | +xx.xppt          |
   | Diluted EPS                     | exact_number                 | exact_number                         | +xx.x%      | exact_number                     | exact_number | +xx.x%            |
   | Free Cash Flow ($M)             | exact_number                 | exact_number                         | +xx.x%      | exact_number                     | exact_number | +xx.x%            |

   At the top of this subsection you MUST write exactly:
   "Latest Quarterly YoY Analysis – Official 10-Q filing (verified via SEC EDGAR XBRL). Numbers are exact filing figures only."

   Clearly state the exact quarter end dates (e.g., "Q2 FY2026 ended Feb 28, 2026 vs Q2 FY2025 ended Feb 28, 2025").
   Use ONLY the exact quarterly and YTD numbers from the verified financial data block.
   If the fiscal year is incomplete, the YTD columns must reflect the partial year-to-date results.
   Highlight any material acceleration or deceleration in growth with clear commentary.

SECTION 4 — Growth Analysis:
Financial Modeling & Projections – Build or update a detailed three-statement model; forecast 5–10 years of key metrics with explicit assumptions; calculate TTM and forward estimates.
→ Total addressable market (TAM) with source
→ Current market share and trajectory
→ Key growth drivers for the next 3-5 years
→ Management guidance vs. analyst consensus — who is more bullish?
→ Is growth organic or acquisition-dependent?

SECTION 5 — Valuation:
Multi-Method Valuation – Perform at least two primary methods:
   - Discounted Cash Flow (DCF) with explicit WACC, terminal growth, and sensitivity analysis.
   - Comparable company analysis (multiples: EV/EBITDA, P/E, etc.) and precedent transactions where relevant, minimum 5 peer comps.
   - Cross-check with any other appropriate method (e.g., sum-of-the-parts).
→ Historical valuation range (5-year P/E band)
→ Bull / Base / Bear price targets with assumptions for each
→ Current price vs. each target — upside or downside %

SECTION 6 — Risk Analysis:
→ Latest developments in the headlines: scan X (old Twitter), and summarize anything new announced in the past 30 days like Gov't investigations, lawsuits, executive turmoil, anything coming out of left field that could negatively affect the company.
→ Top 5 material risks ranked by probability and impact
→ For each risk: what would trigger it, how bad it would be, and what to watch for
→ Short interest and insider activity data (cite source)
→ Accounting quality flags (if any)

SECTION 7 — Catalyst Calendar:
→ Next earnings date
→ Upcoming product launches, regulatory decisions, or strategic events
→ Macro events that specifically impact this stock
→ Timeline of potential catalysts over the next 12 months

SECTION 7.5 — Institutional Analyst Consensus:

TWO PATHS — follow whichever applies:

PATH A — ANALYST_RATINGS_BLOCK IS present in the user message:
→ Use the exact values from the block for every row that has data. Do not substitute or invent different numbers.
→ For rows marked "Not available", fill in the best estimate from your most current training knowledge AND append "(est.)" to the Rating cell to distinguish it from confirmed data.
→ Compute Upside vs Current yourself from the price target and CURRENT_PRICE.
→ If a CONSENSUS_SUMMARY block is present, also cite the aggregate figures (number of analysts, mean/high/low target, consensus rating) in your Street vs DGA paragraph.

PATH B — NO ANALYST_RATINGS_BLOCK in the user message:
→ Use your most up-to-date training knowledge for each firm's rating and 12-month price target.
→ Append "(est.)" to every Rating cell so the reader knows these are model estimates, not confirmed live data.
→ Compute Upside vs Current from the price target and CURRENT_PRICE.

BOTH PATHS — always produce this exact table structure (do not omit it). Use the firm list from the ANALYST_RATINGS_BLOCK if present; otherwise include the 5 most-covered firms you know of:

| Firm | Rating | 12M Price Target | Upside vs Current | Date | Action |
|------|--------|-----------------|-------------------|------|--------|
| Goldman Sachs    | ... | $... | ±xx.x% | YYYY-MM-DD | ... |
| Morgan Stanley   | ... | $... | ±xx.x% | YYYY-MM-DD | ... |
| BofA Securities  | ... | $... | ±xx.x% | YYYY-MM-DD | ... |
| JPMorgan         | ... | $... | ±xx.x% | YYYY-MM-DD | ... |
| Wells Fargo      | ... | $... | ±xx.x% | YYYY-MM-DD | ... |
(and any additional rows provided by the ANALYST_RATINGS_BLOCK)

→ After the table write a "Street vs DGA" paragraph: explain where and why DGA's rating/target diverges from the Street consensus. When CONSENSUS_SUMMARY is provided, quote the aggregate numbers (# analysts, mean target, recommendation key) here.
→ If 3 or more firms disagree with the DGA rating direction, explicitly acknowledge it and explain the thesis divergence before the Section 8 verdict.

SECTION 8 — The Verdict:
→ Bull case: Price target and what has to go right (with probability estimate)
→ Base case: Price target and most likely scenario (with probability estimate)
→ Bear case: Price target and what could go wrong (with probability estimate)
→ Expected value calculation: Probability-weighted price target across all three scenarios
→ Final recommendation with conviction level: High / Medium / Low

SECTION 9 — Institutional Activity:
→ Top 5 institutional holders and their position changes last quarter
→ Any notable hedge fund activity (new positions or exits)

End with a sources section listing every data source used in this report.
"""


def load_system_prompt() -> str:
    """Use override file in /stocks/dga_system_prompt.txt if present."""
    if _DEFAULT_SYSTEM_PROMPT_PATH.exists():
        return _DEFAULT_SYSTEM_PROMPT_PATH.read_text()
    return DEFAULT_DGA_SYSTEM_PROMPT


# ============================================================================
# Market price (free, no-key)
# Primary:  yfinance fast_info — reliably exposes last_price + previous_close
# Fallback: Yahoo Finance chart API via raw requests
# ============================================================================
def fetch_market_snapshot(ticker: str) -> dict:
    """Best-effort current price + previous close + pct_change.

    Uses yfinance fast_info as the primary source (reliable previous_close =
    prior trading day's close) and falls back to the Yahoo chart JSON endpoint
    when yfinance is absent.

    Returns:
        {
            "price":          float | None  — latest trade / current session price
            "previous_close": float | None  — prior trading day's official close
            "pct_change":     float | None  — (price − prev_close) / prev_close × 100
            "market_cap":     float | None
            "source":         str
        }
    """
    out = {"price": None, "previous_close": None, "pct_change": None,
           "market_cap": None, "source": ""}

    # ── Primary: yfinance fast_info ──────────────────────────────────────────
    try:
        import yfinance as yf  # type: ignore
        fi = yf.Ticker(ticker).fast_info
        price = getattr(fi, "last_price", None)
        prev  = getattr(fi, "previous_close", None)
        mcap  = getattr(fi, "market_cap", None)
        if price and float(price) > 0:
            out["price"] = float(price)
            out["source"] = "Yahoo Finance (fast_info)"
        if prev and float(prev) > 0:
            out["previous_close"] = float(prev)
        if mcap:
            out["market_cap"] = mcap
        # If we have both price and previous_close, we're done.
        if out["price"] is not None and out["previous_close"] is not None:
            _attach_pct_change(out)
            return out
    except Exception as exc:  # noqa: BLE001
        print(f"   ⚠️  yfinance fast_info failed for {ticker}: {exc}")

    # ── Fallback: raw Yahoo chart API ────────────────────────────────────────
    try:
        url = f"https://query1.finance.yahoo.com/v8/finance/chart/{ticker}?interval=1d&range=5d"
        resp = requests.get(
            url,
            headers={"User-Agent": "Mozilla/5.0"},
            timeout=15,
        )
        data = resp.json()
        if isinstance(data, dict):
            chart  = data.get("chart", {}) if isinstance(data.get("chart"), dict) else {}
            result = chart.get("result", []) if isinstance(chart.get("result"), list) else []
            if result and isinstance(result[0], dict):
                meta = result[0].get("meta", {}) if isinstance(result[0].get("meta"), dict) else {}
                if out["price"] is None:
                    out["price"] = meta.get("regularMarketPrice") or meta.get("previousClose")
                if out["previous_close"] is None:
                    out["previous_close"] = meta.get("previousClose")
                if not out["source"]:
                    out["source"] = "Yahoo Finance"
    except Exception as exc:  # noqa: BLE001
        print(f"   ⚠️  Market snapshot chart fallback failed for {ticker}: {exc}")

    _attach_pct_change(out)
    return out


def _attach_pct_change(snapshot: dict) -> None:
    """Compute and attach pct_change in-place if price and previous_close are available."""
    price = snapshot.get("price")
    prev  = snapshot.get("previous_close")
    if price is not None and prev and float(prev) != 0:
        snapshot["pct_change"] = round((float(price) - float(prev)) / float(prev) * 100, 2)


# ============================================================================
# Watchlist — persistent list of tickers to scan
# ============================================================================

def load_watchlist() -> list[str]:
    """Return the saved watchlist (uppercase, deduplicated)."""
    try:
        if WATCHLIST_FILE.exists():
            raw = json.loads(WATCHLIST_FILE.read_text())
            tickers = raw.get("tickers", []) if isinstance(raw, dict) else []
            return [t.strip().upper() for t in tickers if isinstance(t, str) and t.strip()]
    except Exception:  # noqa: BLE001
        pass
    return []


def save_watchlist(tickers: list[str]) -> None:
    """Persist the watchlist to disk (atomic write)."""
    clean = list(dict.fromkeys(t.strip().upper() for t in tickers if t.strip()))
    tmp = WATCHLIST_FILE.with_suffix(".tmp")
    tmp.write_text(json.dumps({"tickers": clean}, indent=2))
    tmp.replace(WATCHLIST_FILE)


def add_to_watchlist(ticker: str) -> list[str]:
    current = load_watchlist()
    t = ticker.strip().upper()
    if t and t not in current:
        current.append(t)
        save_watchlist(current)
    return current


def remove_from_watchlist(ticker: str) -> list[str]:
    current = load_watchlist()
    t = ticker.strip().upper()
    current = [x for x in current if x != t]
    save_watchlist(current)
    return current


# ============================================================================
# Market Scan — daily news intelligence per ticker via Grok live search
# ============================================================================

SCAN_SYSTEM_PROMPT = """You are a real-time market intelligence scanner for a professional portfolio manager. Your job is FAST, SPECIFIC, and FACTUAL.

For each stock scan request you receive, use your live web and X (Twitter) search results to produce a concise daily market digest.

Rules:
- Be SPECIFIC: exact dates, exact dollar amounts, exact % figures, and source URLs where possible
- Be CURRENT: items from TODAY and YESTERDAY rank first; do not lead with 30-day-old news
- Be CONCISE: the entire response must fit in under 500 words
- NEVER invent news. If no confirmed live source exists for a fact, mark it "(unconfirmed)"
- For price movement: name the SINGLE most important driver, not generic "market sentiment"
- Tag every news item: [HIGH], [MED], or [LOW] by its market impact on this stock
- [HIGH] = material to a PM today: CEO change, earnings, M&A, regulatory ruling, large guidance revision
- [MED] = noteworthy: analyst action, product launch, partnership, material insider trade
- [LOW] = background: sector commentary, conference attendance, minor data point"""

_SCAN_USER_TEMPLATE = """\
DATE: {today}
TICKER: {ticker}
COMPANY: {company}
CURRENT_PRICE: {price}
PREVIOUS_CLOSE: {prev_close}
PRICE_CHANGE_PCT: {pct_change}%

Using live web and X search, produce a market scan for {ticker}. Use this exact format:

## {ticker} — ${price} ({sign}{pct_change_abs}%)

**📰 Today's Move:** [ONE sentence — the PRIMARY driver of today's price action. Be specific. Cite source.]

### News (newest first):
- **[HIGH|MED|LOW] YYYY-MM-DD — headline** — 1-2 sentence context. Source: URL

(Include 3–7 items. Skip LOW items if space is tight. Latest date first.)

### Earnings:
[If an earnings release occurred this week: Actual vs consensus EPS ($X.XX vs $X.XX est.), revenue ($XB vs $XB est.), and management outlook. Otherwise: "No earnings release this week."]

### Analyst Actions:
[Any upgrades, downgrades, or price-target changes from today or yesterday. Otherwise: "No analyst actions today."]

### Macro / Policy / Legal:
[Any macro event, Fed/central bank action, tariff/trade news, government policy, legislation, or lawsuit that specifically affects {ticker} today. Otherwise: "None identified."]

**🎯 Sentiment: [BULLISH|NEUTRAL|BEARISH]** — one sentence conclusion.
"""


def scan_ticker_news(ticker: str, verbose: bool = True) -> dict:
    """Run a live Grok scan for one ticker and return a structured result dict.

    Returns:
        {
            "ticker": str,
            "ok": bool,
            "scanned_at": str (ISO),
            "price": float | None,
            "previous_close": float | None,
            "pct_change": float | None,
            "markdown": str,   # the Grok-generated digest
            "sentiment": str,  # BULLISH | NEUTRAL | BEARISH | UNKNOWN
            "error": str | None,
        }
    """
    ticker = ticker.strip().upper()
    now_iso = datetime.utcnow().isoformat()

    mkt = fetch_market_snapshot(ticker)
    price = mkt.get("price")
    prev = mkt.get("previous_close")
    pct_change: float | None = None
    if price is not None and prev and prev != 0:
        pct_change = round((price - prev) / prev * 100, 2)

    today = datetime.now().strftime("%Y-%m-%d")
    sign = "+" if (pct_change or 0) >= 0 else ""
    pct_abs = abs(pct_change) if pct_change is not None else 0.0

    user_msg = _SCAN_USER_TEMPLATE.format(
        today=today,
        ticker=ticker,
        company=ticker,
        price=f"{price:.2f}" if price else "N/A",
        prev_close=f"{prev:.2f}" if prev else "N/A",
        pct_change=f"{pct_change:.2f}" if pct_change is not None else "N/A",
        sign=sign,
        pct_change_abs=f"{pct_abs:.2f}",
    )

    if verbose:
        print(f"   📡 Scanning {ticker} (${price}, {sign}{pct_abs:.2f}%)…")

    try:
        markdown = call_grok(SCAN_SYSTEM_PROMPT, user_msg, live_search=True)
    except Exception as exc:  # noqa: BLE001
        error_msg = str(exc)
        if verbose:
            print(f"   ❌ Scan failed for {ticker}: {error_msg}")
        return {
            "ticker": ticker,
            "ok": False,
            "scanned_at": now_iso,
            "price": price,
            "previous_close": prev,
            "pct_change": pct_change,
            "markdown": "",
            "sentiment": "UNKNOWN",
            "error": error_msg,
        }

    # Extract the BULLISH/NEUTRAL/BEARISH tag from the last line.
    sentiment = "NEUTRAL"
    for line in reversed(markdown.splitlines()):
        up = line.upper()
        if "BULLISH" in up:
            sentiment = "BULLISH"
            break
        if "BEARISH" in up:
            sentiment = "BEARISH"
            break
        if "NEUTRAL" in up:
            sentiment = "NEUTRAL"
            break

    if verbose:
        print(f"   ✅ {ticker} → {sentiment}")

    return {
        "ticker": ticker,
        "ok": True,
        "scanned_at": now_iso,
        "price": price,
        "previous_close": prev,
        "pct_change": pct_change,
        "markdown": markdown,
        "sentiment": sentiment,
        "error": None,
    }


def run_portfolio_scan(
    tickers: list[str],
    *,
    on_progress: "Any | None" = None,
    verbose: bool = True,
) -> dict:
    """Scan every ticker in the list and persist results to SCAN_RESULTS_FILE.

    ``on_progress(ticker, result)`` is called after each ticker completes so
    a background job can stream partial updates to the API.

    Returns:
        {
            "ok": bool,
            "scanned_at": str,
            "tickers": [...],
            "results": {TICKER: {...}, ...},
        }
    """
    scanned_at = datetime.utcnow().isoformat()
    results: dict[str, dict] = {}

    for ticker in tickers:
        result = scan_ticker_news(ticker, verbose=verbose)
        results[ticker] = result
        if on_progress is not None:
            try:
                on_progress(ticker, result)
            except Exception:  # noqa: BLE001
                pass

    payload = {
        "ok": bool(results),
        "scanned_at": scanned_at,
        "tickers": list(tickers),
        "results": results,
    }

    # Persist to disk so /api/scan/latest can serve it after a redeploy.
    try:
        tmp = SCAN_RESULTS_FILE.with_suffix(".tmp")
        tmp.write_text(json.dumps(payload, default=str, indent=2))
        tmp.replace(SCAN_RESULTS_FILE)
        if verbose:
            print(f"✅ Market scan complete — {len(results)} tickers. Saved to {SCAN_RESULTS_FILE.name}")
    except Exception as exc:  # noqa: BLE001
        print(f"⚠️  Could not persist scan results: {exc}")

    return payload


# ============================================================================
# Market Intelligence — macro → sector → company idea generation via Grok
# ============================================================================

INTEL_SYSTEM_PROMPT = """You are DGA Capital's chief investment strategist. Your role is to synthesize macro conditions, sector momentum, and individual company catalysts into actionable investment ideas for a long-only U.S. equity portfolio with a 3–5 year investment horizon.

Use your live web and X (Twitter) search access to ground every section in recent, verifiable facts. Be specific: name exact figures, dates, and sources. Do not be generic.

Output format rules:
- Use clean Markdown with the EXACT section headers below
- Under SPECIFIC NAMES, each company must have its ticker in the format **TICKER** on a new line — this is parsed by the app to create tappable links
- Be direct and investment-focused; no disclaimers, no generic risk language
- Aim for the highest expected-value ideas — bold contrarian calls are welcome if grounded in data
- Every section must be populated; if a sector has no genuine opportunity, name a different one"""

_INTEL_SECTOR_USER_TEMPLATE = """\
DATE: {today}
SECTOR_FOCUS: {sector}
INVESTMENT_HORIZON: 3–5 years

You are generating a sector-focused investment intelligence brief for DGA Capital. Use live web and X search to analyze the {sector} sector and identify 10–15 companies with the most asymmetric return potential right now.

"Asymmetric return potential" means: bounded downside, outsized upside from a specific catalyst or structural change that the market is currently underpricing.

Use EXACTLY this format:

---

## {sector_upper}: CURRENT SETUP

*What is driving the {sector} sector right now — macro, regulatory, technological, or competitive changes that create asymmetric opportunity.*

- [KEY DRIVER 1]: [2–3 sentences with specific data, dates, catalysts]
- [KEY DRIVER 2]: [2–3 sentences]
- [KEY DRIVER 3]: [2–3 sentences]

---

## TOP 10–15 NAMES: MOST ASYMMETRIC RETURN POTENTIAL

*Ranked by conviction — highest conviction first. Each name must have a SPECIFIC catalyst or mispricing thesis grounded in recent facts. Be direct and investment-focused.*

**TICKER**
**Company:** [Full name]
**Thesis:** [2–3 sentences — what is the market missing? what changes from here? what is the specific asymmetric setup?]
**Catalyst:** [1 sentence — what triggers the upside in 6–18 months]
**Risk:** [1 sentence — what breaks the thesis]
**Upside / Downside:** [estimated range, e.g. "2–3× upside vs 20% downside if thesis holds"]

[Repeat for each company — always start each block with **TICKER** on its own line]

---

## SECTOR RISK FACTORS

*2–3 risks that could impair the entire sector thesis.*

- [RISK 1]
- [RISK 2]
- [RISK 3 if applicable]
"""

_INTEL_BESTMIX_USER_TEMPLATE = """\
DATE: {today}
PORTFOLIO_TYPE: Best Mix — Cross-Sector High Conviction
INVESTMENT_HORIZON: 3–5 years

You are building DGA Capital's highest-conviction portfolio of 10–15 stocks spanning ALL sectors. These should be the BEST asymmetric opportunities in the entire market right now — not the most popular names, but the most mispriced given current macro, sector, and company-level catalysts.

"Asymmetric return potential" means: bounded downside, outsized upside. You are looking for situations where the market is wrong or slow to price in structural change. Use live web and X search to ground every pick in recent, verifiable facts.

Use EXACTLY this format:

---

## CURRENT MARKET REGIME

*Brief (3–5 bullet) view of macro conditions that frame WHY these are the best cross-sector opportunities today.*

- [KEY THEME 1]: [2 sentences — specific figures, dates]
- [KEY THEME 2]: [2 sentences]
- [KEY THEME 3]: [2 sentences]
- [KEY THEME 4 if relevant]: [2 sentences]

---

## BEST MIX: 10–15 HIGHEST CONVICTION NAMES

*The most asymmetric return opportunities across all sectors. At least 5 different sectors must be represented. Ranked by conviction — highest first.*

**TICKER**
**Company:** [Full name] | **Sector:** [sector name]
**Thesis:** [2–3 sentences — specific mispricing, catalyst, what the market is getting wrong]
**Upside / Downside:** [estimated range]
**Key Catalyst:** [1 sentence]
**Key Risk:** [1 sentence]

[Repeat for each name — always start each block with **TICKER** on its own line]

---

## PORTFOLIO NOTES

*1–2 sentences on sector concentration, correlation, and the single biggest portfolio-level risk in this mix.*
"""


def run_market_intelligence(sector: str = "Tech") -> dict:
    """Run a sector-focused intelligence scan via Grok live search.

    Args:
        sector: Sector to focus on (Tech, Energy, Healthcare, Financials,
                Consumer, Industrials, Materials, Real Estate, Best Mix).
                "Best Mix" generates a cross-sector high-conviction portfolio.

    Returns:
        {
            "ok": bool,
            "sector": str,
            "generated_at": str (ISO),
            "markdown": str,
            "tickers": list[str],   # parsed from **TICKER** tokens
            "error": str | None,
        }
    """
    sector = (sector or "Tech").strip()
    today = datetime.now().strftime("%Y-%m-%d")
    now_iso = datetime.utcnow().isoformat()

    if sector.lower() in ("best mix", "best_mix"):
        user_msg = _INTEL_BESTMIX_USER_TEMPLATE.format(today=today)
    else:
        user_msg = _INTEL_SECTOR_USER_TEMPLATE.format(
            today=today,
            sector=sector,
            sector_upper=sector.upper(),
        )

    print(f"🧠 Running market intelligence (sector: {sector})…")

    try:
        markdown = call_grok(
            INTEL_SYSTEM_PROMPT,
            user_msg,
            model=GROK_INTEL_MODEL,
            live_search=True,
        )
    except Exception as exc:  # noqa: BLE001
        print(f"❌ Intelligence scan failed: {exc}")
        return {
            "ok": False,
            "sector": sector,
            "generated_at": now_iso,
            "markdown": "",
            "tickers": [],
            "error": str(exc),
        }

    # Parse out **TICKER** tokens so the UI can make them tappable.
    import re as _re
    tickers = list(dict.fromkeys(
        m.group(1).upper()
        for m in _re.finditer(r'^\*\*([A-Z]{1,6})\*\*\s*$', markdown, _re.MULTILINE)
    ))

    payload = {
        "ok": True,
        "sector": sector,
        "generated_at": now_iso,
        "markdown": markdown,
        "tickers": tickers,
        "error": None,
    }

    # Persist to disk so the result survives server restarts / Railway redeploys.
    try:
        tmp = INTEL_FILE.with_suffix(".tmp")
        tmp.write_text(json.dumps(payload, default=str, indent=2))
        tmp.replace(INTEL_FILE)
        print(f"✅ Market intelligence complete — {len(tickers)} tickers identified.")
    except Exception as exc:  # noqa: BLE001
        print(f"⚠️  Could not persist intelligence results: {exc}")

    return payload


# ============================================================================
# Daily Brief — Goldman-Sachs-style PM morning note (live web + X search)
# ============================================================================
DAILY_BRIEF_SYSTEM_PROMPT = """You are the chief portfolio manager at DGA Capital, writing your morning brief the way a senior Goldman Sachs equity PM would: dense, specific, action-oriented, no fluff.

Your audience is one person — the firm's PM — who needs to walk into the trading floor with a complete read on the day in 90 seconds. They already know what the S&P is. They want EDGE: who said what overnight, what's mispriced, what to watch into the close, where the consensus is wrong.

Use your live web and X (Twitter) search aggressively. Pull from:
- Bloomberg, Reuters, WSJ, FT, CNBC headlines from the last 18 hours
- X posts from credible market voices (e.g., @zerohedge, @LizAnnSonders, @BobEUnlimited, @TheTranscript_, @SrsResearch, @Stoneocean, @sharkbarbs, @t1alpha, sell-side analysts, fund managers)
- Earnings call snippets, sell-side notes referenced in headlines
- Federal Reserve commentary, ECB/BoJ/PBOC statements
- Sector ETF flows / dark pool prints if mentioned in news
- Asia + Europe overnight close, futures action, FX, crude, gold, 10Y yield, VIX

Hard rules:
- NEVER hedge. Take a view. "Watch X — likely up 3% on the open" beats "X may move."
- Cite SPECIFIC numbers, prices, and times whenever possible
- Name SPECIFIC tickers using **TICKER** format (each on its own line in the names section) — these become tappable in the app
- Skip generic risk disclaimers and "consult your advisor" language
- If a section has nothing genuinely interesting, write "Nothing meaningful overnight" rather than padding with filler
- Total length target: ~700-900 words — dense but readable in under 2 minutes"""


_DAILY_BRIEF_USER_TEMPLATE = """\
DATE: {today}
TIME: Morning brief (pre-market US)

Write your morning brief for the DGA Capital trading floor. Use EXACTLY this format:

---

## ⚡ THE TAKE

*Two to three sentences. The single most important thing to know about today's market and how DGA should be positioned.*

---

## 🌍 OVERNIGHT TAPE

**Asia close:** [Nikkei, Hang Seng, Shanghai Comp — % moves and the why]
**Europe (live):** [STOXX 600, DAX, FTSE — % moves and the why]
**US futures:** [ES, NQ, RTY — % from prior close, key level being tested]
**Cross-asset:** [10Y yield, DXY, WTI crude, gold, BTC — meaningful moves only]
**VIX / spreads:** [Where vol and credit are pricing risk]

---

## 📅 TODAY'S CALENDAR

**Macro releases (with consensus):**
- [HH:MM ET] [Release] — consensus: X.X% / prior: X.X%
- [Repeat for each meaningful print]

**Earnings before/after the bell:**
- **TICKER** [BMO/AMC] — [what the Street is looking for; the one number that moves the stock]
- [Repeat for 3-6 most-watched names]

**Fed / Central bank speak:**
- [Speaker, time, what to listen for]

---

## 🎯 ACTIONABLE NAMES (5-8 tickers)

*Stocks where SOMETHING IS HAPPENING right now — earnings reaction, broker upgrade, news catalyst, technical breakout, or a developing thesis. Each block must start with **TICKER** on its own line.*

**TICKER**
**Setup:** [What's happening — last night's print, a downgrade, a rumor, a chart break]
**The trade:** [Long/short bias, key level, expected magnitude — be specific]
**Risk:** [What kills this in one sentence]

[Repeat for each name]

---

## 🔥 X / NEWS PULSE

*The 4-6 most important headlines or X posts from the last 18 hours, ranked by what actually moves portfolios. Each entry: source, the take, why it matters.*

- **[Source / @handle]** — [The take.] *Matters because…*
- [Repeat]

---

## 🧭 SECTOR & FACTOR ROTATION

*Where money is flowing this week. Be specific about ETFs (XLK, XLF, XLE, XLV, etc.) and factor baskets (momentum, low-vol, quality, value).*

- [Sector/factor]: [direction + the catalyst driving it]
- [Repeat 3-5 lines]

---

## ⚠️ CONTRARIAN WATCH

*One name where the consensus is provably wrong. Be bold.*

**TICKER**
**Consensus says:** [The narrative]
**You say:** [The contrarian view, with the data point that makes you right]
"""


def run_daily_brief() -> dict:
    """Run a Goldman-style morning brief via Grok 4.x with live web + X search.

    Returns:
        {
            "ok": bool,
            "generated_at": str (ISO),
            "markdown": str,
            "tickers": list[str],
            "error": str | None,
        }
    """
    today = datetime.now().strftime("%A, %B %d, %Y")
    now_iso = datetime.utcnow().isoformat()
    user_msg = _DAILY_BRIEF_USER_TEMPLATE.format(today=today)

    print(f"📰 Running Daily Brief ({GROK_INTEL_MODEL}) with live X + web search…")

    try:
        markdown = call_grok(
            DAILY_BRIEF_SYSTEM_PROMPT,
            user_msg,
            model=GROK_INTEL_MODEL,
            live_search=True,
        )
    except Exception as exc:  # noqa: BLE001
        print(f"❌ Daily Brief failed: {exc}")
        return {
            "ok": False,
            "generated_at": now_iso,
            "markdown": "",
            "tickers": [],
            "error": str(exc),
        }

    # Parse out **TICKER** tokens so the UI can make them tappable.
    import re as _re
    tickers = list(dict.fromkeys(
        m.group(1).upper()
        for m in _re.finditer(r'^\*\*([A-Z]{1,6})\*\*\s*$', markdown, _re.MULTILINE)
    ))

    payload = {
        "ok": True,
        "generated_at": now_iso,
        "date_str": today,
        "markdown": markdown,
        "tickers": tickers,
        "error": None,
    }

    # Persist to disk so it survives restarts and can be hydrated.
    try:
        DAILY_BRIEF_FILE.parent.mkdir(parents=True, exist_ok=True)
        tmp = DAILY_BRIEF_FILE.with_suffix(".tmp")
        tmp.write_text(json.dumps(payload, default=str, indent=2))
        tmp.replace(DAILY_BRIEF_FILE)
        # Also push to Dropbox so a Railway redeploy doesn't lose it.
        try:
            push_to_dropbox([str(DAILY_BRIEF_FILE)])
        except Exception:  # noqa: BLE001
            pass
        print(f"✅ Daily Brief complete — {len(tickers)} tickers flagged.")
    except Exception as exc:  # noqa: BLE001
        print(f"⚠️  Could not persist daily brief: {exc}")

    return payload


# ============================================================================
# Paper Portfolio Tracker — forward-track ideas vs SPY and live portfolio
# ----------------------------------------------------------------------------
# Single-file persistence (TRACKER_FILE) holds:
#   - portfolios:    list of paper portfolios (idea baskets)
#   - live_portfolio: auto-promoted from the most recent xlsx upload
#   - spy_series:    daily SPY closes shared across all portfolios
#
# A daily snapshot worker takes end-of-day closes for every tracking
# portfolio + SPY + live and appends them to each series. Snapshots happen
# opportunistically (whenever any tracker endpoint is called and today's
# snapshot hasn't been recorded after market close) plus a once-per-hour
# background thread.
# ============================================================================

import uuid as _uuid

_TRACKER_LOCK = None  # initialised lazily; threading is imported by callers


def _tracker_lock():
    """Lazy module-level lock so we don't depend on threading at import time."""
    global _TRACKER_LOCK
    if _TRACKER_LOCK is None:
        import threading as _threading
        _TRACKER_LOCK = _threading.Lock()
    return _TRACKER_LOCK


def _empty_tracker_state() -> dict:
    return {
        "portfolios":     [],
        "live_portfolio": None,
        "spy_series":     [],   # [{"date": "YYYY-MM-DD", "close": float}, ...]
    }


def _load_tracker_state() -> dict:
    """Read tracker.json (or return empty state)."""
    if not TRACKER_FILE.exists():
        return _empty_tracker_state()
    try:
        data = json.loads(TRACKER_FILE.read_text())
        if not isinstance(data, dict):
            return _empty_tracker_state()
        for k, default in _empty_tracker_state().items():
            data.setdefault(k, default)
        return data
    except Exception as exc:  # noqa: BLE001
        print(f"⚠️  Could not read tracker.json: {exc}")
        return _empty_tracker_state()


def _save_tracker_state(state: dict) -> None:
    """Atomic write + best-effort Dropbox sync."""
    try:
        TRACKER_FILE.parent.mkdir(parents=True, exist_ok=True)
        tmp = TRACKER_FILE.with_suffix(".tmp")
        tmp.write_text(json.dumps(state, default=str, indent=2))
        tmp.replace(TRACKER_FILE)
    except Exception as exc:  # noqa: BLE001
        print(f"⚠️  Could not write tracker.json: {exc}")
        return
    # Dropbox sync — fire-and-forget, never blocks
    try:
        upload_to_dropbox([str(TRACKER_FILE)])
    except Exception:  # noqa: BLE001
        pass


def _today_str() -> str:
    return datetime.now().strftime("%Y-%m-%d")


def _is_after_market_close() -> bool:
    """Return True if it's safely past 4:30pm ET (using a generous 21:00 UTC cutoff
    that works for both EST and EDT). On weekends, returns True so we fold any
    pending Friday close in."""
    now = datetime.utcnow()
    if now.weekday() >= 5:   # Sat/Sun
        return True
    return now.hour >= 21    # 4pm ET (DST) / 5pm ET (standard)


def _fetch_close_price(ticker: str) -> float | None:
    """Return the latest available daily close for *ticker* via yfinance."""
    try:
        import yfinance as yf  # type: ignore
        hist = yf.Ticker(ticker).history(period="5d", auto_adjust=False)
        if hist is None or hist.empty:
            return None
        return float(hist["Close"].dropna().iloc[-1])
    except Exception:  # noqa: BLE001
        # Fallback to fast_info / chart API
        snap = fetch_market_snapshot(ticker)
        return snap.get("price")


def _fetch_history_closes(ticker: str, days: int = 200) -> list[dict]:
    """Return [{"date": "YYYY-MM-DD", "close": float}, ...] going back *days*."""
    try:
        import yfinance as yf  # type: ignore
        hist = yf.Ticker(ticker).history(period=f"{days}d", auto_adjust=False)
        if hist is None or hist.empty:
            return []
        out = []
        for ts, row in hist.iterrows():
            close = row.get("Close")
            if close is None or pd.isna(close):
                continue
            out.append({
                "date":  ts.strftime("%Y-%m-%d"),
                "close": float(close),
            })
        return out
    except Exception:  # noqa: BLE001
        return []


def _ensure_spy_back_to(state: dict, anchor_date: str) -> None:
    """Make sure spy_series goes back at least to *anchor_date*. Fills via yfinance."""
    series = state.get("spy_series") or []
    earliest = series[0]["date"] if series else None
    if earliest and earliest <= anchor_date:
        return
    # Backfill — pull enough history to cover the anchor
    today = datetime.now().date()
    try:
        anchor_dt = datetime.strptime(anchor_date, "%Y-%m-%d").date()
        days_needed = max(30, (today - anchor_dt).days + 10)
    except Exception:  # noqa: BLE001
        days_needed = 200
    fresh = _fetch_history_closes("SPY", days=days_needed)
    if not fresh:
        return
    # Merge by date (latest wins)
    by_date = {p["date"]: p for p in series}
    for p in fresh:
        by_date[p["date"]] = p
    state["spy_series"] = sorted(by_date.values(), key=lambda r: r["date"])


def _portfolio_value_today(holdings: list[dict]) -> float:
    """Compute weighted return value (base 100) using current prices vs entry."""
    total_weight = sum(h.get("weight", 0) for h in holdings) or 1.0
    cum = 0.0
    for h in holdings:
        entry = h.get("entry_price") or h.get("anchor_price")
        cur   = _fetch_close_price(h["ticker"])
        if not entry or not cur or entry <= 0:
            continue
        w = h.get("weight", 0) / total_weight
        cum += w * (cur / entry)
    return round(cum * 100.0, 4) if cum else 100.0


# In-memory cache: year-start close price per (year, ticker). Doesn't change
# during the year, so a single fetch per ticker per session is plenty.
_YEAR_START_CACHE: dict[tuple[int, str], float] = {}


def _year_start_close(ticker: str, year: int | None = None) -> float | None:
    """Return the closing price on the first trading day of *year* (default: this year)."""
    year = year or datetime.now().year
    key = (year, ticker.upper())
    if key in _YEAR_START_CACHE:
        return _YEAR_START_CACHE[key]
    try:
        import yfinance as yf  # type: ignore
        # First two weeks of January is always enough to find the first trading day
        end = datetime(year, 1, 15)
        start = datetime(year, 1, 1)
        hist = yf.Ticker(ticker).history(start=start, end=end, auto_adjust=False)
        if hist is None or hist.empty:
            return None
        first_close = hist["Close"].dropna()
        if first_close.empty:
            return None
        price = float(first_close.iloc[0])
        _YEAR_START_CACHE[key] = price
        return price
    except Exception:  # noqa: BLE001
        return None


def _ytd_history_closes(ticker: str) -> list[dict]:
    """Return [{date, close}, ...] from Jan 1 of the current year through today."""
    year = datetime.now().year
    today = datetime.now()
    days_to_pull = (today.date() - datetime(year, 1, 1).date()).days + 7
    history = _fetch_history_closes(ticker, days=days_to_pull)
    return [p for p in history if p["date"] >= f"{year}-01-01"]


def _build_ytd_series_for_holdings(holdings: list[dict]) -> list[dict]:
    """Daily YTD value series for a basket of weighted holdings.

    Each ticker's "year-start" baseline is its first close on or after Jan 1.
    Forward-fills missing dates so the series is dense across all trading days.
    Returns [{"date", "value", "return_pct"}, ...] rebased so day-1 ≈ 100.
    """
    if not holdings:
        return []

    total_weight = sum(h.get("weight", 0) for h in holdings) or 1.0
    ticker_data: dict[str, dict] = {}
    for h in holdings:
        ytd = _ytd_history_closes(h["ticker"])
        if not ytd:
            continue
        closes = {p["date"]: p["close"] for p in ytd}
        first_price = ytd[0]["close"]
        if first_price <= 0:
            continue
        ticker_data[h["ticker"]] = {
            "weight":      h.get("weight", 0) / total_weight,
            "closes":      closes,
            "first_price": first_price,
        }

    if not ticker_data:
        return []

    all_dates = sorted(set().union(*[set(t["closes"].keys()) for t in ticker_data.values()]))
    last_price = {ticker: None for ticker in ticker_data}

    series: list[dict] = []
    for date in all_dates:
        # Update last-known price for each ticker that traded today
        for ticker, t in ticker_data.items():
            if date in t["closes"]:
                last_price[ticker] = t["closes"][date]
        # Weighted sum (only tickers with at least one close so far contribute)
        cum_value, cum_weight = 0.0, 0.0
        for ticker, t in ticker_data.items():
            if last_price[ticker] is not None:
                cum_value  += t["weight"] * (last_price[ticker] / t["first_price"])
                cum_weight += t["weight"]
        if cum_weight > 0:
            normalized = cum_value / cum_weight  # rescale for any tickers not yet trading
            series.append({
                "date":       date,
                "value":      round(normalized * 100.0, 4),
                "return_pct": round((normalized - 1) * 100.0, 4),
            })
    return series


def _spy_ytd_series() -> list[dict]:
    """SPY's YTD series rebased to 100 on the first trading day of the year."""
    ytd = _ytd_history_closes("SPY")
    if not ytd:
        return []
    base = ytd[0]["close"]
    if base <= 0:
        return []
    return [{
        "date":       p["date"],
        "value":      round(p["close"] / base * 100.0, 4),
        "return_pct": round((p["close"] / base - 1) * 100.0, 4),
    } for p in ytd]


def compute_live_ytd_detail(snapshot_id: str | None = None) -> dict:
    """YTD attribution detail for the auto-promoted live portfolio.

    Uses *year-start prices* as the per-holding baseline (instead of the
    upload-day anchor price) — answers "how is the live book doing YTD?"
    Compares the portfolio against SPY YTD only (vs-live makes no sense
    since this *is* live).

    If `snapshot_id` is provided, opens the YTD detail using that historical
    snapshot's holdings + attribution instead of the current live state.
    This lets the UI re-open any past upload from the snapshot history.

    Response shape mirrors get_idea_portfolio() so the same renderers work,
    with one extra `mode: "live"` field plus `year_start_date`.
    """
    state = _load_tracker_state()
    live = state.get("live_portfolio")
    if not live or not (live.get("holdings") or []):
        return {"ok": False, "error": "No live portfolio yet — run a portfolio rebalance first."}

    today = datetime.now()
    year = today.year

    # If a snapshot_id was requested, swap in that snapshot's holdings + attribution
    # for the rest of the calculation.  We only override what the snapshot stored.
    pinned_snapshot: dict | None = None
    if snapshot_id:
        for s in (live.get("ytd_snapshots") or []):
            if s.get("id") == snapshot_id:
                pinned_snapshot = s
                break
        if pinned_snapshot is None:
            return {"ok": False, "error": f"Snapshot {snapshot_id!r} not found."}
        holdings_for_calc = pinned_snapshot.get("holdings_snapshot") or live["holdings"]
    else:
        holdings_for_calc = live["holdings"]

    # Build a "virtual" holdings list using year-start prices as entry_price.
    # Weights are taken directly from the stored live portfolio — they are the
    # raw values from the uploaded CSV (e.g., INTC 0.0788 for 7.88%).
    virtual: list[dict] = []
    skipped: list[str] = []
    for h in holdings_for_calc:
        ys = _year_start_close(h["ticker"], year=year)
        if not ys or ys <= 0:
            skipped.append(h["ticker"])
            continue
        virtual.append({
            "ticker":      h["ticker"],
            "weight":      h.get("weight", 0),
            "entry_price": round(ys, 4),
            "entry_date":  f"{year}-01-02",
        })

    if not virtual:
        return {"ok": False, "error": "Could not resolve year-start prices for any holdings."}

    # Normalize weights to the tracked universe so portfolio-level return is
    # accurate even if some holdings were excluded (e.g., preferred shares with
    # no yfinance price).  Individual contribution figures still use raw weights
    # so "INTC 7.88% × +140% = 11.0%" reads correctly; we scale the total to
    # the sum of tracked weights so they add up to the displayed portfolio %.
    _tracked_total = sum(h["weight"] for h in virtual) or 1.0
    for h in virtual:
        h["weight"] = round(h["weight"] / _tracked_total, 8)

    bd = _portfolio_breakdown(virtual)
    snapshot_holdings = sorted(
        bd["breakdown"],
        key=lambda h: (h.get("contribution_pct") if h.get("contribution_pct") is not None else -1e9),
        reverse=True,
    )

    # Daily series for the chart
    portfolio_series = _build_ytd_series_for_holdings(virtual)
    spy_series       = _spy_ytd_series()

    spy_ytd = spy_series[-1]["return_pct"] if spy_series else None
    port_ret = bd["weighted_return"]

    # ── Pull stored YTD result (Modified Dietz + transaction-aware attribution) ──
    # If the user has uploaded the unified YTD, prefer those numbers everywhere:
    #  - md_return_pct as the authoritative portfolio return
    #  - per-ticker attribution rows reflecting real partial sales / buys / divs
    #    (replacing the broken weight × snapshot-return approximation)
    history       = pinned_snapshot or live.get("account_history") or {}
    md_return_pct = history.get("md_return_pct")
    stored_attr   = history.get("attribution") or []
    authoritative_return = md_return_pct if md_return_pct is not None else port_ret

    # Build the holdings rows for the Live YTD detail. Use transaction-aware
    # attribution if we have it; otherwise fall back to the snapshot breakdown.
    if stored_attr:
        sorted_holdings = []
        for a in stored_attr:
            end_sh   = a.get("end_shares")    or 0.0
            start_sh = a.get("start_shares")  or 0.0
            ret_pct  = a.get("ticker_return_pct")
            sorted_holdings.append({
                "ticker":             a.get("ticker"),
                "weight":             (a.get("end_value") or 0.0) / max(history.get("end_value") or 1.0, 1.0),
                "entry_price":        a.get("jan1_price"),
                "current_price":      a.get("end_price"),
                "return_pct":         ret_pct,
                "contribution_pct":   a.get("contribution_pct"),
                "dollar_gain":        a.get("dollar_gain"),
                "start_shares":       start_sh,
                "end_shares":         end_sh,
                "trade_count":        a.get("trade_count", 0),
                "dividends_cash":     a.get("dividends_cash", 0.0),
                "total_sold_shares":  a.get("total_sold_shares", 0.0),
                "total_bought_shares": a.get("total_bought_shares", 0.0),
                "vs_avg":             None,  # not meaningful with $-based contribution
            })
        sorted_holdings.sort(
            key=lambda h: abs(h.get("contribution_pct") or 0.0),
            reverse=True,
        )
    else:
        sorted_holdings = snapshot_holdings

    # Days tracked = number of calendar days since year start
    try:
        ys_dt = datetime(year, 1, 2)
        days_tracked = max(0, (today - ys_dt).days)
    except Exception:  # noqa: BLE001
        days_tracked = 0

    # Max drawdown from the portfolio series
    peak = -1e9
    max_dd = 0.0
    for pt in portfolio_series:
        v = float(pt["value"])
        if v > peak:
            peak = v
        dd = (peak - v) / peak * 100.0 if peak > 0 else 0.0
        if dd > max_dd:
            max_dd = dd

    return {
        "ok":                    True,
        "mode":                  "live",
        "name":                  "Live Portfolio · YTD",
        "year_start_date":       f"{year}-01-02",
        "entry_date":            f"{year}-01-02",
        "days_tracked":          days_tracked,
        "n_tickers":             len(sorted_holdings),
        "current_return_pct":    authoritative_return,
        "weighted_avg_return":   authoritative_return,
        "snapshot_return_pct":   port_ret,            # snapshot estimate (always shown)
        "md_return_pct":         md_return_pct,        # None if no history uploaded yet
        "history_uploaded":      md_return_pct is not None,
        "history_meta":          history if history else None,
        "snapshot_id":           (pinned_snapshot or {}).get("id") or (history.get("id") if history else None),
        "is_pinned":             pinned_snapshot is not None,
        "max_drawdown_pct":      round(max_dd, 4),
        "spy_return_pct":        spy_ytd,
        "vs_spy_pct":            round(authoritative_return - spy_ytd, 4) if spy_ytd is not None else None,
        "holdings":              sorted_holdings,
        "series":                portfolio_series,
        "spy_series":            spy_series,
        "live_series":           [],   # this IS live — no separate live benchmark
        "skipped_tickers":       skipped,
        "anchor_date":           live.get("anchor_date"),
        "uploaded_at":           live.get("uploaded_at"),
    }


def _portfolio_breakdown(holdings: list[dict]) -> dict:
    """Compute per-holding return + contribution + vs-portfolio-avg.

    contribution_pct = weight × return_pct  using RAW weights (never normalized).
    A 3.35% IBRX position with +254% YTD return contributes 3.35% × 254% = 8.5%
    to the portfolio's total return — NOT a re-normalized figure that would
    overstate the position's actual impact.
    """
    breakdown: list[dict] = []
    weighted_return = 0.0

    for h in holdings:
        entry = h.get("entry_price") or h.get("anchor_price")
        cur   = _fetch_close_price(h["ticker"])
        raw_w = h.get("weight", 0) or 0
        if entry and cur and entry > 0:
            ret = (cur / entry - 1) * 100
            contribution = raw_w * ret    # raw weight × return — no normalization
            weighted_return += contribution
        else:
            ret = None
            contribution = None
        breakdown.append({
            "ticker":           h["ticker"],
            "weight":           round(raw_w, 6),
            "entry_price":      entry,
            "current_price":    round(cur, 4) if cur else None,
            "return_pct":       round(ret, 4) if ret is not None else None,
            "contribution_pct": round(contribution, 4) if contribution is not None else None,
            "entry_date":       h.get("entry_date"),
        })

    # Each holding's deviation from portfolio's weighted-average return
    for b in breakdown:
        if b["return_pct"] is not None:
            b["vs_avg_pct"] = round(b["return_pct"] - weighted_return, 4)
        else:
            b["vs_avg_pct"] = None

    return {
        "value":           round(weighted_return + 100.0, 4),
        "weighted_return": round(weighted_return, 4),
        "breakdown":       breakdown,
    }


def _spy_value_at(state: dict, target_date: str) -> float | None:
    """Return SPY close on or just before *target_date*."""
    series = state.get("spy_series") or []
    eligible = [r for r in series if r["date"] <= target_date]
    if not eligible:
        return None
    return float(eligible[-1]["close"])


def create_idea_portfolio(name: str, holdings_input: list[dict],
                          source: dict | None = None) -> dict:
    """Lock in a new paper portfolio.

    holdings_input items: {"ticker": str, "weight": float (0..1 or 0..100)}
    Returns the persisted portfolio dict.
    """
    cleaned: list[dict] = []
    today = _today_str()
    seen = set()
    for h in holdings_input:
        t = (h.get("ticker") or "").strip().upper()
        if not t or t in seen:
            continue
        seen.add(t)
        w = float(h.get("weight") or 0)
        if w > 1.5:        # caller used percentages
            w = w / 100.0
        if w <= 0:
            continue
        entry = _fetch_close_price(t)
        if entry is None or entry <= 0:
            print(f"   ⚠️  No price for {t} — skipping")
            continue
        cleaned.append({
            "ticker":      t,
            "weight":      round(w, 6),
            "entry_price": round(entry, 4),
            "entry_date":  today,
        })
    if not cleaned:
        raise ValueError("No valid tickers with prices could be locked in.")

    # Normalise weights to sum to 1.0 (in case caller dropped some)
    total_w = sum(h["weight"] for h in cleaned) or 1.0
    for h in cleaned:
        h["weight"] = round(h["weight"] / total_w, 6)

    portfolio = {
        "id":         str(_uuid.uuid4()),
        "name":       (name or f"Brief — {today}").strip(),
        "created_at": datetime.utcnow().isoformat(),
        "source":     source or {},
        "status":     "tracking",
        "holdings":   cleaned,
        "series":     [{
            "date":       today,
            "value":      100.0,
            "return_pct": 0.0,
        }],
        "notes": "",
    }

    with _tracker_lock():
        state = _load_tracker_state()
        state["portfolios"].append(portfolio)
        _ensure_spy_back_to(state, today)
        _save_tracker_state(state)
    print(f"📌 Locked in paper portfolio '{portfolio['name']}' "
          f"({len(cleaned)} tickers) — id={portfolio['id'][:8]}")
    return portfolio


def list_idea_portfolios() -> list[dict]:
    """Return all portfolios with computed performance metrics (no series)."""
    state = _load_tracker_state()
    out = []
    for p in state.get("portfolios", []):
        out.append(_with_metrics(p, state, include_series=False))
    # Sort: tracking first, then by created_at desc
    out.sort(key=lambda r: (r.get("status") != "tracking", -_to_ts(r.get("created_at"))))
    return out


def get_idea_portfolio(portfolio_id: str) -> dict | None:
    """Return one portfolio with full series + benchmarks for charting."""
    state = _load_tracker_state()
    for p in state.get("portfolios", []):
        if p.get("id") == portfolio_id:
            return _with_metrics(p, state, include_series=True)
    return None


def close_idea_portfolio(portfolio_id: str) -> bool:
    with _tracker_lock():
        state = _load_tracker_state()
        for p in state.get("portfolios", []):
            if p.get("id") == portfolio_id:
                p["status"] = "closed"
                p["closed_at"] = datetime.utcnow().isoformat()
                _save_tracker_state(state)
                return True
    return False


def delete_idea_portfolio(portfolio_id: str) -> bool:
    with _tracker_lock():
        state = _load_tracker_state()
        before = len(state.get("portfolios", []))
        state["portfolios"] = [p for p in state.get("portfolios", [])
                               if p.get("id") != portfolio_id]
        if len(state["portfolios"]) == before:
            return False
        _save_tracker_state(state)
        return True


def _to_ts(iso_str: str | None) -> float:
    if not iso_str:
        return 0.0
    try:
        return datetime.fromisoformat(iso_str.replace("Z", "")).timestamp()
    except Exception:  # noqa: BLE001
        return 0.0


def _with_metrics(p: dict, state: dict, include_series: bool = False) -> dict:
    """Decorate a portfolio dict with current return + benchmark deltas."""
    raw_holdings = p.get("holdings", []) or []
    series = p.get("series", []) or []
    entry_date = (raw_holdings[0].get("entry_date") if raw_holdings
                  else p.get("created_at", "")[:10])

    # Sort holdings by weight desc for display (consistent ordering everywhere)
    sorted_h = sorted(raw_holdings, key=lambda h: h.get("weight", 0), reverse=True)
    holdings_summary = [
        {"ticker": h["ticker"], "weight": h.get("weight", 0)} for h in sorted_h
    ]

    cur_value = _portfolio_value_today(raw_holdings)
    return_pct = round(cur_value - 100.0, 4)

    # Max drawdown from series
    peak = -1e9
    max_dd = 0.0
    for pt in series:
        v = float(pt.get("value", 100.0))
        if v > peak:
            peak = v
        dd = (peak - v) / peak * 100.0 if peak > 0 else 0.0
        if dd > max_dd:
            max_dd = dd

    # SPY benchmark
    spy_anchor = _spy_value_at(state, entry_date)
    spy_today  = _fetch_close_price("SPY")
    spy_return = None
    if spy_anchor and spy_today:
        spy_return = round((spy_today / spy_anchor - 1) * 100.0, 4)

    # Live portfolio benchmark
    live_return = None
    live = state.get("live_portfolio")
    if live and live.get("anchor_date") and live.get("anchor_date") <= entry_date:
        # Live anchor must predate the paper portfolio, otherwise apples-to-oranges
        live_value = _portfolio_value_today(live.get("holdings", []))
        # We compare the live's *return since paper portfolio's entry date* — i.e.
        # we re-anchor live to entry_date by looking up its series at that date.
        live_series = live.get("series", []) or []
        eligible = [r for r in live_series if r["date"] <= entry_date]
        if eligible:
            base_v = eligible[-1].get("value", 100.0)
            if base_v > 0:
                live_return = round((live_value / base_v - 1) * 100.0, 4)
        else:
            # No series back to entry_date — fall back to live's all-time return
            live_return = round(live_value - 100.0, 4)

    days_tracked = 0
    try:
        if entry_date:
            d0 = datetime.strptime(entry_date, "%Y-%m-%d")
            days_tracked = (datetime.utcnow() - d0).days
    except Exception:  # noqa: BLE001
        pass

    out = {
        "id":           p.get("id"),
        "name":         p.get("name"),
        "created_at":   p.get("created_at"),
        "closed_at":    p.get("closed_at"),
        "status":       p.get("status", "tracking"),
        "source":       p.get("source", {}),
        "n_tickers":    len(raw_holdings),
        "entry_date":   entry_date,
        "days_tracked": days_tracked,
        "current_return_pct": return_pct,
        "max_drawdown_pct":   round(max_dd, 4),
        "vs_spy_pct":   None if spy_return is None else round(return_pct - spy_return, 4),
        "vs_live_pct":  None if live_return is None else round(return_pct - live_return, 4),
        "spy_return_pct":   spy_return,
        "live_return_pct":  live_return,
        "milestones": {
            "d30": days_tracked >= 30,
            "d60": days_tracked >= 60,
            "d90": days_tracked >= 90,
        },
        # Lightweight ticker info (always available — used by list view chips)
        "tickers":          [h["ticker"] for h in sorted_h],
        "holdings_summary": holdings_summary,
    }
    if include_series:
        bd = _portfolio_breakdown(raw_holdings)
        # Sort holdings by contribution desc — winners on top, detractors at bottom.
        sorted_breakdown = sorted(
            bd["breakdown"],
            key=lambda h: (h.get("contribution_pct")
                           if h.get("contribution_pct") is not None else -1e9),
            reverse=True,
        )
        out["holdings"]            = sorted_breakdown
        out["weighted_avg_return"] = bd["weighted_return"]
        out["series"]              = series
        out["spy_series"]          = _spy_aligned_series(state, entry_date)
        out["live_series"]         = _live_aligned_series(state, entry_date)
    return out


def _spy_aligned_series(state: dict, anchor_date: str) -> list[dict]:
    """Return SPY's value series rebased to 100 on anchor_date."""
    series = state.get("spy_series") or []
    eligible = [r for r in series if r["date"] >= anchor_date]
    if not eligible:
        return []
    base = float(eligible[0]["close"])
    if base <= 0:
        return []
    return [{
        "date":       r["date"],
        "value":      round(float(r["close"]) / base * 100.0, 4),
        "return_pct": round((float(r["close"]) / base - 1) * 100.0, 4),
    } for r in eligible]


def _live_aligned_series(state: dict, anchor_date: str) -> list[dict]:
    """Return live portfolio's value series rebased to 100 on anchor_date."""
    live = state.get("live_portfolio")
    if not live:
        return []
    series = live.get("series", []) or []
    eligible = [r for r in series if r["date"] >= anchor_date]
    if not eligible:
        return []
    base = float(eligible[0]["value"])
    if base <= 0:
        return []
    return [{
        "date":       r["date"],
        "value":      round(float(r["value"]) / base * 100.0, 4),
        "return_pct": round((float(r["value"]) / base - 1) * 100.0, 4),
    } for r in eligible]


# ── Transaction-aware performance attribution ────────────────────────────────

def _parse_fidelity_positions_extended(raw_text: str) -> dict:
    """Parse a Fidelity positions CSV and return per-holding quantities and prices.

    Extends the basic weight-only parser to also capture:
      - Quantity (number of shares)
      - Last Price (current price per share)
      - Computed Current Value (quantity × price)

    Columns up through Last Price appear BEFORE the "Current Value" column that
    has the unquoted-comma problem, so they parse cleanly via csv.reader.

    Returns:
        {
          "ok": True,
          "holdings": [{"ticker", "quantity", "price", "current_value", "weight"}],
          "total_value": float,         # sum of all position values
          "columns_detected": [str],
        }
    """
    import csv as _csv_m
    from io import StringIO as _StringIO_m

    lines = raw_text.splitlines()

    # Find header row (Symbol + Quantity + optional Percent Of Account)
    header_idx = None
    for i, line in enumerate(lines[:25]):
        low = line.strip().lower()
        if "symbol" in low and "quantity" in low:
            header_idx = i
            break
    if header_idx is None:
        return {"ok": False, "error": "Could not find positions CSV header row (need Symbol + Quantity columns)."}

    # Parse header to locate Quantity and Last Price column indices
    try:
        hdr_fields = next(_csv_m.reader([lines[header_idx]]))
    except Exception:
        return {"ok": False, "error": "Could not parse positions CSV header row."}

    hdr_lower = [h.strip().lower() for h in hdr_fields]
    columns_detected = [h.strip() for h in hdr_fields]

    qty_idx   = next((i for i, h in enumerate(hdr_lower) if h == "quantity"), None)
    price_idx = next((i for i, h in enumerate(hdr_lower)
                      if "last price" in h and "change" not in h), None)
    # "Current Value" column — reliable direct fallback (Fidelity always populates it)
    curval_idx = next((i for i, h in enumerate(hdr_lower)
                       if h in ("current value", "value")), None)

    holdings: list[dict] = []
    total_value = 0.0

    # Money-market / cash tickers — included as positions with is_mm=True.
    # These maintain a ~$1.00 NAV; their "gain" is the interest/dividends they earn.
    _mm_set = {
        "SPAXX",  # Fidelity Government Money Market
        "FZFXX",  # Fidelity Treasury Money Market
        "FZSXX",  # Fidelity Tax-Exempt Money Market
        "FDRXX",  # Fidelity Cash Reserves
        "FZDXX",  # Fidelity Money Market Premium Class
        "VMFXX",  # Vanguard Federal Money Market
        "VMRXX",  # Vanguard Cash Reserves Federal MM
        "SWVXX",  # Schwab Value Advantage Money Fund
        "FCASH",  # Fidelity FCASH (core position cash)
        "CASH",   # Generic cash designation
    }

    for raw_line in lines[header_idx + 1:]:
        line = raw_line.strip()
        if not line:
            break  # blank line = end of positions section

        try:
            fields = next(_csv_m.reader([line]))
        except Exception:
            continue
        if len(fields) < 5:
            continue

        # Ticker: scan first 8 fields (tolerates account-info prefix columns)
        ticker: str | None = None
        is_mm = False
        for fld in fields[:8]:
            candidate = fld.strip().upper().strip("*").strip()
            if candidate in _mm_set:
                ticker = candidate
                is_mm = True
                break
            if _looks_like_ticker(candidate) and candidate not in _FIDELITY_SKIP:
                ticker = candidate
                break
        if not ticker:
            continue

        # Skip non-position pseudo-rows
        if ticker in {"TOTAL", "TOTALS", "SUBTOTAL", "ACCOUNTTOTAL", "PENDINGACTIVITY"}:
            continue

        # Quantity
        quantity: float | None = None
        if qty_idx is not None and qty_idx < len(fields):
            try:
                q_str = fields[qty_idx].strip().replace(",", "")
                if q_str:
                    quantity = float(q_str)
            except (ValueError, TypeError):
                pass

        # Last Price
        price: float | None = None
        if price_idx is not None and price_idx < len(fields):
            try:
                p_str = fields[price_idx].strip().replace("$", "").replace(",", "")
                if p_str and p_str not in ("--", "-", "N/A", "n/a"):
                    price = float(p_str)
            except (ValueError, TypeError):
                pass
        # MM funds maintain a stable $1.00 NAV — use as fallback if price missing
        if is_mm and (price is None or price <= 0):
            price = 1.00

        # Percent Of Account (3rd "%" field — existing reliable approach)
        weight: float | None = None
        pct_count = 0
        for fld in fields:
            fld_s = fld.strip()
            if fld_s.endswith("%"):
                pct_count += 1
                if pct_count == 3:
                    try:
                        val = float(fld_s.replace("%", "").strip())
                        if 0.0 <= val <= 100.0:
                            weight = round(val / 100.0, 8)
                    except (ValueError, TypeError):
                        pass
                    break

        # Current Value — try qty × price first, then read column directly
        current_value: float | None = None
        if quantity is not None and price is not None and quantity > 0 and price > 0:
            current_value = round(quantity * price, 2)
        elif curval_idx is not None and curval_idx < len(fields):
            # Direct "Current Value" column fallback (Fidelity always populates this)
            try:
                cv_str = fields[curval_idx].strip().replace("$", "").replace(",", "")
                if cv_str and cv_str not in ("--", "-", "N/A"):
                    cv_parsed = float(cv_str)
                    if cv_parsed > 0:
                        current_value = round(cv_parsed, 2)
                        # Back-fill price from current value ÷ quantity if we have qty
                        if quantity and quantity > 0 and price is None:
                            price = round(current_value / quantity, 4)
            except (ValueError, TypeError):
                pass
        if current_value is not None and current_value > 0:
            total_value += current_value

        holdings.append({
            "ticker":        ticker,
            "quantity":      round(quantity, 4) if quantity is not None else None,
            "price":         round(price, 4)    if price    is not None else None,
            "current_value": current_value,
            "weight":        weight,
            "is_mm":         is_mm,  # True = money-market / cash position
        })

    return {
        "ok":               True,
        "holdings":         holdings,
        "total_value":      round(total_value, 2),
        "columns_detected": columns_detected,
    }


def parse_activity_for_attribution(raw_text: str) -> dict:
    """Parse Fidelity activity CSV to extract every trade and dividend event.

    Returns:
        {
          "ok": True,
          "trades":    [{"date", "ticker", "type" (BUY|SELL), "shares", "price", "amount"}],
          "dividends": [{"date", "ticker", "amount"}],
        }

    Action parsing:
      "YOU SOLD INTEL CORP COM USD0.001 (INTC) (Cash)"
         → SELL, ticker from Symbol column or (INTC) parenthetical
      "YOU BOUGHT APPLE INC (AAPL) (Cash)"
         → BUY
      "DIVIDEND RECEIVED"
         → dividend using Symbol column
      "REINVESTMENT" with a ticker
         → treated as BUY (reinvested dividend buys more shares)
    """
    import csv as _csv_m2
    import re as _re2
    from io import StringIO as _StringIO2

    lines = raw_text.splitlines()

    header_idx = None
    for i, line in enumerate(lines[:30]):
        low = line.strip().lower()
        if "run date" in low or ("date" in low and "action" in low):
            header_idx = i
            break
    if header_idx is None:
        return {"ok": False, "error": "Could not find activity CSV header row."}

    data_lines = lines[header_idx:]
    try:
        reader = _csv_m2.DictReader(_StringIO2("\n".join(data_lines)))
    except Exception as exc:
        return {"ok": False, "error": f"CSV parse error: {exc}"}

    trades: list[dict]    = []
    dividends: list[dict] = []
    year = datetime.now().year
    year_start = f"{year}-01-01"

    for row in reader:
        if not row:
            continue
        row = {k.strip(): (v or "").strip() for k, v in row.items() if k and k.strip()}
        if not row:
            continue

        raw_date = row.get("Run Date") or row.get("Date") or ""
        try:
            dt = datetime.strptime(raw_date.strip(), "%m/%d/%Y")
            date_str = dt.strftime("%Y-%m-%d")
        except ValueError:
            continue
        if date_str < year_start:
            continue

        action = (row.get("Action") or "").strip().upper()
        if not action:
            continue

        # ── Ticker ──────────────────────────────────────────────────────────
        symbol = (row.get("Symbol") or "").strip().upper()
        if not symbol or not _looks_like_ticker(symbol):
            # Try "(TICKER)" in the action string
            m = _re2.search(r'\(([A-Z]{1,5})\)', action)
            symbol = m.group(1) if m else ""
        if not symbol:
            continue

        # ── Quantity ────────────────────────────────────────────────────────
        raw_qty = (row.get("Quantity") or "").replace(",", "").strip()
        try:
            quantity = abs(float(raw_qty))
        except ValueError:
            quantity = 0.0

        # ── Price ───────────────────────────────────────────────────────────
        raw_price = (row.get("Price ($)") or row.get("Price") or "").replace("$", "").replace(",", "").strip()
        try:
            price = abs(float(raw_price))
        except ValueError:
            price = 0.0

        # ── Amount ──────────────────────────────────────────────────────────
        raw_amt = (row.get("Amount ($)") or row.get("Amount") or "").replace("$", "").replace(",", "").strip()
        if raw_amt.startswith("(") and raw_amt.endswith(")"):
            raw_amt = "-" + raw_amt[1:-1]
        try:
            amount = float(raw_amt)
        except ValueError:
            amount = 0.0

        # ── Classify ────────────────────────────────────────────────────────
        if "YOU SOLD" in action:
            if quantity > 0 and price > 0:
                trades.append({
                    "date":   date_str,
                    "ticker": symbol,
                    "type":   "SELL",
                    "shares": round(quantity, 4),
                    "price":  round(price, 4),
                    "amount": round(abs(amount), 2),  # proceeds received
                })
        elif "YOU BOUGHT" in action:
            if quantity > 0 and price > 0:
                trades.append({
                    "date":   date_str,
                    "ticker": symbol,
                    "type":   "BUY",
                    "shares": round(quantity, 4),
                    "price":  round(price, 4),
                    "amount": round(abs(amount), 2),  # cost paid
                })
        elif "REINVESTMENT" in action:
            # Reinvested dividend buys more fractional shares
            if quantity > 0 and price > 0:
                trades.append({
                    "date":   date_str,
                    "ticker": symbol,
                    "type":   "REINVESTMENT",
                    "shares": round(quantity, 6),
                    "price":  round(price, 4),
                    "amount": round(abs(amount), 2),
                })
        elif "DIVIDEND RECEIVED" in action:
            if amount != 0:
                dividends.append({
                    "date":   date_str,
                    "ticker": symbol,
                    "amount": round(abs(amount), 2),
                })
        elif "INTEREST EARNED" in action or action in ("INTEREST",):
            # Money-market (SPAXX etc.) distributions appear as interest in Fidelity activity.
            # Treat them as income for the ticker (same as a cash dividend).
            if amount != 0 and symbol:
                dividends.append({
                    "date":   date_str,
                    "ticker": symbol,
                    "amount": round(abs(amount), 2),
                })

    return {
        "ok":        True,
        "trades":    sorted(trades,    key=lambda t: t["date"]),
        "dividends": sorted(dividends, key=lambda d: d["date"]),
    }


def compute_position_attribution(
    positions_text: str,
    activity_text: str,
    begin_value: float,
) -> dict:
    """Compute accurate per-stock YTD performance attribution.

    Algorithm per ticker:
        start_shares  = end_shares + Σ(sold_shares) − Σ(bought_shares) − Σ(reinvested_shares)
        start_value   = start_shares × jan1_price
        end_value     = end_shares × current_price   (from positions CSV)
        dollar_gain   = end_value
                      + Σ(sell_proceeds)
                      + dividends_received
                      − start_value
                      − Σ(buy_costs)
                      − Σ(reinvestment_costs)
        contribution% = dollar_gain / begin_value × 100

    This correctly handles:
      • Partial sales: sold shares captured gain at the sale price
      • New purchases: cost basis subtracted, current value included
      • Dividend reinvestments: cash used treated as a buy
      • Cash dividends: credited to that ticker's P&L
      • Positions fully liquidated: end_shares = 0, only sale proceeds count

    Parameters:
        positions_text: Raw text of Fidelity positions CSV (same file uploaded
                        for the live benchmark — provides current shares + prices)
        activity_text:  Raw text of Fidelity activity/history CSV (provides all
                        trades and dividends for the current calendar year)
        begin_value:    Total portfolio value on Jan 1 (for contribution normalisation)
    """
    if begin_value <= 0:
        return {"ok": False, "error": "begin_value must be a positive dollar amount."}

    # ── Step 1: Parse current positions ──────────────────────────────────────
    pos_result = _parse_fidelity_positions_extended(positions_text)
    if not pos_result.get("ok"):
        return pos_result

    holdings      = pos_result["holdings"]
    portfolio_end_value = pos_result["total_value"]  # sum of qty × price

    # Build lookup: ticker → {quantity, price, current_value, is_mm}
    # MM tickers are included so their current_value flows into portfolio_end_value,
    # but they are handled separately in attribution (interest-only, no shares reconstruction).
    current_state: dict[str, dict] = {}
    for h in holdings:
        t = h["ticker"]
        current_state[t] = {
            "end_shares":    h["quantity"]      or 0.0,
            "end_price":     h["price"]         or 0.0,
            "current_value": h["current_value"] or 0.0,
            "is_mm":         h.get("is_mm", False),
        }

    # ── Step 2: Parse activity ────────────────────────────────────────────────
    act_result = parse_activity_for_attribution(activity_text)
    if not act_result.get("ok"):
        return act_result

    trades    = act_result["trades"]
    dividends = act_result["dividends"]

    # Index trades and dividends by ticker
    trades_by_ticker:    dict[str, list] = {}
    dividends_by_ticker: dict[str, float] = {}

    for t in trades:
        trades_by_ticker.setdefault(t["ticker"], []).append(t)
    for d in dividends:
        dividends_by_ticker[d["ticker"]] = dividends_by_ticker.get(d["ticker"], 0.0) + d["amount"]

    # MM tickers get interest-only treatment after the main loop — exclude from
    # shares-reconstruction loop to avoid garbage numbers from cash sweeps.
    mm_tickers: set[str] = {t for t, s in current_state.items() if s.get("is_mm")}

    # All non-MM tickers (positions + sold-out positions from activity)
    all_tickers: set[str] = (
        (set(current_state.keys()) | set(trades_by_ticker.keys())) - mm_tickers
    )

    # ── Step 3: Fetch Jan 1 prices in bulk ───────────────────────────────────
    year         = datetime.now().year
    period_start = datetime(year, 1, 1)
    price_start  = (period_start - timedelta(days=14)).strftime("%Y-%m-%d")
    price_end    = (period_start + timedelta(days=5)).strftime("%Y-%m-%d")

    jan1_prices: dict[str, float] = {}
    try:
        import yfinance as _yf
        tickers_list = list(all_tickers)
        if tickers_list:
            hist = _yf.download(
                tickers_list,
                start=price_start,
                end=price_end,
                auto_adjust=False,
                progress=False,
            )
            close_data = hist.get("Close") if isinstance(hist, dict) else (
                hist["Close"] if "Close" in hist.columns.get_level_values(0) else hist
            )
            for ticker in tickers_list:
                try:
                    if hasattr(close_data, "columns"):
                        series = close_data[ticker].dropna()
                    else:
                        series = close_data.dropna()
                    if not series.empty:
                        jan1_prices[ticker] = float(series.iloc[-1])
                except Exception:
                    pass
    except Exception:
        pass  # fall back to per-ticker below

    # Per-ticker fallback for any that failed the bulk download.
    # Also tries preferred-stock ticker remapping:
    #   Fidelity: NLYPRF  →  yfinance: NLY-PF
    #   Fidelity: BACPRL  →  yfinance: BAC-PL
    #   Fidelity: CPRJ    →  yfinance: C-PJ
    # Pattern: {BASE}PR{SERIES} where BASE is 1-4 uppercase letters,
    # SERIES is a single letter → {BASE}-P{SERIES}
    import re as _re_pref

    def _yf_ticker_variants(t: str) -> list:
        """Return yfinance ticker variants to try, preferred remap first."""
        variants = [t]
        m = _re_pref.match(r'^([A-Z]{1,4})PR([A-Z])$', t)
        if m:
            base, series = m.group(1), m.group(2)
            variants.insert(0, f"{base}-P{series}")   # NLY-PF — most common
            variants.insert(1, f"{base}^{series}")    # NLY^F — alternate
        return variants

    for ticker in all_tickers:
        if ticker not in jan1_prices:
            for yf_sym in _yf_ticker_variants(ticker):
                try:
                    import yfinance as _yf2
                    h = _yf2.Ticker(yf_sym).history(
                        start=price_start, end=price_end, auto_adjust=False
                    )
                    if h is not None and not h.empty:
                        price_val = float(h["Close"].dropna().iloc[-1])
                        if price_val > 0:
                            jan1_prices[ticker] = price_val  # store under original ticker
                            break  # found it — stop trying variants
                except Exception:
                    pass

    # ── Step 4: Compute attribution per ticker ───────────────────────────────
    attribution: list[dict] = []
    total_explained = 0.0

    for ticker in all_tickers:
        state  = current_state.get(ticker, {"end_shares": 0.0, "end_price": 0.0, "current_value": 0.0})
        end_shares    = state["end_shares"]
        end_price     = state["end_price"]
        end_value     = state["current_value"] or (end_shares * end_price)
        ticker_trades = trades_by_ticker.get(ticker, [])

        # Aggregate sells, buys, reinvestments
        sells        = [t for t in ticker_trades if t["type"] == "SELL"]
        buys         = [t for t in ticker_trades if t["type"] == "BUY"]
        reinvestments= [t for t in ticker_trades if t["type"] == "REINVESTMENT"]

        total_sold_shares   = sum(t["shares"] for t in sells)
        total_sell_proceeds = sum(t["amount"] for t in sells)
        total_bought_shares = sum(t["shares"] for t in buys)
        total_buy_cost      = sum(t["amount"] for t in buys)
        total_reinv_shares  = sum(t["shares"] for t in reinvestments)
        total_reinv_cost    = sum(t["amount"] for t in reinvestments)
        dividends_cash      = dividends_by_ticker.get(ticker, 0.0)

        # Reconstruct starting shares
        start_shares = (
            end_shares
            + total_sold_shares
            - total_bought_shares
            - total_reinv_shares
        )
        start_shares = max(0.0, start_shares)  # guard against reconstruction errors

        # Starting value: use Jan 1 price (fallback to end price if unavailable)
        price_fetched = ticker in jan1_prices
        jan1_price    = jan1_prices.get(ticker, end_price)
        start_value   = start_shares * jan1_price

        # ── Unreliable-price detection ──────────────────────────────────────
        # A fully-sold position (end_price = 0) where yfinance also returned
        # nothing will fall back to end_price = $0, making start_value = $0
        # and dollar_gain = full proceeds (completely wrong).
        # Mark such rows so we can warn the user and exclude from totals.
        price_missing = (
            not price_fetched          # yfinance returned nothing
            and start_shares > 0       # position existed at Jan 1
            and jan1_price <= 0        # fallback landed on $0
        )

        if price_missing:
            # We don't know the opening basis — set gain to dividends only
            # (at least dividends are real cash received, capital P&L unknown)
            dollar_gain      = dividends_cash
            contribution_pct = round(dollar_gain / begin_value * 100.0, 4) if begin_value > 0 else 0.0
            ticker_return_pct = None
        else:
            # Total dollar P&L for this ticker this year
            dollar_gain = (
                end_value
                + total_sell_proceeds
                + dividends_cash
                - start_value
                - total_buy_cost
                - total_reinv_cost
            )
            contribution_pct = round(dollar_gain / begin_value * 100.0, 4) if begin_value > 0 else 0.0

            # Ticker-level return (for reference, not used in portfolio attribution)
            if start_value + total_buy_cost + total_reinv_cost > 0:
                invested = start_value + total_buy_cost + total_reinv_cost
                ticker_return_pct = round(
                    (end_value + total_sell_proceeds + dividends_cash - invested) / invested * 100.0, 2
                )
            else:
                ticker_return_pct = None

        total_explained += dollar_gain

        attribution.append({
            "ticker":              ticker,
            "start_shares":        round(start_shares, 4),
            "jan1_price":          round(jan1_price, 4),
            "start_value":         round(start_value, 2),
            "total_sold_shares":   round(total_sold_shares, 4),
            "total_sell_proceeds": round(total_sell_proceeds, 2),
            "total_bought_shares": round(total_bought_shares, 4),
            "total_buy_cost":      round(total_buy_cost, 2),
            "reinvestment_shares": round(total_reinv_shares, 6),
            "reinvestment_cost":   round(total_reinv_cost, 2),
            "dividends_cash":      round(dividends_cash, 2),
            "end_shares":          round(end_shares, 4),
            "end_price":           round(end_price, 4),
            "end_value":           round(end_value, 2),
            "dollar_gain":         round(dollar_gain, 2),
            "contribution_pct":    contribution_pct,
            "ticker_return_pct":   ticker_return_pct,
            "trade_count":         len(ticker_trades),
            "trades":              ticker_trades,
            "jan1_price_source":   "fetched" if price_fetched else "fallback",
            "price_missing":       price_missing,   # True = Jan 1 price unavailable, gain unreliable
        })

    # ── MM tickers: interest/dividends-only attribution ─────────────────────
    # Cash-sweep activity makes shares-reconstruction unreliable for MM funds.
    # Their only real contribution is the interest they earn — $1.00 stable NAV
    # means there is no capital gain to attribute.
    for ticker in sorted(mm_tickers):
        mm_state        = current_state.get(ticker, {})
        end_shares      = mm_state.get("end_shares", 0.0)
        end_price       = mm_state.get("end_price") or 1.0
        end_value       = mm_state.get("current_value") or (end_shares * end_price)
        dividends_cash  = dividends_by_ticker.get(ticker, 0.0)
        dollar_gain     = dividends_cash   # interest earned = only real contribution
        contribution_pct = round(dollar_gain / begin_value * 100.0, 4) if begin_value > 0 else 0.0
        total_explained += dollar_gain
        ticker_return_pct = round(dividends_cash / end_value * 100.0, 2) if end_value > 0 else None

        attribution.append({
            "ticker":              ticker,
            "start_shares":        None,   # not meaningful for MM
            "jan1_price":          1.0,
            "start_value":         None,
            "total_sold_shares":   0.0,
            "total_sell_proceeds": 0.0,
            "total_bought_shares": 0.0,
            "total_buy_cost":      0.0,
            "reinvestment_shares": 0.0,
            "reinvestment_cost":   0.0,
            "dividends_cash":      round(dividends_cash, 2),
            "end_shares":          round(end_shares, 4),
            "end_price":           round(end_price, 4),
            "end_value":           round(end_value, 2),
            "dollar_gain":         round(dollar_gain, 2),
            "contribution_pct":    contribution_pct,
            "ticker_return_pct":   ticker_return_pct,
            "trade_count":         0,
            "trades":              [],
            "jan1_price_source":   "mm_stable_nav",
            "is_mm":               True,
        })

    # Sort: biggest absolute contributors first
    attribution.sort(key=lambda a: abs(a["dollar_gain"]), reverse=True)

    explained_pct = round(total_explained / begin_value * 100.0, 4) if begin_value > 0 else 0.0

    return {
        "ok":                  True,
        "attribution":         attribution,
        "total_dollar_gain":   round(total_explained, 2),
        "explained_pct":       explained_pct,
        "begin_value":         round(begin_value, 2),
        "portfolio_end_value": round(portfolio_end_value, 2),
        "positions_parsed":    len(holdings),  # includes MM/cash positions
        "trades_parsed":       len(trades),
        "dividends_parsed":    len(dividends),
        "jan1_prices":         {k: round(v, 4) for k, v in jan1_prices.items()},
        "columns_positions":   pos_result.get("columns_detected", []),
    }


def parse_fidelity_monthly_perf(raw_text: str) -> dict:
    """Parse a Fidelity monthly performance / account-summary CSV.

    Expected layout (column names are matched loosely):
      Month | Beginning Balance | Market Change | Dividends | Interest |
      Deposits | Withdrawals | Advisory Fees | Ending Balance

    Column matching uses keyword substring search so it works across Fidelity's
    various export formats.

    Returns:
      {
        "ok": True,
        "months": [
          {
            "month":          int,    # 1–12
            "label":          str,    # "Jan", "Feb", …
            "start":          float,
            "market_change":  float,
            "dividends":      float,
            "interest":       float,
            "deposits":       float,  # positive
            "withdrawals":    float,  # negative
            "advisory_fees":  float,  # negative (fee charged)
            "ending":         float,
            "net_flow":       float,  # deposits + withdrawals (for TWRR)
            "hpr":            float,  # HPR_m = ending / (start + net_flow)
          }, ...
        ],
        "year": int,
      }
    """
    import re as _re_mp
    import csv as _csv_mp
    from io import StringIO as _SIO

    # ── Flexible column-keyword mapping ──────────────────────────────────────
    _COL_MAP = {
        "month_label":    ("month", "period", "date"),
        "start":          ("beginning", "start", "opening"),
        "market_change":  ("market", "appreciation", "depreciation",
                           "investment gain", "gain/loss", "change"),
        "dividends":      ("dividend",),
        "interest":       ("interest",),
        "deposits":       ("deposit", "contribution", "receipt",
                          "transfer in", "inflow"),
        "withdrawals":    ("withdrawal", "distribution", "disbursement",
                         "transfer out", "outflow"),
        "advisory_fees":  ("fee", "advisory"),
        "ending":         ("ending", "end value", "closing", "end bal"),
    }

    def _match_col(header_lower: str) -> str | None:
        for canonical, keywords in _COL_MAP.items():
            if any(kw in header_lower for kw in keywords):
                return canonical
        return None

    _MONTH_NAMES = {
        "jan": 1, "feb": 2, "mar": 3, "apr": 4, "may": 5, "jun": 6,
        "jul": 7, "aug": 8, "sep": 9, "oct": 10, "nov": 11, "dec": 12,
    }

    def _parse_month_label(s: str) -> int | None:
        s = s.strip().lower()
        for abbr, num in _MONTH_NAMES.items():
            if s.startswith(abbr):
                return num
        # numeric: 01/2026 or 1/2026 or 2026-01
        m = _re_mp.search(r"(?:^|[-/])(\d{1,2})(?:[-/]\d{4})?$", s)
        if m:
            n = int(m.group(1))
            if 1 <= n <= 12:
                return n
        return None

    def _parse_dollar(s: str) -> float:
        s = str(s or "").strip().replace("$", "").replace(",", "").replace(" ", "")
        if s.startswith("(") and s.endswith(")"):
            s = "-" + s[1:-1]
        try:
            return float(s)
        except ValueError:
            return 0.0

    # ── Find header row ───────────────────────────────────────────────────────
    lines = raw_text.splitlines()
    header_idx = -1
    for i, line in enumerate(lines):
        lower = line.lower()
        if ("beginning" in lower or "starting" in lower or "opening" in lower) and \
           ("end" in lower or "closing" in lower):
            header_idx = i
            break
        if "month" in lower and ("balance" in lower or "value" in lower):
            header_idx = i
            break

    if header_idx < 0:
        return {"ok": False, "error": "Could not find monthly performance header row. "
                "Expected columns: Month, Beginning Balance, …, Ending Balance."}

    data_lines = "\n".join(lines[header_idx:])
    reader = _csv_mp.DictReader(_SIO(data_lines))

    # Map CSV column names to canonical names
    col_map: dict[str, str] = {}
    if reader.fieldnames:
        for raw_col in reader.fieldnames:
            canonical = _match_col(raw_col.strip().lower()) if raw_col else None
            if canonical:
                col_map[raw_col] = canonical

    year = datetime.now().year
    months_out: list[dict] = []

    for row in reader:
        # Flatten row → canonical
        r: dict = {}
        for raw_col, canonical in col_map.items():
            r[canonical] = (row.get(raw_col) or "").strip()

        month_num = _parse_month_label(r.get("month_label", ""))
        if month_num is None:
            continue  # skip non-data rows

        start         = _parse_dollar(r.get("start", ""))
        market_change = _parse_dollar(r.get("market_change", ""))
        dividends     = _parse_dollar(r.get("dividends", ""))
        interest      = _parse_dollar(r.get("interest", ""))
        deposits_raw  = _parse_dollar(r.get("deposits", ""))
        withdraw_raw  = _parse_dollar(r.get("withdrawals", ""))
        advisory_fees = _parse_dollar(r.get("advisory_fees", ""))
        ending        = _parse_dollar(r.get("ending", ""))

        # Normalise signs: deposits = positive, withdrawals = negative
        deposits    = abs(deposits_raw)  if deposits_raw  != 0 else 0.0
        withdrawals = -abs(withdraw_raw) if withdraw_raw  != 0 else 0.0
        advisory    = -abs(advisory_fees) if advisory_fees != 0 else 0.0

        net_flow = deposits + withdrawals   # advisory fees excluded from external flows
        denom    = start + net_flow
        hpr      = (ending / denom - 1.0) if denom > 0 else 0.0

        months_out.append({
            "month":          month_num,
            "label":          datetime(year, month_num, 1).strftime("%b"),
            "start":          round(start, 2),
            "market_change":  round(market_change, 2),
            "dividends":      round(dividends, 2),
            "interest":       round(interest, 2),
            "deposits":       round(deposits, 2),
            "withdrawals":    round(withdrawals, 2),
            "advisory_fees":  round(advisory, 2),
            "ending":         round(ending, 2),
            "net_flow":       round(net_flow, 2),
            "hpr":            round(hpr, 6),
        })

    if not months_out:
        return {"ok": False, "error": "No monthly data rows parsed from performance CSV."}

    months_out.sort(key=lambda x: x["month"])
    return {"ok": True, "months": months_out, "year": year}


def compute_monthly_ytd_chart(
    positions_text: str,
    activity_text: str,
    begin_value: float,
    monthly_perf: dict | None = None,
) -> dict:
    """Compute month-by-month portfolio performance for the YTD chart.

    Returns monthly portfolio values, returns, and top movers per month to
    power the interactive hover chart on the Tracker page.

    Algorithm:
      • Reconstructs share counts at the start of each month by working
        backward from current positions, undoing each month's trades.
      • Fetches end-of-month closing prices via yfinance (interval=1mo).
      • Computes portfolio value each month-end = Σ(shares × price) + cash.
      • Current month is anchored to the actual positions-CSV total (authoritative).
      • Monthly return = (end − start) / start.
      • Top movers per month = per-ticker dollar gain ranked by abs value.
    """
    import yfinance as _yf_m
    from collections import defaultdict

    year         = datetime.now().year
    today        = datetime.now()
    curr_month   = today.month

    # ── Parse positions ───────────────────────────────────────────────────────
    pos_result = _parse_fidelity_positions_extended(positions_text)
    if not pos_result.get("ok"):
        return {"ok": False, "error": "Positions parse failed for monthly chart."}
    holdings   = pos_result["holdings"]
    csv_total  = pos_result["total_value"]   # authoritative end-value

    # ── Parse activity ────────────────────────────────────────────────────────
    act_result = parse_activity_for_attribution(activity_text)
    if not act_result.get("ok"):
        return {"ok": False, "error": "Activity parse failed for monthly chart."}
    trades    = act_result["trades"]
    dividends = act_result["dividends"]

    def _mth(d: str) -> int:
        return int(d[5:7])

    trades_by_month: dict    = defaultdict(list)
    divs_by_month: dict      = defaultdict(list)
    for t in trades:
        trades_by_month[_mth(t["date"])].append(t)
    for d in dividends:
        divs_by_month[_mth(d["date"])].append(d)

    # ── Known MM tickers (excluded from equity reconstruction) ────────────────
    _chart_mm = {"SPAXX","FZFXX","FZSXX","FDRXX","FZDXX",
                 "VMFXX","VMRXX","SWVXX","FCASH","CASH"}

    # ── Current equity share counts ───────────────────────────────────────────
    equity_map: dict = {}                        # ticker → current shares
    for h in holdings:
        if not h.get("is_mm"):
            equity_map[h["ticker"]] = h.get("quantity") or 0.0
    for t in trades:                             # also add fully-sold tickers
        if t["ticker"] not in equity_map and t["ticker"] not in _chart_mm:
            equity_map[t["ticker"]] = 0.0
    all_eq_tickers = set(equity_map.keys()) - _chart_mm

    # ── Reconstruct shares at start of each month (backward from today) ───────
    shares_month_start: dict = {}               # month → {ticker: shares}
    running = dict(equity_map)
    for m in range(curr_month, 0, -1):
        pre: dict = dict(running)
        for t in trades_by_month.get(m, []):
            tk = t["ticker"]
            if t["type"] == "SELL":
                pre[tk] = pre.get(tk, 0.0) + t["shares"]
            elif t["type"] in ("BUY", "REINVESTMENT"):
                pre[tk] = max(0.0, pre.get(tk, 0.0) - t["shares"])
        shares_month_start[m] = pre
        running = pre

    # ── Monthly end-of-month share counts ─────────────────────────────────────
    def _shares_end(m: int) -> dict:
        s = dict(shares_month_start.get(m, {}))
        for t in trades_by_month.get(m, []):
            tk = t["ticker"]
            if t["type"] == "SELL":
                s[tk] = max(0.0, s.get(tk, 0.0) - t["shares"])
            elif t["type"] in ("BUY", "REINVESTMENT"):
                s[tk] = s.get(tk, 0.0) + t["shares"]
        return s

    # ── Fetch monthly prices (Dec prev-year for Jan baseline + all YTD months) ─
    _pref_re = __import__("re").compile(r"^([A-Z]{1,4})PR([A-Z])$")
    def _yf_variants(tk: str) -> list:
        v = [tk]
        pm = _pref_re.match(tk)
        if pm:
            v.insert(0, f"{pm.group(1)}-P{pm.group(2)}")
        return v

    # month_px[ticker][key] where key=0 → Dec prev-year close, key=1..12 → Jan..Dec close
    month_px: dict = defaultdict(dict)
    p_start = f"{year - 1}-12-01"
    p_end   = (today + timedelta(days=10)).strftime("%Y-%m-%d")

    tklist = list(all_eq_tickers) + ["SPY"]
    if tklist:
        try:
            hist = _yf_m.download(
                tklist, start=p_start, end=p_end,
                interval="1mo", auto_adjust=False, progress=False,
            )
            if not hist.empty:
                try:
                    close = hist["Close"]
                except Exception:
                    close = hist
                for tk in tklist:
                    try:
                        if hasattr(close, "columns") and tk in close.columns:
                            s = close[tk].dropna()
                        else:
                            s = close.dropna()
                        for idx, val in s.items():
                            if hasattr(idx, "month") and val == val:  # not NaN
                                if idx.year == year - 1 and idx.month == 12:
                                    month_px[tk][0] = float(val)
                                elif idx.year == year:
                                    month_px[tk][idx.month] = float(val)
                    except Exception:
                        pass
        except Exception:
            pass

    # Per-ticker fallback: try yfinance variants for tickers with missing prices
    for tk in list(all_eq_tickers):
        if not month_px.get(tk):
            for sym in _yf_variants(tk):
                if sym == tk:
                    continue
                try:
                    h2 = _yf_m.download(
                        [sym], start=p_start, end=p_end,
                        interval="1mo", auto_adjust=False, progress=False,
                    )
                    if not h2.empty:
                        try:
                            s = h2["Close"][sym].dropna() if sym in h2["Close"].columns else h2["Close"].dropna()
                        except Exception:
                            s = h2.dropna()
                        for idx, val in s.items():
                            if hasattr(idx, "month") and val == val:
                                key = 0 if (idx.year == year - 1 and idx.month == 12) else (idx.month if idx.year == year else None)
                                if key is not None:
                                    month_px[tk][key] = float(val)
                        if month_px.get(tk):
                            break
                except Exception:
                    pass

    # Current prices as fallback for months with no data
    current_px: dict = {h["ticker"]: h.get("price") or 0.0
                        for h in holdings if not h.get("is_mm")}

    def _price(tk: str, key: int) -> float:
        p = month_px.get(tk, {}).get(key)
        if p and p > 0:
            return p
        # fallback: use current price (approximation)
        return current_px.get(tk, 0.0)

    # ── Cash balance (MM positions — held constant as approximation) ──────────
    cash_bal = sum(h.get("current_value") or 0.0 for h in holdings if h.get("is_mm"))

    # ── Portfolio value at end of each month ──────────────────────────────────
    # Precedence (highest → lowest accuracy):
    #   1. monthly_perf ending balance (exact Fidelity value)
    #   2. current month → CSV total (authoritative)
    #   3. yfinance estimated price × reconstructed shares + cash
    perf_by_month: dict = {}
    if monthly_perf and monthly_perf.get("months"):
        perf_by_month = {p["month"]: p for p in monthly_perf["months"]}

    port_val: dict = {0: float(begin_value)}     # key=0 → Jan 1
    # If monthly_perf provides a Jan start that differs from begin_value, trust it
    if 1 in perf_by_month and perf_by_month[1].get("start", 0) > 0:
        port_val[0] = perf_by_month[1]["start"]

    for m in range(1, curr_month + 1):
        if m == curr_month:
            port_val[m] = float(csv_total)       # authoritative: today's CSV
        elif m in perf_by_month:
            port_val[m] = perf_by_month[m]["ending"]   # exact Fidelity value
        else:
            se = _shares_end(m)
            eq = sum(sh * _price(tk, m) for tk, sh in se.items() if sh > 0)
            port_val[m] = eq + cash_bal

    # ── SPY monthly performance ───────────────────────────────────────────────
    spy_start = month_px["SPY"].get(0) or month_px["SPY"].get(1)   # Dec-prev or Jan
    spy_monthly: dict = {}
    if spy_start:
        for m in range(1, curr_month + 1):
            sp = month_px["SPY"].get(m)
            if sp and spy_start:
                spy_monthly[m] = round((sp / spy_start - 1) * 100.0, 3)

    # ── Build monthly data ────────────────────────────────────────────────────
    months_out = []
    for m in range(1, curr_month + 1):
        sv = port_val[m - 1]
        ev = port_val[m]
        if sv <= 0:
            continue

        dg = ev - sv
        perf = perf_by_month.get(m)
        if perf:
            # Use Fidelity's exact HPR — eliminates cash-flow distortion from bars
            rp = round(perf["hpr"] * 100.0, 4)
        else:
            rp = dg / sv * 100.0

        # Per-ticker movers for this month
        movers = []
        active = set(all_eq_tickers)
        for t in trades_by_month.get(m, []):
            active.add(t["ticker"])

        for tk in active:
            if tk in _chart_mm:
                continue
            shs  = shares_month_start.get(m, {}).get(tk, 0.0)
            she  = _shares_end(m).get(tk, 0.0)
            ps   = _price(tk, m - 1 if m > 1 else 0)
            pe   = _price(tk, m)
            if not ps or not pe:
                continue

            tk_t = [t for t in trades_by_month.get(m, []) if t["ticker"] == tk]
            sell_proc = sum(t["amount"] for t in tk_t if t["type"] == "SELL")
            buy_cost  = sum(t["amount"] for t in tk_t if t["type"] in ("BUY","REINVESTMENT"))
            tk_div    = sum(d["amount"] for d in divs_by_month.get(m, []) if d["ticker"] == tk)

            gain = she * pe + sell_proc + tk_div - shs * ps - buy_cost
            if abs(gain) < 1.0:
                continue
            ret = round((pe / ps - 1) * 100.0, 2) if ps > 0 else None
            movers.append({
                "ticker":          tk,
                "dollar_gain":     round(gain, 2),
                "contribution_pct": round(gain / sv * 100.0, 3),
                "return_pct":      ret,
            })

        movers.sort(key=lambda x: abs(x["dollar_gain"]), reverse=True)

        months_out.append({
            "month":       m,
            "label":       datetime(year, m, 1).strftime("%b"),
            "start_value": round(sv, 2),
            "end_value":   round(ev, 2),
            "dollar_gain": round(dg, 2),
            "return_pct":  round(rp, 4),
            "spy_ytd_pct": spy_monthly.get(m),
            "exact":       perf is not None,   # True = values from Fidelity CSV, not estimated
            "perf_detail": {                   # from monthly performance CSV when available
                "market_change":  perf["market_change"]  if perf else None,
                "dividends":      perf["dividends"]      if perf else None,
                "interest":       perf["interest"]       if perf else None,
                "deposits":       perf["deposits"]       if perf else None,
                "withdrawals":    perf["withdrawals"]    if perf else None,
                "advisory_fees":  perf["advisory_fees"]  if perf else None,
                "net_flow":       perf["net_flow"]       if perf else None,
            } if perf else None,
            "movers":      movers[:8],
            "dividends":   [{"ticker": d["ticker"], "amount": d["amount"]}
                            for d in divs_by_month.get(m, [])],
            "trades":      [{"ticker": t["ticker"], "type": t["type"],
                             "shares": t["shares"], "price": t["price"]}
                            for t in trades_by_month.get(m, [])],
        })

    return {
        "ok":              True,
        "monthly":         months_out,
        "begin_value":     round(float(port_val[0]), 2),
        "end_value":       round(float(csv_total), 2),
        "year":            year,
        "has_exact_perf":  len(perf_by_month) > 0,
    }


# ── Account history / Modified Dietz YTD return ──────────────────────────────

# ── Blocklist: investment-activity action strings that are NEVER external flows ──
# Everything NOT in this list (and not zero-amount) is treated as a potential
# external cash flow (deposit or withdrawal).  This captures transfers in/out,
# credit-card payments, ACH, wire, journal entries, etc. without needing an
# explicit allowlist of every possible Fidelity action string.
_FIDELITY_INTERNAL_ACTIONS = frozenset({
    # ── Trades ──────────────────────────────────────────────────────────────
    "YOU BOUGHT",
    "YOU SOLD",
    # ── Investment income ────────────────────────────────────────────────────
    "DIVIDEND RECEIVED",
    "REINVESTMENT",
    "INTEREST EARNED",
    "INTEREST",                      # money-market / bond interest (shorter variant)
    "LONG-TERM CAP GAIN",
    "SHORT-TERM CAP GAIN",
    "RETURN OF CAPITAL",
    "CAPITAL GAIN DISTRIBUTION",
    # ── Fees & taxes ────────────────────────────────────────────────────────
    "FEE CHARGED",                   # ADR fees, account fees
    "FOREIGN TAX PAID",              # ADR withholding tax
    # ── Corporate actions ───────────────────────────────────────────────────
    "MANDATORY REORGANIZATION",
    "STOCK SPLIT",
    "CONVERSION",
    "EXCHANGE",
    "TENDERED",
    "MERGER",
    "SPINOFF",
    "RIGHTS",
    # ── Fractional / in-lieu proceeds (corporate-action cash, not a deposit) ─
    "IN LIEU",                       # "IN LIEU OF FRACTIONAL SHARES"
    "CASH IN LIEU",
    # ── Options activity ────────────────────────────────────────────────────
    "EXPIRED",
    "ASSIGNED",
    "EXERCISED",
    "OPENING TRANSACTION",
    "CLOSING TRANSACTION",
    # ── Securities lending (Fidelity Fully Paid Lending program) ────────────
    # Bookkeeping entries that move shares in/out of loan; not cash flows.
    # Hidden from the "captured cash flows" UI but retained in
    # parse result `all_transactions` for full audit trail.
    "YOU LOANED",                    # "YOU LOANED SECURITIES"
    "LOAN RETURNED",                 # "LOAN RETURNED" / "LOAN RETURNED SECURITIES"
})


def parse_fidelity_history(raw_text: str) -> dict:
    """Parse a Fidelity account activity/history CSV export.

    Returns:
        {
          "ok": True,
          "flows": [{"date", "amount", "action"}, ...],   # external cash flows only
          "all_transactions": [{"date", "amount", "action", "symbol", "type"}, ...],
          "transaction_count": int,   # total rows parsed
          "net_flow": float,          # sum of external flows (+ = net deposit)
          "columns_detected": [...],  # column names found — for diagnostics
          "unique_actions": [...],    # all unique full action strings seen — for diagnostics
          "skipped_no_amount": int,   # rows where Amount was blank/unparseable
        }

    The CSV has a variable-length header preamble before the data table.
    We scan for the header row containing 'Run Date' or 'Date' then read
    data rows beneath it.
    """
    import csv as _csv_io
    from io import StringIO as _StringIO

    lines = raw_text.splitlines()

    # ── Find the header row ───────────────────────────────────────────────────
    header_idx = None
    for i, line in enumerate(lines[:30]):
        low = line.strip().lower()
        if "run date" in low or ("date" in low and "action" in low):
            header_idx = i
            break
    if header_idx is None:
        return {
            "ok":    False,
            "error": (
                "Could not find header row with 'Run Date' / 'Action' in the CSV. "
                f"First 5 lines were: {lines[:5]}"
            ),
        }

    data_lines = lines[header_idx:]
    try:
        reader = _csv_io.DictReader(_StringIO("\n".join(data_lines)))
    except Exception as exc:  # noqa: BLE001
        return {"ok": False, "error": f"CSV parse error: {exc}"}

    flows: list[dict] = []
    all_transactions: list[dict] = []
    transaction_count = 0
    skipped_no_amount = 0
    unique_actions: set = set()
    columns_detected: list[str] = []
    year = datetime.now().year
    year_start = f"{year}-01-01"

    for row in reader:
        if not row:
            continue
        # Normalise column names; guard against None values (DictReader fills
        # short rows — e.g. footer text — with None for missing columns)
        row = {k.strip(): (v or "").strip() for k, v in row.items() if k and k.strip()}
        if not row:
            continue

        # Capture column list from first real row
        if not columns_detected:
            columns_detected = list(row.keys())

        # Get date — Fidelity uses MM/DD/YYYY
        raw_date = row.get("Run Date") or row.get("Date") or ""
        try:
            dt = datetime.strptime(raw_date.strip(), "%m/%d/%Y")
            date_str = dt.strftime("%Y-%m-%d")
        except ValueError:
            continue

        # Only care about transactions in the current calendar year
        if date_str < year_start:
            continue

        action = (row.get("Action") or "").strip().upper()
        if not action:
            continue

        # Track full action string for diagnostics (shown in UI flows section)
        unique_actions.add(action)

        transaction_count += 1

        # Parse the Amount column
        # Fidelity history CSVs name it "Amount ($)" — fall back to "Amount"
        raw_amt_str = (
            row.get("Amount ($)")
            or row.get("Amount")
            or ""
        ).replace("$", "").replace(",", "").strip()
        if raw_amt_str.startswith("(") and raw_amt_str.endswith(")"):
            raw_amt_str = "-" + raw_amt_str[1:-1]
        try:
            amount = float(raw_amt_str)
        except ValueError:
            skipped_no_amount += 1
            continue

        symbol = (row.get("Symbol") or "").strip().upper()
        tx_type = (row.get("Type") or "").strip()

        # Record every transaction for full history
        all_transactions.append({
            "date":   date_str,
            "amount": round(amount, 2),
            "action": action,
            "symbol": symbol,
            "type":   tx_type,
        })

        if amount == 0:
            continue

        # ── Blocklist: skip known-internal investment activity ─────────────────
        # Everything else (transfers, ACH, wire, journal entries, bill payments,
        # etc.) is an external cash flow regardless of the exact action string.
        is_internal = any(kw in action for kw in _FIDELITY_INTERNAL_ACTIONS)
        if is_internal:
            continue

        flows.append({
            "date":   date_str,
            "amount": round(amount, 2),
            "action": action,
        })

    net_flow = sum(f["amount"] for f in flows)
    return {
        "ok":                True,
        "flows":             sorted(flows, key=lambda f: f["date"]),
        "all_transactions":  sorted(all_transactions, key=lambda t: t["date"]),
        "transaction_count": transaction_count,
        "net_flow":          round(net_flow, 2),
        "columns_detected":  columns_detected,
        "unique_actions":    sorted(unique_actions),
        "skipped_no_amount": skipped_no_amount,
    }


def upload_account_history(
    raw_text: str,
    begin_value: float,
    end_value: float = 0.0,
) -> dict:
    """Attach a Fidelity account history to the live portfolio and compute
    a Modified Dietz YTD return.

    Modified Dietz formula:
        R = (EMV − BMV − CF) / (BMV + Σ(CFᵢ × Wᵢ))

    Where:
        BMV = begin_value   — portfolio value on Jan 1 (user-supplied)
        EMV = end_value     — today's TOTAL account value, user-supplied from
                              Fidelity's account summary (includes money market,
                              all positions, pending activity, etc.)
        CF  = net external cash flows (deposits − withdrawals) from the CSV
        Wᵢ  = (CD − Dᵢ) / CD  — fraction of period remaining after flow date
        CD  = total calendar days from Jan 1 to today
        Dᵢ  = calendar days from Jan 1 to the flow date

    Parameters:
        raw_text:    Raw text of the Fidelity activity/history CSV
        begin_value: Portfolio total value on January 1 of the current year ($)
        end_value:   Portfolio total value TODAY — the full account balance as
                     shown by Fidelity, including ALL positions and money-market.
                     If 0 or not provided, falls back to the stored holdings
                     estimate (less accurate).
    """
    if not raw_text:
        return {"ok": False, "error": "No history CSV content provided."}
    if not begin_value or begin_value <= 0:
        return {"ok": False, "error": "Beginning of year portfolio value is required and must be positive."}

    parsed = parse_fidelity_history(raw_text)
    if not parsed.get("ok"):
        return parsed

    flows    = parsed["flows"]
    net_flow = parsed["net_flow"]

    # ── Determine EMV (End Market Value) ─────────────────────────────────────
    # Preferred: user provides today's actual account total from Fidelity.
    # This includes money-market funds, unrecognized tickers, pending cash, etc.
    # that the live-holdings snapshot cannot see.
    emv_source = "user_provided"
    if end_value and end_value > 0:
        emv = float(end_value)
    else:
        # Fallback: estimate from stored live holdings × price ratios.
        # NOTE: this is unreliable — it misses money market, partial sales, and
        # any position not in the uploaded positions CSV.
        emv_source = "holdings_estimate"
        try:
            state = _load_tracker_state()
        except Exception as exc:  # noqa: BLE001
            return {"ok": False, "error": f"Could not load tracker state: {exc}"}

        live = state.get("live_portfolio")
        if not live or not live.get("holdings"):
            return {
                "ok": False,
                "error": (
                    "No current portfolio value provided and no live portfolio found. "
                    "Please enter today's total account value from Fidelity."
                ),
            }

        emv = 0.0
        for h in (live.get("holdings") or []):
            try:
                ticker = h.get("ticker") or ""
                if not ticker:
                    continue
                price  = _fetch_close_price(ticker)
                w      = float(h.get("weight") or 0)
                if price and price > 0 and w > 0:
                    anchor = float(h.get("anchor_price") or h.get("entry_price") or 0)
                    if anchor > 0:
                        emv += begin_value * w * (price / anchor)
                    else:
                        emv += begin_value * w
            except Exception:  # noqa: BLE001
                continue

        if emv <= 0:
            emv = begin_value  # last-resort: assume flat

    # ── Modified Dietz calculation ────────────────────────────────────────────
    year         = datetime.now().year
    period_start = datetime(year, 1, 1)
    today_dt     = datetime.now()
    cd           = max(1, (today_dt - period_start).days)

    weighted_flows = 0.0
    for f in flows:
        try:
            flow_dt         = datetime.strptime(f["date"], "%Y-%m-%d")
            di              = max(0, (flow_dt - period_start).days)
            wi              = (cd - di) / cd
            weighted_flows += f["amount"] * wi
        except Exception:  # noqa: BLE001
            continue

    denominator = begin_value + weighted_flows
    if denominator <= 0:
        return {"ok": False, "error": "Denominator ≤ 0 — check your beginning value and cash flows."}

    md_return     = (emv - begin_value - net_flow) / denominator
    md_return_pct = round(md_return * 100.0, 4)

    # ── Persist history metadata with the live portfolio ─────────────────────
    try:
        with _tracker_lock():
            state = _load_tracker_state()
            if state.get("live_portfolio"):
                state["live_portfolio"]["account_history"] = {
                    "uploaded_at":       datetime.utcnow().isoformat(),
                    "begin_value":       round(float(begin_value), 2),
                    "end_value":         round(float(emv), 2),
                    "emv_source":        emv_source,
                    "net_flow":          round(float(net_flow), 2),
                    "flow_count":        int(len(flows)),
                    "transaction_count": int(parsed.get("transaction_count", 0)),
                    "md_return_pct":     float(md_return_pct),
                }
                _save_tracker_state(state)
    except Exception as exc:  # noqa: BLE001
        print(f"⚠️  Could not persist account history: {exc}")  # non-fatal

    return {
        "ok":                True,
        "md_return_pct":     md_return_pct,
        "begin_value":       round(float(begin_value), 2),
        "end_value":         round(float(emv), 2),
        "emv_source":        emv_source,
        "net_flow":          round(float(net_flow), 2),
        "flow_count":        len(flows),
        "transaction_count": int(parsed.get("transaction_count", 0)),
        "flows":             flows,
        "all_transactions":  parsed.get("all_transactions", []),
        # ── Diagnostics ──────────────────────────────────────────────────────
        "columns_detected":  parsed.get("columns_detected", []),
        "unique_actions":    parsed.get("unique_actions", []),
        "skipped_no_amount": parsed.get("skipped_no_amount", 0),
    }


def _compute_twrr(
    monthly_data: list,
    flows: list,
) -> float | None:
    """Compute Time-Weighted Rate of Return (TWRR) from monthly sub-periods.

    When monthly_data rows have exact HPRs from the Fidelity monthly performance
    CSV (mo["exact"] = True), those are used directly — giving a result that
    exactly matches Fidelity's reported TWRR.

    Otherwise chains estimated HPRs:
        HPR_m  = EMV_m / (BMV_m + CF_m)   CF from activity CSV
        TWRR   = Π(HPR_m) − 1
    """
    from collections import defaultdict

    flows_by_month: dict = defaultdict(float)
    for f in flows:
        try:
            m = int(f["date"][5:7])
            flows_by_month[m] += float(f["amount"])
        except Exception:  # noqa: BLE001
            pass

    product       = 1.0
    valid_months  = 0
    for mo in monthly_data:
        m = mo["month"]
        if mo.get("exact") and mo.get("perf_detail") and \
                mo["perf_detail"].get("net_flow") is not None:
            # Exact HPR from Fidelity monthly CSV
            bmv      = float(mo.get("start_value") or 0.0)
            emv      = float(mo.get("end_value")   or 0.0)
            net_flow = float(mo["perf_detail"]["net_flow"])
            denom    = bmv + net_flow
        else:
            bmv   = float(mo.get("start_value") or 0.0)
            emv   = float(mo.get("end_value")   or 0.0)
            cf    = flows_by_month.get(m, 0.0)
            denom = bmv + cf

        if denom <= 0 or bmv <= 0:
            continue
        product      *= emv / denom
        valid_months += 1

    if valid_months == 0:
        return None
    return round((product - 1.0) * 100.0, 4)


def compute_unified_ytd(
    positions_text:      str,
    activity_text:       str,
    begin_value:         float | None = None,
    monthly_perf_text:   str | None = None,
) -> dict:
    """Single unified YTD computation: Modified Dietz return + per-stock attribution.

    Replaces the two separate flows (account_history + attribution). One call,
    one set of inputs:
        • positions_text   — Fidelity Positions CSV (current holdings + cash)
        • activity_text    — Fidelity Activity/History CSV (YTD transactions)
        • begin_value      — Jan 1 portfolio total ($).  Optional when
                             monthly_perf_text is supplied — the first month's
                             beginning balance is used automatically.
        • monthly_perf_text — (optional) Fidelity monthly performance summary CSV.
                             When provided, begin_value can be omitted.

    The end portfolio value (EMV) is auto-derived as the sum of every position
    × current price PLUS money-market funds (SPAXX/FZFXX/etc.) — exactly what
    Fidelity shows as the account total. No manual end-value entry needed.
    """
    if not positions_text:
        return {"ok": False, "error": "Positions CSV content is required."}
    if not activity_text:
        return {"ok": False, "error": "Activity CSV content is required."}

    # ── Parse monthly performance CSV early — we may need it for begin_value ──
    monthly_perf_parsed = None
    monthly_perf_error  = None
    if monthly_perf_text and monthly_perf_text.strip():
        try:
            mp = parse_fidelity_monthly_perf(monthly_perf_text)
            if mp.get("ok"):
                monthly_perf_parsed = mp
            else:
                monthly_perf_error = mp.get("error", "monthly perf parse failed")
        except Exception as mp_exc:  # noqa: BLE001
            monthly_perf_error = str(mp_exc)

    # ── Resolve begin_value: manual entry OR first month's start from perf CSV ──
    if not begin_value or begin_value <= 0:
        if monthly_perf_parsed and monthly_perf_parsed.get("months"):
            begin_value = float(monthly_perf_parsed["months"][0]["start"])
        else:
            return {
                "ok": False,
                "error": (
                    "Jan 1 portfolio value is required. "
                    "Either enter it manually or upload a Monthly Performance CSV."
                ),
            }

    # ── Per-ticker attribution (also gives us positions_total_value as EMV) ──
    attr = compute_position_attribution(positions_text, activity_text, begin_value)
    if not attr.get("ok"):
        return attr

    # End market value = total of all positions (incl. money-market) at current px
    end_value = float(attr.get("portfolio_end_value") or 0.0)
    if end_value <= 0:
        return {
            "ok": False,
            "error": (
                "Could not compute end portfolio value from positions CSV — "
                "no positions parsed with both quantity and price."
            ),
        }

    # ── External cash flows ───────────────────────────────────────────────────
    # Always parse the activity CSV so we have individual transaction rows for
    # the flows table display and fallback values.
    flows_parse = parse_fidelity_history(activity_text)
    if not flows_parse.get("ok"):
        return flows_parse

    flows = flows_parse["flows"]   # individual rows — used for display table

    # ── Modified Dietz: R = (EMV − BMV − CF) / (BMV + Σ(CFᵢ × Wᵢ)) ──────────
    year         = datetime.now().year
    period_start = datetime(year, 1, 1)
    today_dt     = datetime.now()
    cd           = max(1, (today_dt - period_start).days)

    # Net flow & weighted flows for MD: use monthly-perf CSV when available
    # (it contains Fidelity's exact deposit/withdrawal totals per month).
    # Fall back to activity CSV parsing when no perf CSV is provided.
    if monthly_perf_parsed and monthly_perf_parsed.get("months"):
        perf_months = monthly_perf_parsed["months"]
        net_flow    = round(sum(float(pm.get("net_flow") or 0.0) for pm in perf_months), 2)
        # Weighted flows: treat each month's net flow as occurring mid-month
        weighted_flows = 0.0
        for pm in perf_months:
            pmf = float(pm.get("net_flow") or 0.0)
            if pmf == 0.0:
                continue
            try:
                mid_month       = datetime(year, pm["month"], 15)
                di              = max(0, (mid_month - period_start).days)
                wi              = (cd - di) / cd
                weighted_flows += pmf * wi
            except Exception:  # noqa: BLE001
                continue
        # flow_count = months with a non-zero deposit or withdrawal
        flow_count_display = sum(
            1 for pm in perf_months
            if (pm.get("deposits") or 0) != 0 or (pm.get("withdrawals") or 0) != 0
        )
    else:
        net_flow       = float(flows_parse["net_flow"])
        weighted_flows = 0.0
        for f in flows:
            try:
                flow_dt         = datetime.strptime(f["date"], "%Y-%m-%d")
                di              = max(0, (flow_dt - period_start).days)
                wi              = (cd - di) / cd
                weighted_flows += float(f["amount"]) * wi
            except Exception:  # noqa: BLE001
                continue
        flow_count_display = len(flows)

    denominator = float(begin_value) + weighted_flows
    if denominator <= 0:
        return {"ok": False, "error": "Modified Dietz denominator ≤ 0 — check your Jan 1 value."}

    md_return     = (end_value - float(begin_value) - net_flow) / denominator
    md_return_pct = round(md_return * 100.0, 4)

    # ── Monthly chart + TWRR (computed before the lock — can be slow) ─────────
    monthly_chart      = None
    monthly_chart_error = None
    twrr_return_pct    = None
    try:
        mc = compute_monthly_ytd_chart(
            positions_text, activity_text, begin_value,
            monthly_perf=monthly_perf_parsed,
        )
        if mc.get("ok"):
            monthly_chart   = mc
            twrr_return_pct = _compute_twrr(mc.get("monthly") or [], flows)
        else:
            monthly_chart_error = mc.get("error", "monthly chart returned not-ok")
    except Exception as mc_exc:  # noqa: BLE001
        monthly_chart_error = str(mc_exc)
        print(f"⚠️  Monthly chart failed: {mc_exc}")

    # ── Persist the MD/TWRR result + attribution + a snapshot for history ─────
    # `account_history` always holds the most-recent run (used by the YTD
    # detail view).  `ytd_snapshots` is an append-only list (cap 50) so the
    # user can browse / re-open / email any past run.
    snapshot_id = datetime.utcnow().strftime("%Y%m%d-%H%M%S")
    snapshot = {
        "id":                  snapshot_id,
        "uploaded_at":         datetime.utcnow().isoformat(),
        "begin_value":         round(float(begin_value), 2),
        "end_value":           round(end_value, 2),
        "emv_source":          "positions_csv",
        "net_flow":            round(net_flow, 2),
        "flow_count":          flow_count_display,
        "flows":               sorted(flows, key=lambda f: f["date"]),  # persisted for page-load render
        "unique_actions":      sorted(flows_parse.get("unique_actions") or []),
        "transaction_count":   int(flows_parse.get("transaction_count", 0)),
        "trade_count":         int(attr.get("trades_parsed", 0)),
        "dividend_count":      int(attr.get("dividends_parsed", 0)),
        "md_return_pct":       float(md_return_pct),
        "twrr_return_pct":     float(twrr_return_pct) if twrr_return_pct is not None else None,
        "attribution":         attr.get("attribution", []),
        "monthly_chart":       monthly_chart,
        "monthly_chart_error": monthly_chart_error,
        "monthly_perf_error":  monthly_perf_error,
        "has_monthly_perf":    monthly_perf_parsed is not None,
        # Capture the live-portfolio holdings at upload time so the snapshot
        # remains a complete record even if the live book changes later.
        "holdings_snapshot":   None,  # filled below from live state
    }
    # Build a holdings list directly from the positions CSV. The Tracker page
    # is independent of any Live Rebalance run — when no live_portfolio exists,
    # we auto-create one from the uploaded positions so the YTD detail and
    # snapshots have something to reference.
    pos_total = float(attr.get("portfolio_end_value") or end_value or 0.0)
    pos_holdings_for_promote: list[dict] = []
    try:
        _pos_again = _parse_fidelity_positions_extended(positions_text)
        if _pos_again.get("ok"):
            for h in _pos_again.get("holdings") or []:
                cv = h.get("current_value") or 0.0
                w = (cv / pos_total) if pos_total > 0 else 0.0
                pos_holdings_for_promote.append({
                    "ticker":       h.get("ticker"),
                    "weight":       round(w, 6),
                    "anchor_price": h.get("price"),
                    "is_mm":        h.get("is_mm", False),
                })
    except Exception:  # noqa: BLE001
        pos_holdings_for_promote = []

    try:
        with _tracker_lock():
            state = _load_tracker_state()
            lp = state.get("live_portfolio")

            # Auto-create a live_portfolio from the positions CSV if none exists.
            # This makes the Tracker page work standalone — no separate Live
            # Rebalance step required.
            if not lp:
                lp = {
                    "anchor_date":  _today_str(),
                    "uploaded_at":  datetime.utcnow().isoformat(),
                    "holdings":     pos_holdings_for_promote,
                    "auto_created": True,
                }
                state["live_portfolio"] = lp
            elif pos_holdings_for_promote:
                # Always refresh holdings from the latest positions CSV so the
                # Live Benchmark card reflects the current upload, not a stale one.
                lp["holdings"] = pos_holdings_for_promote

            # Capture holdings at upload time (deep copy via JSON round-trip
            # — primitives only, so json.loads(json.dumps(...)) is safe).
            # Prefer the freshly-uploaded positions over a stale live snapshot.
            holdings_to_snap = pos_holdings_for_promote or lp.get("holdings") or []
            try:
                snapshot["holdings_snapshot"] = json.loads(
                    json.dumps(holdings_to_snap, default=str)
                )
                snapshot["anchor_date"] = lp.get("anchor_date")
            except Exception:  # noqa: BLE001
                snapshot["holdings_snapshot"] = holdings_to_snap

            lp["account_history"] = snapshot

            # Append-only snapshot history (capped at 50 most recent)
            snaps = lp.get("ytd_snapshots") or []
            snaps.append(snapshot)
            if len(snaps) > 50:
                snaps = snaps[-50:]
            lp["ytd_snapshots"] = snaps

            _save_tracker_state(state)
    except Exception as exc:  # noqa: BLE001
        print(f"⚠️  Could not persist unified YTD: {exc}")  # non-fatal

    return {
        "ok":                True,
        "snapshot_id":       snapshot_id,
        "md_return_pct":     md_return_pct,
        "twrr_return_pct":   twrr_return_pct,
        "begin_value":       round(float(begin_value), 2),
        "end_value":         round(end_value, 2),
        "emv_source":        "positions_csv",
        "net_flow":          round(net_flow, 2),
        "flow_count":        flow_count_display,
        "trade_count":       int(attr.get("trades_parsed", 0)),
        "dividend_count":    int(attr.get("dividends_parsed", 0)),
        "transaction_count": int(flows_parse.get("transaction_count", 0)),
        "flows":             flows,
        "unique_actions":    sorted(flows_parse.get("unique_actions") or []),
        "attribution":       attr.get("attribution", []),
        "total_dollar_gain": attr.get("total_dollar_gain"),
        "explained_pct":     attr.get("explained_pct"),
        "positions_parsed":  attr.get("positions_parsed"),
        "monthly_chart":       monthly_chart,
        "monthly_chart_error": monthly_chart_error,
        "monthly_perf_error":  monthly_perf_error,
        "has_monthly_perf":    monthly_perf_parsed is not None,
    }


# ── YTD snapshot history + email ────────────────────────────────────────────

def list_ytd_snapshots() -> dict:
    """Return all stored YTD snapshots with summary metadata, newest first."""
    state = _load_tracker_state()
    lp = state.get("live_portfolio") or {}
    snaps = lp.get("ytd_snapshots") or []
    summaries = []
    for s in reversed(snaps):  # newest first
        attr = s.get("attribution") or []
        summaries.append({
            "id":              s.get("id"),
            "uploaded_at":     s.get("uploaded_at"),
            "begin_value":     s.get("begin_value"),
            "end_value":       s.get("end_value"),
            "md_return_pct":   s.get("md_return_pct"),
            "net_flow":        s.get("net_flow"),
            "trade_count":     s.get("trade_count", 0),
            "dividend_count":  s.get("dividend_count", 0),
            "positions_count": len(attr),
            "anchor_date":     s.get("anchor_date"),
        })
    return {"ok": True, "snapshots": summaries}


def get_ytd_snapshot(snapshot_id: str) -> dict:
    """Return a full YTD snapshot by id, including attribution + holdings."""
    state = _load_tracker_state()
    lp = state.get("live_portfolio") or {}
    for s in (lp.get("ytd_snapshots") or []):
        if s.get("id") == snapshot_id:
            return {"ok": True, "snapshot": s}
    return {"ok": False, "error": f"Snapshot {snapshot_id!r} not found."}


def delete_ytd_snapshot(snapshot_id: str) -> dict:
    """Remove a snapshot from history (irreversible)."""
    with _tracker_lock():
        state = _load_tracker_state()
        lp = state.get("live_portfolio")
        if not lp:
            return {"ok": False, "error": "No live portfolio."}
        snaps = lp.get("ytd_snapshots") or []
        new = [s for s in snaps if s.get("id") != snapshot_id]
        if len(new) == len(snaps):
            return {"ok": False, "error": f"Snapshot {snapshot_id!r} not found."}
        lp["ytd_snapshots"] = new
        # If we just deleted the most-recent snapshot, also reset
        # account_history so the YTD detail view doesn't keep showing it.
        if (lp.get("account_history") or {}).get("id") == snapshot_id:
            lp["account_history"] = new[-1] if new else None
        _save_tracker_state(state)
    return {"ok": True, "deleted": snapshot_id, "remaining": len(new)}


def set_current_ytd_snapshot(snapshot_id: str) -> dict:
    """Promote a past snapshot to live_portfolio.account_history.

    This makes the Live Benchmark card and YTD detail view reflect a specific
    past run rather than always showing the most-recently uploaded one.
    Also refreshes live_portfolio.holdings from the snapshot's holdings_snapshot
    so the Live Benchmark holdings chips update immediately.
    """
    with _tracker_lock():
        state = _load_tracker_state()
        lp = state.get("live_portfolio")
        if not lp:
            return {"ok": False, "error": "No live portfolio found."}
        snaps = lp.get("ytd_snapshots") or []
        snap = next((s for s in snaps if s.get("id") == snapshot_id), None)
        if not snap:
            return {"ok": False, "error": f"Snapshot {snapshot_id!r} not found."}
        lp["account_history"] = snap
        # Also update holdings from this snapshot so live benchmark chips update
        snap_holdings = snap.get("holdings_snapshot")
        if snap_holdings:
            lp["holdings"] = snap_holdings
        _save_tracker_state(state)
    return {"ok": True, "snapshot_id": snapshot_id}


def email_ytd_report(to_email: str, snapshot_id: str | None = None) -> dict:
    """Email the full YTD report (live benchmark + attribution + YTD detail).

    If snapshot_id is None, uses the most recent stored snapshot.
    Builds an HTML email styled to mirror the in-app YTD detail view.
    """
    if not to_email or "@" not in to_email:
        return {"ok": False, "error": "A valid email address is required."}

    state = _load_tracker_state()
    lp = state.get("live_portfolio") or {}
    snaps = lp.get("ytd_snapshots") or []
    if not snaps:
        return {"ok": False, "error": "No YTD snapshots found — run the YTD calculator first."}

    snap = None
    if snapshot_id:
        for s in snaps:
            if s.get("id") == snapshot_id:
                snap = s
                break
        if snap is None:
            return {"ok": False, "error": f"Snapshot {snapshot_id!r} not found."}
    else:
        snap = snaps[-1]

    # Pull the full YTD detail (chart series, SPY benchmark, contribution-sorted
    # holdings) so the email can render every section of the on-screen view.
    try:
        detail = compute_live_ytd_detail(snapshot_id=snap.get("id"))
        if not detail.get("ok"):
            detail = None
    except Exception as exc:  # noqa: BLE001
        print(f"⚠️  email_ytd_report: could not load YTD detail: {exc}")
        detail = None

    msg = _compose_ytd_report_email(to_email, snap, lp, detail)
    return send_portfolio_email(msg)


def _build_monthly_table_html(
    snap: dict,
    pct_fmt,   # callable: float -> "+X.XX%"
    color_fmt, # callable: float -> hex color
    usd0_fmt,  # callable: float -> "$X,XXX"
) -> str:
    """Return an HTML <tr><td> block with the YTD-by-month performance table.

    Uses snap["monthly_chart"]["monthly"] when available; returns "" otherwise.
    Rendered as an email-safe table with one row per completed month.
    """
    mc = snap.get("monthly_chart") or {}
    months = mc.get("monthly") or []
    if not months:
        return ""

    # Header row
    header = (
        "<tr style='background:#F4F6F8'>"
        "<th style='padding:7px 8px;text-align:left;font-size:9px;letter-spacing:0.8px;"
        "color:#6B7280;border-bottom:1px solid #DCE3EB'>MONTH</th>"
        "<th style='padding:7px 8px;text-align:right;font-size:9px;letter-spacing:0.8px;"
        "color:#6B7280;border-bottom:1px solid #DCE3EB'>RETURN</th>"
        "<th style='padding:7px 8px;text-align:right;font-size:9px;letter-spacing:0.8px;"
        "color:#6B7280;border-bottom:1px solid #DCE3EB'>SPY</th>"
        "<th style='padding:7px 8px;text-align:right;font-size:9px;letter-spacing:0.8px;"
        "color:#6B7280;border-bottom:1px solid #DCE3EB'>VS SPY</th>"
        "<th style='padding:7px 8px;text-align:right;font-size:9px;letter-spacing:0.8px;"
        "color:#6B7280;border-bottom:1px solid #DCE3EB'>END VALUE</th>"
        "<th style='padding:7px 8px;text-align:right;font-size:9px;letter-spacing:0.8px;"
        "color:#6B7280;border-bottom:1px solid #DCE3EB'>$ GAIN</th>"
        "</tr>"
    )

    # Convert SPY YTD-end-of-month series → SPY MONTHLY return for apples-to-apples
    # comparison with the portfolio's monthly return_pct.
    spy_monthly_returns: list[float | None] = []
    prev_spy_ytd_decimal = 0.0   # Jan 1 baseline = 0% YTD
    for mo in months:
        sy = mo.get("spy_ytd_pct")
        if sy is None:
            spy_monthly_returns.append(None)
            continue
        cur_decimal = sy / 100.0
        # SPY return for this month alone, derived from cumulative YTD chain
        if (1.0 + prev_spy_ytd_decimal) > 0:
            spy_m_pct = ((1.0 + cur_decimal) / (1.0 + prev_spy_ytd_decimal) - 1.0) * 100.0
        else:
            spy_m_pct = None
        spy_monthly_returns.append(spy_m_pct)
        prev_spy_ytd_decimal = cur_decimal

    body_rows = ""
    for idx, mo in enumerate(months):
        # NOTE: mo["return_pct"] is ALREADY in percent form (e.g. 6.40 means +6.40%).
        # Do NOT multiply by 100 again — that produced the 640% display bug.
        rp_pct  = mo.get("return_pct") or 0.0
        spy_m   = spy_monthly_returns[idx] if idx < len(spy_monthly_returns) else None
        vs_spy  = (rp_pct - spy_m) if spy_m is not None else None
        ev      = mo.get("end_value")
        dg      = mo.get("dollar_gain")
        exact   = mo.get("exact", False)
        label   = mo.get("label", "")

        exact_dot = (
            "<span style='display:inline-block;width:5px;height:5px;border-radius:50%;"
            "background:#C9A84C;margin-left:3px;vertical-align:middle;' "
            "title='Exact from Fidelity CSV'></span>"
            if exact else ""
        )

        spy_str = pct_fmt(spy_m) if spy_m is not None else "—"
        spy_col = color_fmt(spy_m) if spy_m is not None else "#9CA3AF"
        vs_spy_str = pct_fmt(vs_spy) if vs_spy is not None else "—"
        vs_spy_col = color_fmt(vs_spy) if vs_spy is not None else "#9CA3AF"

        body_rows += (
            f"<tr>"
            f"<td style='padding:7px 8px;border-bottom:1px solid #EEE;font-weight:700;"
            f"font-family:Menlo,monospace;font-size:11px'>{label}{exact_dot}</td>"
            f"<td style='padding:7px 8px;border-bottom:1px solid #EEE;text-align:right;"
            f"color:{color_fmt(rp_pct)};font-weight:700;font-family:Menlo,monospace;"
            f"font-size:12px'>{pct_fmt(rp_pct)}</td>"
            f"<td style='padding:7px 8px;border-bottom:1px solid #EEE;text-align:right;"
            f"color:{spy_col};font-family:Menlo,monospace;font-size:11px'>{spy_str}</td>"
            f"<td style='padding:7px 8px;border-bottom:1px solid #EEE;text-align:right;"
            f"color:{vs_spy_col};font-family:Menlo,monospace;font-size:11px'>{vs_spy_str}</td>"
            f"<td style='padding:7px 8px;border-bottom:1px solid #EEE;text-align:right;"
            f"font-family:Menlo,monospace;font-size:11px;color:#374151'>{usd0_fmt(ev) if ev is not None else '—'}</td>"
            f"<td style='padding:7px 8px;border-bottom:1px solid #EEE;text-align:right;"
            f"color:{color_fmt(dg) if dg is not None else '#9CA3AF'};"
            f"font-family:Menlo,monospace;font-size:11px'>{usd0_fmt(dg) if dg is not None else '—'}</td>"
            f"</tr>"
        )

    # Totals row — md_return_pct is also already in percent form (do NOT × 100).
    total_rp   = (snap.get("md_return_pct") or 0.0)
    total_ev   = (snap.get("end_value") or 0.0)
    total_gain = sum((mo.get("dollar_gain") or 0.0) for mo in months)
    # SPY YTD = last month's spy_ytd_pct (cumulative, already %)
    total_spy_ytd = None
    for mo in reversed(months):
        sy = mo.get("spy_ytd_pct")
        if sy is not None:
            total_spy_ytd = sy
            break
    total_vs_spy = (total_rp - total_spy_ytd) if total_spy_ytd is not None else None

    tfoot = (
        f"<tr style='background:#F4F6F8'>"
        f"<td style='padding:8px;font-weight:800;font-size:11px;color:#0A1628'>YTD</td>"
        f"<td style='padding:8px;text-align:right;color:{color_fmt(total_rp)};"
        f"font-weight:800;font-family:Menlo,monospace'>{pct_fmt(total_rp)}</td>"
        f"<td style='padding:8px;text-align:right;"
        f"color:{color_fmt(total_spy_ytd) if total_spy_ytd is not None else '#9CA3AF'};"
        f"font-weight:700;font-family:Menlo,monospace;font-size:11px'>"
        f"{pct_fmt(total_spy_ytd) if total_spy_ytd is not None else '—'}</td>"
        f"<td style='padding:8px;text-align:right;"
        f"color:{color_fmt(total_vs_spy) if total_vs_spy is not None else '#9CA3AF'};"
        f"font-weight:800;font-family:Menlo,monospace;font-size:11px'>"
        f"{pct_fmt(total_vs_spy) if total_vs_spy is not None else '—'}</td>"
        f"<td style='padding:8px;text-align:right;font-family:Menlo,monospace;"
        f"font-weight:700;color:#0A1628'>{usd0_fmt(total_ev)}</td>"
        f"<td style='padding:8px;text-align:right;color:{color_fmt(total_gain)};"
        f"font-weight:800;font-family:Menlo,monospace'>{usd0_fmt(total_gain)}</td>"
        f"</tr>"
    )

    exact_note = (
        "<div style='font-size:9px;color:#9CA3AF;margin-top:6px'>"
        "<span style='display:inline-block;width:5px;height:5px;border-radius:50%;"
        "background:#C9A84C;vertical-align:middle;margin-right:3px'></span>"
        "Gold dot = exact monthly return from Investment Income Balance CSV"
        "</div>"
        if any(mo.get("exact") for mo in months) else ""
    )

    return (
        f"<tr><td style='background:#fff;padding:18px 24px;border-top:1px solid #E5E7EB'>"
        f"<div style='font-size:11px;font-weight:800;letter-spacing:1.5px;color:#C9A84C;"
        f"margin-bottom:10px'>YTD PERFORMANCE BY MONTH</div>"
        f"<table width='100%' cellpadding='0' cellspacing='0' "
        f"style='border-collapse:collapse;font-size:12px'>"
        f"<thead>{header}</thead>"
        f"<tbody>{body_rows}</tbody>"
        f"<tfoot>{tfoot}</tfoot>"
        f"</table>"
        f"{exact_note}"
        f"</td></tr>"
    )


def _compose_ytd_report_email(
    to_email: str,
    snap: dict,
    live_portfolio: dict,
    detail: dict | None = None,
) -> EmailMessage:
    """Build the HTML email for a YTD snapshot.

    Sections:
      1. Live Benchmark positions (chips)
      2. Modified Dietz YTD summary stats
      3. Per-stock attribution table (ticker, jan1, activity, now, P&L, contrib)
      4. SPY-vs-portfolio summary line + skipped tickers note
    """
    import base64 as _b64

    from_addr = _optional_env("GMAIL_USER", to_email)
    uploaded_dt = snap.get("uploaded_at", "")
    try:
        uploaded_human = datetime.fromisoformat(uploaded_dt.replace("Z","")).strftime("%Y-%m-%d %H:%M UTC")
    except Exception:  # noqa: BLE001
        uploaded_human = uploaded_dt

    md_pct      = snap.get("md_return_pct") or 0.0
    begin_value = snap.get("begin_value")   or 0.0
    end_value   = snap.get("end_value")     or 0.0
    net_flow    = snap.get("net_flow")      or 0.0
    attribution = snap.get("attribution") or []
    holdings    = snap.get("holdings_snapshot") or live_portfolio.get("holdings") or []

    subject = (
        f"DGA Live Portfolio — YTD Report "
        f"{md_pct:+.2f}% — {uploaded_human}"
    )

    # ── Plain-text fallback ──────────────────────────────────────────────────
    plain_lines = [
        "DGA Capital — Live Portfolio YTD Report",
        f"Generated: {uploaded_human}",
        "",
        f"YTD Return (Modified Dietz): {md_pct:+.2f}%",
        f"Jan 1 Value:  ${begin_value:,.0f}",
        f"Today Value:  ${end_value:,.0f}",
        f"Net Flows:    ${net_flow:+,.0f}",
        f"Trades:       {snap.get('trade_count', 0)} ({snap.get('dividend_count', 0)} dividends)",
        "",
        "LIVE BENCHMARK POSITIONS",
        "=" * 60,
    ]
    for h in sorted(holdings, key=lambda x: -(x.get("weight") or 0)):
        plain_lines.append(f"  {h.get('ticker','?'):8s}  {(h.get('weight') or 0)*100:6.2f}%")
    plain_lines += ["", "PER-STOCK ATTRIBUTION", "=" * 60]
    plain_lines.append(f"  {'Ticker':<8} {'Jan 1 sh':>12} {'$ P&L':>14} {'% Contrib':>10}")
    for a in attribution:
        plain_lines.append(
            f"  {a.get('ticker','?'):<8} {(a.get('start_shares') or 0):>12,.2f} "
            f"${(a.get('dollar_gain') or 0):>13,.0f}  {(a.get('contribution_pct') or 0):>9.2f}%"
        )

    plain_body = "\n".join(plain_lines)

    # ── Logo (inline base64) ─────────────────────────────────────────────────
    logo_img_tag = ""
    for logo_name in ("DGAlogo-webFINAL-68.png", "dga_logo_small.png", "DGAlogo-web184.png", "dga_logo.png"):
        logo_path = SCRIPT_DIR / "branding" / logo_name
        if logo_path.exists():
            logo_b64 = _b64.b64encode(logo_path.read_bytes()).decode()
            logo_img_tag = (
                "<div style='background:#ffffff;border-radius:8px;padding:6px 14px;"
                "display:inline-block;line-height:0'>"
                f"<img src='data:image/png;base64,{logo_b64}' "
                f"alt='DGA Capital' style='height:36px;width:auto;display:block'>"
                "</div>"
            )
            break

    # ── HTML helpers ─────────────────────────────────────────────────────────
    def color(v):
        if v is None:
            return "#666"
        return "#16A34A" if v >= 0 else "#DC2626"

    def usd0(v):
        if v is None:
            return "—"
        sign = "−" if v < 0 else ""
        return sign + "$" + f"{abs(v):,.0f}"

    def shares(v):
        if v is None:
            return "—"
        return f"{v:,.2f}"

    def pct(v, d=2):
        if v is None:
            return "—"
        return f"{'+' if v >= 0 else ''}{v:.{d}f}%"

    # Live benchmark chips (sorted by weight desc)
    sorted_hold = sorted(holdings, key=lambda h: -(h.get("weight") or 0))
    chips_html = "".join(
        f"<span style='display:inline-block;background:#F4F6F8;"
        f"border:1px solid #DCE3EB;border-radius:6px;padding:4px 9px;"
        f"margin:2px 3px 2px 0;font-family:Menlo,monospace;font-size:11px;color:#0A1628'>"
        f"<strong>{h.get('ticker','?')}</strong> "
        f"<span style='color:#677788;font-weight:400'>{(h.get('weight') or 0)*100:.1f}%</span>"
        f"</span>"
        for h in sorted_hold
    )

    # Per-stock attribution rows
    attr_rows_html = ""
    for a in attribution:
        sells_chip = ""
        buys_chip  = ""
        divs_chip  = ""
        if (a.get("total_sold_shares") or 0) > 0:
            avgp = (a["total_sell_proceeds"] or 0) / a["total_sold_shares"]
            sells_chip = (
                f"<span style='background:#FEE2E2;color:#DC2626;border-radius:3px;"
                f"padding:1px 5px;margin-right:3px;font-size:10px;font-weight:700;"
                f"font-family:Menlo,monospace;display:inline-block;margin-bottom:2px'>"
                f"▼ {shares(a['total_sold_shares'])} @ ${avgp:.2f}</span>"
            )
        if (a.get("total_bought_shares") or 0) > 0:
            avgp = (a["total_buy_cost"] or 0) / a["total_bought_shares"]
            buys_chip = (
                f"<span style='background:#DCFCE7;color:#16A34A;border-radius:3px;"
                f"padding:1px 5px;margin-right:3px;font-size:10px;font-weight:700;"
                f"font-family:Menlo,monospace;display:inline-block;margin-bottom:2px'>"
                f"▲ {shares(a['total_bought_shares'])} @ ${avgp:.2f}</span>"
            )
        if (a.get("dividends_cash") or 0) > 0:
            divs_chip = (
                f"<span style='background:#FEF3C7;color:#B45309;border-radius:3px;"
                f"padding:1px 5px;margin-right:3px;font-size:10px;font-weight:700;"
                f"font-family:Menlo,monospace;display:inline-block;margin-bottom:2px'>"
                f"div ${(a['dividends_cash']):,.0f}</span>"
            )
        activity_html = (sells_chip + buys_chip + divs_chip) or "<span style='color:#9CA3AF'>—</span>"

        start_cell = (
            f"<strong>{shares(a.get('start_shares'))}</strong> sh<br>"
            f"<span style='color:#9CA3AF;font-size:10px'>"
            f"@ ${a.get('jan1_price') or 0:.2f} = {usd0(a.get('start_value'))}</span>"
            if (a.get("start_shares") or 0) > 0
            else "<span style='color:#9CA3AF'>—</span>"
        )
        end_cell = (
            f"<strong>{shares(a.get('end_shares'))}</strong> sh<br>"
            f"<span style='color:#9CA3AF;font-size:10px'>"
            f"@ ${a.get('end_price') or 0:.2f} = {usd0(a.get('end_value'))}</span>"
            if (a.get("end_shares") or 0) > 0
            else "<span style='color:#9CA3AF;font-size:10px'>fully sold</span>"
        )

        gain = a.get("dollar_gain") or 0.0
        contrib = a.get("contribution_pct") or 0.0
        attr_rows_html += (
            f"<tr>"
            f"<td style='padding:7px 8px;border-bottom:1px solid #EEE;font-weight:700'>{a.get('ticker','?')}</td>"
            f"<td style='padding:7px 8px;border-bottom:1px solid #EEE;font-family:Menlo,monospace;font-size:11px'>{start_cell}</td>"
            f"<td style='padding:7px 8px;border-bottom:1px solid #EEE'>{activity_html}</td>"
            f"<td style='padding:7px 8px;border-bottom:1px solid #EEE;font-family:Menlo,monospace;font-size:11px'>{end_cell}</td>"
            f"<td style='padding:7px 8px;border-bottom:1px solid #EEE;text-align:right;color:{color(gain)};font-weight:700;font-family:Menlo,monospace'>"
            f"{('+' if gain >= 0 else '−') + '$' + f'{abs(gain):,.0f}'}</td>"
            f"<td style='padding:7px 8px;border-bottom:1px solid #EEE;text-align:right;color:{color(contrib)};font-weight:800;font-family:Menlo,monospace'>{pct(contrib)}</td>"
            f"</tr>"
        )

    total_gain = sum(a.get("dollar_gain") or 0.0 for a in attribution)
    total_contrib = (total_gain / begin_value * 100.0) if begin_value else 0.0

    # ─────────────────────────────────────────────────────────────────────────
    # Sections that mirror the in-app YTD detail view (chart + bars + table)
    # ─────────────────────────────────────────────────────────────────────────
    detail = detail or {}
    port_series = detail.get("series")     or []
    spy_series  = detail.get("spy_series") or []
    spy_ret     = detail.get("spy_return_pct")
    vs_spy      = detail.get("vs_spy_pct")
    max_dd      = detail.get("max_drawdown_pct") or 0.0
    n_tickers   = detail.get("n_tickers") or len(holdings)
    sorted_h    = detail.get("holdings") or []

    # ── Section: Live Portfolio YTD vs SPY (inline SVG chart) ────────────────
    chart_html = ""
    if port_series and spy_series:
        # SVG dimensions
        W, H = 680, 220
        L, R, T, B = 50, 18, 18, 28  # inner padding
        iw, ih = W - L - R, H - T - B

        # Combined Y range across both series (in % terms)
        all_y = [float(p.get("return_pct") or 0.0) for p in port_series] \
              + [float(p.get("return_pct") or 0.0) for p in spy_series]
        if all_y:
            y_min = min(all_y + [0.0])
            y_max = max(all_y + [0.0])
            if y_max - y_min < 0.5:
                y_max = y_min + 1.0
            y_pad = (y_max - y_min) * 0.10
            y_min -= y_pad
            y_max += y_pad
        else:
            y_min, y_max = -1.0, 1.0

        # X over the longer of the two series (they share dates)
        n_pts = max(len(port_series), len(spy_series))
        if n_pts < 2:
            n_pts = 2

        def to_xy(idx, val, n):
            x = L + (idx / (n - 1)) * iw if n > 1 else L
            y = T + ih - ((float(val) - y_min) / (y_max - y_min)) * ih
            return x, y

        def build_path(series):
            if not series:
                return ""
            pts = []
            n = len(series)
            for i, p in enumerate(series):
                x, y = to_xy(i, p.get("return_pct") or 0.0, n)
                pts.append((x, y))
            d = f"M {pts[0][0]:.1f} {pts[0][1]:.1f}"
            for x, y in pts[1:]:
                d += f" L {x:.1f} {y:.1f}"
            return d

        port_path = build_path(port_series)
        spy_path  = build_path(spy_series)

        # Y-axis grid lines & labels at min, mid, max, and 0% if in range
        ticks = sorted({y_min, (y_min + y_max) / 2, y_max, 0.0 if y_min <= 0 <= y_max else y_min})
        ticks = [t for t in ticks if y_min <= t <= y_max]
        grid_html = ""
        for t in ticks:
            ty = T + ih - ((t - y_min) / (y_max - y_min)) * ih
            grid_html += (
                f'<line x1="{L}" y1="{ty:.1f}" x2="{W-R}" y2="{ty:.1f}" '
                f'stroke="rgba(255,255,255,0.06)" stroke-width="1"/>'
                f'<text x="{L-6}" y="{ty+3:.1f}" fill="rgba(255,255,255,0.45)" '
                f'font-size="9" font-family="Menlo,monospace" text-anchor="end">'
                f'{("+" if t >= 0 else "")}{t:.1f}%</text>'
            )

        # X-axis date labels (start + end)
        first_date = (port_series[0] if port_series else {}).get("date", "")
        last_date  = (port_series[-1] if port_series else {}).get("date", "")
        x_labels = (
            f'<text x="{L}" y="{H-8}" fill="rgba(255,255,255,0.45)" '
            f'font-size="9" font-family="Menlo,monospace">{first_date}</text>'
            f'<text x="{W-R}" y="{H-8}" fill="rgba(255,255,255,0.45)" '
            f'font-size="9" font-family="Menlo,monospace" text-anchor="end">{last_date}</text>'
        )

        port_ytd_str = pct(md_pct)
        spy_ytd_str  = pct(spy_ret)
        vs_str       = pct(vs_spy)

        chart_html = f"""
  <tr><td style="background:#fff;padding:18px 24px;border-top:1px solid #E5E7EB">
    <div style="font-size:11px;font-weight:800;letter-spacing:1.5px;color:#C9A84C;margin-bottom:6px">
      LIVE PORTFOLIO YTD vs SPY YTD
    </div>
    <div style="font-size:11px;color:#6B7280;margin-bottom:10px;font-family:Menlo,monospace">
      Portfolio <strong style="color:{color(md_pct)}">{port_ytd_str}</strong> ·
      SPY <strong style="color:{color(spy_ret)}">{spy_ytd_str}</strong> ·
      vs SPY <strong style="color:{color(vs_spy)}">{vs_str}</strong> ·
      Max DD <strong style="color:#6B7280">−{max_dd:.1f}%</strong>
    </div>
    <div style="background:#0A1628;border-radius:8px;padding:10px;text-align:center">
      <svg width="100%" viewBox="0 0 {W} {H}" preserveAspectRatio="xMidYMid meet" xmlns="http://www.w3.org/2000/svg">
        <rect x="{L}" y="{T}" width="{iw}" height="{ih}" fill="rgba(255,255,255,0.02)" rx="4"/>
        {grid_html}
        <path d="{spy_path}"  fill="none" stroke="#60A5FA" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"/>
        <path d="{port_path}" fill="none" stroke="#C9A84C" stroke-width="2.5" stroke-linecap="round" stroke-linejoin="round"/>
        {x_labels}
        <g transform="translate({L+8},{T+8})">
          <rect x="-2" y="-10" width="124" height="22" fill="rgba(0,0,0,0.4)" rx="4"/>
          <circle cx="6" cy="0" r="3" fill="#C9A84C"/>
          <text x="14" y="3" fill="#fff" font-size="10" font-family="Helvetica,Arial">Live YTD</text>
          <circle cx="64" cy="0" r="3" fill="#60A5FA"/>
          <text x="72" y="3" fill="#fff" font-size="10" font-family="Helvetica,Arial">SPY YTD</text>
        </g>
      </svg>
    </div>
  </td></tr>"""

    # ── Section: Performance Attribution bars (one row per stock) ────────────
    bars_html = ""
    if attribution:
        max_abs = max((abs(a.get("contribution_pct") or 0.0) for a in attribution), default=0.01)
        if max_abs <= 0:
            max_abs = 0.01

        bar_rows = ""
        for a in attribution:
            v   = float(a.get("contribution_pct") or 0.0)
            ret = a.get("ticker_return_pct")
            is_pos = v >= 0
            width_pct = abs(v) / max_abs * 50.0  # each side gets up to 50% of bar track
            bar_color = "#16A34A" if is_pos else "#DC2626"

            # Tornado-style bar: positive grows right of center, negative grows left
            if is_pos:
                bar_inner = (
                    f'<div style="position:absolute;left:50%;top:5px;height:18px;'
                    f'width:{width_pct:.1f}%;background:{bar_color};border-radius:2px;"></div>'
                )
            else:
                bar_inner = (
                    f'<div style="position:absolute;right:50%;top:5px;height:18px;'
                    f'width:{width_pct:.1f}%;background:{bar_color};border-radius:2px;"></div>'
                )

            ret_str = "" if ret is None else f"&nbsp;<span style='color:#9CA3AF;font-size:10px'>(ret {pct(ret, 1)})</span>"

            bar_rows += (
                f"<tr>"
                f"<td style='padding:5px 8px;font-family:Menlo,monospace;font-size:11px;font-weight:700;color:#0A1628;width:60px;white-space:nowrap'>"
                f"{a.get('ticker','?')}</td>"
                f"<td style='padding:5px 4px;width:60%;'>"
                f"<div style='position:relative;height:28px;background:rgba(0,0,0,0.02);border-radius:3px'>"
                f"<div style='position:absolute;left:50%;top:0;width:1px;height:28px;background:rgba(0,0,0,0.15)'></div>"
                f"{bar_inner}"
                f"</div></td>"
                f"<td style='padding:5px 8px;text-align:right;font-family:Menlo,monospace;font-size:11px;font-weight:800;color:{bar_color};white-space:nowrap'>"
                f"{pct(v)}{ret_str}</td>"
                f"</tr>"
            )

        bars_html = f"""
  <tr><td style="background:#fff;padding:18px 24px;border-top:1px solid #E5E7EB">
    <div style="font-size:11px;font-weight:800;letter-spacing:1.5px;color:#C9A84C;margin-bottom:4px">
      PERFORMANCE ATTRIBUTION — BARS
    </div>
    <div style="font-size:10px;color:#9CA3AF;margin-bottom:10px;font-style:italic">
      Each position's contribution to the portfolio total · sorted by absolute impact
    </div>
    <table width="100%" cellpadding="0" cellspacing="0" style="border-collapse:collapse">
      <tbody>{bar_rows}</tbody>
    </table>
  </td></tr>"""

    # ── Section: Holdings sorted by contribution ─────────────────────────────
    holdings_table_html = ""
    if sorted_h:
        rows_html = ""
        for h in sorted_h:
            ent = h.get("entry_price")
            cur = h.get("current_price")
            ret = h.get("return_pct")
            con = h.get("contribution_pct")
            wt  = (h.get("weight") or 0) * 100
            ent_str = f"${ent:.2f}" if ent else "—"
            cur_str = f"${cur:.2f}" if cur else "—"
            tdbase = "padding:7px 8px;border-bottom:1px solid #EEE;text-align:right;font-family:Menlo,monospace;font-size:11px"
            rows_html += (
                f"<tr>"
                f"<td style='padding:7px 8px;border-bottom:1px solid #EEE;font-weight:800;color:#0A1628'>{h.get('ticker','?')}</td>"
                f"<td style='{tdbase}'>{wt:.1f}%</td>"
                f"<td style='{tdbase}'>{ent_str}</td>"
                f"<td style='{tdbase}'>{cur_str}</td>"
                f"<td style='{tdbase};font-weight:700;color:{color(ret)}'>{pct(ret, 1)}</td>"
                f"<td style='padding:7px 8px;border-bottom:1px solid #EEE;text-align:right;font-family:Menlo,monospace;font-size:12px;font-weight:800;color:{color(con)}'>{pct(con)}</td>"
                f"</tr>"
            )

        holdings_table_html = f"""
  <tr><td style="background:#fff;padding:18px 24px;border-top:1px solid #E5E7EB">
    <div style="font-size:11px;font-weight:800;letter-spacing:1.5px;color:#C9A84C;margin-bottom:10px">
      HOLDINGS — SORTED BY CONTRIBUTION
    </div>
    <table width="100%" cellpadding="0" cellspacing="0" style="border-collapse:collapse;font-size:12px">
      <thead>
        <tr style="background:#F4F6F8">
          <th style="padding:8px;text-align:left;font-size:9px;letter-spacing:0.8px;color:#6B7280;border-bottom:1px solid #DCE3EB">TICKER</th>
          <th style="padding:8px;text-align:right;font-size:9px;letter-spacing:0.8px;color:#6B7280;border-bottom:1px solid #DCE3EB">WEIGHT</th>
          <th style="padding:8px;text-align:right;font-size:9px;letter-spacing:0.8px;color:#6B7280;border-bottom:1px solid #DCE3EB">YEAR START</th>
          <th style="padding:8px;text-align:right;font-size:9px;letter-spacing:0.8px;color:#6B7280;border-bottom:1px solid #DCE3EB">CURRENT</th>
          <th style="padding:8px;text-align:right;font-size:9px;letter-spacing:0.8px;color:#6B7280;border-bottom:1px solid #DCE3EB">RETURN</th>
          <th style="padding:8px;text-align:right;font-size:9px;letter-spacing:0.8px;color:#6B7280;border-bottom:1px solid #DCE3EB">CONTRIB</th>
        </tr>
      </thead>
      <tbody>{rows_html}</tbody>
    </table>
  </td></tr>"""

    # ── HTML body ────────────────────────────────────────────────────────────
    html_body = f"""<!DOCTYPE html>
<html><head><meta charset="UTF-8"></head>
<body style="margin:0;padding:0;background:#f4f6f8;font-family:'Helvetica Neue',Arial,sans-serif;color:#0A1628">
<table width="100%" cellpadding="0" cellspacing="0" style="background:#f4f6f8;padding:24px 0">
<tr><td align="center">
<table width="720" cellpadding="0" cellspacing="0" style="max-width:720px;width:100%">

  <!-- Header -->
  <tr><td style="background:#0A1628;padding:18px 24px;border-radius:10px 10px 0 0;color:#fff">
    <table width="100%" cellpadding="0" cellspacing="0">
    <tr>
      <td style="vertical-align:middle">{logo_img_tag}</td>
      <td style="vertical-align:middle;text-align:right">
        <div style="font-size:11px;color:#C9A84C;font-weight:800;letter-spacing:1.5px">LIVE PORTFOLIO YTD REPORT</div>
        <div style="font-size:11px;color:#94A3B8;margin-top:3px">{uploaded_human}</div>
      </td>
    </tr>
    </table>
  </td></tr>

  <!-- Hero metric -->
  <tr><td style="background:#fff;padding:24px;text-align:center;border-bottom:1px solid #E5E7EB">
    <div style="font-size:11px;font-weight:800;letter-spacing:1.2px;color:#9CA3AF">YTD RETURN — CASH-FLOW ADJUSTED (MODIFIED DIETZ)</div>
    <div style="font-size:48px;font-weight:800;font-family:Menlo,monospace;color:{color(md_pct)};margin-top:8px">{pct(md_pct)}</div>
  </td></tr>

  <!-- Stat grid -->
  <tr><td style="background:#fff;padding:18px 24px">
    <table width="100%" cellpadding="0" cellspacing="0">
    <tr>
      <td style="width:25%;padding:10px;border:1px solid #E5E7EB;border-radius:6px;vertical-align:top">
        <div style="font-size:9px;font-weight:800;letter-spacing:1px;color:#9CA3AF">JAN 1</div>
        <div style="font-size:18px;font-weight:800;font-family:Menlo,monospace;margin-top:4px">${begin_value:,.0f}</div>
      </td>
      <td style="width:8px"></td>
      <td style="width:25%;padding:10px;border:1px solid #E5E7EB;border-radius:6px;vertical-align:top">
        <div style="font-size:9px;font-weight:800;letter-spacing:1px;color:#9CA3AF">TODAY</div>
        <div style="font-size:18px;font-weight:800;font-family:Menlo,monospace;margin-top:4px">${end_value:,.0f}</div>
        <div style="font-size:9px;color:#16A34A;margin-top:2px">✓ from positions CSV</div>
      </td>
      <td style="width:8px"></td>
      <td style="width:25%;padding:10px;border:1px solid #E5E7EB;border-radius:6px;vertical-align:top">
        <div style="font-size:9px;font-weight:800;letter-spacing:1px;color:#9CA3AF">NET FLOWS</div>
        <div style="font-size:18px;font-weight:800;font-family:Menlo,monospace;color:{color(net_flow)};margin-top:4px">{usd0(net_flow)}</div>
        <div style="font-size:9px;color:#9CA3AF;margin-top:2px">{snap.get('flow_count', 0)} events</div>
      </td>
      <td style="width:8px"></td>
      <td style="width:25%;padding:10px;border:1px solid #E5E7EB;border-radius:6px;vertical-align:top">
        <div style="font-size:9px;font-weight:800;letter-spacing:1px;color:#9CA3AF">TRADES</div>
        <div style="font-size:18px;font-weight:800;font-family:Menlo,monospace;margin-top:4px">{snap.get('trade_count', 0)}</div>
        <div style="font-size:9px;color:#9CA3AF;margin-top:2px">{snap.get('dividend_count', 0)} dividends</div>
      </td>
    </tr>
    </table>
  </td></tr>

  <!-- Live benchmark positions -->
  <tr><td style="background:#fff;padding:18px 24px;border-top:1px solid #E5E7EB">
    <div style="font-size:11px;font-weight:800;letter-spacing:1.5px;color:#C9A84C;margin-bottom:10px">
      LIVE BENCHMARK · {len(sorted_hold)} POSITIONS
    </div>
    <div>{chips_html}</div>
  </td></tr>

  <!-- YTD by month table (right after positions) -->
{_build_monthly_table_html(snap, pct, color, usd0)}

  {chart_html}
  {bars_html}
  {holdings_table_html}

  <!-- Attribution table (LAST) -->
  <tr><td style="background:#fff;padding:18px 24px;border-top:1px solid #E5E7EB">
    <div style="font-size:11px;font-weight:800;letter-spacing:1.5px;color:#C9A84C;margin-bottom:10px">
      PERFORMANCE ATTRIBUTION — BY HOLDING (TRANSACTION-AWARE)
    </div>
    <table width="100%" cellpadding="0" cellspacing="0" style="border-collapse:collapse;font-size:12px">
      <thead>
        <tr style="background:#F4F6F8">
          <th style="padding:8px;text-align:left;font-size:9px;letter-spacing:0.8px;color:#6B7280;border-bottom:1px solid #DCE3EB">TICKER</th>
          <th style="padding:8px;text-align:left;font-size:9px;letter-spacing:0.8px;color:#6B7280;border-bottom:1px solid #DCE3EB">JAN 1</th>
          <th style="padding:8px;text-align:left;font-size:9px;letter-spacing:0.8px;color:#6B7280;border-bottom:1px solid #DCE3EB">YTD ACTIVITY</th>
          <th style="padding:8px;text-align:left;font-size:9px;letter-spacing:0.8px;color:#6B7280;border-bottom:1px solid #DCE3EB">NOW</th>
          <th style="padding:8px;text-align:right;font-size:9px;letter-spacing:0.8px;color:#6B7280;border-bottom:1px solid #DCE3EB">$ P&amp;L</th>
          <th style="padding:8px;text-align:right;font-size:9px;letter-spacing:0.8px;color:#6B7280;border-bottom:1px solid #DCE3EB">% CONTRIB</th>
        </tr>
      </thead>
      <tbody>{attr_rows_html}</tbody>
      <tfoot>
        <tr style="background:#F4F6F8">
          <td colspan="4" style="padding:10px 8px;text-align:right;font-weight:700;font-size:11px;color:#0A1628">TOTAL</td>
          <td style="padding:10px 8px;text-align:right;color:{color(total_gain)};font-weight:800;font-family:Menlo,monospace">{usd0(total_gain)}</td>
          <td style="padding:10px 8px;text-align:right;color:{color(total_contrib)};font-weight:800;font-family:Menlo,monospace">{pct(total_contrib)}</td>
        </tr>
      </tfoot>
    </table>
  </td></tr>

  <!-- Footer -->
  <tr><td style="background:#0A1628;padding:14px 24px;border-radius:0 0 10px 10px;text-align:center;color:#94A3B8;font-size:10px">
    DGA Capital · Generated by the Research Analyst · Snapshot ID {snap.get('id','')}
  </td></tr>

</table></td></tr></table></body></html>"""

    msg = EmailMessage()
    msg["Subject"] = subject
    msg["From"]    = from_addr
    msg["To"]      = to_email
    msg.set_content(plain_body)
    msg.add_alternative(html_body, subtype="html")
    return msg


def promote_live_portfolio(holdings_input: list[dict]) -> dict:
    """Persist *holdings_input* as the auto-promoted live portfolio.

    Called automatically after a successful portfolio rebalance. Captures
    today's close as anchor_price for each ticker.

    holdings_input items: {"ticker": str, "weight": float (0..1 or 0..100)}
    """
    cleaned: list[dict] = []
    today = _today_str()
    for h in holdings_input or []:
        t = (h.get("ticker") or h.get("Ticker") or "").strip().upper()
        if not t:
            continue
        w = h.get("weight") if h.get("weight") is not None else h.get("Weight")
        try:
            w = float(w)
        except Exception:  # noqa: BLE001
            continue
        if w > 1.5:
            w = w / 100.0
        if w <= 0:
            continue
        anchor = _fetch_close_price(t)
        if not anchor or anchor <= 0:
            continue
        cleaned.append({
            "ticker":       t,
            "weight":       round(w, 6),
            "anchor_price": round(anchor, 4),
        })
    if not cleaned:
        return {"ok": False, "error": "No valid tickers with prices."}

    # ── Store raw weights exactly as in the input CSV ─────────────────────────
    # Do NOT normalize here — the user's weights already reflect their actual
    # portfolio fractions (e.g., INTC 7.88%).  Some tickers may be excluded
    # because yfinance can't price them (preferred shares, etc.), but that is
    # NOT a reason to inflate the remaining weights.  Calculations that need
    # weights to sum to 1.0 (e.g., _portfolio_breakdown) normalize internally.
    #
    # CRITICAL: when there's already a live_portfolio with prior account_history
    # and ytd_snapshots, we MUST preserve those — they represent the user's
    # uploaded YTD runs and should never be silently wiped by a fresh rebalance.
    # Only `holdings`, `anchor_date`, `uploaded_at`, and `series` are replaced.
    with _tracker_lock():
        state = _load_tracker_state()
        existing = state.get("live_portfolio") or {}
        live = dict(existing)   # shallow copy so we keep every existing key
        live["uploaded_at"] = datetime.utcnow().isoformat()
        live["anchor_date"] = today
        live["holdings"]    = cleaned
        live["series"]      = [{"date": today, "value": 100.0, "return_pct": 0.0}]
        # account_history and ytd_snapshots flow through untouched if they exist
        # (existing.get("account_history") / existing.get("ytd_snapshots"))
        state["live_portfolio"] = live
        _ensure_spy_back_to(state, today)
        _save_tracker_state(state)

    preserved = []
    if existing.get("account_history"): preserved.append("account_history")
    if existing.get("ytd_snapshots"):   preserved.append(f"{len(existing['ytd_snapshots'])} YTD snapshots")
    suffix = f" (preserved {', '.join(preserved)})" if preserved else ""
    print(f"📊 Live portfolio promoted ({len(cleaned)} tickers, anchored {today}){suffix}")
    return {"ok": True, "live_portfolio": live}


def take_daily_snapshot(force: bool = False) -> dict:
    """Append today's closing values to every tracking portfolio + SPY + live.

    Idempotent: if today's snapshot already exists for a portfolio, skip it
    unless *force=True*. Safe to call multiple times per day.
    """
    if not force and not _is_after_market_close():
        return {"ok": True, "skipped": True, "reason": "before market close"}

    today = _today_str()
    updated_portfolios = 0

    with _tracker_lock():
        state = _load_tracker_state()
        # 1. SPY series — append today's close if not present
        spy_series = state.get("spy_series") or []
        if not any(r["date"] == today for r in spy_series):
            spy_close = _fetch_close_price("SPY")
            if spy_close:
                spy_series.append({"date": today, "close": round(spy_close, 4)})
                spy_series.sort(key=lambda r: r["date"])
                state["spy_series"] = spy_series

        # 2. Each tracking portfolio
        for p in state.get("portfolios", []):
            if p.get("status") != "tracking":
                continue
            series = p.get("series", []) or []
            if any(r["date"] == today for r in series):
                continue
            value = _portfolio_value_today(p.get("holdings", []))
            series.append({
                "date":       today,
                "value":      value,
                "return_pct": round(value - 100.0, 4),
            })
            p["series"] = sorted(series, key=lambda r: r["date"])
            updated_portfolios += 1

        # 3. Live portfolio
        live = state.get("live_portfolio")
        if live:
            live_series = live.get("series", []) or []
            if not any(r["date"] == today for r in live_series):
                value = _portfolio_value_today(live.get("holdings", []))
                live_series.append({
                    "date":       today,
                    "value":      value,
                    "return_pct": round(value - 100.0, 4),
                })
                live["series"] = sorted(live_series, key=lambda r: r["date"])

        _save_tracker_state(state)

    print(f"📸 Daily snapshot: {updated_portfolios} portfolio(s) updated for {today}")
    return {"ok": True, "date": today, "portfolios_updated": updated_portfolios}


def _start_tracker_snapshot_worker() -> None:
    """Background thread: take a daily snapshot once after market close."""
    import threading as _threading

    def loop():
        last_snapshot_date = None
        while True:
            try:
                today = _today_str()
                if last_snapshot_date != today and _is_after_market_close():
                    take_daily_snapshot()
                    last_snapshot_date = today
            except Exception as exc:  # noqa: BLE001
                print(f"⚠️  Snapshot worker error: {exc}")
            time.sleep(3600)  # check hourly

    _threading.Thread(target=loop, daemon=True).start()


# ============================================================================
# Analyst ratings — Yahoo Finance (primary) with GuruFocus fallback
# ============================================================================

# Canonical target-firm labels mapped to name-fragments we match against
# whatever the upstream provider returns. Order defines table-row order.
# These firms are deliberately chosen because they are syndicated publicly by
# Yahoo's upgrade/downgrade feed — Fidelity and Morningstar aren't (their
# ratings are subscription-only and not in any free feed).
_TARGET_FIRMS: list[tuple[str, list[str]]] = [
    ("Goldman Sachs",   ["goldman sachs", "goldman"]),
    ("Morgan Stanley",  ["morgan stanley"]),
    ("BofA Securities", ["b of a", "bofa", "bank of america", "merrill"]),
    ("JPMorgan",        ["jpmorgan", "j.p. morgan", "jp morgan"]),
    ("Wells Fargo",     ["wells fargo"]),
    ("Jefferies",       ["jefferies"]),
    ("Citi",            ["citigroup", "citi ", "citi"]),
    ("Barclays",        ["barclays"]),
    ("UBS",             ["ubs"]),
    ("Evercore ISI",    ["evercore"]),
]


def _fetch_yahoo_analyst_ratings(ticker: str) -> dict | None:
    """Pull per-firm upgrades/downgrades + aggregate consensus from Yahoo.

    Returns a dict with {"firms": [...], "consensus": {...}} or None if
    Yahoo is unreachable or yfinance is not installed. Never raises.
    """
    try:
        import yfinance as yf  # type: ignore
    except Exception as exc:  # noqa: BLE001
        print(f"   ⚠️  yfinance not available: {exc}")
        return None

    try:
        t = yf.Ticker(ticker)
    except Exception as exc:  # noqa: BLE001
        print(f"   ⚠️  yfinance Ticker() failed for {ticker}: {exc}")
        return None

    # --- Per-firm upgrades/downgrades DataFrame (already sorted newest first).
    firm_rows: list[dict] = []
    try:
        ud = t.upgrades_downgrades
    except Exception as exc:  # noqa: BLE001
        print(f"   ⚠️  yfinance upgrades_downgrades failed for {ticker}: {exc}")
        ud = None

    # Ratings older than 6 months are stale — ignore them.
    _cutoff_date = (datetime.utcnow() - timedelta(days=183)).strftime("%Y-%m-%d")

    if ud is not None and hasattr(ud, "empty") and not ud.empty:
        try:
            df = ud.reset_index() if "GradeDate" not in ud.columns else ud.copy()
        except Exception:  # noqa: BLE001
            df = None
        if df is not None and "Firm" in df.columns:
            for label, fragments in _TARGET_FIRMS:
                try:
                    mask = df["Firm"].astype(str).str.lower().apply(
                        lambda s, frags=fragments: any(f in s for f in frags)
                    )
                    sub = df[mask]
                    if len(sub) == 0:
                        continue
                    row = sub.iloc[0]
                    date_val = row.get("GradeDate", "")
                    try:
                        import pandas as pd  # local import; already project dep
                        if isinstance(date_val, pd.Timestamp):
                            date_val = date_val.strftime("%Y-%m-%d")
                    except Exception:
                        pass
                    date_str = str(date_val)[:10]
                    # Skip ratings older than 6 months — no longer actionable.
                    if date_str and date_str < _cutoff_date:
                        print(f"   ⏭️  Skipping stale {label} rating from {date_str} (>6 months old)")
                        continue
                    target_val = row.get("currentPriceTarget", 0) or 0
                    try:
                        target_val = float(target_val)
                    except (TypeError, ValueError):
                        target_val = 0.0
                    firm_rows.append({
                        "firm":   label,
                        "rating": str(row.get("ToGrade", "") or "—").strip() or "—",
                        "target": target_val,
                        "date":   date_str,
                        "action": str(row.get("Action", "") or "—").strip() or "—",
                    })
                except Exception as exc:  # noqa: BLE001
                    print(f"   ⚠️  Yahoo firm parse failed ({label}): {exc}")
                    continue

    # --- Aggregate consensus (targetMeanPrice, #analysts, recommendationKey).
    consensus: dict = {}
    try:
        info = t.info if isinstance(t.info, dict) else {}
    except Exception as exc:  # noqa: BLE001
        print(f"   ⚠️  yfinance info failed for {ticker}: {exc}")
        info = {}
    for k in ("targetMeanPrice", "targetHighPrice", "targetLowPrice",
              "targetMedianPrice", "numberOfAnalystOpinions",
              "recommendationKey", "recommendationMean", "currentPrice"):
        v = info.get(k)
        if v is not None:
            consensus[k] = v

    if not firm_rows and not consensus:
        return None
    return {"firms": firm_rows, "consensus": consensus}


def _fetch_gurufocus_analyst_ratings(ticker: str) -> list[dict]:
    """Legacy fallback — returns a list of normalized firm dicts (may be empty)."""
    token = _optional_env("GURUFOCUS_TOKEN")
    if not token:
        return []

    records: list = []
    try:
        url = (
            f"https://api.gurufocus.com/public/user/{token}/stock/{ticker}/upgrades_downgrades"
        )
        resp = requests.get(
            url,
            headers={"User-Agent": "DGA Research Analyst"},
            timeout=20,
        )
        if resp.status_code == 200:
            raw = resp.json()
            if isinstance(raw, list):
                records = raw
            elif isinstance(raw, dict):
                for key in ("upgrades_downgrades", "data", "result", "results"):
                    if isinstance(raw.get(key), list):
                        records = raw[key]
                        break
        else:
            print(f"   ⚠️  GuruFocus API returned {resp.status_code} for {ticker}")
    except Exception as exc:  # noqa: BLE001
        print(f"   ⚠️  GuruFocus API error for {ticker}: {exc}")

    def _f(rec: dict, *keys: str) -> str:
        for k in keys:
            v = rec.get(k)
            if v is not None and str(v).strip() not in ("", "None", "null"):
                return str(v).strip()
        return ""

    rows: list[dict] = []
    seen: set[str] = set()
    for rec in records:
        if not isinstance(rec, dict):
            continue
        firm_raw = _f(rec, "analyst", "analyst_firm", "firm", "firm_name",
                      "company", "broker", "institution").lower()
        for label, fragments in _TARGET_FIRMS:
            if label in seen:
                continue
            if any(f in firm_raw for f in fragments):
                target_raw = _f(rec, "price_target", "new_target", "target_price",
                                "pt", "new_price_target", "adj_price_target")
                try:
                    target_val = float(target_raw.replace("$", "").replace(",", ""))
                except (ValueError, AttributeError):
                    target_val = 0.0
                rows.append({
                    "firm":   label,
                    "rating": _f(rec, "current_rating", "new_rating", "rating",
                                 "recommendation", "action_pt", "adj_pt_rating") or "—",
                    "target": target_val,
                    "date":   (_f(rec, "date", "action_date", "updated_date",
                                  "created_at") or "—")[:10],
                    "action": _f(rec, "action", "type", "action_type",
                                 "change_type", "event") or "—",
                })
                seen.add(label)
                break
    return rows


def fetch_analyst_ratings(ticker: str) -> str:
    """Return a pre-formatted markdown block of analyst ratings for Section 7.5.

    Source priority:
      1. Yahoo Finance via yfinance (free, no API key, rich data)
      2. GuruFocus (legacy, requires GURUFOCUS_TOKEN)
      3. Empty string → Grok uses training-data fallback (PATH B)

    Always returns a block with the full firm table (rows for firms we
    couldn't get are marked "Not available") when ANY source succeeded,
    so Grok renders the section deterministically.
    """
    firm_rows: list[dict] = []
    consensus: dict = {}
    source_label = ""

    # --- Primary: Yahoo Finance
    yahoo = _fetch_yahoo_analyst_ratings(ticker)
    if yahoo:
        firm_rows = yahoo.get("firms") or []
        consensus = yahoo.get("consensus") or {}
        if firm_rows or consensus:
            source_label = "Yahoo Finance"
            print(f"   📊 Yahoo: {len(firm_rows)} firm ratings + consensus for {ticker}")

    # --- Fallback: GuruFocus (only if Yahoo yielded nothing)
    if not firm_rows and not consensus:
        gf_rows = _fetch_gurufocus_analyst_ratings(ticker)
        if gf_rows:
            firm_rows = gf_rows
            source_label = "GuruFocus"
            print(f"   📊 GuruFocus: {len(firm_rows)} firm ratings for {ticker}")

    # --- Nothing worked: signal Grok to use training-data fallback.
    if not firm_rows and not consensus:
        return ""

    # --- Build the markdown block. Keep row order = _TARGET_FIRMS order.
    by_firm = {r["firm"]: r for r in firm_rows}
    cutoff_display = (datetime.utcnow() - timedelta(days=183)).strftime("%Y-%m-%d")
    lines = [
        f"ANALYST_RATINGS_BLOCK (source: {source_label} — ratings within last 6 months only, cutoff {cutoff_display} — use these exact values in Section 7.5):",
        "| Firm | Rating | 12M Price Target | Date | Action |",
        "|------|--------|-----------------|------|--------|",
    ]
    for label, _frags in _TARGET_FIRMS:
        r = by_firm.get(label)
        if r:
            tgt = r.get("target") or 0
            tgt_fmt = f"${tgt:,.2f}" if isinstance(tgt, (int, float)) and tgt > 0 else "—"
            lines.append(
                f"| {label} | {r.get('rating','—')} | {tgt_fmt} | "
                f"{r.get('date','—')} | {r.get('action','—')} |"
            )
        else:
            lines.append(f"| {label} | Not available | — | — | — |")

    # --- Aggregate consensus block (richer signal for Grok).
    if consensus:
        lines.append("")
        lines.append("CONSENSUS_SUMMARY (Yahoo Finance aggregate across ALL covering analysts):")
        n = consensus.get("numberOfAnalystOpinions")
        if n:
            lines.append(f"- Analysts covering: {n}")
        rk = consensus.get("recommendationKey")
        rm = consensus.get("recommendationMean")
        if rk:
            rm_txt = f" (mean {rm:.2f}/5, 1=Strong Buy)" if isinstance(rm, (int, float)) else ""
            lines.append(f"- Consensus rating: {rk.upper()}{rm_txt}")
        tm = consensus.get("targetMeanPrice")
        th = consensus.get("targetHighPrice")
        tl = consensus.get("targetLowPrice")
        tmed = consensus.get("targetMedianPrice")
        if tm:
            parts = [f"mean ${tm:,.2f}"]
            if tmed: parts.append(f"median ${tmed:,.2f}")
            if th:   parts.append(f"high ${th:,.2f}")
            if tl:   parts.append(f"low ${tl:,.2f}")
            lines.append(f"- 12M price target: {', '.join(parts)}")

    return "\n".join(lines)


# ============================================================================
# Grok (xAI) call
# ============================================================================
def _client():
    # Imported lazily so the module can be loaded without the openai package
    # (e.g. for parser/word-render unit tests) — only needed at call time.
    from openai import OpenAI  # type: ignore
    return OpenAI(api_key=get_grok_api_key(), base_url="https://api.x.ai/v1")


def _extract_responses_text(resp) -> str:
    """Extract the assistant text from an xAI Responses API reply.

    Shape of the reply (OpenAI-compatible Responses API):
        resp.output = [ { "type": "message", "content": [ {"type":"output_text","text":"..."} ] } ]
    We also accept ``resp.output_text`` (SDK convenience) when present.
    """
    # Convenience aggregate that newer SDK builds expose.
    text = getattr(resp, "output_text", None)
    if text:
        return text

    output = getattr(resp, "output", None) or []
    chunks: list[str] = []
    for item in output:
        # Pydantic model -> dict coercion for either shape.
        if hasattr(item, "model_dump"):
            item = item.model_dump()
        if not isinstance(item, dict):
            continue
        if item.get("type") == "message":
            for c in item.get("content") or []:
                if hasattr(c, "model_dump"):
                    c = c.model_dump()
                if isinstance(c, dict) and c.get("type") in ("output_text", "text"):
                    t = c.get("text") or ""
                    if t:
                        chunks.append(t)
    return "\n".join(chunks)


def call_grok(system_prompt: str, user_content: str,
              model: str = GROK_MODEL,
              *,
              live_search: bool = False,
              search_from_date: str | None = None) -> str:
    """Call xAI Grok.

    When ``live_search=True`` we use xAI's Responses API with the server-side
    ``web_search`` and ``x_search`` tools so Grok can pull fresh news, X
    (Twitter) posts, and web results while writing the report. This is how
    Section 2 (Recent Developments) surfaces breaking news like a CEO
    stepping down two days ago — past the model's training cutoff.

    When ``live_search=False`` (default) we use plain chat.completions, which
    is faster and cheaper for non-time-sensitive calls (e.g. the portfolio
    roll-up where every per-ticker report already has the news baked in).

    If the Responses API path trips for any reason (old SDK, API drift,
    quota), we fall back to chat.completions so a bad parameter never breaks
    the pipeline.
    """
    client = _client()

    if live_search:
        # xAI Responses API with built-in web + X search tools.
        # The model decides when to call the tools; the server runs them
        # and folds the results back in before returning the final message.
        try:
            resp = client.responses.create(
                model=model,
                input=[
                    {"role": "system", "content": system_prompt},
                    {"role": "user", "content": user_content},
                ],
                tools=[
                    {"type": "web_search"},
                    {"type": "x_search"},
                ],
            )
            text = _extract_responses_text(resp)
            if text:
                return text
            # Empty text is treated as a soft failure — fall through.
            print("   ⚠️  Grok live-search returned empty text; retrying without search.")
        except Exception as exc:  # noqa: BLE001
            print(f"   ⚠️  Grok live-search call failed ({exc}); retrying without search.")

    # Fallback / default path.
    resp = client.chat.completions.create(
        model=model,
        messages=[
            {"role": "system", "content": system_prompt},
            {"role": "user", "content": user_content},
        ],
    )
    return resp.choices[0].message.content or ""


# ============================================================================
# Ratings / price-target extraction for portfolio ranking
# ============================================================================
_RATING_RE = re.compile(
    r"\b(Strong Buy|Buy|Hold|Sell|Strong Sell)\b", re.IGNORECASE
)
_PRICE_TARGET_RE = re.compile(
    r"price target[^\$]{0,60}\$([0-9]{1,4}(?:,[0-9]{3})*(?:\.[0-9]+)?)", re.IGNORECASE,
)
# Stronger anchors that reliably appear in DGA reports. Handles bold markdown,
# colons, dashes, unicode punctuation, and multiple section-title variations.
_PT_STRONG_RE = re.compile(
    r"(?:12[-\s–]?Month\s+Price\s+Target|Base\s+Case\s+Price\s+Target|"
    r"12[-\s–]?month\s+target|Price\s+Target|Fair\s+Value|Target\s+Price|"
    r"Intrinsic\s+Value)"
    r"[\*\s:]{0,20}"  # allow "**: " etc between label and number
    r"\$\s*([0-9]{1,5}(?:,[0-9]{3})*(?:\.[0-9]+)?)",
    re.IGNORECASE,
)
_PT_TABLE_RE = re.compile(
    # e.g. "| **12M Price Target** | $38.46 |" from the Price Target Derivation table.
    r"\|\s*\**\s*12[Mm]?\s+Price\s+Target\s*\**\s*\|[^\|]*\|[^\|]*\|\s*\**\s*"
    r"\$\s*([0-9]{1,5}(?:,[0-9]{3})*(?:\.[0-9]+)?)"
)
_SECTOR_REPORT_RE = re.compile(
    r"\*{0,2}\s*Sector\s*:?\s*\*{0,2}\s*([A-Z][A-Za-z &/\-]+?)(?:\s*\*{0,2}\s*(?:\||\n|Industry))",
    re.IGNORECASE,
)
_CURRENT_PRICE_RE = re.compile(
    r"(?:Current\s+Price|CURRENT_PRICE|Current\s+market\s+price)"
    r"[^\$\d-]{0,15}\$?\s*([0-9]{1,4}(?:,[0-9]{3})*(?:\.[0-9]+)?)",
    re.IGNORECASE,
)
_UPSIDE_RE = re.compile(
    r"(?:Implied\s+Return|Upside|Expected\s+Return|Implied\s+Upside)"
    r"[^\d\-\+]{0,15}([+-]?\d+(?:\.\d+)?)\s*%",
    re.IGNORECASE,
)
_RATING_ANCHORED_RE = re.compile(
    r"(?:"
    r"(?:Overall\s+[Rr]ating|Final\s+Recommendation|Recommendation|Conviction\s+Level)"
    r"[:\*\s]{0,8}"
    r"|We\s+rate\s+[A-Z.]+\s+"
    r"|[Rr]ating[:\s]+\**\s*"
    r")\**\s*(Strong Buy|Buy|Hold|Sell|Strong Sell)\**",
    re.IGNORECASE,
)


def extract_summary_from_report(report_text: str) -> dict:
    """Pull rating + 12-month price target + current price + upside + thesis.

    Uses strong anchors when present (the DGA report template is consistent),
    falls back to looser matching for older reports.
    """
    # --- Rating: prefer anchored matches (Overall Rating / We rate XXX BUY).
    rating = None
    m = _RATING_ANCHORED_RE.search(report_text[:6000])
    if m:
        rating = m.group(1).title()
    else:
        m2 = _RATING_RE.search(report_text[:4000])
        if m2:
            rating = m2.group(1).title()

    # --- 12M Price Target.
    price_target = None
    m = _PT_STRONG_RE.search(report_text)
    if not m:
        m = _PT_TABLE_RE.search(report_text)
    if not m:
        m = _PRICE_TARGET_RE.search(report_text)
    if m:
        try:
            price_target = float(m.group(1).replace(",", ""))
        except ValueError:
            pass

    # --- Sector (from the report header table).
    sector = None
    sm = _SECTOR_REPORT_RE.search(report_text[:4000])
    if sm:
        sector = sm.group(1).strip(" *|\t")

    # --- Current Price.
    current_price = None
    m = _CURRENT_PRICE_RE.search(report_text)
    if m:
        try:
            current_price = float(m.group(1).replace(",", ""))
        except ValueError:
            pass

    # --- Upside / Implied Return.
    upside_pct = None
    m = _UPSIDE_RE.search(report_text)
    if m:
        try:
            upside_pct = float(m.group(1))
        except ValueError:
            pass
    if upside_pct is None and price_target and current_price:
        try:
            upside_pct = round((price_target - current_price) / current_price * 100, 2)
        except ZeroDivisionError:
            pass

    # Thesis hint.
    thesis = ""
    for line in report_text.split("\n"):
        s = line.strip()
        if not s or s.startswith("#") or s.startswith("|") or s.startswith(">"):
            continue
        thesis = s[:400]
        break

    return {
        "rating": rating,
        "price_target": price_target,
        "current_price": current_price,
        "upside_pct": upside_pct,
        "sector": sector,
        "thesis": thesis,
    }


# ============================================================================
# Gamma.app integration
# ============================================================================
def _load_gamma_index() -> dict:
    """Load the on-disk Gamma metadata index (ticker → { gamma_url, generated_at })."""
    try:
        if GAMMA_INDEX_FILE.exists():
            return json.loads(GAMMA_INDEX_FILE.read_text())
    except Exception:
        pass
    return {}


def _save_gamma_index(idx: dict) -> None:
    """Persist the Gamma index to disk (best-effort)."""
    try:
        GAMMA_INDEX_FILE.write_text(json.dumps(idx, indent=2, sort_keys=True))
    except Exception as exc:  # noqa: BLE001
        print(f"   ⚠️  Could not persist gamma index: {exc}")


def _record_gamma(key: str, gamma_url: str | None, *, pptx_filename: str | None = None,
                  credits: int | None = None) -> None:
    """Record a Gamma generation in the index (key is usually the ticker, or
    'PORTFOLIO' for the portfolio summary)."""
    if not gamma_url:
        return
    idx = _load_gamma_index()
    idx[key.upper()] = {
        "gamma_url": gamma_url,
        "generated_at": datetime.utcnow().isoformat(),
        "pptx_filename": pptx_filename,
        "credits": credits,
    }
    _save_gamma_index(idx)


def _existing_fresh_gamma(key: str) -> dict | None:
    """Return the existing Gamma entry for `key` IF it's < GAMMA_FRESH_DAYS old
    AND the local PPTX file still exists.  Used to skip duplicate generation
    on rapid re-runs of the same ticker."""
    idx = _load_gamma_index()
    entry = idx.get(key.upper())
    if not entry:
        return None
    try:
        gen_at = datetime.fromisoformat(entry.get("generated_at", ""))
    except Exception:
        return None
    age_days = (datetime.utcnow() - gen_at).total_seconds() / 86400.0
    if age_days > GAMMA_FRESH_DAYS:
        return None
    pptx_name = entry.get("pptx_filename") or ""
    if pptx_name and not (STOCKS_FOLDER / pptx_name).exists():
        return None  # Stale index entry — local PPTX is gone
    return entry


def _purge_stale_local_gamma(key: str) -> None:
    """If the existing local PPTX for `key` is older than GAMMA_FRESH_DAYS,
    delete it so the next generation can replace it cleanly.  Also removes
    the matching index entry."""
    idx = _load_gamma_index()
    entry = idx.get(key.upper())
    if not entry:
        return
    try:
        gen_at = datetime.fromisoformat(entry.get("generated_at", ""))
    except Exception:
        return
    age_days = (datetime.utcnow() - gen_at).total_seconds() / 86400.0
    if age_days <= GAMMA_FRESH_DAYS:
        return  # still fresh
    pptx_name = entry.get("pptx_filename") or ""
    if pptx_name:
        old_pptx = STOCKS_FOLDER / pptx_name
        try:
            if old_pptx.exists():
                old_pptx.unlink()
                print(f"   🗑  Removed stale Gamma PPTX (>{GAMMA_FRESH_DAYS}d): {pptx_name}")
        except Exception:
            pass
    idx.pop(key.upper(), None)
    _save_gamma_index(idx)


def _gamma_design_block() -> str:
    return """IMPORTANT DESIGN RULES (enforce strictly):
- Branding: DGA CAPITAL. Place a bold "DGA CAPITAL" wordmark in gold
  (#C9A84C) on a deep navy (#0A1628) background on the title card and as
  a small header on every subsequent card. Use "DGA Capital Research" as
  the recurring footer line on every slide.
- Color palette: navy (#0A1628 / #132040), gold (#C9A84C / #D9BE6E),
  off-white (#F5F7FA), dark gray (#3D4A5C). Use gold sparingly as accent.
- Clean, modern corporate-finance theme (Chisel or equivalent).
- ALL TEXT minimum 12pt (titles 28–32pt, headings 20–24pt, body 14–18pt).
- CHARTS: solid fills, bold borders, high contrast, ample white space.
  • Pie / donut charts: show percentage labels ON each segment (e.g. "45%").
  • Legends 12pt+.
- Tables: large numbers, professional black borders.
- Must look like a top-tier Goldman Sachs research deck.
"""


def create_gamma_for_stock(report_text: str, ticker: str, latest_filing_type: str,
                           out_pptx: Path | None = None) -> tuple[str | None, int]:
    print(f"   📤 Gamma: generating presentation for {ticker}…")
    include_qtr = latest_filing_type == "10-Q"
    num_cards = 19 if include_qtr else 18
    qtr_card = (
        "8. Latest Quarterly Financial Deep Dive – latest Q vs prior-year Q (include both the quarterly and YTD comparison tables with YoY % changes)\n"
        if include_qtr
        else ""
    )
    shortened = report_text[:18000]
    title = f"DGA Capital Research — {ticker} | {datetime.now().strftime('%B %d, %Y')}"
    input_text = f"""Create a professional institutional-quality investment research **PRESENTATION** for {ticker}.

Title: "{title}"

{_gamma_design_block()}

Create at least {num_cards} cards structured exactly like this:
1. Title + Key Thesis
2. Executive Summary + Rating + Price Target
3. Bull Case
4. Business Overview
5. Revenue Breakdown (by segment + by geography — use clean donut charts with % labels)
6. Competitive Moat
7. Financial Deep Dive + Key Metrics Table (Annual + TTM)
{qtr_card}8. Income Statement & Margins
9. Balance Sheet & Cash Flow
10. Growth Analysis
11. 5-Year Projections
12. DCF Valuation
13. Comps Valuation
14. Historical Valuation + Bull/Base/Bear Targets
15. Risk Analysis
16. Latest Developments
17. Catalyst Calendar
18. Final Verdict & Recommendation

Use the full report content below. Every chart, legend and table must be large and professional.

{shortened}
"""
    return _gamma_generate(input_text, num_cards=num_cards, out_pptx=out_pptx)


def create_gamma_portfolio_summary(summary_markdown: str, ranked_rows: list[dict],
                                   out_pptx: Path | None = None) -> tuple[str | None, int]:
    print("   📤 Gamma: generating portfolio summary deck…")
    title = f"DGA Capital Research — Portfolio Summary | {datetime.now().strftime('%B %d, %Y')}"
    def _action(r: dict) -> str:
        return str(r.get("action", "")).strip().upper()
    top_trim = [r for r in ranked_rows if _action(r) in ("TRIM", "SELL", "STRONG SELL")]
    top_buy = [r for r in ranked_rows if _action(r) in ("ADD", "BUY", "STRONG BUY")]
    header = (
        f"Portfolio under review: {len(ranked_rows)} positions.\n"
        f"Candidates to TRIM / SELL: {', '.join(r['ticker'] for r in top_trim) or 'None'}\n"
        f"Candidates to ADD / BUY: {', '.join(r['ticker'] for r in top_buy) or 'None'}\n"
    )
    input_text = f"""Create a professional **PORTFOLIO-LEVEL** research deck.

Title: "{title}"

{_gamma_design_block()}

{header}

Include the following cards (in order):
1. Title + portfolio snapshot
2. Summary ranking — all tickers with rating, price target, upside %, action
3. TOP TRIM / SELL candidates — why each, with catalysts
4. TOP ADD / BUY candidates — why each, with catalysts
5. Valuation comparison table across all names (P/E, EV/EBITDA, FCF yield, upside)
6. Sector / concentration risk
7. Rebalancing action plan (concrete recommended moves)
8. Key catalysts to watch across the portfolio (next 12 months)

Use the portfolio write-up below verbatim where applicable.

{summary_markdown}
"""
    return _gamma_generate(input_text, num_cards=8, out_pptx=out_pptx)


def _gamma_generate(input_text: str, num_cards: int,
                    out_pptx: Path | None = None) -> tuple[str | None, int]:
    """Generate a Gamma deck. Raises RuntimeError with the actual API
    response on any non-2xx so the caller (and the mobile UI) can surface
    the real reason — invalid key, exhausted credits, bad folder id, etc.
    """
    headers = {"Content-Type": "application/json", "X-API-KEY": get_gamma_api_key()}
    payload = {
        "inputText": input_text,
        "textMode": "generate",
        "format": "presentation",
        "numCards": max(8, num_cards),
        "exportAs": "pptx",
        "folderIds": [GAMMA_FOLDER_ID] if GAMMA_FOLDER_ID else None,
    }
    try:
        resp = requests.post(
            "https://public-api.gamma.app/v1.0/generations",
            json=payload,
            headers=headers,
            timeout=60,
        )
    except requests.RequestException as exc:
        raise RuntimeError(f"Gamma POST failed (network): {exc}") from exc

    if resp.status_code not in (200, 201):
        body = (resp.text or "")[:500]
        print(f"   ❌ Gamma error {resp.status_code}: {body}")
        # Detect "out of credits" specifically — Gamma returns 402 or a body
        # containing 'credit' / 'insufficient' when the workspace is empty.
        body_lower = body.lower()
        if (resp.status_code == 402
                or "credit" in body_lower
                or "insufficient" in body_lower
                or "quota" in body_lower
                or "billing" in body_lower):
            raise RuntimeError(
                "Gamma is out of credits. Top up at https://gamma.app/account → "
                "Plan & Billing, then re-run the analysis."
            )
        if resp.status_code in (401, 403):
            raise RuntimeError(
                "Gamma API key was rejected. Check GAMMA_API_KEY in Railway → "
                "Variables (get a fresh key at https://gamma.app/account)."
            )
        # Surface the API's actual error message so the mobile UI shows
        # something actionable for any other error.
        raise RuntimeError(f"Gamma API {resp.status_code}: {body}")

    gen_id = resp.json().get("generationId")
    print(f"   ✅ Gamma generation started ({gen_id})")

    for attempt in range(200):
        time.sleep(6)
        try:
            status = requests.get(
                f"https://public-api.gamma.app/v1.0/generations/{gen_id}",
                headers=headers,
                timeout=30,
            ).json()
        except Exception as exc:  # noqa: BLE001
            print(f"   ⚠️  Gamma status poll error: {exc}")
            continue
        st = status.get("status")
        if st == "completed":
            gamma_url = status.get("gammaUrl")
            export_url = status.get("exportUrl")
            credits = status.get("credits", {})
            used = credits.get("deducted", 0)
            remaining = credits.get("remaining", "?")
            print(f"   ✅ PPTX ready: {gamma_url}  (credits used: {used}, remaining: {remaining})")
            if export_url and out_pptx is not None:
                try:
                    r = requests.get(export_url, stream=True, timeout=60)
                    with open(out_pptx, "wb") as fh:
                        for chunk in r.iter_content(8192):
                            fh.write(chunk)
                    print(f"   💾 Saved {out_pptx}")
                except Exception as exc:  # noqa: BLE001
                    print(f"   ⚠️  Could not save PPTX: {exc}")
            return gamma_url, used
        if st == "failed":
            err_msg = str(status.get("error") or status.get("message") or "unknown")
            print(f"   ❌ Gamma generation failed: {err_msg}")
            err_lower = err_msg.lower()
            if "credit" in err_lower or "insufficient" in err_lower or "billing" in err_lower:
                raise RuntimeError(
                    "Gamma is out of credits. Top up at https://gamma.app/account → "
                    "Plan & Billing, then re-run the analysis."
                )
            raise RuntimeError(f"Gamma generation failed: {err_msg}")
        if attempt % 10 == 0:
            print(f"   ⏳ Gamma still generating… ({attempt+1}/200)  status={st}")
    print("   ❌ Gamma timeout")
    raise RuntimeError("Gamma generation timeout (>20 min)")


# ============================================================================
# Per-ticker analysis pipeline
# ============================================================================
def analyze_ticker(ticker: str, *, system_prompt: str, generate_gamma: bool,
                   verbose: bool = True, reuse_existing: bool = False,
                   on_progress=None) -> dict:
    """Public wrapper around :func:`_analyze_ticker_impl` that never raises.

    Any uncaught exception inside the pipeline is converted to a structured
    ``{"ok": False, "error": "<msg>", "traceback": "..."}`` result and the
    full traceback is printed so Railway logs show the exact file+line of
    the crash (including cryptic errors like 'list' object has no attribute
    'get' that previously bubbled up with no location info).

    Progress reporting:
        ``on_progress`` is an optional callable ``(step, pct, label)``
        invoked at meaningful pipeline checkpoints. ``step`` is one of:
        ``"sec_filings" | "financials" | "market_data" | "grok" |
        "rendering" | "gamma" | "upload" | "done"``. ``pct`` is in [0, 1].
        Exceptions inside the callback are swallowed so progress bookkeeping
        can never crash the analysis itself.
    """
    try:
        return _analyze_ticker_impl(
            ticker,
            system_prompt=system_prompt,
            generate_gamma=generate_gamma,
            verbose=verbose,
            reuse_existing=reuse_existing,
            on_progress=on_progress,
        )
    except BaseException as exc:  # noqa: BLE001
        tb_str = traceback.format_exc()
        print(f"\n❌ analyze_ticker({ticker}) CRASHED:\n{tb_str}", flush=True)
        # Extract the last traceback frame (the actual crash site) for the UI.
        tb_lines = tb_str.strip().splitlines()
        last_frame = ""
        for i, line in enumerate(tb_lines):
            if line.lstrip().startswith("File "):
                last_frame = line.strip()
        tail = " @ " + last_frame if last_frame else ""
        return {
            "ticker": ticker.strip().upper(),
            "ok": False,
            "error": f"{type(exc).__name__}: {exc}{tail}",
            "traceback": tb_str,
        }


def _emit_progress(on_progress, step: str, pct: float, label: str = "") -> None:
    """Safe wrapper around the user-provided progress callback."""
    if on_progress is None:
        return
    try:
        on_progress(step, max(0.0, min(1.0, float(pct))), label)
    except Exception:  # noqa: BLE001
        pass  # progress reporting must never crash the pipeline


def _analyze_ticker_impl(ticker: str, *, system_prompt: str, generate_gamma: bool,
                         verbose: bool = True, reuse_existing: bool = False,
                         on_progress=None) -> dict:
    """Analyze a single ticker end-to-end.

    When ``reuse_existing`` is True and a cached markdown report already exists
    in /stocks, we load it instead of re-calling Grok. This is what the
    portfolio-rebalance flow uses by default so we don't burn 20+ API calls
    every time we re-optimize weights.
    """
    ticker = ticker.strip().upper()
    result = {"ticker": ticker, "ok": False}

    # --- Fast path: reuse an existing report if present and requested.
    md_path = STOCKS_FOLDER / f"{ticker}_DGA_Report.md"
    if reuse_existing and not md_path.exists():
        # Local cache miss — try the shared 'DGA Research Reports' Drive folder.
        # This survives Railway redeploys, which wipe the local /stocks folder.
        drive_md = fetch_report_from_drive(ticker)
        if drive_md:
            try:
                md_path.write_text(drive_md)
                print(f"☁️   {ticker}: hydrated cached report from Google Drive")
            except Exception:  # noqa: BLE001
                pass
    if reuse_existing and md_path.exists():
        print(f"♻️  {ticker}: reusing cached report at {md_path.name}")
        _emit_progress(on_progress, "done", 1.0, "Loaded from cache")
        report_text = md_path.read_text()
        mkt = fetch_market_snapshot(ticker)
        audit_path = STOCKS_FOLDER / f"{ticker}_xbrl_extract.json"
        entity_name, latest_filing_type = ticker, "10-K"
        if audit_path.exists():
            try:
                with open(audit_path) as fh:
                    d = json.load(fh)
                entity_name = d.get("entity_name", ticker) or ticker
                latest_filing_type = d.get("latest_filing_type", "10-K") or "10-K"
            except Exception:
                pass
        summary = extract_summary_from_report(report_text)
        market_price = mkt.get("price") or summary.get("current_price")
        return {
            "ok": True,
            "ticker": ticker,
            "entity_name": entity_name,
            "latest_filing_type": latest_filing_type,
            "market_price": market_price,
            "report_text": report_text,
            "docx": str(STOCKS_FOLDER / f"{ticker}_DGA_Report.docx"),
            "md": str(md_path),
            "xbrl_json": str(audit_path) if audit_path.exists() else None,
            "gamma_url": None,
            "gamma_credits": 0,
            "summary": summary,
            "cached": True,
        }

    # --- Step 1: download the latest 10-K and 10-Q into stock-financials/{TICKER}/
    # This parses the actual XBRL instance documents from each filing, so the
    # columns we read later map 1-to-1 onto the filing's own period contexts.
    print(f"\n🚀 {ticker}: downloading latest 10-K + 10-Q Excel workbooks…")
    _emit_progress(on_progress, "sec_filings", 0.05,
                   "Downloading SEC filings (10-K, 10-Q)")
    data: dict | None = None
    try:
        pull_sec_financials.download_financials(ticker)
    except Exception as exc:  # noqa: BLE001
        print(f"   ⚠️  Could not download fresh Excel files: {exc}")
        print("   Falling back to existing workbooks (if any) or companyfacts API.")

    # --- Step 2: read the Excel workbooks and build the verified data dict
    _emit_progress(on_progress, "financials", 0.20,
                   "Extracting filing-accurate financials")
    try:
        data = xlsx_edgar.extract_financials(ticker)
        verified_block = xlsx_edgar.format_verified_block(data)
        print(f"   ✅ Loaded filing-accurate financials from Excel workbooks.")
    except Exception as exc:  # noqa: BLE001
        print(f"   ⚠️  Excel reader failed: {exc}")
        print(f"   Falling back to SEC companyfacts API…")
        try:
            data = edgar.extract_financials(ticker, user_agent=get_sec_user_agent())
            verified_block = edgar.format_verified_block(data)
        except Exception as exc2:  # noqa: BLE001
            print(f"   ❌ EDGAR fallback also failed: {exc2}")
            traceback.print_exc()
            # Last-resort: continue with no verified financials so Grok can still
            # produce a qualitative report (it will note the data is unavailable).
            print(f"   ⚠️  Proceeding without verified financials for {ticker}.")
            verified_block = (
                f"## ⚠️ Financial Data Unavailable\n\n"
                f"Automated extraction failed for **{ticker}**. "
                f"Error: {exc2}\n\n"
                f"The analysis below relies on Grok's training data and publicly "
                f"available information only. No SEC XBRL figures have been verified."
            )
            data = {"ticker": ticker, "errors": [str(exc2)]}

    # Cache the raw extract for auditing.
    audit_path = STOCKS_FOLDER / f"{ticker}_xbrl_extract.json"
    try:
        with open(audit_path, "w") as f:
            json.dump(data, f, indent=2, default=str)
    except Exception:
        pass

    if verbose:
        print(verified_block)

    # Current price
    _emit_progress(on_progress, "market_data", 0.30,
                   "Fetching live price + analyst ratings")
    mkt = fetch_market_snapshot(ticker)

    # Fetch live analyst ratings from GuruFocus (best-effort, non-blocking).
    analyst_block = fetch_analyst_ratings(ticker)
    if analyst_block:
        print(f"   📊 Analyst ratings block built for {ticker} ({len(analyst_block)} chars)")
    else:
        print(f"   ⚠️  No live analyst data — Grok will use training-data estimates")

    # Compose Grok user message
    today = datetime.now().strftime("%Y-%m-%d")
    user_msg = (
        f"DATE: {today}\n"
        f"TICKER: {ticker}\n"
        f"ENTITY: {data.get('entity_name','')}\n"
        f"CURRENT_PRICE: {mkt.get('price')}\n"
        f"PREVIOUS_CLOSE: {mkt.get('previous_close')}\n"
        f"LATEST_FILING_TYPE: {data.get('latest_filing_type')}\n\n"
        f"{verified_block}\n\n"
        + (f"{analyst_block}\n\n" if analyst_block else "")
        + f"Generate the full research report for {ticker} following every rule in your system prompt."
    )

    # live_search=True makes Grok scan X/news/web for the last ~90 days of
    # developments on this ticker, which is how Section 2 (Recent
    # Developments) can surface a CEO departure from two days ago even
    # though it's past the model's training cutoff.
    print(f"   🧠 Calling Grok ({GROK_MODEL}) with live X/news/web search…")
    _emit_progress(on_progress, "grok", 0.40,
                   f"Grok ({GROK_MODEL}) — analyzing + live X/news search")
    try:
        report_text = call_grok(system_prompt, user_msg, live_search=True)
    except Exception as exc:  # noqa: BLE001
        print(f"   ❌ Grok API error: {exc}")
        result["error"] = f"Grok: {exc}"
        return result

    # Save markdown too, for debugging / iteration.
    md_path = STOCKS_FOLDER / f"{ticker}_DGA_Report.md"
    md_path.write_text(report_text)

    # Render Word
    _emit_progress(on_progress, "rendering", 0.85,
                   "Rendering Word document")
    out_docx = STOCKS_FOLDER / f"{ticker}_DGA_Report.docx"
    summary = extract_summary_from_report(report_text)
    rating_hint = summary.get("rating") or ""
    render_report(
        report_text,
        ticker=ticker,
        entity_name=data.get("entity_name", ticker),
        output_path=str(out_docx),
        price=mkt.get("price"),
        rating_hint=rating_hint,
    )
    print(f"   💾 Word: {out_docx}")

    # Gamma — wrapped in try/except so a missing API key or Gamma API error
    # never kills the whole analysis job.  gamma_error is set when generation
    # was requested but failed so the mobile/web UI can surface the reason.
    gamma_url = None
    gamma_credits = 0
    gamma_error: str | None = None
    if generate_gamma:
        _emit_progress(on_progress, "gamma", 0.92, "Generating Gamma presentation")
        # 1. Reuse a fresh existing deck if one was generated < GAMMA_FRESH_DAYS
        #    ago — saves credits and avoids piling up duplicate decks in the
        #    Gamma.app workspace for rapid re-runs of the same ticker.
        existing = _existing_fresh_gamma(ticker)
        if existing:
            gamma_url     = existing.get("gamma_url")
            gamma_credits = existing.get("credits", 0) or 0
            print(f"   ♻️  Reusing existing Gamma deck (<{GAMMA_FRESH_DAYS}d old) for {ticker}: {gamma_url}")
        else:
            # 2. Older than the freshness window → purge the stale local PPTX
            #    + index entry before regenerating, so we don't keep a stale
            #    file lying around with the same name.
            _purge_stale_local_gamma(ticker)
            try:
                out_pptx = STOCKS_FOLDER / f"{ticker}_DGA_Presentation.pptx"
                gamma_url, gamma_credits = create_gamma_for_stock(
                    report_text, ticker, data.get("latest_filing_type", "10-K"), out_pptx=out_pptx
                )
                if gamma_url is None:
                    gamma_error = "Gamma generation failed (API error or timeout — check server logs)"
                else:
                    _record_gamma(ticker, gamma_url,
                                  pptx_filename=out_pptx.name, credits=gamma_credits)
            except Exception as _gamma_exc:  # noqa: BLE001
                gamma_error = str(_gamma_exc)
                print(f"   ⚠️  Gamma skipped for {ticker}: {_gamma_exc}")

    # Upload report files to Google Drive / Dropbox (best-effort, non-blocking).
    # Include the gamma index so the URL survives Railway redeploys.
    _emit_progress(on_progress, "upload", 0.97,
                   "Uploading to Dropbox / Drive")
    drive_files = [p for p in [md_path, out_docx] if p.exists()]
    if generate_gamma:
        pptx_path = STOCKS_FOLDER / f"{ticker}_DGA_Presentation.pptx"
        if pptx_path.exists():
            drive_files.append(pptx_path)
        if GAMMA_INDEX_FILE.exists():
            drive_files.append(GAMMA_INDEX_FILE)
    gdrive_status: dict = {"ok": False, "skipped": True}
    try:
        gdrive_status = push_to_google_drive(drive_files)
    except Exception:  # noqa: BLE001
        pass

    _emit_progress(on_progress, "done", 1.0, "Report ready")
    result.update({
        "ok": True,
        "entity_name": data.get("entity_name", ticker),
        "latest_filing_type": data.get("latest_filing_type"),
        "market_price": mkt.get("price"),
        "report_text": report_text,
        "docx": str(out_docx),
        "md": str(md_path),
        "xbrl_json": str(audit_path),
        "gamma_url": gamma_url,
        "gamma_credits": gamma_credits,
        "gamma_error": gamma_error,
        "summary": summary,
        "gdrive": gdrive_status,
    })
    return result


# ============================================================================
# Portfolio roll-up (Grok call over per-stock summaries)
# ============================================================================
PORTFOLIO_SYSTEM_PROMPT = """You are DGA Capital's portfolio strategist. You are given per-stock
analyses already produced by the DGA equity research team. Your job is to:

1) Rank all positions into ACTION buckets:
   - SELL   (overvalued, broken thesis, cut exposure entirely)
   - TRIM   (overvalued vs targets, reduce weight)
   - HOLD   (fairly valued / on-thesis)
   - ADD    (attractive, increase weight)
   - BUY    (high conviction, initiate or materially add)

2) For each name provide: rating, 12-month price target, % upside / downside vs current
   price, and a one-line reason.

3) Produce:
   a. A Summary Ranking Table (markdown) with columns:
      | Ticker | Name | Rating | Current Price | 12M Target | Upside % | Action | One-line Reason |
   b. A Top Trim / Sell write-up section (what to reduce and why)
   c. A Top Add / Buy write-up section (what to increase and why)
   d. A Rebalancing Action Plan section with concrete moves
   e. A portfolio-level Key Catalysts Calendar (next 12 months)

Use ONLY the information you were given. Do not fabricate numbers. Keep the tone
institutional and decision-oriented. Output must be pure markdown, no preamble.
"""


def run_portfolio_summary(ticker_results: list[dict], *, generate_gamma: bool) -> dict:
    """Build the portfolio roll-up after all tickers have been analyzed."""
    usable = [r for r in ticker_results if r.get("ok")]
    if not usable:
        return {"ok": False, "error": "No successful per-stock analyses to summarize."}

    # Summarize each stock's report in ~800-1200 tokens to stay under context limits.
    per_stock_blobs = []
    for r in usable:
        s = r.get("summary", {}) or {}
        rep = r.get("report_text", "") or ""
        # Feed back Grok the Executive Summary + Verdict sections specifically.
        exec_part = _extract_section(rep, r"executive summary", max_chars=3000)
        verdict_part = _extract_section(rep, r"verdict", max_chars=3000)
        per_stock_blobs.append(
            f"### {r['ticker']} — {r.get('entity_name','')}\n"
            f"Rating (extracted): {s.get('rating') or 'N/A'}\n"
            f"12M Price Target (extracted): {s.get('price_target') or 'N/A'}\n"
            f"Current Price: {r.get('market_price')}\n\n"
            f"**Executive Summary excerpt:**\n{exec_part}\n\n"
            f"**Verdict excerpt:**\n{verdict_part}\n"
        )

    today = datetime.now().strftime("%Y-%m-%d")
    user_msg = (
        f"DATE: {today}\n"
        f"PORTFOLIO SIZE: {len(usable)} positions\n\n"
        + "\n\n".join(per_stock_blobs)
    )

    print("\n🧠 Portfolio summary: calling Grok for rebalancing analysis…")
    try:
        summary_md = call_grok(PORTFOLIO_SYSTEM_PROMPT, user_msg)
    except Exception as exc:  # noqa: BLE001
        return {"ok": False, "error": f"Grok portfolio call: {exc}"}

    # Parse the ranking table so we can tell Gamma top buys vs sells.
    ranked_rows = _parse_action_table(summary_md)

    # Save markdown
    md_path = STOCKS_FOLDER / "Portfolio_Summary.md"
    md_path.write_text(summary_md)

    # Word
    out_docx = STOCKS_FOLDER / "Portfolio_Summary.docx"
    render_report(
        summary_md,
        ticker="PORTFOLIO",
        entity_name="Portfolio Review",
        output_path=str(out_docx),
    )
    print(f"💾 Portfolio Word summary: {out_docx}")

    # Gamma — wrapped so a missing API key never kills the portfolio job
    gamma_url = None
    gamma_credits = 0
    gamma_error: str | None = None
    if generate_gamma:
        existing = _existing_fresh_gamma("PORTFOLIO")
        if existing:
            gamma_url     = existing.get("gamma_url")
            gamma_credits = existing.get("credits", 0) or 0
            print(f"   ♻️  Reusing portfolio Gamma deck (<{GAMMA_FRESH_DAYS}d old): {gamma_url}")
        else:
            _purge_stale_local_gamma("PORTFOLIO")
            try:
                out_pptx = STOCKS_FOLDER / "Portfolio_Summary.pptx"
                gamma_url, gamma_credits = create_gamma_portfolio_summary(
                    summary_md, ranked_rows, out_pptx=out_pptx
                )
                if gamma_url is None:
                    gamma_error = "Gamma generation failed (API error or timeout — check server logs)"
                else:
                    _record_gamma("PORTFOLIO", gamma_url,
                                  pptx_filename=out_pptx.name, credits=gamma_credits)
            except Exception as _gamma_exc:  # noqa: BLE001
                gamma_error = str(_gamma_exc)
                print(f"   ⚠️  Gamma skipped for portfolio: {_gamma_exc}")

    return {
        "ok": True,
        "docx": str(out_docx),
        "md": str(md_path),
        "summary_md": summary_md,
        "gamma_url": gamma_url,
        "gamma_credits": gamma_credits,
        "gamma_error": gamma_error,
        "ranked_rows": ranked_rows,
    }


def _extract_section(markdown_text: str, keyword_regex: str,
                     max_chars: int = 2500) -> str:
    """Rough section extractor: finds a heading matching `keyword_regex` and takes
    text up to the next heading (or max_chars)."""
    lines = markdown_text.split("\n")
    out: list[str] = []
    capturing = False
    heading_re = re.compile(r"^#{1,3}\s+.*(" + keyword_regex + ")", re.IGNORECASE)
    next_heading_re = re.compile(r"^#{1,3}\s+")
    for ln in lines:
        if not capturing:
            if heading_re.search(ln):
                capturing = True
                out.append(ln)
            continue
        if next_heading_re.match(ln) and len("\n".join(out)) > 200:
            break
        out.append(ln)
        if len("\n".join(out)) > max_chars:
            break
    return "\n".join(out).strip() or "(section not found)"


def _parse_action_table(summary_md: str) -> list[dict]:
    """Parse the first markdown table we can find with columns including Ticker + Action."""
    lines = summary_md.split("\n")
    rows: list[dict] = []
    for i, line in enumerate(lines):
        if not line.strip().startswith("|"):
            continue
        if i + 1 >= len(lines):
            continue
        if "---" not in lines[i + 1]:
            continue
        header_cells = [c.strip() for c in line.strip().strip("|").split("|")]
        idx_lookup = {h.lower(): j for j, h in enumerate(header_cells)}
        j = i + 2
        while j < len(lines) and lines[j].strip().startswith("|"):
            cells = [c.strip() for c in lines[j].strip().strip("|").split("|")]
            d: dict = {}
            for key_name, canonical in [
                ("ticker", "ticker"),
                ("name", "name"),
                ("rating", "rating"),
                ("current price", "current_price"),
                ("12m target", "price_target"),
                ("target", "price_target"),
                ("upside %", "upside"),
                ("upside", "upside"),
                ("action", "action"),
                ("one-line reason", "reason"),
                ("reason", "reason"),
            ]:
                for h, jdx in idx_lookup.items():
                    if key_name in h and canonical not in d and jdx < len(cells):
                        d[canonical] = cells[jdx]
            if d.get("ticker"):
                rows.append(d)
            j += 1
        if rows:
            break
    return rows


# ============================================================================
# Portfolio file loader
# ============================================================================
def load_portfolio_file(path: str) -> list[dict]:
    """Load a portfolio file.

    Expected schema (new):
        | Ticker | Weight | Optimized |
        |--------|--------|-----------|
        | AAPL   | 0.05   |           |

    - The "Optimized" column is intentionally ignored so the same file can be
      re-used as input on the NEXT run, where the rebalancer will write a
      fresh Optimized column.
    - Weight can be expressed as a decimal (0.05) or a whole-number percent (5).
    - Falls back to legacy two-column (ticker, Allocation) and single-column
      files for backward compatibility.

    Returns a list of dicts: [{"ticker": "AAPL", "weight": 0.05}, ...]
    """
    p = Path(path).expanduser()
    if not p.exists():
        raise FileNotFoundError(path)

    # ── Header-row detection ──────────────────────────────────────────────────
    # Fidelity / Schwab exports often prepend metadata rows ("Brokerage Services",
    # "As of <date>", blank lines, account header) before the real CSV table.
    # Scan for the first row that looks like an actual portfolio header and tell
    # pandas to skip everything above it.
    if p.suffix.lower() in (".xlsx", ".xls", ".xlsm"):
        # Excel: detect a header row that contains a recognisable column.
        df_full = pd.read_excel(p, header=None)
        header_idx = _detect_header_row(df_full)
        df = pd.read_excel(p, header=header_idx)
    elif p.suffix.lower() in (".csv", ".tsv"):
        sep = "\t" if p.suffix.lower() == ".tsv" else ","
        # Pre-scan the raw text for the header row
        try:
            raw_text = p.read_text(errors="replace")
            raw_lines = raw_text.splitlines()
        except Exception:  # noqa: BLE001
            raw_text, raw_lines = "", []
        header_idx = _detect_header_row_from_lines(raw_lines, sep)

        # ── Fidelity CSV: bypass pandas entirely ──────────────────────────────
        # Fidelity exports have UNQUOTED commas in numeric values (quantities
        # like "3,025.317", dollar values like "$37,584.59"), which shift all
        # subsequent columns in pandas.  Instead of trying to pre-clean the
        # text, we use a content-based extractor that finds each field by what
        # it looks like, not where it sits positionally.
        try:
            _hdr_fields = next(_fidelity_csv.reader([raw_lines[header_idx]]))
        except Exception:
            _hdr_fields = []

        if _is_fidelity_format(_hdr_fields):
            _raw_records = _parse_fidelity_lines(raw_lines, header_idx)
            # ── Consolidate duplicates (same ticker in multiple lots/baskets)
            _consolidated: dict[str, float | None] = {}
            _order: list[str] = []
            for _r in _raw_records:
                _t = _r["ticker"]
                if _t not in _consolidated:
                    _consolidated[_t] = _r["weight"]
                    _order.append(_t)
                else:
                    _ex = _consolidated[_t]
                    _nw = _r["weight"]
                    if _ex is None and _nw is None:
                        pass
                    elif _ex is None:
                        _consolidated[_t] = _nw
                    elif _nw is None:
                        pass
                    else:
                        _consolidated[_t] = round(_ex + _nw, 8)
            _records = [{"ticker": _t, "weight": _consolidated[_t]} for _t in _order]
            if _records and all(_r["weight"] is None for _r in _records):
                _n = len(_records)
                for _r in _records:
                    _r["weight"] = round(1.0 / _n, 6)
            if not _records:
                raise ValueError(
                    f"No positions found in Fidelity CSV "
                    f"(header at line {header_idx}: {_hdr_fields[:6]}...)"
                )
            return _records

        # ── Non-Fidelity CSV: pandas path (regex pre-clean) ───────────────────
        # For non-Fidelity CSVs we still try to strip unquoted numeric commas
        # before handing to pandas.
        from io import StringIO as _StringIO
        df = pd.read_csv(
            _StringIO(raw_text), sep=sep, skiprows=header_idx,
            engine="python",
            on_bad_lines="skip",
        )
        # Drop completely empty rows (Fidelity sometimes has blank separators
        # between Stocks / Options / Mutual Funds tables — we keep rows up to
        # the first all-NaN row, which corresponds to the end of the equities table).
        if df.notna().any(axis=1).any():
            # Find the first all-NaN row after at least one valid row.
            valid_mask = df.notna().any(axis=1)
            first_valid_idx = valid_mask.idxmax()
            after_first = valid_mask.loc[first_valid_idx:]
            if (~after_first).any():
                end_idx = (~after_first).idxmax()
                df = df.loc[first_valid_idx:end_idx - 1] if end_idx > first_valid_idx else df.loc[first_valid_idx:]
            else:
                df = df.loc[first_valid_idx:]
    else:
        raise ValueError(f"Unsupported portfolio file: {p.suffix}")

    cols_lower = {str(c).strip().lower(): c for c in df.columns}

    # ── Ticker column: header-name match (validated by content) ──────────────
    ticker_col = None
    for key in ("ticker", "tickers", "symbol", "symbols"):
        if key in cols_lower and _column_contains_tickers(df[cols_lower[key]]):
            ticker_col = cols_lower[key]
            break
    if ticker_col is None:
        for col in df.columns:
            if _column_contains_tickers(df[col]):
                ticker_col = col
                break
    if ticker_col is None:
        ticker_col = df.columns[0]

    # ── Weight column: ONLY use the brokerage's explicit "Percent Of Account"
    # column. Take values directly. No dollar-fallback, no normalization, no
    # equal-weight default. The user's brokerage already tells us the weight
    # in a single authoritative column — that's the truth, use it.
    weight_col = None
    for key in ("weight", "weights", "allocation", "alloc",
                "weight %", "weight (%)",
                "allocation %", "allocation (%)",
                "percent of account", "% of account", "% account",
                "percent of portfolio", "% of portfolio"):
        if key in cols_lower:
            weight_col = cols_lower[key]
            break

    # ── Iterate rows, collect (ticker, weight) pairs ─────────────────────────
    raw_records: list[dict] = []
    for _, row in df.iterrows():
        raw_t = row[ticker_col]
        if pd.isna(raw_t):
            continue
        ticker = str(raw_t).strip().upper().strip("*").strip()
        if not ticker or ticker in ("NAN", "NONE"):
            continue
        # Skip footer rows
        if ticker in ("TOTAL", "TOTALS", "SUBTOTAL", "CASH",
                      "ACCOUNTTOTAL", "ACCOUNT", "PENDINGACTIVITY"):
            continue
        # Skip CUSIPs (start with a digit — these are bonds like 36966TKX9)
        if ticker[0].isdigit():
            continue
        # Skip money market / cash (handled by extended parser; not rebalanceable)
        if ticker in ("SPAXX", "FZFXX", "FZSXX", "FDRXX", "FZDXX",
                      "VMFXX", "VMRXX", "SWVXX", "FCASH", "CASH"):
            continue
        # Tickers are alphanumeric (dots/dashes allowed)
        if not all(c.isalnum() or c in (".", "-") for c in ticker):
            continue

        # Weight: read directly from the brokerage's percent column.
        # No dollar fallback, no equal-weight default — if the column is
        # missing or unparseable, we keep the row with weight=None and the
        # downstream optimizer can decide what to do.
        weight: float | None = None
        if weight_col is not None and pd.notna(row[weight_col]):
            raw = row[weight_col]
            try:
                had_percent_sign = isinstance(raw, str) and "%" in raw
                if isinstance(raw, str):
                    cleaned = raw.replace("%", "").replace("$", "").replace(",", "").strip()
                    if cleaned.startswith("(") and cleaned.endswith(")"):
                        cleaned = "-" + cleaned[1:-1]
                    weight = float(cleaned)
                else:
                    weight = float(raw)
                # An explicit % sign means "this number is already a percent".
                # 0.62% → 0.0062 ; 10.07% → 0.1007.
                if had_percent_sign:
                    weight = weight / 100.0
                elif weight > 1.5:
                    weight = weight / 100.0
            except (TypeError, ValueError):
                weight = None
        raw_records.append({"ticker": ticker, "weight": weight})

    # ── Consolidate duplicates: same ticker (e.g. INTC bought directly +
    # INTC held inside a basket) → sum their percentages into ONE position.
    consolidated: dict[str, float | None] = {}
    order: list[str] = []
    for r in raw_records:
        t = r["ticker"]
        if t not in consolidated:
            consolidated[t] = r["weight"]
            order.append(t)
        else:
            existing = consolidated[t]
            new = r["weight"]
            if existing is None and new is None:
                consolidated[t] = None
            elif existing is None:
                consolidated[t] = new
            elif new is None:
                pass  # keep existing
            else:
                consolidated[t] = round(existing + new, 8)

    records: list[dict] = [{"ticker": t, "weight": consolidated[t]} for t in order]

    # If absolutely no weights were provided ANYWHERE, fall back to equal-weight.
    # (Only applies when the file truly has no Percent Of Account column.)
    if records and all(r["weight"] is None for r in records):
        n = len(records)
        for r in records:
            r["weight"] = round(1.0 / n, 6)

    # Better diagnostics if zero rows were extracted — surface the columns we
    # actually saw plus a sample of the matched columns so the failure mode is
    # diagnosable from the error message alone.
    if not records:
        cols_seen = list(df.columns)[:20]
        n_rows = len(df)
        # Sample first 3 raw values from the matched ticker / weight columns
        sample: list[str] = []
        if ticker_col is not None:
            try:
                vals = df[ticker_col].head(5).tolist()
                sample.append(f"sample {ticker_col} values: {vals}")
            except Exception:  # noqa: BLE001
                pass
        if weight_col is not None:
            try:
                vals = df[weight_col].head(5).tolist()
                sample.append(f"sample {weight_col} values: {vals}")
            except Exception:  # noqa: BLE001
                pass
        raise ValueError(
            f"No portfolio positions could be parsed from {n_rows} data rows. "
            f"Columns detected: {cols_seen}. "
            f"Ticker column matched: {ticker_col!r}. "
            f"Weight column matched: {weight_col!r}. "
            + (" | ".join(sample) if sample else "")
        )

    return records


def portfolio_tickers(records: list[dict]) -> list[str]:
    return [r["ticker"] for r in records]


# Header-row detection helpers — handles Fidelity / Schwab metadata preambles
_HEADER_HINTS = (
    "ticker", "tickers", "symbol", "symbols",
    "weight", "weights", "allocation",
    "percent of account", "% of account", "percent of portfolio",
)


def _row_looks_like_header(values) -> bool:
    """A row is a portfolio table header if any cell matches a known label."""
    for v in values:
        if v is None:
            continue
        try:
            s = str(v).strip().lower()
        except Exception:  # noqa: BLE001
            continue
        if not s or s == "nan":
            continue
        for hint in _HEADER_HINTS:
            if s == hint or hint in s:
                return True
    return False


def _detect_header_row_from_lines(lines: list[str], sep: str) -> int:
    """Return the 0-indexed line that looks like the CSV header. Defaults to 0."""
    import csv as _csv
    for i, line in enumerate(lines[:30]):  # only scan first 30 lines
        if not line.strip():
            continue
        try:
            row = next(_csv.reader([line], delimiter=sep))
        except Exception:  # noqa: BLE001
            continue
        if _row_looks_like_header(row):
            return i
    return 0


def _detect_header_row(df_full) -> int:
    """Excel-equivalent of _detect_header_row_from_lines."""
    for i in range(min(30, len(df_full))):
        row_vals = df_full.iloc[i].tolist()
        if _row_looks_like_header(row_vals):
            return i
    return 0


# ── Fidelity-specific CSV parser ──────────────────────────────────────────────
# Fidelity brokerage exports have UNQUOTED commas inside numeric fields
# (dollar amounts, share quantities like "3,025.317", current values like
# "$37,584.59").  These extra commas fool pandas and csv.reader into treating
# them as field separators, shifting all subsequent columns to the right.
#
# Rather than trying to pre-clean the text with fragile regex, we use content-
# based extraction:
#
#   • Ticker (Symbol): first field in columns 0–7 that matches the ticker
#     pattern.  Symbol is column D (index 3) and comes before all problematic
#     numeric columns, so position 3 is usually correct; the scan handles any
#     edge case where an earlier text field contains a comma.
#
#   • Percent Of Account: the 3rd field in the row that ends with "%".
#     In Fidelity's column layout the "%" fields appear in this fixed order:
#       1st: Today's Gain/Loss %   (col K, index 10)
#       2nd: Total Gain/Loss %     (col M, index 12)
#       3rd: Percent Of Account    (col N, index 13)  ← what we want
#     Extra commas inside numeric values never produce a "%"-suffixed fragment,
#     so this count is immune to column shifts.

import csv as _fidelity_csv


def _is_fidelity_format(header_fields: list[str]) -> bool:
    """True if the header row contains both 'Symbol' and 'Percent Of Account'."""
    lower = [f.strip().lower() for f in header_fields]
    return "percent of account" in lower and "symbol" in lower


_FIDELITY_SKIP = frozenset({
    "SPAXX", "FZFXX", "FZSXX", "FDRXX", "FZDXX",
    "VMFXX", "VMRXX", "SWVXX", "FCASH", "CASH",
    "TOTAL", "TOTALS", "SUBTOTAL", "ACCOUNTTOTAL", "PENDINGACTIVITY",
})


def _parse_fidelity_lines(lines: list[str], header_idx: int) -> list[dict]:
    """
    Parse a Fidelity brokerage CSV export, bypassing pandas entirely.

    Returns a list of raw (ticker, weight) dicts suitable for the same
    consolidation step used by load_portfolio_file().
    """
    records: list[dict] = []

    for raw_line in lines[header_idx + 1:]:
        line = raw_line.strip()
        if not line:
            # Blank line = section separator; stop here (equities section done).
            break

        try:
            fields = next(_fidelity_csv.reader([line]))
        except Exception:  # noqa: BLE001
            continue

        if len(fields) < 5:
            continue

        # ── Ticker ────────────────────────────────────────────────────────────
        # Symbol is nominally at column index 3 but scan indices 0-7 to
        # tolerate any commas in the Account Number / Account Name / Basket
        # Name fields that precede Symbol.
        ticker: str | None = None
        for fld in fields[:8]:
            candidate = fld.strip().upper().strip("*").strip()
            if _looks_like_ticker(candidate):
                ticker = candidate
                break
        if ticker is None:
            continue

        if ticker in _FIDELITY_SKIP:
            continue
        if ticker[0].isdigit():
            continue
        if not all(c.isalnum() or c in (".", "-") for c in ticker):
            continue

        # ── Percent Of Account (3rd "%" field in the row) ─────────────────────
        # In Fidelity's layout the only three fields that end with "%" are
        # Today's G/L %, Total G/L %, and Percent Of Account — in that order.
        # Extra commas in dollar / quantity values never produce a "%" fragment.
        weight: float | None = None
        pct_count = 0
        for fld in fields:
            fld_s = fld.strip()
            if fld_s.endswith("%"):
                pct_count += 1
                if pct_count == 3:
                    try:
                        val = float(fld_s.replace("%", "").strip())
                        # Sanity-check: portfolio weight must be 0–100 %
                        if 0.0 <= val <= 100.0:
                            weight = round(val / 100.0, 8)
                    except (ValueError, TypeError):
                        pass
                    break

        records.append({"ticker": ticker, "weight": weight})

    return records


# ── Content-based column detection ───────────────────────────────────────────
# When header names are unreliable (Fidelity sometimes puts security
# descriptions under a "Symbol" header, or dollar values under a "Percent Of
# Account" header), fall back to identifying columns by what their values
# look like.

import re as _re


_TICKER_PATTERN = _re.compile(r"^[A-Z][A-Z0-9.\-]{0,5}$")


def _looks_like_ticker(val) -> bool:
    """A ticker is 1–6 chars, starts with a letter, all uppercase alphanumeric."""
    if val is None:
        return False
    try:
        s = str(val).strip().upper()
    except Exception:  # noqa: BLE001
        return False
    if not s or s in ("NAN", "NONE"):
        return False
    if s in ("TOTAL", "TOTALS", "SUBTOTAL", "CASH", "ACCOUNT"):
        return False
    return bool(_TICKER_PATTERN.match(s))


def _column_contains_tickers(series) -> bool:
    """Return True if at least 50% of non-null values look like ticker symbols."""
    non_null = series.dropna()
    if len(non_null) == 0:
        return False
    matches = sum(1 for v in non_null.head(20) if _looks_like_ticker(v))
    return matches >= max(2, int(0.5 * min(len(non_null), 20)))


def _looks_like_percentage(val) -> bool:
    """Numbers in [0, 100] range, optionally with a % suffix or whitespace."""
    if val is None:
        return False
    try:
        s = str(val).strip().replace("%", "").replace(",", "")
        if s.startswith("(") and s.endswith(")"):
            s = "-" + s[1:-1]
        f = float(s)
        # Reject dollar-like values (anything with a leading $)
        if "$" in str(val):
            return False
        return -50.0 <= f <= 100.0
    except (TypeError, ValueError):
        return False


def _column_contains_percentages(series) -> bool:
    """Column likely holds percentages: most values are numeric 0..100, no $."""
    non_null = series.dropna()
    if len(non_null) == 0:
        return False
    matches = 0
    for v in non_null.head(20):
        s = str(v)
        if "$" in s:           # any $ → not a percentage column
            return False
        if _looks_like_percentage(v):
            matches += 1
    return matches >= max(2, int(0.6 * min(len(non_null), 20)))


def _looks_like_money(val) -> bool:
    """Strings with $ prefix or a numeric value > 100 (likely dollars)."""
    if val is None:
        return False
    try:
        s = str(val).strip()
        if s.startswith("$") or "$" in s:
            return True
        # Bare numbers: only if they're clearly larger than typical percentages
        f = float(s.replace(",", ""))
        return abs(f) >= 100.0
    except (TypeError, ValueError):
        return False


def _column_contains_dollars(series) -> bool:
    """Most values look like dollar amounts (have $ or are numerically large)."""
    non_null = series.dropna()
    if len(non_null) == 0:
        return False
    matches = sum(1 for v in non_null.head(20) if _looks_like_money(v))
    return matches >= max(2, int(0.6 * min(len(non_null), 20)))


def _parse_money(val) -> float:
    """Convert "$1,234.56" / "(123.45)" / "1234" to float."""
    if val is None:
        return 0.0
    s = str(val).strip().replace("$", "").replace(",", "")
    if s.startswith("(") and s.endswith(")"):
        s = "-" + s[1:-1]
    return float(s) if s else 0.0


# ============================================================================
# Sector lookup (Yahoo Finance quoteSummary, best-effort)
# ============================================================================
_SECTOR_OVERRIDES = {
    # Fall-backs for tickers that Yahoo sometimes misses / labels oddly.
    "FNMA": "Financial Services",
    "IBRX": "Healthcare",
    "MOH":  "Healthcare",
    "HHH":  "Real Estate",
    "SPG":  "Real Estate",
    "ASML": "Technology",
    "TSM":  "Technology",
    "SMCI": "Technology",
    "INTC": "Technology",
    "CSCO": "Technology",
    "PYPL": "Financial Services",
    "C":    "Financial Services",
    "WFC":  "Financial Services",
    "NFLX": "Communication Services",
    "CMCSA": "Communication Services",
    "DIS":  "Communication Services",
    "T":    "Communication Services",
    "TSLA": "Consumer Cyclical",
    "HAL":  "Energy",
    "VALE": "Basic Materials",
}


# SEC SIC codes that classify a company as biotech / pharma. Authoritative
# fallback when Yahoo Finance returns "Unknown" — SEC classifications don't
# change with API throttling.
#   2833 = Medicinal Chemicals & Botanical Products
#   2834 = Pharmaceutical Preparations
#   2835 = In Vitro & In Vivo Diagnostic Substances
#   2836 = Biological Products, Except Diagnostic
#   8731 = Commercial Physical & Biological Research
_BIOTECH_SIC_CODES: set[int] = {2833, 2834, 2835, 2836, 8731}

# Module-level cache of resolved (sector, industry). Yahoo throttles burst
# requests in a portfolio rebalance, so caching successful resolutions
# means we hit each ticker at most once per process. Failed resolutions are
# NOT cached so we retry on the next call.
_SECTOR_INDUSTRY_CACHE: dict[str, tuple[str, str]] = {}


def _industry_from_sec_sic(ticker: str) -> tuple[str, str] | None:
    """Authoritative fallback: derive (sector, industry) from SEC SIC code.

    SEC classifications never get throttled or return empty like Yahoo does.
    For biotech/pharma SIC codes, returns ("Healthcare", "Biotechnology")
    so the existing _is_biotech() string-match rule fires correctly.
    """
    try:
        from edgar import Company  # type: ignore
        c = Company(ticker)
        sic = c.sic
        if sic is None:
            return None
        sic_int = int(sic)
        if sic_int in _BIOTECH_SIC_CODES:
            return ("Healthcare", "Biotechnology")
    except Exception:  # noqa: BLE001
        return None
    return None


def fetch_sector_and_industry(ticker: str) -> tuple[str, str]:
    """Return (sector, industry) for *ticker*.

    Resolution order (each step is tried only if the previous fails):
      1. In-memory cache (most rebalance runs hit each ticker repeatedly)
      2. yfinance.info — usually reliable, but Yahoo throttles bursts
      3. raw quoteSummary endpoint — often returns empty; legacy fallback
      4. SEC SIC code via edgartools — authoritative, never throttled,
         catches biotech/pharma even when Yahoo is fully blocked

    Returning "Unknown" for industry causes the biotech 7% hard-cap to be
    silently skipped — the SIC fallback prevents that regression.
    """
    t = ticker.strip().upper()

    # ── 1. Cache ───────────────────────────────────────────────────────────
    if t in _SECTOR_INDUSTRY_CACHE:
        return _SECTOR_INDUSTRY_CACHE[t]

    sector_override = _SECTOR_OVERRIDES.get(t)

    # ── 2. yfinance.info ───────────────────────────────────────────────────
    try:
        import yfinance as yf  # type: ignore
        info = yf.Ticker(t).info or {}
        industry = (info.get("industry") or info.get("industryDisp") or "").strip()
        sector   = (info.get("sector")   or "").strip()
        if industry:   # only cache if we got real industry data
            result = (sector_override or sector or "Unknown", industry)
            _SECTOR_INDUSTRY_CACHE[t] = result
            return result
    except Exception:  # noqa: BLE001
        pass

    # ── 3. Raw quoteSummary endpoint (legacy fallback) ─────────────────────
    try:
        url = (
            f"https://query1.finance.yahoo.com/v10/finance/quoteSummary/{t}"
            f"?modules=assetProfile,summaryProfile"
        )
        resp = requests.get(
            url, headers={"User-Agent": "Mozilla/5.0"}, timeout=10,
        )
        data = resp.json()
        results = data.get("quoteSummary", {}).get("result", []) or []
        if results:
            profile = results[0].get("assetProfile") or results[0].get("summaryProfile") or {}
            sector_raw   = profile.get("sector",   "") or ""
            industry_raw = profile.get("industry", "") or ""
            if industry_raw:
                result = (sector_override or sector_raw or "Unknown", industry_raw)
                _SECTOR_INDUSTRY_CACHE[t] = result
                return result
    except Exception:  # noqa: BLE001
        pass

    # ── 4. SEC SIC code (authoritative biotech/pharma classification) ──────
    sic_result = _industry_from_sec_sic(t)
    if sic_result is not None:
        # SIC says biotech → use its sector unless the user has an override
        result = (sector_override or sic_result[0], sic_result[1])
        _SECTOR_INDUSTRY_CACHE[t] = result
        return result

    # ── 5. Last resort — don't cache so we retry on the next call ──────────
    return (sector_override or "Unknown"), "Unknown"


def fetch_sector(ticker: str) -> str:
    """Backward-compat wrapper — returns sector string only."""
    s, _ = fetch_sector_and_industry(ticker)
    return s


# ============================================================================
# Rebalancer — strategies
# ============================================================================
STRATEGIES = {
    "current": {
        "label": "Current Portfolio",
        "description": (
            "Keeps every uploaded position — no selling. "
            "Optimizes weights for risk-adjusted expected return. "
            "Position caps vary by market cap tier: large caps up to 20%, "
            "mid caps 15%, small caps 10%, micro caps 7%. "
            "Sector cap 30%. Works with any number of positions."
        ),
        "min_names": 1,
        "target_names": 999,   # all positions kept
        "max_names": 999,
        "max_position": 0.20,  # baseline; overridden per-ticker by market cap
        "min_position": 0.01,  # 1% floor — everyone stays in
        "max_sector": 0.30,
        "score_exponent": 1.5,
        "no_drop": True,       # never drop a position regardless of rating
    },
    "pro": {
        "label": "High Conviction",
        "description": "8–15 best ideas, max 15% each, min 3% if held, sector cap 25%. Institutional risk/reward.",
        "min_names": 8,
        "target_names": 12,
        "max_names": 15,
        "max_position": 0.15,
        "min_position": 0.03,
        "max_sector": 0.25,
        "score_exponent": 1.6,
    },
    "allin": {
        "label": "All In — Top 3",
        "description": "Only the 3 highest-conviction names, proportional to score up to 40% cap.",
        "min_names": 3,
        "target_names": 3,
        "max_names": 3,
        "max_position": 0.40,
        "min_position": 0.20,
        "max_sector": 1.00,  # effectively unconstrained
        "score_exponent": 2.0,
    },
}


def _market_cap_max_position(market_cap: float | None) -> float:
    """Dynamic per-position cap based on market capitalisation.

    Larger, more liquid names can absorb heavier allocation.
    Small and micro caps are capped tighter to control idiosyncratic risk.
    """
    if market_cap is None or market_cap <= 0:
        return 0.10  # Unknown — conservative default
    b = market_cap / 1_000_000_000  # convert to billions
    if b >= 10:
        return 0.20  # Large / mega cap
    if b >= 2:
        return 0.15  # Mid cap
    if b >= 0.3:
        return 0.10  # Small cap
    return 0.07       # Micro cap / speculative name

# ── Biotech hard cap ─────────────────────────────────────────────────────────
# Small / mid-cap biotech (<$15B market cap) is volatile enough to wipe a
# portfolio if the thesis is wrong.  Cap those positions at 7% regardless
# of the strategy's standard maximum.
_BIOTECH_CAP_THRESHOLD_B = 15.0   # $15 billion
_BIOTECH_MAX_POSITION    = 0.07   # 7% hard cap


def _is_biotech(sector: str, industry: str) -> bool:
    """Return True if sector/industry indicates biotech or biopharma."""
    s = (sector   or "").lower()
    i = (industry or "").lower()
    return (
        "biotech" in i
        or "biopharmaceutical" in i
        or "biopharma" in i
        # "Drug Manufacturers - Specialty & Generic" under Healthcare
        or ("drug manufacturer" in i and "healthcare" in s)
        # Generic pharma — high binary risk, small cap
        or ("pharmaceutical" in i and "healthcare" in s
            and "large" not in i and "major" not in i)
    )


def _effective_max_position(
    market_cap: float | None,
    sector:     str = "",
    industry:   str = "",
) -> float:
    """Per-ticker position cap: market-cap tier, overridden by biotech rule.

    Biotech / biopharma companies below $15B market cap are hard-capped at 7%
    regardless of what the strategy's standard maximum would allow.
    """
    base    = _market_cap_max_position(market_cap)
    mcap_b  = (market_cap or 0) / 1_000_000_000
    if mcap_b < _BIOTECH_CAP_THRESHOLD_B and _is_biotech(sector, industry):
        return min(base, _BIOTECH_MAX_POSITION)
    return base


_RATING_SCORE = {
    "strong buy": 5.0,
    "buy": 3.5,
    "hold": 1.0,
    "sell": -1.0,
    "strong sell": -3.0,
}


def _score_ticker(result: dict) -> dict:
    """Composite rebalance score for a single analyzed ticker."""
    s = result.get("summary") or {}
    rating = (s.get("rating") or "Hold").lower()
    rating_score = _RATING_SCORE.get(rating, 1.0)

    price = result.get("market_price")
    pt = s.get("price_target")

    # Prefer a pre-computed upside from the portfolio Grok roll-up if available.
    upside = s.get("upside_pct")
    if upside is None:
        if isinstance(price, (int, float)) and isinstance(pt, (int, float)) and price:
            upside = (pt - price) / price * 100.0
        else:
            upside = 0.0
    # Clip extreme outliers so one +500% fantasy target doesn't dominate.
    upside = max(-60.0, min(120.0, float(upside)))

    # Composite: rating dominates, upside fine-tunes the ranking within a rating.
    composite = rating_score + (upside / 10.0)
    return {
        "ticker": result["ticker"],
        "rating": rating.title(),
        "price": price,
        "price_target": pt,
        "upside_pct": round(upside, 2),
        "score": round(composite, 4),
        "sector":   result.get("sector",   "Unknown"),
        "industry": result.get("industry", "Unknown"),
        "action": s.get("action", ""),
    }


def _waterfall_cap(items: list[dict], key: str, cap: float) -> None:
    """Iteratively cap item[key] at `cap` and redistribute excess to uncapped items."""
    for _ in range(60):
        over = [i for i in items if i[key] > cap + 1e-9]
        if not over:
            return
        excess = sum(i[key] - cap for i in over)
        for i in over:
            i[key] = cap
        uncapped = [i for i in items if i[key] < cap - 1e-9]
        pool = sum(i[key] for i in uncapped)
        if pool <= 0:
            # Everyone is at the cap — renormalize and bail.
            total = sum(i[key] for i in items) or 1.0
            for i in items:
                i[key] /= total
            return
        for i in uncapped:
            i[key] += excess * (i[key] / pool)


def _apply_floor(items: list[dict], key: str, floor: float, cap: float) -> None:
    """Bump sub-floor items up to `floor`, take proportionally from items above floor."""
    for _ in range(30):
        below = [i for i in items if i[key] < floor - 1e-9]
        if not below:
            return
        deficit = sum(floor - i[key] for i in below)
        for i in below:
            i[key] = floor
        above = [i for i in items if i[key] > floor + 1e-9]
        pool = sum(i[key] for i in above)
        if pool <= 0:
            return
        for i in above:
            i[key] -= deficit * (i[key] / pool)
        _waterfall_cap(items, key, cap)


def _apply_sector_cap(items: list[dict], key: str, max_sector: float, max_pos: float) -> None:
    """Scale down any sector whose aggregate weight exceeds the cap, redistribute excess."""
    if max_sector >= 1.0:
        return
    for _ in range(10):
        sector_totals: dict[str, float] = {}
        for i in items:
            sector_totals[i["sector"]] = sector_totals.get(i["sector"], 0.0) + i[key]
        violators = [s for s, w in sector_totals.items() if w > max_sector + 1e-9]
        if not violators:
            return
        for sect in violators:
            sw = sector_totals[sect]
            scale = max_sector / sw
            excess = sw - max_sector
            # Scale down over-weight sector
            for i in items:
                if i["sector"] == sect:
                    i[key] *= scale
            # Distribute excess to names in other sectors, proportional to current weight.
            others = [i for i in items if i["sector"] != sect and i[key] > 0]
            pool = sum(i[key] for i in others)
            if pool <= 0:
                continue
            for i in others:
                i[key] += excess * (i[key] / pool)
        _waterfall_cap(items, key, max_pos)


def _merge_ranked_rows(
    ticker_results: list[dict], ranked_rows: list[dict] | None
) -> list[dict]:
    """Attach Grok-synthesized rating/target/upside/action to each ticker result.

    Any field already present on the ticker result is preserved; ranked_rows
    only fills in gaps.
    """
    if not ranked_rows:
        return ticker_results
    rr_by_tkr = {}
    for r in ranked_rows:
        tk = (r.get("ticker") or "").strip().upper()
        if tk:
            rr_by_tkr[tk] = r

    import re as _re
    pct_re = _re.compile(r"(-?\d+(?:\.\d+)?)")

    for tr in ticker_results:
        tk = tr["ticker"]
        rr = rr_by_tkr.get(tk)
        if not rr:
            continue
        s = tr.setdefault("summary", {}) or {}
        # Rating
        if rr.get("rating") and not s.get("rating"):
            s["rating"] = rr["rating"]
        # Price target
        pt_raw = rr.get("price_target")
        if pt_raw and not s.get("price_target"):
            m = pct_re.search(str(pt_raw).replace(",", ""))
            if m:
                try:
                    s["price_target"] = float(m.group(1))
                except ValueError:
                    pass
        # Current price (if Yahoo was unavailable)
        cp_raw = rr.get("current_price")
        if cp_raw and tr.get("market_price") in (None, 0):
            m = pct_re.search(str(cp_raw).replace(",", ""))
            if m:
                try:
                    tr["market_price"] = float(m.group(1))
                except ValueError:
                    pass
        # Upside
        up_raw = rr.get("upside")
        if up_raw:
            m = pct_re.search(str(up_raw).replace(",", ""))
            if m:
                try:
                    s["upside_pct"] = float(m.group(1))
                except ValueError:
                    pass
        # Action (SELL/TRIM/HOLD/ADD/BUY)
        action = (rr.get("action") or "").strip().upper()
        if action:
            s["action"] = action
            # If Grok says SELL/STRONG SELL, fold that into the rating.
            if "STRONG SELL" in action:
                s["rating"] = "Strong Sell"
            elif action == "SELL":
                s["rating"] = "Sell"
    return ticker_results


def compute_rebalance(
    ticker_results: list[dict],
    strategy: str = "pro",
    ranked_rows: list[dict] | None = None,
) -> dict:
    """Produce an optimized weight vector for the given strategy.

    If ``ranked_rows`` (from the portfolio Grok roll-up) is supplied, it is
    used to fill in rating / price / target / upside / action on each ticker
    result before scoring — that signal is strictly better than the quick
    regex scraped from the per-stock report.

    Returns a dict with:
      - strategy: key (e.g. "pro")
      - label / description
      - weights: {ticker: fraction_0_to_1}
      - rows: list of {ticker, score, rating, upside, sector, weight} (all tickers)
    """
    cfg = STRATEGIES.get(strategy) or STRATEGIES["current"]
    usable = [r for r in ticker_results if r.get("ok")]

    # Hydrate sector + industry for each ticker.
    # Industry is needed for the biotech position cap.
    for r in usable:
        needs_sector   = not r.get("sector")
        needs_industry = not r.get("industry")
        if needs_sector or needs_industry:
            report_sector = (r.get("summary") or {}).get("sector")
            if report_sector and not needs_industry:
                # Sector already known from report; no API call needed.
                if needs_sector:
                    r["sector"] = report_sector
            else:
                try:
                    sec, ind = fetch_sector_and_industry(r["ticker"])
                    if needs_sector:
                        r["sector"]   = report_sector or sec or "Unknown"
                    if needs_industry:
                        r["industry"] = ind or "Unknown"
                except Exception:
                    if needs_sector:   r["sector"]   = report_sector or "Unknown"
                    if needs_industry: r["industry"] = "Unknown"

    # Enrich with portfolio Grok roll-up if available.
    _merge_ranked_rows(usable, ranked_rows)

    scored = [_score_ticker(r) for r in usable]

    # ── "Current Portfolio" strategy — keep every position, no selling ────────
    if cfg.get("no_drop"):
        # Sell/Strong Sell tickers get a small positive score (minimum weight)
        # rather than zero — they stay in the portfolio.
        selected = [dict(s) for s in scored]
        for s in selected:
            if s["rating"].lower() in ("sell", "strong sell"):
                s["score"] = 0.15  # kept at minimum weight
            elif s["score"] <= 0:
                s["score"] = 0.20  # small positive floor for unknowns

        # Fetch market cap for dynamic per-position caps.
        # Uses yfinance fast_info (already a dependency); never raises.
        for s in selected:
            mcap = None
            try:
                import yfinance as yf  # type: ignore
                fi = yf.Ticker(s["ticker"]).fast_info
                raw = getattr(fi, "market_cap", None)
                if raw and float(raw) > 0:
                    mcap = float(raw)
            except Exception:  # noqa: BLE001
                pass
            s["_market_cap"] = mcap
            s["_max_pos"] = _effective_max_position(
                mcap, s.get("sector", ""), s.get("industry", "")
            )

        # Initial allocation proportional to score^exponent.
        exponent = cfg["score_exponent"]
        total = sum(max(s["score"], 0) ** exponent for s in selected) or 1.0
        for s in selected:
            s["weight"] = (max(s["score"], 0) ** exponent) / total

        # Apply per-ticker position caps (vary by market cap tier).
        for _ in range(80):
            over = [s for s in selected if s["weight"] > s["_max_pos"] + 1e-9]
            if not over:
                break
            excess = sum(s["weight"] - s["_max_pos"] for s in over)
            for s in over:
                s["weight"] = s["_max_pos"]
            uncapped = [s for s in selected if s["weight"] < s["_max_pos"] - 1e-9]
            pool = sum(s["weight"] for s in uncapped)
            if pool <= 0:
                break
            for s in uncapped:
                s["weight"] += excess * (s["weight"] / pool)

        # 1% floor so every position stays in.
        global_max = max((s["_max_pos"] for s in selected), default=cfg["max_position"])
        _apply_floor(selected, "weight", cfg["min_position"], global_max)

        # Sector cap.
        _apply_sector_cap(selected, "weight", cfg["max_sector"], global_max)

    # ── Standard strategies — select best N, can drop positions ──────────────
    else:
        # Drop SELL / Strong Sell and non-positive scores.
        eligible = [s for s in scored if s["rating"].lower() not in ("sell", "strong sell")]
        eligible = [s for s in eligible if s["score"] > 0]

        if not eligible:
            return {
                "strategy": strategy,
                "label": cfg["label"],
                "description": cfg["description"],
                "weights": {s["ticker"]: 0.0 for s in scored},
                "rows": [dict(s, weight=0.0, in_portfolio=False) for s in scored],
            }

        eligible.sort(key=lambda x: -x["score"])
        n_target = min(
            max(cfg["min_names"], cfg["target_names"]), cfg["max_names"], len(eligible)
        )
        selected = [dict(s) for s in eligible[:n_target]]

        # Fetch market cap for selected tickers so biotech caps can be applied.
        for s in selected:
            mcap = None
            try:
                import yfinance as yf  # type: ignore
                fi = yf.Ticker(s["ticker"]).fast_info
                raw = getattr(fi, "market_cap", None)
                if raw and float(raw) > 0:
                    mcap = float(raw)
            except Exception:  # noqa: BLE001
                pass
            s["_market_cap"] = mcap
            s["_max_pos"] = _effective_max_position(
                mcap, s.get("sector", ""), s.get("industry", "")
            )

        exponent = cfg["score_exponent"]
        total = sum(max(s["score"], 0) ** exponent for s in selected) or 1.0
        for s in selected:
            s["weight"] = (max(s["score"], 0) ** exponent) / total

        # Standard uniform cap first, then per-ticker biotech override.
        _waterfall_cap(selected, "weight", cfg["max_position"])
        # Apply per-ticker effective caps (biotech < $15B hard-capped at 7%).
        for _ in range(80):
            over = [s for s in selected if s["weight"] > s["_max_pos"] + 1e-9]
            if not over:
                break
            excess = sum(s["weight"] - s["_max_pos"] for s in over)
            for s in over:
                s["weight"] = s["_max_pos"]
            uncapped = [s for s in selected if s["weight"] < s["_max_pos"] - 1e-9]
            pool = sum(s["weight"] for s in uncapped)
            if pool <= 0:
                break
            for s in uncapped:
                s["weight"] += excess * (s["weight"] / pool)

        _apply_floor(selected, "weight", cfg["min_position"], cfg["max_position"])
        _apply_sector_cap(selected, "weight", cfg["max_sector"], cfg["max_position"])

    # ── Shared finalisation ───────────────────────────────────────────────────
    # Renormalize for rounding drift.
    total = sum(s["weight"] for s in selected) or 1.0
    for s in selected:
        s["weight"] = s["weight"] / total

    # Build final weights dict over ALL analyzed tickers (zero for dropped).
    selected_by_tkr = {s["ticker"]: s for s in selected}
    weights = {}
    rows = []
    for s in scored:
        sel = selected_by_tkr.get(s["ticker"])
        w = round(float(sel["weight"]), 4) if sel else 0.0
        weights[s["ticker"]] = w
        rows.append({**s, "weight": w, "in_portfolio": bool(sel)})

    # Small residual correction so weights sum to exactly 1.0.
    total_w = sum(weights.values())
    if total_w > 0 and abs(total_w - 1.0) > 1e-6:
        scale = 1.0 / total_w
        weights = {k: round(v * scale, 4) for k, v in weights.items()}
        for row in rows:
            row["weight"] = weights[row["ticker"]]

    return {
        "strategy": strategy,
        "label": cfg["label"],
        "description": cfg["description"],
        "weights": weights,
        "rows": rows,
    }


# ============================================================================
# DGA-portfolio.xlsx writer
# ============================================================================
DGA_PORTFOLIO_FILENAME = "DGA-portfolio.xlsx"


def write_dga_portfolio_xlsx(
    *,
    output_path: Path,
    input_records: list[dict],
    primary_strategy: str,
    strategy_results: dict[str, dict],
) -> Path:
    """Write the DGA-portfolio.xlsx output.

    Layout:
      Columns: Ticker | Weight | Optimized | <other strategy 1> | <other strategy 2>
      - "Weight" is the user's INPUT weight (what they held coming in).
      - "Optimized" is the PRIMARY strategy's new weights — this is the column
        the loader will IGNORE on the next input run, as required.
      - The remaining strategies appear as extra comparison columns so the user
        can see all three side by side.
      - A second sheet "Summary" lists per-ticker rating, price target, upside,
        sector, and the weight under each strategy for auditability.
      - A third sheet "Strategies" documents the constraint definitions.
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    # Column order: primary first, then the other two in a stable order.
    other_order = [s for s in ("current", "pro", "allin") if s != primary_strategy]
    order = [primary_strategy] + other_order

    # Build a ticker list starting with the input file order, then append any
    # extras (shouldn't happen normally, but covers edge cases).
    input_tickers = [r["ticker"] for r in input_records]
    input_weight_lookup = {r["ticker"]: r.get("weight", 0.0) or 0.0 for r in input_records}

    all_tickers_in_play = list(input_tickers)
    for key in order:
        for t in strategy_results[key]["weights"].keys():
            if t not in all_tickers_in_play:
                all_tickers_in_play.append(t)

    wb = openpyxl.Workbook()

    # -----------------------------------------------------------------------
    # Sheet 1: Portfolio (the thing the user cares about)
    # -----------------------------------------------------------------------
    ws = wb.active
    ws.title = "Portfolio"

    # Styling
    navy_fill = PatternFill("solid", fgColor="0A1628")
    gold_fill = PatternFill("solid", fgColor="C9A84C")
    header_font = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
    primary_header_font = Font(name="Calibri", size=12, bold=True, color="0A1628")
    cell_font = Font(name="Calibri", size=11)
    bold_font = Font(name="Calibri", size=11, bold=True)
    thin = Side(border_style="thin", color="3D4A5C")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Headers
    primary_label = strategy_results[primary_strategy]["label"]
    headers = ["Ticker", "Weight", "Optimized"]
    for key in other_order:
        headers.append(strategy_results[key]["label"])

    for col_idx, name in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=name)
        cell.font = primary_header_font if col_idx == 3 else header_font
        cell.fill = gold_fill if col_idx == 3 else navy_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    # Subheader row showing what strategy each column represents.
    sub_cell = ws.cell(row=2, column=1, value="")
    sub_cell.font = cell_font
    sub_cell.alignment = Alignment(horizontal="center")
    ws.cell(row=2, column=2, value="Previous (input)").font = Font(
        name="Calibri", size=9, italic=True, color="3D4A5C"
    )
    ws.cell(row=2, column=2).alignment = Alignment(horizontal="center")
    ws.cell(row=2, column=3, value=f"[{primary_label}]").font = Font(
        name="Calibri", size=9, italic=True, color="0A1628", bold=True
    )
    ws.cell(row=2, column=3).alignment = Alignment(horizontal="center")
    for i, key in enumerate(other_order, start=4):
        c = ws.cell(row=2, column=i, value=f"[{strategy_results[key]['label']}]")
        c.font = Font(name="Calibri", size=9, italic=True, color="3D4A5C")
        c.alignment = Alignment(horizontal="center")

    # Data rows
    for r_idx, ticker in enumerate(all_tickers_in_play, start=3):
        row_values = [ticker, input_weight_lookup.get(ticker, 0.0)]
        row_values.append(strategy_results[primary_strategy]["weights"].get(ticker, 0.0))
        for key in other_order:
            row_values.append(strategy_results[key]["weights"].get(ticker, 0.0))
        for c_idx, v in enumerate(row_values, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=v)
            cell.font = bold_font if c_idx == 1 else cell_font
            if c_idx >= 2:
                cell.number_format = "0.00%"
                cell.alignment = Alignment(horizontal="right")
            else:
                cell.alignment = Alignment(horizontal="center")
            cell.border = border

    # Totals row
    last_row = 3 + len(all_tickers_in_play)
    total_cell = ws.cell(row=last_row, column=1, value="TOTAL")
    total_cell.font = bold_font
    total_cell.alignment = Alignment(horizontal="center")
    total_cell.border = border
    for col_idx in range(2, len(headers) + 1):
        col_letter = get_column_letter(col_idx)
        cell = ws.cell(
            row=last_row,
            column=col_idx,
            value=f"=SUM({col_letter}3:{col_letter}{last_row-1})",
        )
        cell.font = bold_font
        cell.number_format = "0.00%"
        cell.alignment = Alignment(horizontal="right")
        cell.fill = PatternFill("solid", fgColor="F5F7FA")
        cell.border = border

    # Column widths
    ws.column_dimensions["A"].width = 12
    for c_idx in range(2, len(headers) + 1):
        ws.column_dimensions[get_column_letter(c_idx)].width = max(18, len(headers[c_idx-1]) + 6)
    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A3"

    # -----------------------------------------------------------------------
    # Sheet 2: Summary / Audit
    # -----------------------------------------------------------------------
    ws2 = wb.create_sheet("Summary")
    summary_headers = [
        "Ticker", "Rating", "Current Price", "12M Target", "Upside %", "Sector",
        f"{strategy_results[primary_strategy]['label']} (Primary)",
    ]
    for key in other_order:
        summary_headers.append(strategy_results[key]["label"])
    for col_idx, name in enumerate(summary_headers, start=1):
        c = ws2.cell(row=1, column=col_idx, value=name)
        c.font = header_font
        c.fill = navy_fill
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = border

    # Build a lookup of per-ticker row data from the primary strategy (it has all
    # metadata since every ticker is in `rows`).
    primary_rows_by_tkr = {row["ticker"]: row for row in strategy_results[primary_strategy]["rows"]}
    for r_idx, ticker in enumerate(all_tickers_in_play, start=2):
        meta = primary_rows_by_tkr.get(ticker, {})
        vals = [
            ticker,
            meta.get("rating", "—"),
            meta.get("price"),
            meta.get("price_target"),
            meta.get("upside_pct"),
            meta.get("sector", "Unknown"),
            strategy_results[primary_strategy]["weights"].get(ticker, 0.0),
        ]
        for key in other_order:
            vals.append(strategy_results[key]["weights"].get(ticker, 0.0))
        for c_idx, v in enumerate(vals, start=1):
            cell = ws2.cell(row=r_idx, column=c_idx, value=v)
            cell.font = cell_font
            cell.border = border
            if c_idx == 1:
                cell.font = bold_font
                cell.alignment = Alignment(horizontal="center")
            elif c_idx in (3, 4):  # prices
                cell.number_format = "$#,##0.00"
                cell.alignment = Alignment(horizontal="right")
            elif c_idx == 5:
                cell.number_format = "0.00"
                cell.alignment = Alignment(horizontal="right")
            elif c_idx >= 7:
                cell.number_format = "0.00%"
                cell.alignment = Alignment(horizontal="right")
            else:
                cell.alignment = Alignment(horizontal="left")

    for c_idx, name in enumerate(summary_headers, start=1):
        ws2.column_dimensions[get_column_letter(c_idx)].width = max(14, len(name) + 4)
    ws2.freeze_panes = "B2"

    # -----------------------------------------------------------------------
    # Sheet 3: Strategy definitions
    # -----------------------------------------------------------------------
    ws3 = wb.create_sheet("Strategies")
    cfg_headers = [
        "Strategy", "Label", "Min names", "Target", "Max names",
        "Max per position", "Min per position", "Max per sector", "Description",
    ]
    for col_idx, name in enumerate(cfg_headers, start=1):
        c = ws3.cell(row=1, column=col_idx, value=name)
        c.font = header_font
        c.fill = navy_fill
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = border
    for r_idx, skey in enumerate(order, start=2):
        cfg = STRATEGIES[skey]
        tag = " (PRIMARY)" if skey == primary_strategy else ""
        vals = [
            skey + tag,
            cfg["label"],
            cfg["min_names"],
            cfg["target_names"],
            cfg["max_names"],
            cfg["max_position"],
            cfg["min_position"],
            cfg["max_sector"],
            cfg["description"],
        ]
        for c_idx, v in enumerate(vals, start=1):
            cell = ws3.cell(row=r_idx, column=c_idx, value=v)
            cell.font = bold_font if skey == primary_strategy else cell_font
            cell.border = border
            if c_idx in (6, 7, 8):
                cell.number_format = "0.00%"
    for c_idx, name in enumerate(cfg_headers, start=1):
        ws3.column_dimensions[get_column_letter(c_idx)].width = max(14, len(name) + 4)

    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return output_path


# ============================================================================
# Google Drive upload (uses MCP-connected Drive from the cowork session).
# The actual MCP call is performed from the outer agent — this helper just
# builds the payload path and metadata.
# ============================================================================
def _gsheets_upsert_sheet(sh, title: str, rows: list[list]) -> None:
    """Clear and rewrite a worksheet, creating it if it doesn't exist."""
    import gspread
    try:
        ws = sh.worksheet(title)
        ws.clear()
    except gspread.exceptions.WorksheetNotFound:
        ws = sh.add_worksheet(title=title, rows=max(len(rows) + 20, 50), cols=30)
    if rows:
        ws.update(rows)


def _gsheets_append_log(sh, title: str, headers: list, row: list) -> None:
    """Append a row to a log sheet, creating with headers if it doesn't exist."""
    import gspread
    try:
        ws = sh.worksheet(title)
    except gspread.exceptions.WorksheetNotFound:
        ws = sh.add_worksheet(title=title, rows=200, cols=10)
        ws.append_row(headers)
    ws.append_row(row)


def push_to_google_sheets(
    *,
    input_records: list[dict],
    primary_strategy: str,
    strategy_results: dict[str, dict],
    run_timestamp: str,
) -> dict:
    """Push portfolio results to Google Sheets via a service account.

    Required env vars:
      GOOGLE_SERVICE_ACCOUNT_JSON  — path to the service account key JSON file
      GOOGLE_SHEETS_SPREADSHEET_ID — ID from the spreadsheet URL

    Creates / updates three sheets: Portfolio, Summary, Run Log.
    Returns {"ok": True, "url": "..."} or {"ok": False, "error": "..."}.
    """
    creds_path = _optional_env("GOOGLE_SERVICE_ACCOUNT_JSON")
    spreadsheet_id = _optional_env("GOOGLE_SHEETS_SPREADSHEET_ID")

    if not creds_path or not spreadsheet_id:
        return {
            "ok": False,
            "skipped": True,
            "error": "GOOGLE_SERVICE_ACCOUNT_JSON or GOOGLE_SHEETS_SPREADSHEET_ID not configured",
        }

    try:
        import gspread
        from google.oauth2.service_account import Credentials
    except ImportError:
        return {"ok": False, "error": "gspread not installed; run: pip install gspread google-auth"}

    try:
        scopes = [
            "https://www.googleapis.com/auth/spreadsheets",
            "https://www.googleapis.com/auth/drive",
        ]
        creds = Credentials.from_service_account_file(creds_path, scopes=scopes)
        gc = gspread.authorize(creds)
        sh = gc.open_by_key(spreadsheet_id)
    except Exception as exc:
        return {"ok": False, "error": f"Auth/open failed: {exc}"}

    try:
        other_order = [s for s in ("pro", "concentrated", "allin") if s != primary_strategy]
        order = [primary_strategy] + other_order

        input_tickers = [r["ticker"] for r in input_records]
        input_weight_lookup = {r["ticker"]: r.get("weight", 0.0) or 0.0 for r in input_records}

        all_tickers: list[str] = list(input_tickers)
        for key in order:
            for t in strategy_results[key]["weights"]:
                if t not in all_tickers:
                    all_tickers.append(t)

        primary_label = strategy_results[primary_strategy]["label"]

        # ---- Sheet: Portfolio ----
        port_headers = ["Ticker", "Weight (Prior)", f"Optimized [{primary_label}]"]
        for key in other_order:
            port_headers.append(strategy_results[key]["label"])

        port_rows: list[list] = [port_headers]
        for ticker in all_tickers:
            row: list = [
                ticker,
                round(input_weight_lookup.get(ticker, 0.0), 4),
                round(strategy_results[primary_strategy]["weights"].get(ticker, 0.0), 4),
            ]
            for key in other_order:
                row.append(round(strategy_results[key]["weights"].get(ticker, 0.0), 4))
            port_rows.append(row)

        total_row: list = ["TOTAL", round(sum(input_weight_lookup.values()), 4)]
        total_row.append(round(sum(strategy_results[primary_strategy]["weights"].get(t, 0.0) for t in all_tickers), 4))
        for key in other_order:
            total_row.append(round(sum(strategy_results[key]["weights"].get(t, 0.0) for t in all_tickers), 4))
        port_rows.append(total_row)

        _gsheets_upsert_sheet(sh, "Portfolio", port_rows)

        # ---- Sheet: Summary ----
        sum_headers = [
            "Ticker", "Rating", "Current Price", "12M Target", "Upside %", "Sector",
            f"{primary_label} (Primary)",
        ]
        for key in other_order:
            sum_headers.append(strategy_results[key]["label"])

        primary_rows_by_tkr = {
            r["ticker"]: r for r in strategy_results[primary_strategy]["rows"]
        }
        sum_rows: list[list] = [sum_headers]
        for ticker in all_tickers:
            meta = primary_rows_by_tkr.get(ticker, {})
            srow: list = [
                ticker,
                meta.get("rating", "—"),
                meta.get("price"),
                meta.get("price_target"),
                meta.get("upside_pct"),
                meta.get("sector", "Unknown"),
                round(strategy_results[primary_strategy]["weights"].get(ticker, 0.0), 4),
            ]
            for key in other_order:
                srow.append(round(strategy_results[key]["weights"].get(ticker, 0.0), 4))
            sum_rows.append(srow)

        _gsheets_upsert_sheet(sh, "Summary", sum_rows)

        # ---- Sheet: Run Log (append-only audit trail) ----
        log_headers = ["Timestamp", "Tickers", "Strategy", "Top Picks", "# Positions"]
        top_picks = sorted(
            strategy_results[primary_strategy]["weights"].items(), key=lambda x: -x[1]
        )[:3]
        log_row = [
            run_timestamp,
            ", ".join(all_tickers),
            primary_label,
            ", ".join(f"{t} {w:.1%}" for t, w in top_picks),
            sum(1 for v in strategy_results[primary_strategy]["weights"].values() if v > 0),
        ]
        _gsheets_append_log(sh, "Run Log", log_headers, log_row)

        url = f"https://docs.google.com/spreadsheets/d/{spreadsheet_id}"
        return {"ok": True, "url": url}

    except Exception as exc:
        return {"ok": False, "error": str(exc)}


# ============================================================================
# Google Drive upload — real Drive folder this time.
#
# A service account has zero personal Drive quota, BUT if *you* share a Drive
# folder with the service-account email (same way you shared the spreadsheet),
# then files the service account creates inside that folder count against your
# storage, not its own. That's how we persist report .md / .docx / .pptx / .xlsx
# files across Railway restarts (the local /stocks cache gets wiped on every
# redeploy).
#
# Folder resolution order:
#   1. GOOGLE_DRIVE_FOLDER_ID env var (explicit ID wins).
#   2. A folder named GOOGLE_DRIVE_FOLDER_NAME (default "DGA Research Reports")
#      that is shared with the service account.
# ============================================================================
DGA_DRIVE_FOLDER_NAME = "DGA Research Reports"

_DRIVE_CACHE: dict[str, Any] = {"svc": None, "folder_id": None, "checked": False}


def _drive_service():
    """Build (and cache) a Google Drive v3 service from the service-account creds."""
    if _DRIVE_CACHE["svc"] is not None:
        return _DRIVE_CACHE["svc"]
    creds_src = _optional_env("GOOGLE_SERVICE_ACCOUNT_JSON")
    if not creds_src:
        return None
    try:
        from google.oauth2.service_account import Credentials
        from googleapiclient.discovery import build
    except ImportError:
        return None
    try:
        scopes = ["https://www.googleapis.com/auth/drive"]
        if creds_src.strip().startswith("{"):
            info = json.loads(creds_src)
            creds = Credentials.from_service_account_info(info, scopes=scopes)
        else:
            creds = Credentials.from_service_account_file(creds_src, scopes=scopes)
        svc = build("drive", "v3", credentials=creds, cache_discovery=False)
    except Exception:
        return None
    _DRIVE_CACHE["svc"] = svc
    return svc


def _drive_folder_id() -> str | None:
    """Resolve (and cache) the target Drive folder ID."""
    if _DRIVE_CACHE["folder_id"]:
        return _DRIVE_CACHE["folder_id"]

    explicit = _optional_env("GOOGLE_DRIVE_FOLDER_ID")
    if explicit:
        _DRIVE_CACHE["folder_id"] = explicit
        return explicit

    svc = _drive_service()
    if svc is None:
        return None

    folder_name = _optional_env("GOOGLE_DRIVE_FOLDER_NAME", DGA_DRIVE_FOLDER_NAME)
    # Escape single quotes in the name for the Drive query DSL.
    safe_name = folder_name.replace("'", "\\'")
    try:
        resp = svc.files().list(
            q=(f"mimeType='application/vnd.google-apps.folder' "
               f"and name='{safe_name}' and trashed=false"),
            fields="files(id, name)",
            pageSize=5,
            supportsAllDrives=True,
            includeItemsFromAllDrives=True,
        ).execute()
    except Exception:
        return None
    files = resp.get("files") or []
    if not files:
        return None
    folder_id = files[0]["id"]
    _DRIVE_CACHE["folder_id"] = folder_id
    return folder_id


def _drive_find_file(svc, folder_id: str, filename: str) -> str | None:
    """Return the Drive file ID of *filename* in *folder_id*, or None."""
    safe_name = filename.replace("'", "\\'")
    try:
        resp = svc.files().list(
            q=f"name='{safe_name}' and '{folder_id}' in parents and trashed=false",
            fields="files(id, name, modifiedTime)",
            pageSize=1,
            supportsAllDrives=True,
            includeItemsFromAllDrives=True,
        ).execute()
    except Exception:
        return None
    files = resp.get("files") or []
    return files[0]["id"] if files else None


# ============================================================================
# Dropbox storage — primary cache backend.
#
# Requires three env vars set in Railway (or .env):
#   DROPBOX_APP_KEY      — from dropbox.com/developers
#   DROPBOX_APP_SECRET   — from dropbox.com/developers
#   DROPBOX_REFRESH_TOKEN — obtained once via the /tmp/dropbox_auth.py helper
#
# Files land at: /DGA Research Reports/<filename>
# (or DROPBOX_FOLDER_PATH if you override it)
# ============================================================================
# With App Folder permission type the SDK root IS the app folder
# (/Apps/DGA Research/ in your Dropbox). We write directly to that root.
DROPBOX_DEFAULT_FOLDER = ""

_DROPBOX_CLIENT_CACHE: dict[str, Any] = {"client": None}


def _dropbox_client():
    """Return a cached Dropbox client, or None if not configured."""
    if _DROPBOX_CLIENT_CACHE["client"] is not None:
        return _DROPBOX_CLIENT_CACHE["client"]
    try:
        import dropbox  # type: ignore
    except ImportError:
        return None
    refresh_token = _optional_env("DROPBOX_REFRESH_TOKEN")
    app_key = _optional_env("DROPBOX_APP_KEY")
    app_secret = _optional_env("DROPBOX_APP_SECRET")
    if not (refresh_token and app_key and app_secret):
        return None
    try:
        dbx = dropbox.Dropbox(
            oauth2_refresh_token=refresh_token,
            app_key=app_key,
            app_secret=app_secret,
        )
        dbx.users_get_current_account()  # validate credentials on first use
        _DROPBOX_CLIENT_CACHE["client"] = dbx
        return dbx
    except Exception:
        return None


def _dropbox_folder() -> str:
    # Empty string = app folder root (correct for "App Folder" permission type).
    # Set DROPBOX_FOLDER_PATH to a subfolder name (e.g. "Reports") if you want
    # a subfolder inside the app folder.
    raw = _optional_env("DROPBOX_FOLDER_PATH", DROPBOX_DEFAULT_FOLDER).strip("/")
    return f"/{raw}" if raw else ""


# Each file type gets its own Dropbox subfolder for tidy browsing.
DROPBOX_PRESENTATIONS_SUBFOLDER = "Presentations"   # .pptx
DROPBOX_REPORTS_SUBFOLDER       = "Reports"         # .docx
DROPBOX_MD_SUBFOLDER            = "MD cached"       # .md  (markdown reports)
DROPBOX_REBALANCED_SUBFOLDER    = "Rebalanced"      # .xlsx (portfolio rebalance files)


def _dropbox_dest_for(file_name: str) -> str:
    """Pick the Dropbox destination path for a given file name.

    Routes each file type to its own subfolder:
      .pptx  → <base>/Presentations/
      .docx  → <base>/Reports/
      .md    → <base>/MD cached/
      .xlsx  → <base>/Rebalanced/
      other  → <base>/   (e.g. .json metadata files)
    """
    base = _dropbox_folder()
    name_lower = file_name.lower()

    if name_lower.endswith(".pptx"):
        sub = DROPBOX_PRESENTATIONS_SUBFOLDER
    elif name_lower.endswith(".docx"):
        sub = DROPBOX_REPORTS_SUBFOLDER
    elif name_lower.endswith(".md"):
        sub = DROPBOX_MD_SUBFOLDER
    elif name_lower.endswith(".xlsx"):
        sub = DROPBOX_REBALANCED_SUBFOLDER
    else:
        return f"{base}/{file_name}" if base else f"/{file_name}"

    return f"{base}/{sub}/{file_name}" if base else f"/{sub}/{file_name}"


def push_to_dropbox(file_paths: list[Path | str]) -> dict:
    """Upload files to the Dropbox 'DGA Research Reports' folder.

    Returns {"ok": True, "uploaded": [...], "folder": "..."} or
    {"ok": False, "skipped"?: bool, "error": "..."}.
    """
    dbx = _dropbox_client()
    if dbx is None:
        return {"ok": False, "skipped": True,
                "error": "Dropbox not configured (need DROPBOX_APP_KEY, "
                         "DROPBOX_APP_SECRET, DROPBOX_REFRESH_TOKEN)"}
    try:
        import dropbox  # type: ignore
    except ImportError:
        return {"ok": False, "error": "dropbox package not installed"}

    folder = _dropbox_folder()
    uploaded: list[str] = []
    errors: list[str] = []
    for fp in file_paths:
        p = Path(fp)
        if not p.exists():
            continue
        # Route .pptx files to the dedicated /Presentations subfolder.
        dest = _dropbox_dest_for(p.name)
        try:
            dbx.files_upload(
                p.read_bytes(),
                dest,
                mode=dropbox.files.WriteMode.overwrite,
                mute=True,
            )
            uploaded.append(p.name)
        except Exception as exc:  # noqa: BLE001
            errors.append(f"{p.name}: {exc}")

    def _sub(name: str) -> str:
        return f"{folder}/{name}" if folder else f"/{name}"

    return {
        "ok": bool(uploaded) or not [Path(f) for f in file_paths if Path(f).exists()],
        "uploaded": uploaded,
        "folder": folder,
        "presentations_folder": _sub(DROPBOX_PRESENTATIONS_SUBFOLDER),
        "reports_folder":       _sub(DROPBOX_REPORTS_SUBFOLDER),
        "md_folder":            _sub(DROPBOX_MD_SUBFOLDER),
        "rebalanced_folder":    _sub(DROPBOX_REBALANCED_SUBFOLDER),
        "errors": errors or None,
    }


def fetch_from_dropbox(ticker: str) -> str | None:
    """Download `{TICKER}_DGA_Report.md` from Dropbox, or None.

    Looks first in the `MD cached/` subfolder (current layout), then falls
    back to the base folder so old reports stored before the reorganisation
    are still found.
    """
    dbx = _dropbox_client()
    if dbx is None:
        return None
    folder = _dropbox_folder()
    filename = f"{ticker}_DGA_Report.md"

    # Primary: new subfolder location
    md_sub = (f"{folder}/{DROPBOX_MD_SUBFOLDER}/{filename}"
              if folder else f"/{DROPBOX_MD_SUBFOLDER}/{filename}")
    # Fallback: legacy base-folder location (pre-reorganisation)
    base_path = f"{folder}/{filename}" if folder else f"/{filename}"

    for path in (md_sub, base_path):
        try:
            _, response = dbx.files_download(path)
            return response.content.decode("utf-8", errors="replace")
        except Exception:
            continue
    return None


def _is_drive_quota_error(exc: Exception) -> bool:
    s = str(exc).lower()
    return "storagequotaexceeded" in s or "do not have storage quota" in s


def _sheets_archive_handle():
    """Return a gspread spreadsheet handle for the markdown archive fallback.

    We store each ticker report as a tab in the existing DGA-portfolio sheet
    because a service account CAN write to cells of a sheet owned-and-shared by
    a real user (no quota cost), but it CANNOT create new binary files in a
    personal Drive folder (the service account has zero storage quota, and
    personal Drive folders don't proxy to the owner's quota — only Google
    Workspace Shared Drives do).
    """
    creds_src = _optional_env("GOOGLE_SERVICE_ACCOUNT_JSON")
    spreadsheet_id = _optional_env("GOOGLE_SHEETS_SPREADSHEET_ID")
    if not creds_src or not spreadsheet_id:
        return None
    try:
        import gspread
        from google.oauth2.service_account import Credentials
    except ImportError:
        return None
    try:
        scopes = [
            "https://www.googleapis.com/auth/spreadsheets",
            "https://www.googleapis.com/auth/drive",
        ]
        if creds_src.strip().startswith("{"):
            info = json.loads(creds_src)
            creds = Credentials.from_service_account_info(info, scopes=scopes)
        else:
            creds = Credentials.from_service_account_file(creds_src, scopes=scopes)
        gc = gspread.authorize(creds)
        return gc.open_by_key(spreadsheet_id)
    except Exception:
        return None


def _sheets_archive_write(ticker: str, report_text: str) -> bool:
    """Write *report_text* into a ticker-named tab of the portfolio sheet."""
    sh = _sheets_archive_handle()
    if sh is None:
        return False
    try:
        import gspread
    except ImportError:
        return False
    tab_title = f"{ticker[:45]} (Report)"
    now = datetime.utcnow().strftime("%Y-%m-%d %H:%M UTC")
    # Google Sheets hard-caps a cell at 50 000 characters; split the report
    # across column-A rows in ~45 000 char chunks so we never lose text.
    chunk_size = 45_000
    chunks = [report_text[i:i + chunk_size]
              for i in range(0, max(len(report_text), 1), chunk_size)] or [""]
    try:
        try:
            ws = sh.worksheet(tab_title)
            ws.clear()
        except gspread.exceptions.WorksheetNotFound:
            ws = sh.add_worksheet(title=tab_title, rows=max(len(chunks) + 5, 20), cols=2)
        ws.update("A1", [[f"Report: {ticker}", f"Updated: {now}"]])
        # Each chunk goes in its own row starting at A2.
        body = [[c] for c in chunks]
        end_row = 1 + len(body)
        ws.update(f"A2:A{end_row}", body)
        return True
    except Exception:
        return False


def _sheets_archive_read(ticker: str) -> str | None:
    """Read a cached report out of the portfolio sheet's ticker tab."""
    sh = _sheets_archive_handle()
    if sh is None:
        return None
    try:
        import gspread
    except ImportError:
        return None
    # Prefer the new, clearer tab name but also fall back to the older
    # tab-named-exactly-after-the-ticker layout that earlier runs produced.
    for tab_title in (f"{ticker[:45]} (Report)", ticker[:50]):
        try:
            ws = sh.worksheet(tab_title)
        except gspread.exceptions.WorksheetNotFound:
            continue
        except Exception:
            continue
        try:
            values = ws.col_values(1)  # column A only
        except Exception:
            continue
        # Row 1 is the header ("Report: TKR"); content is row 2 onward.
        body = "".join(values[1:]).strip()
        if body:
            return body
    return None


def push_to_google_drive(
    file_paths: list[Path | str],
    *,
    folder_name: str = DGA_DRIVE_FOLDER_NAME,
) -> dict:
    """Persist report files so they survive Railway redeploys.

    Primary path: real Drive upload into the shared DGA folder. This path
    ONLY works when:
      - the folder lives on a Google Workspace Shared Drive, OR
      - GOOGLE_SERVICE_ACCOUNT_JSON holds OAuth user-delegated creds.
    With a vanilla personal-Gmail + service-account setup, Google returns
    403 storageQuotaExceeded on every upload.

    Fallback path: write the .md reports (and DGA-portfolio.xlsx row metadata)
    into ticker-named tabs of the portfolio Google Sheet. A service account
    CAN mutate cells in a sheet that's been shared with it, so this works
    everywhere. Sheet URL is returned as `sheets_url`.

    Returns a dict with {ok, drive_uploaded, sheets_archived, folder_url?,
    sheets_url?, errors?}.
    """
    result: dict[str, Any] = {
        "ok": False,
        "drive_uploaded": [],
        "sheets_archived": [],
        "errors": [],
    }

    md_paths: list[Path] = []
    non_md_paths: list[Path] = []
    for fp in file_paths:
        p = Path(fp)
        if not p.exists():
            continue
        (md_paths if p.suffix.lower() == ".md" else non_md_paths).append(p)

    # --- Preferred: Dropbox upload (works on personal accounts, no quota issues) ---
    dbx_result = push_to_dropbox(md_paths + non_md_paths)
    if dbx_result.get("ok") and dbx_result.get("uploaded"):
        result["ok"] = True
        result["dropbox_uploaded"] = dbx_result["uploaded"]
        result["dropbox_folder"] = dbx_result.get("folder")
        if not result["errors"]:
            result.pop("errors")
        return result
    if not dbx_result.get("skipped"):
        # Dropbox was configured but failed — surface errors.
        for e in (dbx_result.get("errors") or []):
            result["errors"].append(f"dropbox: {e}")

    # --- Secondary: Drive upload ---
    svc = _drive_service()
    folder_id = _drive_folder_id()
    quota_hit = False
    if svc is not None and folder_id:
        try:
            from googleapiclient.http import MediaFileUpload
        except ImportError:
            svc = None  # fall through to sheets-only
        if svc is not None:
            for p in md_paths + non_md_paths:
                try:
                    mime, _ = mimetypes.guess_type(str(p))
                    media = MediaFileUpload(
                        str(p), mimetype=mime or "application/octet-stream",
                        resumable=False,
                    )
                    existing = _drive_find_file(svc, folder_id, p.name)
                    if existing:
                        svc.files().update(
                            fileId=existing,
                            media_body=media,
                            supportsAllDrives=True,
                        ).execute()
                    else:
                        svc.files().create(
                            body={"name": p.name, "parents": [folder_id]},
                            media_body=media,
                            fields="id",
                            supportsAllDrives=True,
                        ).execute()
                    result["drive_uploaded"].append(p.name)
                except Exception as exc:  # noqa: BLE001
                    if _is_drive_quota_error(exc):
                        quota_hit = True
                        break  # don't bother retrying the rest — quota won't change mid-run
                    result["errors"].append(f"drive:{p.name}: {exc}")
        result["folder_id"] = folder_id
        result["folder_url"] = f"https://drive.google.com/drive/folders/{folder_id}"

    # --- Fallback: Sheets-tab archive for the .md reports ---
    # We do this when Drive upload hit quota, OR when Drive isn't configured.
    # The xlsx / docx / pptx files are skipped here — they can't be reconstructed
    # from sheet cells and live in Sheets as dedicated uploads would need quota too.
    if quota_hit or svc is None or not folder_id:
        sh = _sheets_archive_handle()
        if sh is not None:
            for p in md_paths:
                ticker = p.stem.replace("_DGA_Report", "").replace("_DGA_report", "")
                ok = _sheets_archive_write(ticker, p.read_text(encoding="utf-8",
                                                              errors="replace"))
                if ok:
                    result["sheets_archived"].append(ticker)
                else:
                    result["errors"].append(f"sheets:{p.name}: write failed")
            sid = _optional_env("GOOGLE_SHEETS_SPREADSHEET_ID")
            if sid:
                result["sheets_url"] = f"https://docs.google.com/spreadsheets/d/{sid}"

    if quota_hit:
        result["errors"].append(
            "Drive folder upload hit the service-account 0-quota wall. "
            "Using Sheets-tab archive as fallback. To enable real Drive "
            "folder uploads, move the folder to a Google Workspace Shared "
            "Drive or switch to OAuth user-delegated credentials."
        )

    result["ok"] = bool(result["drive_uploaded"] or result["sheets_archived"])
    # Drop the errors key if there were none — keeps the status response tidy.
    if not result["errors"]:
        result.pop("errors")
    return result


def fetch_report_from_drive(ticker: str) -> str | None:
    """Try to load a cached `{TICKER}_DGA_Report.md`.

    Checks Dropbox first (preferred, works on personal accounts), then the
    shared Drive folder (requires Workspace), then the Sheets-tab archive.
    Returns markdown text if found, else None.
    """
    # --- Dropbox first (preferred) ---
    dbx_text = fetch_from_dropbox(ticker)
    if dbx_text:
        return dbx_text

    # --- Drive second ---
    svc = _drive_service()
    folder_id = _drive_folder_id()
    if svc is not None and folder_id:
        filename = f"{ticker}_DGA_Report.md"
        file_id = _drive_find_file(svc, folder_id, filename)
        if file_id:
            try:
                from googleapiclient.http import MediaIoBaseDownload
                import io
                buf = io.BytesIO()
                request = svc.files().get_media(fileId=file_id, supportsAllDrives=True)
                downloader = MediaIoBaseDownload(buf, request)
                done = False
                while not done:
                    _, done = downloader.next_chunk()
                return buf.getvalue().decode("utf-8", errors="replace")
            except Exception:
                pass  # fall through to sheets

    # --- Sheets fallback ---
    return _sheets_archive_read(ticker)


# ============================================================================
# High-level orchestration (used by CLI *and* API)
# ============================================================================
def run_portfolio_rebalance(
    portfolio_records: list[dict],
    *,
    primary_strategy: str = "current",
    generate_gamma: bool = False,
    reuse_existing: bool = True,
    output_path: Path | str | None = None,
    system_prompt: str | None = None,
    verbose: bool = False,
    on_progress=None,
) -> dict:
    """Analyze every ticker in *portfolio_records* and produce DGA-portfolio.xlsx.

    Returns a dict with:
      - ok: bool
      - tickers_ok / tickers_failed
      - strategy_results: { strategy_key: { weights, held, strategy, ... } }
      - xlsx_path: str path to the generated xlsx
      - primary_strategy: key of the primary strategy shown first
      - summary: short roll-up for API responses

    Progress reporting:
        ``on_progress`` is an optional callable
        ``(step, pct, label, extra)`` invoked at each pipeline boundary.
        ``step`` is one of:
            ``"queued" | "analyzing" | "rollup" | "weights" | "writing"
             | "email" | "sheets" | "drive" | "done"``
        ``extra`` is a dict with per-step context — most importantly
        ``{"ticker_index": i, "ticker_total": n, "ticker": "AAPL",
        "ok": [list of OK so far], "failed": [list of failed so far]}``
        during ``"analyzing"``. The frontend uses this to render a
        per-ticker counter ("3 / 12 — analyzing AAPL") instead of a
        single opaque spinner.
    """
    if primary_strategy not in STRATEGIES:
        raise ValueError(f"Unknown strategy: {primary_strategy}")

    if system_prompt is None:
        system_prompt = load_system_prompt()

    tickers = portfolio_tickers(portfolio_records)
    n_tickers = len(tickers)

    def _emit(step, pct, label, extra=None):
        if on_progress is None:
            return
        try:
            on_progress(step, max(0.0, min(1.0, float(pct))), label, extra or {})
        except Exception:  # noqa: BLE001
            pass

    # Per-ticker analyses dominate the runtime. We allocate 80% of the
    # progress bar to this phase, then split the remaining 20% across
    # roll-up, weight computation, xlsx write, and side-effects.
    ANALYZE_BUDGET = 0.80

    _emit("analyzing", 0.0, f"Analyzing {n_tickers} tickers…",
          {"ticker_total": n_tickers, "ticker_index": 0,
           "ok": [], "failed": []})

    ticker_results: list[dict] = []
    ok_so_far: list[str] = []
    failed_so_far: list[str] = []
    for i, ticker in enumerate(tickers):
        # Announce the ticker at the START of its analysis so the UI shows
        # the right ticker label *while* it's running, not after.
        _emit("analyzing",
              ANALYZE_BUDGET * (i / max(1, n_tickers)),
              f"Analyzing {ticker} ({i + 1}/{n_tickers})",
              {"ticker_total": n_tickers, "ticker_index": i + 1,
               "ticker": ticker,
               "ok": list(ok_so_far), "failed": list(failed_so_far)})
        try:
            r = analyze_ticker(
                ticker,
                system_prompt=system_prompt,
                generate_gamma=generate_gamma,
                verbose=verbose,
                reuse_existing=reuse_existing,
            )
        except Exception as exc:  # noqa: BLE001
            r = {"ticker": ticker, "ok": False, "error": str(exc)}
        ticker_results.append(r)
        if r.get("ok"):
            ok_so_far.append(ticker)
        else:
            failed_so_far.append(ticker)

    ok_results = [r for r in ticker_results if r.get("ok")]
    failed = [r for r in ticker_results if not r.get("ok")]

    _emit("rollup", ANALYZE_BUDGET + 0.04,
          f"Roll-up commentary ({len(ok_results)} tickers)",
          {"ticker_total": n_tickers, "ticker_index": n_tickers,
           "ok": [r["ticker"] for r in ok_results],
           "failed": [r.get("ticker", "?") for r in failed]})

    # Run the Grok roll-up (best-effort — gives us ranked rows if reachable).
    ranked_rows = None
    roll = {"ok": False}
    if len(ok_results) > 1:
        try:
            roll = run_portfolio_summary(ok_results, generate_gamma=generate_gamma)
            if roll.get("ok"):
                ranked_rows = roll.get("ranked_rows")
        except Exception:  # noqa: BLE001
            roll = {"ok": False}

    _emit("weights", ANALYZE_BUDGET + 0.10,
          "Computing portfolio weights (3 strategies)")

    # Always compute ALL three strategies so the xlsx shows comparisons.
    strategy_results: dict[str, dict] = {}
    for skey in STRATEGIES:
        strategy_results[skey] = compute_rebalance(
            ok_results, strategy=skey, ranked_rows=ranked_rows,
        )

    # Choose where to write the xlsx.
    if output_path is None:
        output_path = Path.cwd() / DGA_PORTFOLIO_FILENAME
    else:
        output_path = Path(output_path)

    _emit("writing", ANALYZE_BUDGET + 0.13,
          "Writing DGA-portfolio.xlsx")

    write_dga_portfolio_xlsx(
        output_path=output_path,
        input_records=portfolio_records,
        primary_strategy=primary_strategy,
        strategy_results=strategy_results,
    )

    # Build a lightweight JSON-safe summary for the API.
    def _slim(res: dict) -> dict:
        w = res.get("weights", {})
        return {
            "strategy": res.get("strategy"),
            "label": res.get("label"),
            "held": sum(1 for v in w.values() if v > 0),
            "weights": {k: round(v, 4) for k, v in w.items() if v > 0},
        }

    run_ts = datetime.utcnow().strftime("%Y-%m-%d %H:%M UTC")

    _emit("email", ANALYZE_BUDGET + 0.16, "Sending portfolio email")

    # Email portfolio results — only fires for multi-ticker (portfolio) runs.
    email_status: dict = {"ok": False, "skipped": True}
    if len(ok_results) > 1:
        try:
            email_msg = build_portfolio_email(
                tickers_ok=[r["ticker"] for r in ok_results],
                tickers_failed=[r["ticker"] for r in failed],
                summary_markdown=(roll.get("summary_md") if roll.get("ok") else "") or "",
                ranked_rows=ranked_rows,
                strategy_results=strategy_results,
                output_xlsx=output_path,
                portfolio_docx=Path(roll["docx"]) if roll.get("ok") and roll.get("docx") else None,
                gamma_url=roll.get("gamma_url") if roll.get("ok") else None,
            )
            email_status = send_portfolio_email(email_msg)
        except Exception as exc:  # noqa: BLE001
            email_status = {"ok": False, "error": str(exc)}

    _emit("sheets", ANALYZE_BUDGET + 0.18, "Pushing to Google Sheets")

    # Google Sheets push — only fires for multi-ticker runs.
    gsheets_status: dict = {"ok": False, "skipped": True}
    if len(ok_results) > 1:
        try:
            gsheets_status = push_to_google_sheets(
                input_records=portfolio_records,
                primary_strategy=primary_strategy,
                strategy_results=strategy_results,
                run_timestamp=run_ts,
            )
        except Exception as exc:  # noqa: BLE001
            gsheets_status = {"ok": False, "error": str(exc)}

    _emit("drive", ANALYZE_BUDGET + 0.19, "Uploading to Dropbox / Drive")

    # Google Drive upload — portfolio xlsx + all per-ticker reports + the
    # Grok portfolio roll-up (Portfolio_Summary.md/.docx) so the Research
    # page's "Last Portfolio Summary" card can hydrate from Dropbox after
    # a Railway redeploy.
    gdrive_status: dict = {"ok": False, "skipped": True}
    try:
        drive_files: list[Path] = [output_path]
        for r in ok_results:
            for key in ("docx", "md"):
                p = r.get(key)
                if p and Path(p).exists():
                    drive_files.append(Path(p))
        # Add the portfolio roll-up files if we produced them.
        if roll.get("ok"):
            for key in ("md", "docx"):
                p = roll.get(key)
                if p and Path(p).exists():
                    drive_files.append(Path(p))
        gdrive_status = push_to_google_drive(drive_files)
    except Exception as exc:  # noqa: BLE001
        gdrive_status = {"ok": False, "error": str(exc)}

    _emit("done", 1.0,
          f"Done — {len(ok_results)}/{n_tickers} tickers analyzed",
          {"ticker_total": n_tickers, "ticker_index": n_tickers,
           "ok": [r["ticker"] for r in ok_results],
           "failed": [r.get("ticker", "?") for r in failed]})

    return {
        "ok": bool(ok_results),
        "primary_strategy": primary_strategy,
        "tickers_ok": [r["ticker"] for r in ok_results],
        "tickers_failed": [{"ticker": r["ticker"], "error": r.get("error")}
                            for r in failed],
        "xlsx_path": str(output_path),
        "portfolio_roll_up_ok": bool(roll.get("ok")),
        "strategies": {k: _slim(v) for k, v in strategy_results.items()},
        "email": email_status,
        "gsheets": gsheets_status,
        "gdrive": gdrive_status,
    }


# ============================================================================
# CLI
# ============================================================================
def _prompt_yes_no(prompt: str, default: bool = False) -> bool:
    default_str = "Y/n" if default else "y/N"
    resp = input(f"{prompt} [{default_str}]: ").strip().lower()
    if not resp:
        return default
    return resp in ("y", "yes")


def main() -> int:
    ap = argparse.ArgumentParser(description="DGA Capital Research — Claude Edition")
    ap.add_argument("ticker", nargs="?", help="Single ticker (e.g. AAPL)")
    ap.add_argument("--portfolio", help="Path to a CSV or XLSX portfolio file "
                                        "(columns: Ticker | Weight | Optimized)")
    ap.add_argument("--scan", nargs="*", metavar="TICKER",
                    help="Run live news scan. Pass tickers (e.g. --scan AAPL MSFT) or "
                         "omit to use the saved watchlist (stocks/watchlist.json). "
                         "Results are printed and saved to stocks/scan_results.json.")
    ap.add_argument("--gamma", action="store_true", help="Force Gamma deck generation")
    ap.add_argument("--no-gamma", action="store_true", help="Skip Gamma deck generation")
    ap.add_argument("--strategy", choices=list(STRATEGIES.keys()), default="current",
                    help="Primary rebalance strategy (default: current)")
    ap.add_argument("--reuse", action="store_true",
                    help="Reuse cached markdown reports from /stocks where present")
    ap.add_argument("--out",
                    help=f"Output xlsx path (defaults to ./{DGA_PORTFOLIO_FILENAME})")
    args = ap.parse_args()

    # ── SCAN mode ─────────────────────────────────────────────────────────────
    if args.scan is not None:
        print("╔══════════════════════════════════════════════════╗")
        print("║  DGA MARKET SCAN — Live News Intelligence        ║")
        print("╚══════════════════════════════════════════════════╝")
        tickers_to_scan: list[str] = [t.strip().upper() for t in args.scan if t.strip()]
        if not tickers_to_scan:
            tickers_to_scan = load_watchlist()
        if not tickers_to_scan:
            print("❌ No tickers to scan. Either pass them after --scan or add them "
                  "to the watchlist (stocks/watchlist.json).")
            return 2
        print(f"\n📡 Scanning {len(tickers_to_scan)} ticker(s): {', '.join(tickers_to_scan)}\n")
        scan = run_portfolio_scan(tickers_to_scan, verbose=True)
        print("\n" + "─" * 60)
        for ticker, r in scan["results"].items():
            print(f"\n{'█' * 50}")
            if r.get("ok") and r.get("markdown"):
                print(r["markdown"])
            else:
                print(f"⚠️  {ticker}: {r.get('error', 'No data')}")
        print("\n" + "─" * 60)
        print(f"\n✅ Scan complete. Results saved → {SCAN_RESULTS_FILE}")
        return 0

    print("╔══════════════════════════════════════════════════╗")
    print("║  DGA CAPITAL RESEARCH ANALYST — Claude Edition  ║")
    print("╚══════════════════════════════════════════════════╝")

    # Resolve input: CLI takes precedence; else prompt.
    portfolio_records: list[dict] = []
    single_ticker_mode = False
    if args.portfolio:
        portfolio_records = load_portfolio_file(args.portfolio)
    elif args.ticker:
        portfolio_records = [{"ticker": args.ticker.strip().upper(), "weight": None}]
        single_ticker_mode = True
    else:
        print("\nChoose input mode:")
        print("  1) Single ticker")
        print("  2) Portfolio CSV or XLSX (Ticker | Weight | Optimized)")
        mode = input("Select 1 or 2 (or paste a ticker directly): ").strip()
        if mode == "1":
            t = input("Enter ticker (e.g. AAPL): ").strip().upper()
            if t:
                portfolio_records = [{"ticker": t, "weight": None}]
                single_ticker_mode = True
        elif mode == "2":
            pf = input("Path to portfolio file (.csv or .xlsx): ").strip()
            try:
                portfolio_records = load_portfolio_file(pf)
            except Exception as exc:  # noqa: BLE001
                print(f"❌ Could not load portfolio: {exc}")
                return 2
        else:
            if mode:
                portfolio_records = [{"ticker": mode.upper(), "weight": None}]
                single_ticker_mode = True

    if not portfolio_records:
        print("❌ No tickers to analyze.")
        return 2

    tickers = portfolio_tickers(portfolio_records)

    # Gamma decision
    if args.gamma:
        generate_gamma = True
    elif args.no_gamma:
        generate_gamma = False
    else:
        generate_gamma = _prompt_yes_no(
            "Generate Gamma.app presentations as well? (uses Gamma credits)",
            default=False,
        )

    system_prompt = load_system_prompt()

    print(f"\n📋 Tickers to analyze ({len(tickers)}): {', '.join(tickers)}")
    print(f"📁 Output folder: {STOCKS_FOLDER}")
    print(f"🎨 Gamma generation: {'ON' if generate_gamma else 'OFF'}")
    if not single_ticker_mode:
        print(f"⚖️  Primary rebalance strategy: {STRATEGIES[args.strategy]['label']}")
        print(f"♻️  Reuse cached reports: {'ON' if args.reuse else 'OFF'}")

    results: list[dict] = []
    for ticker in tickers:
        try:
            res = analyze_ticker(
                ticker,
                system_prompt=system_prompt,
                generate_gamma=generate_gamma,
                verbose=False,
                reuse_existing=args.reuse,
            )
        except Exception as exc:  # noqa: BLE001
            print(f"❌ {ticker} failed: {exc}")
            res = {"ticker": ticker, "ok": False, "error": str(exc)}
        results.append(res)

    ok = [r for r in results if r.get("ok")]
    fail = [r for r in results if not r.get("ok")]

    print("\n==============================================")
    print(f"  SUMMARY: {len(ok)} succeeded, {len(fail)} failed")
    print("==============================================")
    for r in ok:
        tag = " [cached]" if r.get("cached") else ""
        print(f"  ✅ {r['ticker']}{tag}  →  {r.get('docx','')}")
        if r.get("gamma_url"):
            print(f"       📽️   {r['gamma_url']}")
    for r in fail:
        print(f"  ❌ {r['ticker']}  {r.get('error','')}")

    # Portfolio flow (only if we have more than 1 ticker OR a portfolio file).
    if len(ok) > 1 and not single_ticker_mode:
        print("\n==============================================")
        print("  PORTFOLIO ROLL-UP + REBALANCE")
        print("==============================================")
        roll = run_portfolio_summary(ok, generate_gamma=generate_gamma)
        ranked_rows = None
        if roll.get("ok"):
            print(f"  ✅ Portfolio Word: {roll['docx']}")
            if roll.get("gamma_url"):
                print(f"  📽️   {roll['gamma_url']}")
            ranked_rows = roll.get("ranked_rows")
        else:
            print(f"  ⚠️  Portfolio roll-up failed: {roll.get('error')}")

        # Compute all three strategies, using the Grok roll-up's ranked table
        # (rating + upside) to enrich the scoring signal when available.
        strategy_results: dict[str, dict] = {}
        for skey in STRATEGIES:
            strategy_results[skey] = compute_rebalance(
                ok, strategy=skey, ranked_rows=ranked_rows,
            )
            held = sum(1 for w in strategy_results[skey]["weights"].values() if w > 0)
            print(f"  📊 {STRATEGIES[skey]['label']}: {held} positions")

        # Write the xlsx.
        out_path = Path(args.out) if args.out else (SCRIPT_DIR.parent / DGA_PORTFOLIO_FILENAME)
        # If user didn't override --out, default to the working directory the
        # script was launched from (so it lands next to the portfolio file).
        if not args.out:
            out_path = Path.cwd() / DGA_PORTFOLIO_FILENAME

        write_dga_portfolio_xlsx(
            output_path=out_path,
            input_records=portfolio_records,
            primary_strategy=args.strategy,
            strategy_results=strategy_results,
        )
        print(f"  💾 Optimized portfolio: {out_path}")

        # Email the results — portfolio runs only (single-ticker analyses
        # never email; that branch never reaches here).
        try:
            email_msg = build_portfolio_email(
                tickers_ok=[r["ticker"] for r in ok],
                tickers_failed=[r["ticker"] for r in fail],
                summary_markdown=(roll.get("summary_md") if roll.get("ok") else "") or "",
                ranked_rows=ranked_rows,
                strategy_results=strategy_results,
                output_xlsx=out_path,
                portfolio_docx=Path(roll["docx"]) if roll.get("ok") and roll.get("docx") else None,
                gamma_url=roll.get("gamma_url") if roll.get("ok") else None,
            )
            email_res = send_portfolio_email(email_msg)
            if email_res.get("ok"):
                print(f"  📧 Emailed portfolio results to {email_res['sent_to']}")
            else:
                print(f"  📧 Email pending — {email_res.get('error', 'unknown error')}")
        except Exception as exc:
            print(f"  ⚠️  Could not send portfolio email: {exc}")

    return 0 if ok else 1


if __name__ == "__main__":
    try:
        sys.exit(main())
    except RuntimeError as exc:
        # _require_env raises RuntimeError for missing env vars so background
        # threads can catch them. Re-surface as a clean CLI exit here.
        print(f"❌ {exc}", file=sys.stderr)
        sys.exit(1)
